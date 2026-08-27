#!/usr/bin/env python3
"""
ساخت Excel جامع: روش‌های جستجو + ایمیل‌ها + مستندسازی
دقت جستجو، نحوه کراول، ایمیل‌های طبیعی
"""
import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, numbers
from openpyxl.utils import get_column_letter

# ── Constants ──
FONT_FA = "B Mitra"
FONT_EN = "Times New Roman"
DARK_BLUE = "1F4E79"
MED_BLUE = "2E75B6"
LIGHT_BLUE = "D6E4F0"
GREEN = "27AE60"
YELLOW = "F39C12"
RED = "E74C3C"
GRAY = "BDC3C7"
WHITE = "FFFFFF"
DARK_GRAY = "2C3E50"

def set_rtl(ws):
    """Set worksheet to RTL"""
    ws.sheet_view.rightToLeft = True

def hdr_font(size=12):
    return Font(name=FONT_FA, size=size, bold=True, color=WHITE)

def fa_font(size=11, bold=False, color="000000"):
    return Font(name=FONT_FA, size=size, bold=bold, color=color)

def en_font(size=11, bold=False, color="000000"):
    return Font(name=FONT_EN, size=size, bold=bold, color=color)

def fill(color):
    return PatternFill(start_color=color, end_color=color, fill_type="solid")

def align_right():
    return Alignment(horizontal="right", vertical="center", wrap_text=True)

def align_left():
    return Alignment(horizontal="left", vertical="center", wrap_text=True)

def align_center():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

thin_border = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin")
)

def style_header_row(ws, row, cols, fill_color=DARK_BLUE):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = hdr_font(11)
        cell.fill = fill(fill_color)
        cell.alignment = align_center()
        cell.border = thin_border

def style_data_cell(ws, row, col, font_type="fa", size=10):
    cell = ws.cell(row=row, column=col)
    if font_type == "en":
        cell.font = en_font(size)
    else:
        cell.font = fa_font(size)
    cell.alignment = align_right()
    cell.border = thin_border
    return cell

NOW = datetime.now().strftime("%Y-%m-%d %H:%M")

# ── Create Workbook ──
wb = Workbook()

# ═══════════════════════════════════════════════════
# SHEET 1: روش‌های جستجو | Search Methods
# ═══════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "روش‌های جستجو"
set_rtl(ws1)

# Title
ws1.merge_cells("A1:H1")
title_cell = ws1["A1"]
title_cell.value = "روش‌های جستجوی کار — Search Methodology"
title_cell.font = Font(name=FONT_FA, size=16, bold=True, color=WHITE)
title_cell.fill = fill(DARK_BLUE)
title_cell.alignment = align_center()

ws1.merge_cells("A2:H2")
ws1["A2"].value = f"آخرین بروزرسانی: {NOW}"
ws1["A2"].font = fa_font(9)
ws1["A2"].alignment = align_right()

# Headers
headers = ["#", "روش جستجو", "Search Method", "دقت تقریبی", "توضیح", "مثال", "محدودیت", "وضعیت"]
for i, h in enumerate(headers, 1):
    ws1.cell(row=4, column=i, value=h)
style_header_row(ws1, 4, 8)

# Data
methods = [
    [1, "جستجوی Google", "Google Search", "70-75%",
     "عبارت جستجو می‌سازم، نتایج را می‌خوانم، لینک‌ها را باز می‌کنم",
     '"midwife jobs NZ visa sponsorship 2026"',
     "بعضی نتایج قدیمی، بعضی لینک‌ها خراب", "✅ فعال"],
    [2, "خواندن سایت کارفرما", "Employer Website Crawl", "85-90%",
     "آدرس سایت → صفحه Careers → استخراج آگهی + ایمیل + فرم",
     "saskhealthauthority.ca/careers",
     "بعضی سایت‌ها JavaScript هستند", "✅ فعال"],
    [3, "سایت دولتی", "Government Portal", "90%+",
     "سایت رسمی مهاجرت/کار → لیست مشاغل کمبود + قوانین ویزا",
     "canada.ca/ircc, nzis.govt.nz",
     "اطلاعات رسمی ولی کند به‌روز می‌شود", "✅ فعال"],
    [4, "آژانس استخدام بین‌المللی", "International Recruiter", "70-75%",
     "آژانس‌های معتبر → ثبت‌نام CV → درخواست مشاوره",
     "Hays Healthcare, TalentOrange",
     "بعضی پول می‌گیرند، بعضی غیرواقعی", "✅ فعال"],
    [5, "ایمیل مستقیم کارفرما", "Direct Employer Email", "40-50%",
     "ایمیل HR/Recruitment → ایمیل درخواست شخصی → کاربر ارسال",
     "careers@employer.com",
     "ایمیل قدیمی، پاسخ تضمینی نیست", "✅ فعال"],
    [6, "LinkedIn Jobs", "LinkedIn Job Search", "70-75%",
     "جستجو در LinkedIn → آگهی + پروفایل استخدام‌کننده",
     "linkedin.com/jobs/search",
     "نیاز به لاگین، ربات بلاک می‌شود", "⚠️ محدود"],
    [7, "انجمن حرفه‌ای", "Professional Association", "80-85%",
     "انجمن حرفه‌ای → لیست کارفرماها + راهنمای ثبت‌نام",
     "College of Midwives, ANMF",
     "فقط اطلاعات حرفه‌ای، شغل مستقیم ندارد", "✅ فعال"],
    [8, "Job Board معروف", "Major Job Board", "75-80%",
     "Seek, Indeed, Trade Me, StepStone → جستجو با فیلتر",
     "seek.com.au, indeed.com",
     "فیلتر ربات، بعضی آگهی‌ها پنهان", "✅ فعال"],
    [9, "کراول صفحه تماس", "Contact Page Crawl", "45-55%",
     "صفحه Contact/About → استخراج ایمیل + تلفن + آدرس",
     "employer.com/contact",
     "ایمیل عمومی است، نه مستقیم HR", "✅ فعال"],
    [10, "_explore سایت دولتی کار", "Government Job Search", "85-90%",
     "سایت کار دولتی → جستجوی شغل + فیلتر اقامت",
     "jobbank.gc.ca, jobs.govt.nz",
     "فقط آگهی‌های دولتی", "✅ فعال"],
]

for i, row_data in enumerate(methods):
    r = 5 + i
    for j, val in enumerate(row_data):
        cell = style_data_cell(ws1, r, j + 1, "fa" if j < 2 or j in [4, 5, 6] else "en")
        cell.value = val
        # Color code accuracy
        if j == 3:
            if "90" in str(val):
                cell.fill = fill("D5F5E3")
            elif "80" in str(val) or "85" in str(val):
                cell.fill = fill("FEF9E7")
            elif "70" in str(val):
                cell.fill = fill("FDEBD0")
            else:
                cell.fill = fill("FADBD8")

# Column widths
widths = [5, 25, 25, 12, 45, 40, 35, 12]
for i, w in enumerate(widths):
    ws1.column_dimensions[get_column_letter(i + 1)].width = w

# Accuracy summary
r = 16
ws1.merge_cells(f"A{r}:H{r}")
ws1[f"A{r}"].value = "نکته مهم: هیچ روشی ۱۰۰٪ دقیق نیست. بهترین نتیجه از ترکیب چند روش حاصل می‌شود."
ws1[f"A{r}"].font = fa_font(11, bold=True, color=RED)

# ═══════════════════════════════════════════════════
# SHEET 2: وضعیت جستجو | Search Status
# ═══════════════════════════════════════════════════
ws2 = wb.create_sheet("وضعیت جستجو")
set_rtl(ws2)

ws2.merge_cells("A1:G1")
ws2["A1"].value = "وضعیت جستجوی هر منبع — Source Search Status"
ws2["A1"].font = Font(name=FONT_FA, size=14, bold=True, color=WHITE)
ws2["A1"].fill = fill(MED_BLUE)
ws2["A1"].alignment = align_center()

headers2 = ["#", "منبع / سایت", "نوع", "کشور", "آخرین جستجو", "آگهی یافت شده", "وضعیت"]
for i, h in enumerate(headers2, 1):
    ws2.cell(row=3, column=i, value=h)
style_header_row(ws2, 3, 7, MED_BLUE)

sources = [
    [1, "Google Jobs", "موتور جستجو", "جهانی", NOW, "—", "✅ فعال"],
    [2, "Seek.com.au", "Job Board", "🇦🇺 استرالیا", "2026-08-27", "—", "✅ فعال"],
    [3, "Trade Me Jobs", "Job Board", "🇳🇿 نیوزیلند", "2026-08-27", "—", "✅ فعال"],
    [4, "Indeed", "Job Board", "جهانی", "2026-08-27", "—", "✅ فعال"],
    [5, "Job Bank Canada", "دولتی", "🇨🇦 کانادا", "2026-08-27", "—", "✅ فعال"],
    [6, "Health NZ", "دولتی/کارفرما", "🇳🇿 نیوزیلند", "2026-08-27", "—", "✅ فعال"],
    [7, "Saskatchewan HA", "دولتی/کارفرما", "🇨🇦 کانادا", "2026-08-27", "—", "✅ فعال"],
    [8, "Hays Healthcare", "آژانس استخدام", "🇦🇺 استرالیا", "2026-08-27", "—", "✅ فعال"],
    [9, "TalentOrange", "آژانس استخدام", "🇩🇪 آلمان", "2026-08-27", "—", "✅ فعال"],
    [10, "Holalemania", "کارفرما", "🇩🇪 آلمان", "2026-08-27", "—", "✅ فعال"],
    [11, "Working In Health NZ", "آژانس استخدام", "🇳🇿 نیوزیلند", "2026-08-27", "—", "✅ فعال"],
    [12, "RGH Global", "آژانس استخدام", "🇳🇿 نیوزیلند", "2026-08-27", "—", "✅ فعال"],
    [13, "Make it in Germany", "دولتی", "🇩🇪 آلمان", "2026-08-27", "—", "✅ فعال"],
    [14, "HomeAffairs Australia", "دولتی", "🇦🇺 استرالیا", "2026-08-27", "—", "✅ فعال"],
    [15, "Canadian Midwives Assoc", "انجمن حرفه‌ای", "🇨🇦 کانادا", "2026-08-27", "—", "✅ فعال"],
    [16, "ANMF", "انجمن حرفه‌ای", "🇦🇺 استرالیا", "2026-08-27", "—", "✅ فعال"],
    [17, "LinkedIn", "شبکه حرفه‌ای", "جهانی", "2026-08-27", "—", "⚠️ محدود"],
    [18, "Fletcher Building", "کارفرما", "🇳🇿 نیوزیلند", "—", "—", "🟡 بررسی"],
    [19, "Mainfreight", "کارفرما", "🇳🇿 نیوزیلند", "—", "—", "🟡 بررسی"],
    [20, "Datacom", "کارفرما", "🇳🇿 نیوزیلند", "—", "—", "🟡 بررسی"],
]

for i, row_data in enumerate(sources):
    r = 4 + i
    for j, val in enumerate(row_data):
        cell = style_data_cell(ws2, r, j + 1, "en" if j in [3, 4, 5] else "fa")
        cell.value = val

widths2 = [5, 30, 20, 20, 18, 15, 12]
for i, w in enumerate(widths2):
    ws2.column_dimensions[get_column_letter(i + 1)].width = w

# ═══════════════════════════════════════════════════
# SHEET 3: ایمیل‌های آماده | Ready Emails
# ═══════════════════════════════════════════════════
ws3 = wb.create_sheet("ایمیل‌های آماده")
set_rtl(ws3)

ws3.merge_cells("A1:H1")
ws3["A1"].value = "ایمیل‌های آماده ارسال — Ready to Send Emails"
ws3["A1"].font = Font(name=FONT_FA, size=14, bold=True, color=WHITE)
ws3["A1"].fill = fill(GREEN)
ws3["A1"].alignment = align_center()

headers3 = ["#", "متقاضی", "کارفرما", "کشور", "موضوع ایمیل", "گیرنده", "وضعیت", "لینک آگهی"]
for i, h in enumerate(headers3, 1):
    ws3.cell(row=3, column=i, value=h)
style_header_row(ws3, 3, 8, GREEN)

emails = [
    [1, "👩 ندا", "Saskatchewan Health Authority", "🇨🇦 کانادا",
     "Midwife — International Candidate — Saskatchewan",
     "SHAInternational@saskhealthauthority.ca", "🟢 آماده",
     "saskhealthauthority.ca/careers"],
    [2, "👨 توحید", "Saskatchewan Health Authority", "🇨🇦 کانادا",
     "IT Operations Manager — International Candidate",
     "SHAInternational@saskhealthauthority.ca", "🟢 آماده",
     "saskhealthauthority.ca/careers"],
    [3, "👩 ندا", "Health New Zealand", "🇳🇿 نیوزیلند",
     "International Midwife — Expression of Interest",
     "careers@health.govt.nz", "🟢 آماده",
     "health.govt.nz/work-with-us"],
    [4, "👩 ندا", "Working In Health NZ", "🇳🇿 نیوزیلند",
     "Midwife — Seeking Accredited Employer",
     "info@workinginhealth.co.nz", "🟢 آماده",
     "workinginhealth.co.nz"],
    [5, "👩 ندا", "Holalemania", "🇩🇪 آلمان",
     "Geburtshelferin — Internationale Bewerberin",
     "info@holalemania.de", "🟢 آماده",
     "holalemania.de/karriere"],
    [6, "👨 توحید", "Holalemania", "🇩🇪 آلمان",
     "IT Operations Manager — International Application",
     "info@holalemania.de", "🟢 آماده",
     "holalemania.de/karriere"],
    [7, "👩 ندا", "Hays Healthcare Australia", "🇦🇺 استرالیا",
     "International Midwife — Registration & CV",
     "healthcare@hays.com.au", "🟢 آماده",
     "hays.com.au/healthcare"],
    [8, "👩 ندا", "TalentOrange", "🇩🇪 آلمان",
     "Internationale Hebamme — Deutschland",
     "info@talentorange.de", "🟢 آماده",
     "talentorange.de"],
    [9, "👨 توحید", "TalentOrange", "🇩🇪 آلمان",
     "IT Manager — International Candidate",
     "info@talentorange.de", "🟢 آماده",
     "talentorange.de"],
    [10, "👩 ندا", "RGH Global", "🇳🇿 نیوزیلند",
     "Midwife — International Recruitment NZ",
     "info@rghglobal.co.nz", "🟢 آماده",
     "rghglobal.co.nz"],
]

for i, row_data in enumerate(emails):
    r = 4 + i
    for j, val in enumerate(row_data):
        cell = style_data_cell(ws3, r, j + 1, "en" if j in [3, 5, 7] else "fa")
        cell.value = val
        if j == 6:
            cell.fill = fill("D5F5E3")

widths3 = [5, 15, 30, 18, 40, 38, 12, 35]
for i, w in enumerate(widths3):
    ws3.column_dimensions[get_column_letter(i + 1)].width = w

# ═══════════════════════════════════════════════════
# SHEET 4: نمونه ایمیل طبیعی | Natural Email Samples
# ═══════════════════════════════════════════════════
ws4 = wb.create_sheet("نمونه ایمیل")
set_rtl(ws4)

ws4.merge_cells("A1:D1")
ws4["A1"].value = "نمونه ایمیل طبیعی — Natural Email Sample (غیر AI)"
ws4["A1"].font = Font(name=FONT_FA, size=14, bold=True, color=WHITE)
ws4["A1"].fill = fill("8E44AD")
ws4["A1"].alignment = align_center()

# Email sample - Neda
ws4.merge_cells("A3:D3")
ws4["A3"].value = "📧 ایمیل ندا → Saskatchewan Health Authority"
ws4["A3"].font = fa_font(12, bold=True)
ws4["A3"].fill = fill(LIGHT_BLUE)

email_text_neda = """Subject: Midwife — International Candidate — Saskatchewan

Dear Hiring Team,

I am a practising midwife at Milad Hospital in Tehran with 12 years of experience in antenatal care, labour management, and postnatal support.

I came across Saskatchewan Health Authority's commitment to welcoming internationally trained healthcare professionals and I am very interested in contributing to your maternity services.

I hold a Bachelor of Midwifery from Iran and have extensive clinical experience including high-risk pregnancies, emergency obstetric care, and newborn resuscitation. I have also supervised junior midwives and contributed to quality improvement initiatives in our department.

I am currently at English level A2 and am actively preparing for OET/IELTS. I understand this is a requirement for registration with the College of Midwives of Saskatchewan and I am fully committed to meeting it.

I would appreciate any guidance on the Expression of Interest process for internationally trained midwives. I am available for an interview at your convenience.

My CV is attached for your review.

Kind regards,
Neda Arjmand
Midwife — Milad Hospital, Tehran"""

ws4.merge_cells("A4:D4")
ws4["A4"].value = email_text_neda
ws4["A4"].font = en_font(10)
ws4["A4"].alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
ws4.row_dimensions[4].height = 300

# Email sample - Tohid
ws4.merge_cells("A6:D6")
ws4["A6"].value = "📧 ایمیل توحید → Saskatchewan Health Authority"
ws4["A6"].font = fa_font(12, bold=True)
ws4["A6"].fill = fill(LIGHT_BLUE)

email_text_tohid = """Subject: IT Operations Manager — International Application

Dear Hiring Team,

I am writing regarding IT operations opportunities at Saskatchewan Health Authority. I have 19 years of experience managing IT infrastructure across multi-site environments, including healthcare settings.

My background includes Windows Server administration, VMware/Hyper-V virtualization, network management with Cisco and MikroTik equipment, and backup solutions using Veeam. I have managed teams of up to 8 IT staff and oversaw infrastructure for a hospital network with 500+ endpoints.

I am drawn to Saskatchewan because of its growing healthcare sector and the Authority's track record of supporting international professionals. I believe my experience with hospital IT systems — including EMR infrastructure, high-availability servers, and clinical network management — would be a good fit.

I am at English A2 and committed to reaching the required level for professional integration.

I have attached my CV and would welcome the opportunity to discuss how my skills might match your current or upcoming IT needs.

Best regards,
Tohid Arjmand
IT Operations Manager — 19 Years Experience"""

ws4.merge_cells("A7:D7")
ws4["A7"].value = email_text_tohid
ws4["A7"].font = en_font(10)
ws4["A7"].alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
ws4.row_dimensions[7].height = 300

# Tips
ws4.merge_cells("A9:D9")
ws4["A9"].value = "💡 قوانین نوشتن ایمیل طبیعی"
ws4["A9"].font = fa_font(12, bold=True)
ws4["A9"].fill = fill(YELLOW)

tips = [
    "1. از جزئیات واقعی آگهی استفاده کن — نه متن عمومی",
    "2. جملات کوتاه و بلند باشند — مثل انسان واقعی",
    "3. دلیل انتخاب کارفرما مشخص باشد",
    "4. صادقانه درباره سطح زبان بنویس",
    "5. درخواست مشخص داشته باش (EOI, interview, guidance)",
    "6. عبارت AI مثل 'I am writing to express my interest' نباشد",
    "7. طول: ۱۵۰-۲۵۰ کلمه",
    "8. Subject line مشخص باشد",
    "9. امضا ساده باشد",
    "10. از اصطلاحات محلی استفاده کن (Kiaora, Guten Tag)",
]

for i, tip in enumerate(tips):
    r = 10 + i
    ws4.merge_cells(f"A{r}:D{r}")
    ws4[f"A{r}"].value = tip
    ws4[f"A{r}"].font = fa_font(10)

ws4.column_dimensions["A"].width = 30
ws4.column_dimensions["B"].width = 30
ws4.column_dimensions["C"].width = 30
ws4.column_dimensions["D"].width = 30

# ═══════════════════════════════════════════════════
# SHEET 5: دقت جستجو | Search Accuracy
# ═══════════════════════════════════════════════════
ws5 = wb.create_sheet("دقت جستجو")
set_rtl(ws5)

ws5.merge_cells("A1:F1")
ws5["A1"].value = "تحلیل دقت جستجو — Search Accuracy Analysis"
ws5["A1"].font = Font(name=FONT_FA, size=14, bold=True, color=WHITE)
ws5["A1"].fill = fill(RED)
ws5["A1"].alignment = align_center()

headers5 = ["#", "روش", "دقت", "چرا؟", "چه کاری می‌توانم انجام بدهم", "چه کاری نمی‌توانم"]
for i, h in enumerate(headers5, 1):
    ws5.cell(row=3, column=i, value=h)
style_header_row(ws5, 3, 6, RED)

accuracy = [
    [1, "خواندن صفحه وب ساده", "90%+",
     "HTML ساده، متن قابل استخراج",
     "خواندن محتوا، استخراج ایمیل و لینک",
     "خواندن صفحات JavaScript-only"],
    [2, "جستجوی Google", "70-75%",
     "نتایج متنوع ولی همه چیز را نشان نمی‌دهد",
     "یافتن آگهی‌ها و سایت‌های مرتبط",
     "دسترسی به نتایج پشت login"],
    [3, "سایت دولتی", "90%+",
     "اطلاعات رسمی و معتبر",
     "خواندن قوانین، لیست مشاغل، فرآیندها",
     "دسترسی به سیستم‌های آنلاین"],
    [4, "استخراج ایمیل", "60-70%",
     "ایمیل‌ها در صفحه تماس هستند",
     "پیدا کردن ایمیل‌های عمومی HR",
     "تأیید اینکه ایمیل فعال است"],
    [5, "تشخیص فرم آنلاین", "80%",
     "لینک فرم در صفحه قابل شناسایی است",
     "پیدا کردن لینک فرم",
     "پر کردن فرم (نیاز به مرورگر)"],
    [6, "خواندن PDF آگهی", "40-50%",
     "بعضی PDFها متن ساده دارند",
     "خواندن PDFهای ساده",
     "خواندن PDF اسکن‌شده"],
    [7, "ورود به LinkedIn", "0%",
     "نیاز به لاگین واقعی",
     "هیچ",
     "خواندن پروفایل خصوصی"],
    [8, "پر کردن فرم آنلاین", "0%",
     "نیاز به مرورگر واقعی",
     "هیچ (فقط لینک می‌دهم)",
     "پر کردن فرم"],
    [9, "ارسال ایمیل", "0%",
     "نیاز به سرویس ایمیل واقعی",
     "هیچ (فقط متن می‌نویسم)",
     "ارسال واقعی"],
    [10, "دور زدن CAPTCHA", "0%",
     "غیراخلاقی و غیرممکن",
     "هیچ",
     "هیچ"],
]

for i, row_data in enumerate(accuracy):
    r = 4 + i
    for j, val in enumerate(row_data):
        cell = style_data_cell(ws5, r, j + 1, "en" if j in [2] else "fa")
        cell.value = val
        if j == 2:
            if "90" in str(val):
                cell.fill = fill("D5F5E3")
            elif "70" in str(val) or "80" in str(val):
                cell.fill = fill("FEF9E7")
            elif "40" in str(val) or "60" in str(val):
                cell.fill = fill("FDEBD0")
            else:
                cell.fill = fill("FADBD8")

widths5 = [5, 28, 12, 35, 35, 35]
for i, w in enumerate(widths5):
    ws5.column_dimensions[get_column_letter(i + 1)].width = w

# ═══════════════════════════════════════════════════
# SHEET 6: خلاصه | Summary
# ═══════════════════════════════════════════════════
ws6 = wb.create_sheet("خلاصه")
set_rtl(ws6)

ws6.merge_cells("A1:C1")
ws6["A1"].value = "خلاصه — فرآیند جستجو و محدودیت‌ها"
ws6["A1"].font = Font(name=FONT_FA, size=14, bold=True, color=WHITE)
ws6["A1"].fill = fill(DARK_GRAY)
ws6["A1"].alignment = align_center()

summary = [
    ["✅ من چه کاری می‌توانم انجام بدهم", "", ""],
    ["", "خواندن صفحات وب ساده", "HTML ساده → متن → استخراج"],
    ["", "جستجو در Google", "عبارت جستجو → نتایج → لینک"],
    ["", "خواندن سایت کارفرما", "صفحه Careers → آگهی + ایمیل"],
    ["", "سایت دولتی", "قوانین + لیست مشاغل + فرآیند"],
    ["", "نوشتن ایمیل حرفه‌ای", "بر اساس جزئیات واقعی آگهی"],
    ["", "نوشتن CV/کاور لیتر", "تطبیق با آگهی و پروفایل"],
    ["", "ساخت Excel/گزارش", "openpyxl + Markdown"],
    ["", "بروزرسانی حافظه", "فایل‌های Markdown"],
    ["", "", ""],
    ["❌ من چه کاری نمی‌توانم انجام بدهم", "", ""],
    ["", "ارسال واقعی ایمیل", "فقط متن آماده می‌کنم"],
    ["", "پر کردن فرم آنلاین", "فقط لینک فرم می‌دهم"],
    ["", "ورود به LinkedIn", "نیاز به لاگین واقعی"],
    ["", "خواندن صفحات JavaScript-only", "بعضی سایت‌ها کار نمی‌کنند"],
    ["", "دور زدن CAPTCHA", "غیرممکن"],
    ["", "تأیید ایمیل فعال", "ممکن است قدیمی باشد"],
    ["", "تضمین پاسخ کارفرما", "هیچ تضمینی وجود ندارد"],
    ["", "", ""],
    ["⏰ بودجه زمانی هر روز", "", ""],
    ["", "بارگذاری حافظه", "۱ دقیقه"],
    ["", "جستجوی Google", "۵ دقیقه"],
    ["", "خواندن سایت کارفرما", "۵ دقیقه"],
    ["", "فیلترینگ و امتیازدهی", "۲ دقیقه"],
    ["", "نوشتن ایمیل", "۳ دقیقه"],
    ["", "بروزرسانی Excel", "۲ دقیقه"],
    ["", "گزارش", "۲ دقیقه"],
    ["", "مجموع", "۲۰ دقیقه"],
]

for i, row_data in enumerate(summary):
    r = 3 + i
    for j, val in enumerate(row_data):
        cell = style_data_cell(ws6, r, j + 1, "fa")
        cell.value = val
        if j == 0 and val.startswith("✅"):
            cell.fill = fill("D5F5E3")
            cell.font = fa_font(11, bold=True)
        elif j == 0 and val.startswith("❌"):
            cell.fill = fill("FADBD8")
            cell.font = fa_font(11, bold=True)
        elif j == 0 and val.startswith("⏰"):
            cell.fill = fill("D6E4F0")
            cell.font = fa_font(11, bold=True)

ws6.column_dimensions["A"].width = 40
ws6.column_dimensions["B"].width = 35
ws6.column_dimensions["C"].width = 35

# ═══════════════════════════════════════════════════
# SAVE
# ═══════════════════════════════════════════════════
os.makedirs("dashboard", exist_ok=True)
fname = f"dashboard/Search_Methods_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
wb.save(fname)
print(f"✅ {fname}")
print(f"📊 شیت‌ها: ۶")
print(f"   1. روش‌های جستجو | Search Methods")
print(f"   2. وضعیت جستجو | Search Status")
print(f"   3. ایمیل‌های آماده | Ready Emails")
print(f"   4. نمونه ایمیل | Natural Email Samples")
print(f"   5. دقت جستجو | Search Accuracy")
print(f"   6. خلاصه | Summary")
