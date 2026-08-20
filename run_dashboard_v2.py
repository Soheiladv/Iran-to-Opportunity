#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Migration Hunter Dashboard v2
- آرشیو خودکار Excel قدیمی
- شیت ایمیل و لینک آگهی‌ها
- ۱۰ آگهی برتر در روز
- امتیازدهی بر اساس نقاط مثبت و امکان تحقق
- هر کشور جستجو شده به شیت اضافه می‌شود
"""
import os
import shutil
from datetime import datetime
from pathlib import Path

BASE_DIR = Path(__file__).parent
DASHBOARD_DIR = BASE_DIR / "dashboard"
ARCHIVE_DIR = BASE_DIR / "archive"
NOW = datetime.now()
DATE_STR = NOW.strftime("%Y-%m-%d %H:%M")

# ════════════════════════════════════════════════════════════
# ۱. آرشیو خودکار Excel قدیمی
# ════════════════════════════════════════════════════════════

def archive_old_excels():
    """انتقال فایل‌های Excel قدیمی به archive"""
    ARCHIVE_DIR.mkdir(exist_ok=True)
    
    archived = 0
    for f in DASHBOARD_DIR.glob("*.xlsx"):
        # فقط فایل‌هایی که Final یا Full نیستند (یا قدیمی‌تر از ۱ ساعت)
        age_hours = (NOW.timestamp() - f.stat().st_mtime) / 3600
        if age_hours > 1 or "Final" not in f.name:
            dest = ARCHIVE_DIR / f.name
            shutil.move(str(f), str(dest))
            archived += 1
            print(f"  📦 آرشیو: {f.name}")
    
    print(f"\n📦 {archived} فایل آرشیو شد → archive/")
    return archived

# ════════════════════════════════════════════════════════════
# ۲. سیستم امتیازدهی آگهی
# ════════════════════════════════════════════════════════════

def score_job(job):
    """
    امتیازدهی بر اساس:
    - نقاط مثبت (حمایت مالی، بین‌المللی، زبان)
    - امکان تحقق (واقعی بودن، دسترسی، سرعت)
    """
    score = 0
    positives = []
    feasibility = []
    
    # ── نقاط مثبت (60 امتیاز) ──
    if job.get("sponsorship") == "Confirmed":
        score += 20
        positives.append("حمایت مالی تأیید شده +20")
    elif job.get("sponsorship") == "Likely":
        score += 12
        positives.append("حمایت مالی محتمل +12")
    elif job.get("sponsorship") == "Possible":
        score += 6
        positives.append("حمایت مالی ممکن +6")
    
    if job.get("international"):
        score += 15
        positives.append("استخدام بین‌المللی +15")
    
    if job.get("free_service"):
        score += 5
        positives.append("خدمات رایگان +5")
    
    if job.get("relocation"):
        score += 5
        positives.append("کمک انتقال +5")
    
    if job.get("language_training"):
        score += 5
        positives.append("آموزش زبان +5")
    
    # ── امکان تحقق (40 امتیاز) ──
    if job.get("url_works"):
        score += 10
        feasibility.append("لینک فعال +10")
    elif job.get("url_works") == False:
        score -= 5
        feasibility.append("لینک خراب -5")
    
    if job.get("contact_email"):
        score += 8
        feasibility.append("ایمیل تماس +8")
    
    if job.get("form_available"):
        score += 7
        feasibility.append("فرم آنلاین +7")
    
    if job.get("phone"):
        score += 5
        feasibility.append("تلفن تماس +5")
    
    if job.get("fresh"):
        score += 10
        feasibility.append("تازه و فعال +10")
    
    return {
        "score": min(score, 100),
        "positives": positives,
        "feasibility": feasibility
    }

# ════════════════════════════════════════════════════════════
# ۳. داده‌های واقعی
# ════════════════════════════════════════════════════════════

ALL_JOBS = [
    # ── نیوزیلند ──
    {
        "id": 1, "applicant": "NEDA", "employer": "Health New Zealand",
        "country": "nz", "country_fa": "نیوزیلند",
        "title": "Midwife / Registered Midwife",
        "url": "https://www.healthnz.govt.nz/careers",
        "email": "careers@healthnz.govt.nz",
        "phone": None, "form": "https://www.healthnz.govt.nz/careers/international",
        "sponsorship": "Confirmed", "international": True, "free_service": False,
        "relocation": False, "language_training": False,
        "url_works": True, "contact_email": True, "form_available": True,
        "phone": False, "fresh": True, "green_list": True,
        "lang_req": "IELTS Academic / OET — Registration",
        "visa_path": "AEWV + Green List",
        "notes": "Green List occupation — اقامت فوری"
    },
    {
        "id": 2, "applicant": "NEDA", "employer": "RGH Global",
        "country": "nz", "country_fa": "نیوزیلند",
        "title": "Midwife with Sponsorship",
        "url": "https://www.rgh-global.com/jobs/midwife-with-sponsorship/",
        "email": "info@rgh-global.com",
        "phone": None, "form": None,
        "sponsorship": "Confirmed", "international": True, "free_service": False,
        "relocation": True, "language_training": False,
        "url_works": True, "contact_email": True, "form_available": False,
        "phone": False, "fresh": True, "green_list": False,
        "lang_req": "IELTS/OET",
        "visa_path": "AEWV",
        "notes": "متخصص استخدام ماما"
    },
    {
        "id": 3, "applicant": "NEDA", "employer": "Working In Health NZ",
        "country": "nz", "country_fa": "نیوزیلند",
        "title": "International Midwife",
        "url": "https://www.workingin-health.co.nz",
        "email": None,
        "phone": None, "form": "https://www.workingin-health.co.nz/midwifery-jobs/",
        "sponsorship": "Confirmed", "international": True, "free_service": True,
        "relocation": True, "language_training": False,
        "url_works": True, "contact_email": False, "form_available": True,
        "phone": False, "fresh": True, "green_list": True,
        "lang_req": "Registration requirement",
        "visa_path": "AEWV + Green List",
        "notes": "خدمات رایگان — آژانس تخصصی"
    },
    # ── آلمان ──
    {
        "id": 4, "applicant": "NEDA", "employer": "Holalemania GmbH",
        "country": "de", "country_fa": "آلمان",
        "title": "Geburtshelfer/in (Midwife)",
        "url": "https://holalemania.de/en/midwives/",
        "email": "info@holalemania.de",
        "phone": "+49-40-41 49 65 05",
        "form": "https://holalemania.de/en/application/",
        "sponsorship": "Confirmed", "international": True, "free_service": True,
        "relocation": True, "language_training": True,
        "url_works": True, "contact_email": True, "form_available": True,
        "phone": True, "fresh": True, "green_list": False,
        "lang_req": "German B1-B2 (provided by Holalemania)",
        "visa_path": "Work Visa + Recognition",
        "notes": "۱۳ سال تجربه — ۹۲۱ استخدام موفق — آموزش زبان"
    },
    {
        "id": 5, "applicant": "NEDA", "employer": "TalentOrange",
        "country": "de", "country_fa": "آلمان",
        "title": "Healthcare Professional",
        "url": "https://www.talentorange.com/en/",
        "email": None,
        "phone": None, "form": "https://www.talentorange.com/en/for-candidates/",
        "sponsorship": "Confirmed", "international": True, "free_service": True,
        "relocation": True, "language_training": True,
        "url_works": True, "contact_email": False, "form_available": True,
        "phone": False, "fresh": True, "green_list": False,
        "lang_req": "German B2 (borseh provided)",
        "visa_path": "Work Visa",
        "notes": "بورسیه زبان آلمانی B2"
    },
    # ── کانادا ──
    {
        "id": 6, "applicant": "NEDA", "employer": "Saskatchewan Health Authority",
        "country": "ca", "country_fa": "کانادا",
        "title": "Midwife / Healthcare Professional",
        "url": "https://www.saskhealthauthority.ca/careers",
        "email": "SHAInternational@saskhealthauthority.ca",
        "phone": None,
        "form": "https://www.saskhealthauthority.ca/careers-volunteering/careers/internationally-trained-health-care-professionals",
        "sponsorship": "Likely", "international": True, "free_service": False,
        "relocation": True, "language_training": False,
        "url_works": True, "contact_email": True, "form_available": True,
        "phone": False, "fresh": True, "green_list": False,
        "lang_req": "CLB 7 (IELTS 6.0+)",
        "visa_path": "Provincial Nominee Program",
        "notes": "فعالانه استخدام بین‌المللی — EOI system"
    },
    # ── استرالیا ──
    {
        "id": 7, "applicant": "NEDA", "employer": "Hays Healthcare Australia",
        "country": "au", "country_fa": "استرالیا",
        "title": "Midwife",
        "url": "https://www.hays.com.au/jobs/healthcare",
        "email": None,
        "phone": None, "form": "https://www.hays.com.au/register-your-cv",
        "sponsorship": "Possible", "international": True, "free_service": True,
        "relocation": False, "language_training": False,
        "url_works": True, "contact_email": False, "form_available": True,
        "phone": False, "fresh": True, "green_list": False,
        "lang_req": "IELTS 7.0",
        "visa_path": "482 / 189 / 190",
        "notes": "آژانس استخدام بزرگ — حمایت ویزا ممکن"
    },
    # ── توحید IT ──
    {
        "id": 8, "applicant": "TOHID", "employer": "Arbeitnow",
        "country": "de", "country_fa": "آلمان",
        "title": "IT Manager / Infrastructure Manager",
        "url": "https://www.arbeitnow.com/visa-sponsorship-jobs",
        "email": None,
        "phone": None, "form": None,
        "sponsorship": "Likely", "international": True, "free_service": True,
        "relocation": False, "language_training": False,
        "url_works": True, "contact_email": False, "form_available": False,
        "phone": False, "fresh": True, "green_list": False,
        "lang_req": "English — varies",
        "visa_path": "Work Visa / Blue Card",
        "notes": "مشاغل IT با حمایت ویزا"
    },
]

# ════════════════════════════════════════════════════════════
# ۴. ساخت Excel نهایی
# ════════════════════════════════════════════════════════════

def build_dashboard_v2():
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    except ImportError:
        print("❌ pip install openpyxl")
        return

    wb = Workbook()

    # Styles
    FA = Font(name='B Mitra', size=11)
    FA_T = Font(name='B Mitra', size=14, bold=True)
    FA_S = Font(name='B Mitra', size=10, italic=True, color='666666')
    HD = Font(name='B Mitra', size=11, bold=True, color='FFFFFF')
    GF = PatternFill(start_color='92D050', end_color='92D050', fill_type='solid')
    YF = PatternFill(start_color='FFD966', end_color='FFD966', fill_type='solid')
    OF = PatternFill(start_color='F4B084', end_color='F4B084', fill_type='solid')
    BF = PatternFill(start_color='9BC2E6', end_color='9BC2E6', fill_type='solid')
    RF = PatternFill(start_color='FF6B6B', end_color='FF6B6B', fill_type='solid')
    HF = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    GR = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
    BD = Border(left=Side('thin'), right=Side('thin'), top=Side('thin'), bottom=Side('thin'))

    def sh(ws, r, c):
        cell = ws.cell(row=r, column=c)
        cell.font = HD; cell.fill = HF
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = BD

    def sc(ws, r, c, v=None, f=FA, fill=None):
        cell = ws.cell(row=r, column=c)
        if v is not None: cell.value = v
        cell.font = f
        cell.alignment = Alignment(vertical='center', wrap_text=True, horizontal='right')
        cell.border = BD
        if fill: cell.fill = fill
        return cell

    # ── شیت ۱: داشبورد اصلی ──
    ws1 = wb.active
    ws1.title = "📊 Dashboard | داشبورد"
    ws1.sheet_view.rightToLeft = True
    ws1.merge_cells('A1:G1')
    ws1.cell(1, 1, f"📊 داشبورد جامع — {DATE_STR}").font = FA_T
    ws1.cell(2, 1, "آرشیو خودکار + شیت ایمیل/لینک + ۱۰ آگهی برتر").font = FA_S

    h1 = ["شاخص | Metric", "مقدار | Value", "توضیح | Notes"]
    for i, x in enumerate(h1, 1): ws1.cell(4, i, x)
    sh(ws1, 4, 3)

    # آمار
    scored = [score_job(j) for j in ALL_JOBS]
    summary = [
        ("کل آگهی‌ها | Total Jobs", len(ALL_JOBS), ""),
        ("آگهی‌های با امتیاز ۸۰+", sum(1 for s in scored if s["score"] >= 80), "بالاترین اولویت"),
        ("آگهی‌های با امتیاز ۶۰-۷۹", sum(1 for s in scored if 60 <= s["score"] < 80), "متوسط"),
        ("آگهی‌های با امتیاز زیر ۶۰", sum(1 for s in scored if s["score"] < 60), "نیاز به بررسی"),
        ("کشورهای جستجو شده", 4, "nz, de, ca, au"),
        ("ایمیل‌های تماس", sum(1 for j in ALL_JOBS if j.get("email")), ""),
        ("فرم‌های آنلاین", sum(1 for j in ALL_JOBS if j.get("form")), ""),
        ("تلفن‌های تماس", sum(1 for j in ALL_JOBS if j.get("phone")), ""),
        ("آرشیو شده | Archived", "خودکار", "فایل‌های قدیمی → archive/"),
    ]
    for r, (l, v, n) in enumerate(summary, 5):
        sc(ws1, r, 1, l); sc(ws1, r, 2, v); sc(ws1, r, 3, n)
    ws1.column_dimensions['A'].width = 35
    ws1.column_dimensions['B'].width = 20
    ws1.column_dimensions['C'].width = 30

    # ── شیت ۲: ۱۰ آگهی برتر ──
    ws2 = wb.create_sheet("🏆 Top 10 | ۱۰ برتر")
    ws2.sheet_view.rightToLeft = True
    ws2.merge_cells('A1:J1')
    ws2.cell(1, 1, "🏆 ۱۰ آگهی برتر بر اساس امتیاز | Top 10 Jobs by Score").font = FA_T

    h2 = ["# | ردیف", "کارفرما | Employer", "کشور | Country", "شغل | Job", "امتیاز | Score", "نقاط مثبت | Positives", "امکان تحقق | Feasibility", "لینک | URL", "ایمیل | Email", "وضعیت | Status"]
    for i, x in enumerate(h2, 1): ws2.cell(3, i, x)
    sh(ws2, 3, 10)

    # مرتب‌سازی بر اساس امتیاز
    scored_jobs = sorted(
        [(j, score_job(j)) for j in ALL_JOBS],
        key=lambda x: x[1]["score"],
        reverse=True
    )[:10]

    for r, (job, sc_data) in enumerate(scored_jobs, 4):
        sc(ws2, r, 1, r - 3)
        sc(ws2, r, 2, job["employer"])
        sc(ws2, r, 3, job["country_fa"])
        sc(ws2, r, 4, job["title"])
        
        # امتیاز با رنگ
        score_cell = sc(ws2, r, 5, sc_data["score"])
        if sc_data["score"] >= 80: score_cell.fill = GF
        elif sc_data["score"] >= 60: score_cell.fill = YF
        else: score_cell.fill = OF
        
        sc(ws2, r, 6, "\n".join(sc_data["positives"][:3]))
        sc(ws2, r, 7, "\n".join(sc_data["feasibility"][:3]))
        sc(ws2, r, 8, job["url"])
        sc(ws2, r, 9, job.get("email") or "—")
        
        status = "🟢 فعال" if job["url_works"] else "🔴 خراب"
        sc(ws2, r, 10, status)

    for col in 'ABCDEFGHIJ': ws2.column_dimensions[col].width = 28

    # ── شیت ۳: ایمیل و لینک آگهی‌ها ──
    ws3 = wb.create_sheet("📧 Emails & Links | ایمیل و لینک")
    ws3.sheet_view.rightToLeft = True
    ws3.merge_cells('A1:J1')
    ws3.cell(1, 1, "📧 ایمیل، لینک و فرم آگهی‌ها | Job Ads Email, Links & Forms").font = FA_T

    h3 = ["# | ردیف", "کارفرما | Employer", "کشور | Country", "ایمیل | Email", "تلفن | Phone", "لینک آگهی | Job URL", "فرم آنلاین | Form", "لینک فعال | URL Works", "وضعیت | Status", "یادداشت | Notes"]
    for i, x in enumerate(h3, 1): ws3.cell(3, i, x)
    sh(ws3, 3, 10)

    for r, job in enumerate(ALL_JOBS, 4):
        sc(ws3, r, 1, r - 3)
        sc(ws3, r, 2, job["employer"])
        sc(ws3, r, 3, job["country_fa"])
        
        email_cell = sc(ws3, r, 4, job.get("email") or "—")
        if job.get("email"): email_cell.fill = GR
        
        sc(ws3, r, 5, job.get("phone") or "—")
        
        url_cell = sc(ws3, r, 6, job.get("url") or "—")
        if job.get("url_works"): url_cell.fill = GR
        
        form_cell = sc(ws3, r, 7, job.get("form") or "—")
        if job.get("form"): form_cell.fill = GR
        
        works_cell = sc(ws3, r, 8, "✅ فعال" if job["url_works"] else "❌ خراب")
        works_cell.fill = GF if job["url_works"] else RF
        
        sc(ws3, r, 9, "🟢 آماده" if job.get("email") or job.get("form") else "🟡 نیاز به بررسی")
        sc(ws3, r, 10, job.get("notes") or "")

    for col in 'ABCDEFGHIJ': ws3.column_dimensions[col].width = 28

    # ── شیت ۴: فرصت‌های ندا ──
    ws4 = wb.create_sheet("👩 Neda | ندا")
    ws4.sheet_view.rightToLeft = True
    ws4.merge_cells('A1:I1')
    ws4.cell(1, 1, "👩 فرصت‌های مامایی ندا | Neda Midwifery Opportunities").font = FA_T

    h4 = ["#", "کارفرما | Employer", "کشور | Country", "شغل | Job", "حمایت | Sponsor", "امتیاز | Score", "لینک | URL", "زبان | Language", "ویزا | Visa"]
    for i, x in enumerate(h4, 1): ws4.cell(3, i, x)
    sh(ws4, 3, 9)

    neda_jobs = [(j, score_job(j)) for j in ALL_JOBS if j["applicant"] == "NEDA"]
    for r, (job, sc_data) in enumerate(neda_jobs, 4):
        sc(ws4, r, 1, r - 3)
        sc(ws4, r, 2, job["employer"])
        sc(ws4, r, 3, job["country_fa"])
        sc(ws4, r, 4, job["title"])
        sc(ws4, r, 5, job["sponsorship"])
        c = sc(ws4, r, 6, sc_data["score"])
        c.fill = GF if sc_data["score"] >= 80 else YF if sc_data["score"] >= 60 else OF
        sc(ws4, r, 7, job["url"])
        sc(ws4, r, 8, job["lang_req"])
        sc(ws4, r, 9, job["visa_path"])

    for col in 'ABCDEFGHI': ws4.column_dimensions[col].width = 25

    # ── شیت ۵: فرصت‌های توحید ──
    ws5 = wb.create_sheet("👨 Tohid | توحید")
    ws5.sheet_view.rightToLeft = True
    ws5.merge_cells('A1:I1')
    ws5.cell(1, 1, "👨 فرصت‌های IT توحید | Tohid IT Opportunities").font = FA_T
    for i, x in enumerate(h4, 1): ws5.cell(3, i, x)
    sh(ws5, 3, 9)

    tohid_jobs = [(j, score_job(j)) for j in ALL_JOBS if j["applicant"] == "TOHID"]
    for r, (job, sc_data) in enumerate(tohid_jobs, 4):
        sc(ws5, r, 1, r - 3)
        sc(ws5, r, 2, job["employer"])
        sc(ws5, r, 3, job["country_fa"])
        sc(ws5, r, 4, job["title"])
        sc(ws5, r, 5, job["sponsorship"])
        c = sc(ws5, r, 6, sc_data["score"])
        c.fill = GF if sc_data["score"] >= 80 else YF if sc_data["score"] >= 60 else OF
        sc(ws5, r, 7, job["url"])
        sc(ws5, r, 8, job["lang_req"])
        sc(ws5, r, 9, job["visa_path"])

    for col in 'ABCDEFGHI': ws5.column_dimensions[col].width = 25

    # ── شیت ۶: مقایسه کشورها ──
    ws6 = wb.create_sheet("🌍 Countries | کشورها")
    ws6.sheet_view.rightToLeft = True
    ws6.merge_cells('A1:F1')
    ws6.cell(1, 1, "🌍 کشورهای جستجو شده | Countries Searched").font = FA_T

    h6 = ["کشور | Country", "تعداد آگهی | Jobs", "بهترین امتیاز | Best Score", "حمایت مالی | Sponsor", "زبان | Language", "ویزا | Visa"]
    for i, x in enumerate(h6, 1): ws6.cell(3, i, x)
    sh(ws6, 3, 6)

    countries_stats = {}
    for job in ALL_JOBS:
        c = job["country"]
        if c not in countries_stats:
            countries_stats[c] = {"name": job["country_fa"], "jobs": 0, "best": 0, "sponsor": False}
        countries_stats[c]["jobs"] += 1
        s = score_job(job)["score"]
        countries_stats[c]["best"] = max(countries_stats[c]["best"], s)
        if job["sponsorship"] in ["Confirmed", "Likely"]:
            countries_stats[c]["sponsor"] = True

    country_info = {
        "nz": ("نیوزیلند", "English", "AEWV/Green List"),
        "de": ("آلمان", "German B1-B2", "Blue Card/Work Visa"),
        "ca": ("کانادا", "CLB 7", "Express Entry/PNP"),
        "au": ("استرالیا", "IELTS 7", "482/189/190"),
    }
    for r, (code, stats) in enumerate(countries_stats.items(), 4):
        ci = country_info.get(code, ("—", "—", "—"))
        sc(ws6, r, 1, f"{stats['name']} ({code.upper()})")
        sc(ws6, r, 2, stats["jobs"])
        c = sc(ws6, r, 3, stats["best"])
        c.fill = GF if stats["best"] >= 80 else YF
        sc(ws6, r, 4, "✅" if stats["sponsor"] else "⚠️")
        sc(ws6, r, 5, ci[1])
        sc(ws6, r, 6, ci[2])

    for col in 'ABCDEF': ws6.column_dimensions[col].width = 28

    # ── شیت ۷: نقاط ضعف لینکدین ──
    ws7 = wb.create_sheet("⚠️ Weaknesses | نقاط ضعف")
    ws7.sheet_view.rightToLeft = True
    ws7.merge_cells('A1:E1')
    ws7.cell(1, 1, "⚠️ نقاط ضعف لینکدین | LinkedIn Weaknesses").font = FA_T

    h7 = ["متقاضی | Applicant", "دسته | Category", "مشکل | Issue", "راه‌حل | Fix", "اولویت | Priority"]
    for i, x in enumerate(h7, 1): ws7.cell(3, i, x)
    sh(ws7, 3, 5)

    weaknesses = [
        ("ندا", "عکس", "عدم عکس حرفه‌ای", "عکس پرتره ساده", "🔴 فوری"),
        ("ندا", "Headline", "عنوان ساده", "Registered Midwife | International", "🔴 فوری"),
        ("ندا", "About", "ندارد", "۳-۵ پاراگراف تجربه", "🔴 فوری"),
        ("ندا", "Languages", "فقط فارسی", "English A2 + German A1", "🔴 فوری"),
        ("ندا", "Skills", "کمتر از ۵", "۱۰-۱۵ مهارت مامایی", "🟠 مهم"),
        ("توحید", "عکس", "عدم عکس حرفه‌ای", "عکس پرتره ساده", "🔴 فوری"),
        ("توحید", "Headline", "غیراستاندارد", "IT Operations Manager | Infrastructure", "🔴 فوری"),
        ("توحید", "About", "فنی و خشک", "خلاصه حرفه‌ای مدیریتی", "🔴 فوری"),
        ("توحید", "Recommendations", "ندارد", "درخواست از مدیران", "🔴 فوری"),
        ("توحید", "Open to Work", "فعال نیست", "فعال کردن badge", "🔴 فوری"),
    ]
    for r, (app, cat, issue, fix, pri) in enumerate(weaknesses, 4):
        sc(ws7, r, 1, app)
        sc(ws7, r, 2, cat)
        sc(ws7, r, 3, issue)
        sc(ws7, r, 4, fix)
        sc(ws7, r, 5, pri)

    for col in 'ABCDE': ws7.column_dimensions[col].width = 30

    # ── شیت ۸: اقدامات ──
    ws8 = wb.create_sheet("🎯 Actions | اقدامات")
    ws8.sheet_view.rightToLeft = True
    ws8.merge_cells('A1:D1')
    ws8.cell(1, 1, "🎯 ۱۰ اقدام برتر امروز | Top 10 Actions").font = FA_T

    h8 = ["#", "اقدام | Action", "کشور | Country", "اولویت | Priority"]
    for i, x in enumerate(h8, 1): ws8.cell(3, i, x)
    sh(ws8, 3, 4)

    actions = [
        (1, "ارسال CV به Saskatchewan Health Authority", "🇨🇦 کانادا", "🔴 فوری"),
        (2, "ثبت‌نام Hays Healthcare Australia", "🇦🇺 استرالیا", "🔴 فوری"),
        (3, "ایمیل Holalemania + فرم آنلاین", "🇩🇪 آلمان", "🔴 فوری"),
        (4, "پیگیری Working In Health NZ", "🇳🇿 نیوزیلند", "🟠 مهم"),
        (5, "اصلاح لینکدین ندا (Headline + About)", "🌍 بین‌المللی", "🟠 مهم"),
        (6, "اصلاح لینکدین توحید (Headline + About)", "🌍 بین‌المللی", "🟠 مهم"),
        (7, "شروع آمادگی OET", "🌍 بین‌المللی", "🟠 مهم"),
        (8, "جستجوی IT آلمان/اتریش", "🇩🇪🇦🇹 اروپا", "🟡 متوسط"),
        (9, "ثبت‌نام TalentOrange", "🇩🇪 آلمان", "🟡 متوسط"),
        (10, "بررسی پاسخ ایمیل‌ها", "🌍 بین‌المللی", "🟡 متوسط"),
    ]
    for r, (n, a, c, p) in enumerate(actions, 4):
        sc(ws8, r, 1, n); sc(ws8, r, 2, a)
        sc(ws8, r, 3, c); sc(ws8, r, 4, p)

    for col in 'ABCD': ws8.column_dimensions[col].width = 38

    # ── ذخیره ──
    DASHBOARD_DIR.mkdir(exist_ok=True)
    fn = f"MigrationHunter_v2_{NOW.strftime('%Y%m%d_%H%M')}.xlsx"
    fp = DASHBOARD_DIR / fn
    wb.save(fp)
    print(f"\n✅ داشبورد v2: {fn}")
    print(f"   شیت‌ها: {len(wb.sheetnames)}")
    for s in wb.sheetnames:
        print(f"   • {s}")
    return fp

# ── اجرا ──
if __name__ == "__main__":
    print("=" * 60)
    print("🔄 داشبورد v2 — آرشیو + امتیازدهی + ۱۰ آگهی")
    print(f"📅 {DATE_STR}")
    print("=" * 60)
    
    print("\n📦 آرشیو فایل‌های قدیمی...")
    archive_old_excels()
    
    print("\n📊 ساخت داشبورد...")
    build_dashboard_v2()
    
    print("\n✅ تمام!")
