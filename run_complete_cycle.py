#!/usr/bin/env python3
"""
چرخه کامل: ایمیل طبیعی + نقاط ضعف لینکدین + لینک‌های جستجو + فونت فارسی RTL
"""
import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# ═══════════════════════════════════════════════════
# CONSTANTS
# ═══════════════════════════════════════════════════
FONT_FA = "B Mitra"
FONT_EN = "Times New Roman"
DARK_BLUE = "1B4F72"
MED_BLUE = "2E86C1"
LIGHT_BLUE = "D6EAF8"
GREEN = "27AE60"
LIGHT_GREEN = "D5F5E3"
YELLOW = "F39C12"
LIGHT_YELLOW = "FEF9E7"
RED = "E74C3C"
LIGHT_RED = "FADBD8"
PURPLE = "8E44AD"
LIGHT_PURPLE = "E8DAEF"
GRAY = "BDC3C7"
LIGHT_GRAY = "F2F3F4"
DARK = "2C3E50"
WHITE = "FFFFFF"
NOW = datetime.now().strftime("%Y-%m-%d %H:%M")

def set_rtl(ws):
    ws.sheet_view.rightToLeft = True

def make_font(name=FONT_FA, size=11, bold=False, color="000000"):
    return Font(name=name, size=size, bold=bold, color=color)

def make_fill(color):
    return PatternFill(start_color=color, end_color=color, fill_type="solid")

def rtl_align(h="right", v="center", wrap=True):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def center_align():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

thin = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin")
)

def write_header(ws, row, cols, fill_color=DARK_BLUE):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = make_font(size=11, bold=True, color=WHITE)
        cell.fill = make_fill(fill_color)
        cell.alignment = center_align()
        cell.border = thin

def write_cell(ws, row, col, value, font_type="fa", size=10, bold=False, color="000000", bg=None, h_align="right"):
    cell = ws.cell(row=row, column=col, value=value)
    fname = FONT_FA if font_type == "fa" else FONT_EN
    cell.font = make_font(name=fname, size=size, bold=bold, color=color)
    cell.alignment = rtl_align(h=h_align)
    cell.border = thin
    if bg:
        cell.fill = make_fill(bg)
    return cell

# ═══════════════════════════════════════════════════
# 1. EMAILS (Natural, non-AI)
# ═══════════════════════════════════════════════════
emails = {
    "neda_health_nz": {
        "to": "international.recruitment@health.govt.nz",
        "subject": "Registered Midwife — International Candidate — New Zealand",
        "body": """Dear Health New Zealand International Recruitment Team,

My name is Neda Arjmand. I am a registered midwife at Milad Hospital in Tehran, Iran, with 12 years of hands-on experience in maternity care.

I have been following Health New Zealand's international recruitment programme and I am very interested in contributing to your maternity services across New Zealand.

My clinical background includes:

— Antenatal assessment and care for low and high-risk pregnancies
— Intrapartum care including normal and assisted deliveries
— Management of obstetric emergencies (postpartum haemorrhage, cord prolapse, shoulder dystocia)
— Neonatal resuscitation and newborn care
— Postnatal care and breastfeeding support
— Supervision of junior midwifery staff

I hold a Bachelor of Midwifery from Iran and I am currently preparing for OET to meet the registration requirements of the Midwifery Council of New Zealand.

I understand that registration is a prerequisite for practising in New Zealand, and I would greatly appreciate any guidance on the process for internationally qualified midwives.

I am available for an interview at your convenience and have attached my CV for your review.

Thank you for your time and consideration.

Kind regards,
Neda Arjmand
Midwife — Milad Hospital, Tehran
Email: [YOUR_EMAIL]
Phone: [YOUR_PHONE]""",
    },
    "tohid_health_nz": {
        "to": "international.recruitment@health.govt.nz",
        "subject": "IT Operations Manager — International Candidate — Health NZ",
        "body": """Dear Health New Zealand International Recruitment Team,

My name is Tohid Arjmand. I am an IT operations professional with 19 years of experience managing infrastructure across multi-site environments, including healthcare settings.

I am writing to express my interest in IT operations or infrastructure management roles within Health New Zealand.

My background includes:

— Windows Server administration and Active Directory management
— VMware, Hyper-V, and Proxmox virtualization
— Network infrastructure (Cisco, MikroTik, Ubiquiti/UniFi)
— Backup and disaster recovery (Veeam)
— IT service management and team leadership (up to 8 staff)
— Hospital IT systems including EMR infrastructure and clinical network management
— Multi-site IT operations across hospitality and healthcare environments

I have managed IT infrastructure for organisations with 500+ endpoints and I understand the critical nature of high-availability systems in healthcare.

I am currently at English A2 and am committed to reaching the level required for professional integration in New Zealand.

I would welcome the opportunity to discuss how my skills might match your current or upcoming IT requirements.

My CV is attached for your review.

Best regards,
Tohid Arjmand
IT Operations Manager — 19 Years Experience
Email: [YOUR_EMAIL]
Phone: [YOUR_PHONE]""",
    },
    "neda_rgh": {
        "to": "info@rgh-global.com",
        "subject": "Midwife — International Candidate — Sponsorship Programme",
        "body": """Dear RGH-Global Team,

I am a registered midwife from Iran with 12 years of clinical experience at Milad Hospital in Tehran. I came across your midwifery sponsorship programme for New Zealand and I am very interested.

My experience covers the full scope of midwifery care: antenatal, intrapartum, and postnatal services, including high-risk pregnancies and obstetric emergencies. I hold a Bachelor of Midwifery and I am currently preparing for OET.

I understand that registration with the Midwifery Council of New Zealand is required and I am committed to completing this process.

I would appreciate any information on the next steps for internationally qualified midwives through your programme.

My CV is attached.

Thank you,
Neda Arjmand""",
    },
    "neda_alberta": {
        "to": "careers@albertahealthservices.ca",
        "subject": "Registered Midwife — International Application — Alberta Health Services",
        "body": """Dear Alberta Health Services Recruitment Team,

My name is Neda Arjmand. I am a registered midwife from Iran with 12 years of clinical experience at Milad Hospital in Tehran.

I am interested in midwifery positions with Alberta Health Services and I would like to understand the process for internationally educated midwives.

My clinical experience includes antenatal care, labour and delivery, postnatal care, and management of obstetric emergencies. I hold a Bachelor of Midwifery from Iran.

I would appreciate information on:
— The process for international midwifery credential assessment
— Current midwifery vacancies in Alberta
— Any relocation support available for international candidates

My CV is attached for your review.

Kind regards,
Neda Arjmand""",
    },
    "tohid_sask_it": {
        "to": "SHAInternational@saskhealthauthority.ca",
        "subject": "IT Operations Manager — International Candidate — Saskatchewan",
        "body": """Dear Saskatchewan Health Authority International Team,

My name is Tohid Arjmand. I am an IT operations professional with 19 years of experience, including healthcare IT environments.

I am writing to enquire about IT infrastructure or operations management opportunities within the Saskatchewan Health Authority.

My skills include Windows Server administration, VMware/Hyper-V virtualization, network management (Cisco, MikroTik), backup solutions (Veeam), and team leadership. I have managed IT infrastructure for hospital networks with 500+ endpoints.

I am at English A2 and working towards the required level for professional integration.

I would appreciate any guidance on current or upcoming IT positions and the process for international candidates.

My CV is attached.

Best regards,
Tohid Arjmand""",
    },
}

os.makedirs("output/emails", exist_ok=True)
for key, email in emails.items():
    path = f"output/emails/{key}.txt"
    with open(path, "w", encoding="utf-8") as f:
        f.write(f"To: {email['to']}\n")
        f.write(f"Subject: {email['subject']}\n")
        f.write(f"{'='*60}\n\n")
        f.write(email['body'])
    print(f"📧 {path}")

# ═══════════════════════════════════════════════════
# 2. LINKEDIN WEAKNESSES
# ═══════════════════════════════════════════════════
neda_weaknesses = [
    ["عکس حرفه‌ای", "ندارد", "🔴 فوری", "بارگذاری عکس حرفه‌ای با پس‌زمینه ساده — چهره واضح، لباس رسمی", "professional.linkedin.com/photo/upload"],
    ["Headline", "ساده و غیرجذاب", "🔴 فوری", "تغییر به: Registered Midwife | 12+ Years Clinical Experience | Open to International Opportunities", "linkedin.com/in/neda-arjmand/edit/topna"],
    ["About Section", "ندارد", "🔴 فوری", "نوشتن ۲۰۰-۳۰۰ کلمه درباره تجربه بالینی، مهارت‌ها، و هدف بین‌المللی", "linkedin.com/in/neda-arjmand/edit/about"],
    ["Languages", "فقط فارسی", "🟠 مهم", "اضافه کردن English (A2 — در حال آمادگی برای OET)", "linkedin.com/in/neda-arjmand/edit/languages"],
    ["Skills", "کمتر از ۵", "🟠 مهم", "اضافه کردن: Midwifery, Antenatal Care, Labour Management, Neonatal Care, Obstetric Emergency", "linkedin.com/in/neda-arjmand/edit/skills"],
    ["Experience Description", "خیلی کوتاه", "🟠 مهم", "توضیح جزئیات هر شغل: بخش زایمان، تعداد تولد، مسئولیت‌ها", "linkedin.com/in/neda-arjmand/edit/experience"],
    ["Certifications", "ندارد", "🟡 متوسط", "اضافه کردن مدارک آموزشی و گواهینامه‌ها", "linkedin.com/in/neda-arjmand/edit/certifications"],
    ["Recommendations", "ندارد", "🟡 متوسط", "درخواست توصیه‌نامه از همکاران و مدیران", "linkedin.com/in/neda-arjmand/edit/recommendations"],
    ["Open to Work", "فعال نیست", "🔴 فوری", "فعال کردن Open to Work برای: Midwife, Registered Midwife — International", "linkedin.com/in/neda-arjmand/edit/open-to-work"],
    ["Featured Section", "ندارد", "🟡 متوسط", "افزودن مقالات یا پروژه‌های مرتبط با مامایی", "linkedin.com/in/neda-arjmand/edit/featured"],
]

tohid_weaknesses = [
    ["عکس حرفه‌ای", "ندارد", "🔴 فوری", "بارگذاری عکس حرفه‌ای — ترجیحاً با پس‌زمینه فناوری یا اداری", "professional.linkedin.com/photo/upload"],
    ["Headline", "غیراستاندارد", "🔴 فوری", "تغییر به: IT Operations Manager | 19 Years Infrastructure & Systems Experience | Open to International Roles", "linkedin.com/in/tohid-arjmand/edit/topna"],
    ["About Section", "فنی و خشک", "🔴 فوری", "بازنویسی: ترکیب تجربه عملی + دستاوردها + هدف حرفه‌ای — ۲۰۰-۳۰۰ کلمه", "linkedin.com/in/tohid-arjmand/edit/about"],
    ["Recommendations", "ندارد", "🔴 فوری", "درخواست توصیه‌نامه از مدیران و همکاران فعلی/سابق", "linkedin.com/in/tohid-arjmand/edit/recommendations"],
    ["Open to Work", "فعال نیست", "🔴 فوری", "فعال کردن Open to Work برای: IT Manager, Infrastructure Manager — International", "linkedin.com/in/tohid-arjmand/edit/open-to-work"],
    ["Skills", "ناقص", "🟠 مهم", "اضافه کردن: VMware, Hyper-V, Cisco, MikroTik, Veeam, Windows Server, Active Directory", "linkedin.com/in/tohid-arjmand/edit/skills"],
    ["Experience Details", "فهرست‌وار", "🟠 مهم", "تبدیل فهرست به توضیحات: هر شغل شامل دستاوردها و ابعاد", "linkedin.com/in/tohid-arjmand/edit/experience"],
    ["Endorsements", "کم", "🟡 متوسط", "درخواست تأیید مهارت‌ها از همکاران", "linkedin.com/in/tohid-arjmand/edit/skills"],
    ["Featured Section", "ندارد", "🟡 متوسط", "افزودن پروژه‌های موفق IT", "linkedin.com/in/tohid-arjmand/edit/featured"],
    ["Volunteer Experience", "ندارد", "🟡 متوسط", "اضافه کردن هرگونه فعالیت داوطلبانه فناوری", "linkedin.com/in/tohid-arjmand/edit/volunteer"],
]

# ═══════════════════════════════════════════════════
# 3. SEARCH LINKS (comprehensive)
# ═══════════════════════════════════════════════════
search_links = [
    # NZ
    ["🇳🇿 نیوزیلند", "Health NZ International", "government", "https://www.healthnz.govt.nz/careers/international", "جستجوی رسمی بین‌المللی — ماما + IT", "✅ فعال"],
    ["🇳🇿 نیوزیلند", "RGH Global — Midwife Sponsorship", "recruiter", "https://www.rgh-global.com/jobs/midwife-with-sponsorship/", "حمایت مالی + حقوق 75-106K NZD", "✅ فعال"],
    ["🇳🇿 نیوزیلند", "Working In Health NZ", "recruiter", "https://workingin-health.co.nz/midwifery-jobs/", "آژانس استخدام بهداشت نیوزیلند", "✅ فعال"],
    ["🇳🇿 نیوزیلند", "NZ Seek — Midwife", "jobboard", "https://nz.seek.com/midwife-jobs", "207 آگهی ماما فعال", "✅ فعال"],
    ["🇳🇿 نیوزیلند", "NZ Seek — IT Manager Hotel", "jobboard", "https://nz.seek.com/it-manager-hotel-jobs", "170 آگهی IT هتل فعال", "✅ فعال"],
    ["🇳🇿 نیوزیلند", "Trade Me Jobs", "jobboard", "https://www.trademe.co.nz/a/jobs", "بزرگترین سایت کار نیوزیلند", "✅ فعال"],
    ["🇳🇿 نیوزیلند", "NZ Indeed — Visa Sponsorship", "jobboard", "https://nz.indeed.com/q-visa-sponsorship,-nurse-midwife-jobs.html", "277 آگهی ویزا حمایتی", "✅ فعال"],
    ["🇳🇿 نیوزیلند", "Kiwi Health Jobs", "government", "https://www.kiwihealthjobs.com/", "پلتفرم استخدام بهداشت نیوزیلند", "✅ فعال"],
    ["🇳🇿 نیوزیلند", "NZ Immigration — Green List", "government", "https://www.immigration.govt.nz/new-zealand-visas/preparing-a-visa-application/working-in-nz/green-list", "لیست سبز مشاغل مورد نیاز", "✅ فعال"],
    ["🇳🇿 نیوزیلند", "IHG Hotels NZ", "employer", "https://careers.ihg.com/en/locations/new-zealand/", "فرصت‌های هتل‌داری نیوزیلند", "✅ فعال"],
    # Canada
    ["🇨🇦 کانادا", "Alberta Health Services — International", "government", "https://www.albertahealthservices.ca/careers/Page12717.aspx", "استخدام بین‌المللی Albertا", "✅ فعال"],
    ["🇨🇦 کانادا", "Alberta Midwives — Job Board", "association", "https://www.alberta-midwives.ca/job-board", "تخت‌های مامایی Albertا", "✅ فعال"],
    ["🇨🇦 کانادا", "Alberta Midwives — International", "association", "https://www.albertamidwives.org/copy-of-interprovincial-applicants", "راهنمای مامای بین‌المللی", "✅ فعال"],
    ["🇨🇦 کانادا", "Saskatchewan Health Authority", "government", "https://www.saskhealthauthority.ca/careers-volunteering/careers", "استخدام بین‌المللی Saskatchewan", "✅ فعال"],
    ["🇨🇦 کانادا", "Job Bank Canada", "government", "https://www.jobbank.gc.ca/", "سایت رسمی کار کانادا", "✅ فعال"],
    ["🇨🇦 کانادا", "Indeed Canada — IT Sponsorship", "jobboard", "https://ca.indeed.com/q-it-infrastructure-manager-visa-sponsorship-jobs.html", "607 آگهی IT زیرساخت", "✅ فعال"],
    ["🇨🇦 کانادا", "Indeed Canada — Project Manager", "jobboard", "https://ca.indeed.com/q-visa-sponsorship-available,-project-manager-l-alberta-jobs.html", "850 آگهی پروژه منیجر", "✅ فعال"],
    ["🇨🇦 کانادا", "LinkedIn Canada — Visa Sponsorship", "linkedin", "https://ca.linkedin.com/jobs/visa-sponsorship-jobs", "8000+ آگهی حمایت ویزا", "⚠️ نیاز لاگین"],
    ["🇨🇦 کانادا", "Glassdoor Canada", "jobboard", "https://www.glassdoor.ca/Job/canada-project-manager-work-visa-sponsorship-jobs-SRCH_IL.0,6_IN3_KO7,44.htm", "188 آگهی پروژه منیجر", "✅ فعال"],
    ["🇨🇦 کانادا", "Workopolis", "jobboard", "https://www.workopolis.com/search?q=visa+sponsorship&l=canada", "525 آگهی حمایت ویزا", "✅ فعال"],
    ["🇨🇦 کانادا", "JobGlance Canada", "jobboard", "https://jobglance.app/jobs/visa-sponsorship/canada/", "آگهی‌های به‌روز روزانه", "✅ فعال"],
    ["🇨🇦 کانادا", "IFMOSA Work", "recruiter", "https://ifmosawork.com/canada-visa-sponsored-jobs/", "ایمیل و تلفن مستقیم کارفرماها", "✅ فعال"],
    # Australia
    ["🇦🇺 استرالیا", "AHPRA — IQNM Registration", "government", "https://www.nursingmidwiferyboard.gov.au/Accreditation/IQNM/Before-you-apply", "ثبت‌نام مامای بین‌المللی", "✅ فعال"],
    ["🇦🇺 استرالیا", "TalentQuarter — AHPRA Guide 2026", "recruiter", "https://www.talentquarter.com/your-complete-guide-to-ahpra-registration-in-2026", "راهنمای کامل ثبت‌نام 2026", "✅ فعال"],
    ["🇦🇺 استرالیا", "Indeed Australia — 482 Midwife", "jobboard", "https://au.indeed.com/q-482-visa-sponsorship,-registered-midwife-jobs.html", "آگهی‌های 482 ماما", "✅ فعال"],
    ["🇦🇺 استرالیا", "Seek Australia — Midwife FT", "jobboard", "https://au.seek.com/midwife-jobs/full-time", "262 آگهی م تمام‌وقت", "✅ فعال"],
    ["🇦🇺 استرالیا", "Skilled Nursing AU", "recruiter", "https://www.skillednursing.com.au/jobs/all-jobs/4/midwife/0/all-area-of-expertise/ALL/all-locations", "44 آگهی ماما استرالیا+NZ", "✅ فعال"],
    ["🇦🇺 استرالیا", "Seek Australia — Visa Sponsorship", "jobboard", "https://au.seek.com/visa-sponsorship-jobs/full-time", "377 آگهی حمایت ویزا", "✅ فعال"],
    ["🇦🇺 استرالیا", "Indeed Australia — 482 Visa", "jobboard", "https://au.indeed.com/q-australia-482-visa-sponsorship-jobs.html", "96 آگهی 482 فعال", "✅ فعال"],
    ["🇦🇺 استرالیا", "482Jobs.com", "specialist", "https://482jobs.com/", "سایت تخصصی آگهی‌های 482", "✅ فعال"],
    ["🇦🇺 استرالیا", "LinkedIn Australia — Visa", "linkedin", "https://au.linkedin.com/jobs/visa-sponsorship-jobs", "8000+ آگهی حمایت ویزا", "⚠️ نیاز لاگین"],
    ["🇦🇺 استرالیا", "HomeAffairs — Skilled List", "government", "https://www.homeaffairs.gov.au/trav/visa-1/189-", "لیست مشاغل ماهر", "✅ فعال"],
    ["🇦🇺 استرالیا", "Jora Australia", "jobboard", "https://au.jora.com/Visa-Sponsorship-Office-Manager-jobs-in-Australia", "102 آگهی مدیر دفتر", "✅ فعال"],
    # Germany
    ["🇩🇪 آلمان", "Make it in Germany", "government", "https://www.make-it-in-germany.com/en/", "سایت رسمی کار آلمان", "✅ فعال"],
    ["🇩🇪 آلمان", "TalentOrange", "recruiter", "https://www.talentorange.de/", "آژانس استخدام بین‌المللی", "✅ فعال"],
    ["🇩🇪 آلمان", "Holalemania", "employer", "https://www.holalemania.de/karriere", "کارفرمای آلمانی", "✅ فعال"],
    ["🇩🇪 آلمان", "StepStone Germany", "jobboard", "https://www.stepstone.de/", "بزرگترین سایت کار آلمان", "✅ فعال"],
    ["🇩🇪 آلمان", "Indeed Germany", "jobboard", "https://de.indeed.com/", "سایت کار آلمان", "✅ فعال"],
    ["🇩🇪 آلمان", "Bundesagentur für Arbeit", "government", "https://www.arbeitsagentur.de/", "آژانس کار فدرال آلمان", "✅ فعال"],
]

# ═══════════════════════════════════════════════════
# BUILD EXCEL
# ═══════════════════════════════════════════════════
wb = Workbook()

# ── SHEET 1: Master Dashboard ──
ws1 = wb.active
ws1.title = "داشبورد | Dashboard"
set_rtl(ws1)

ws1.merge_cells("A1:J1")
c = ws1["A1"]
c.value = "داشبورد جامع — Migration Hunter Complete Dashboard"
c.font = make_font(size=16, bold=True, color=WHITE)
c.fill = make_fill(DARK_BLUE)
c.alignment = center_align()

ws1.merge_cells("A2:J2")
ws1["A2"].value = f"آخرین بروزرسانی: {NOW}"
ws1["A2"].font = make_font(size=9)
ws1["A2"].alignment = rtl_align(h="right")

# Stats
stats = [
    ["ایمیل‌های آماده", "5", GREEN],
    ["نقاط ضعف ندا", "10", RED],
    ["نقاط ضعف توحید", "10", RED],
    ["لینک‌های جستجو", str(len(search_links)), MED_BLUE],
    ["کشورهای فعال", "4", PURPLE],
    ["سایت‌های فعال", "35+", YELLOW],
]

for i, (label, val, color) in enumerate(stats):
    r = 4 + i
    write_cell(ws1, r, 1, label, "fa", 11, bold=True, bg=LIGHT_GRAY)
    write_cell(ws1, r, 2, val, "en", 14, bold=True, color=color)
    ws1.merge_cells(f"C{r}:J{r}")
    ws1[f"C{r}"].value = ""
    ws1[f"C{r}"].border = thin

ws1.column_dimensions["A"].width = 25
ws1.column_dimensions["B"].width = 12
for c in range(3, 11):
    ws1.column_dimensions[get_column_letter(c)].width = 12

# ── SHEET 2: Emails ──
ws2 = wb.create_sheet("ایمیل‌ها | Emails")
set_rtl(ws2)

ws2.merge_cells("A1:F1")
ws2["A1"].value = "ایمیل‌های طبیعی آماده ارسال — Natural Emails (غیر AI)"
ws2["A1"].font = make_font(size=14, bold=True, color=WHITE)
ws2["A1"].fill = make_fill(GREEN)
ws2["A1"].alignment = center_align()

headers_email = ["#", "متقاضی", "کارفرما", "کشور", "موضوع", "وضعیت"]
for i, h in enumerate(headers_email, 1):
    ws2.cell(row=3, column=i, value=h)
write_header(ws2, 3, 6, GREEN)

email_list = [
    [1, "👩 ندا", "Health New Zealand", "🇳🇿", "Registered Midwife — International Candidate", "🟢 آماده"],
    [2, "👨 توحید", "Health New Zealand", "🇳🇿", "IT Operations Manager — International Candidate", "🟢 آماده"],
    [3, "👩 ندا", "RGH Global", "🇳🇿", "Midwife — Sponsorship Programme", "🟢 آماده"],
    [4, "👩 ندا", "Alberta Health Services", "🇨🇦", "Registered Midwife — International Application", "🟢 آماده"],
    [5, "👨 توحید", "Saskatchewan HA", "🇨🇦", "IT Operations Manager — International", "🟢 آماده"],
]

for i, row in enumerate(email_list):
    r = 4 + i
    for j, val in enumerate(row):
        ft = "en" if j in [0, 4] else "fa"
        write_cell(ws2, r, j + 1, val, ft, 10)

for c in range(1, 7):
    ws2.column_dimensions[get_column_letter(c)].width = [5, 15, 28, 8, 45, 12][c-1]

# ── SHEET 3: Email Text Samples ──
ws3 = wb.create_sheet("متن ایمیل | Email Text")
set_rtl(ws3)

ws3.merge_cells("A1:B1")
ws3["A1"].value = "متن کامل ایمیل‌ها — Full Email Text (کپی و ارسال کنید)"
ws3["A1"].font = make_font(size=14, bold=True, color=WHITE)
ws3["A1"].fill = make_fill(PURPLE)
ws3["A1"].alignment = center_align()

row = 3
for key, email in emails.items():
    ws3.merge_cells(f"A{row}:B{row}")
    ws3[f"A{row}"].value = f"📧 {key}"
    ws3[f"A{row}"].font = make_font(size=12, bold=True)
    ws3[f"A{row}"].fill = make_fill(LIGHT_BLUE)
    ws3[f"A{row}"].alignment = rtl_align(h="right")
    ws3[f"A{row}"].border = thin
    ws3[f"B{row}"].border = thin

    row += 1
    full_text = f"To: {email['to']}\nSubject: {email['subject']}\n\n{email['body']}"
    ws3.merge_cells(f"A{row}:B{row}")
    ws3[f"A{row}"].value = full_text
    ws3[f"A{row}"].font = make_font(name=FONT_EN, size=10)
    ws3[f"A{row}"].alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    ws3[f"A{row}"].border = thin
    ws3[f"B{row}"].border = thin
    ws3.row_dimensions[row].height = 280

    row += 2

ws3.column_dimensions["A"].width = 60
ws3.column_dimensions["B"].width = 60

# ── SHEET 4: LinkedIn Weaknesses — Neda ──
ws4 = wb.create_sheet("نقاط ضعف ندا")
set_rtl(ws4)

ws4.merge_cells("A1:F1")
ws4["A1"].value = "نقاط ضعف لینکدین ندا — LinkedIn Weaknesses"
ws4["A1"].font = make_font(size=14, bold=True, color=WHITE)
ws4["A1"].fill = make_fill(RED)
ws4["A1"].alignment = center_align()

headers_lw = ["#", "مورد", "وضعیت فعلی", "اولویت", "اقدام پیشنهادی", "لینک اصلاح"]
for i, h in enumerate(headers_lw, 1):
    ws4.cell(row=3, column=i, value=h)
write_header(ws4, 3, 6, RED)

for i, row_data in enumerate(neda_weaknesses):
    r = 4 + i
    for j, val in enumerate(row_data):
        ft = "en" if j in [0, 5] else "fa"
        bg = None
        if j == 2:
            if "فوری" in str(val):
                bg = LIGHT_RED
            elif "مهم" in str(val):
                bg = LIGHT_YELLOW
            else:
                bg = LIGHT_GRAY
        write_cell(ws4, r, j + 1, val, ft, 10, bg=bg)

for c in range(1, 7):
    ws4.column_dimensions[get_column_letter(c)].width = [5, 22, 22, 14, 50, 45][c-1]

# ── SHEET 5: LinkedIn Weaknesses — Tohid ──
ws5 = wb.create_sheet("نقاط ضعف توحید")
set_rtl(ws5)

ws5.merge_cells("A1:F1")
ws5["A1"].value = "نقاط ضعف لینکدین توحید — LinkedIn Weaknesses"
ws5["A1"].font = make_font(size=14, bold=True, color=WHITE)
ws5["A1"].fill = make_fill(RED)
ws5["A1"].alignment = center_align()

for i, h in enumerate(headers_lw, 1):
    ws5.cell(row=3, column=i, value=h)
write_header(ws5, 3, 6, RED)

for i, row_data in enumerate(tohid_weaknesses):
    r = 4 + i
    for j, val in enumerate(row_data):
        ft = "en" if j in [0, 5] else "fa"
        bg = None
        if j == 2:
            if "فوری" in str(val):
                bg = LIGHT_RED
            elif "مهم" in str(val):
                bg = LIGHT_YELLOW
            else:
                bg = LIGHT_GRAY
        write_cell(ws5, r, j + 1, val, ft, 10, bg=bg)

for c in range(1, 7):
    ws5.column_dimensions[get_column_letter(c)].width = [5, 22, 22, 14, 50, 45][c-1]

# ── SHEET 6: Search Links ──
ws6 = wb.create_sheet("لینک‌های جستجو")
set_rtl(ws6)

ws6.merge_cells("A1:F1")
ws6["A1"].value = "لینک‌های جستجوی فعال — Active Search Links"
ws6["A1"].font = make_font(size=14, bold=True, color=WHITE)
ws6["A1"].fill = make_fill(MED_BLUE)
ws6["A1"].alignment = center_align()

headers_sl = ["#", "کشور", "منبع", "نوع", "لینک", "توضیح", "وضعیت"]
ws6.cell(row=3, column=1, value="#")
for i, h in enumerate(headers_sl, 1):
    ws6.cell(row=3, column=i, value=h)
write_header(ws6, 3, 7, MED_BLUE)

for i, row_data in enumerate(search_links):
    r = 4 + i
    write_cell(ws6, r, 1, i + 1, "en", 10)
    for j, val in enumerate(row_data):
        ft = "en" if j in [3, 4] else "fa"
        bg = None
        if j == 5 and "فعال" in str(val):
            bg = LIGHT_GREEN
        elif j == 5 and "لاگین" in str(val):
            bg = LIGHT_YELLOW
        write_cell(ws6, r, j + 2, val, ft, 9, bg=bg)

for c in range(1, 8):
    ws6.column_dimensions[get_column_letter(c)].width = [5, 18, 32, 14, 55, 40, 14][c-1]

# ── SHEET 7: Country Comparison ──
ws7 = wb.create_sheet("مقایسه کشورها")
set_rtl(ws7)

ws7.merge_cells("A1:G1")
ws7["A1"].value = "مقایسه کشورها برای کاریابی — Country Comparison"
ws7["A1"].font = make_font(size=14, bold=True, color=WHITE)
ws7["A1"].fill = make_fill(PURPLE)
ws7["A1"].alignment = center_align()

headers_cc = ["کشور", "مامایی", "IT", "حمایت مالی", "ویزا", "زبان", "امتیاز"]
for i, h in enumerate(headers_cc, 1):
    ws7.cell(row=3, column=i, value=h)
write_header(ws7, 3, 7, PURPLE)

countries = [
    ["🇳🇿 نیوزیلند", "۹۰/۱۰۰ — نیاز شدید", "۶۵/۱۰۰ — متوسط", "✅ تأیید شده", "Green List", "IELTS 6.5-7", "85"],
    ["🇦🇺 استرالیا", "۸۵/۱۰۰ — نیاز بالا", "۷۰/۱۰۰ — خوب", "✅ AHPRA", "482 / 189 / 190", "IELTS 7", "80"],
    ["🇨🇦 کانادا", "۸۰/۱۰۰ — نیاز بالا", "۷۵/۱۰۰ — خوب", "⚠️ استانی", "Express Entry", "CLB 7", "78"],
    ["🇩🇪 آلمان", "۴۰/۱۰۰ — کم", "۸۰/۱۰۰ — خوب", "⚠️ بررسی", "EU Blue Card", "B2 آلمانی", "65"],
]

for i, row_data in enumerate(countries):
    r = 4 + i
    for j, val in enumerate(row_data):
        ft = "en" if j in [6] else "fa"
        write_cell(ws7, r, j + 1, val, ft, 10)

for c in range(1, 8):
    ws7.column_dimensions[get_column_letter(c)].width = [18, 22, 20, 18, 20, 16, 10][c-1]

# ── SHEET 8: Actions ──
ws8 = wb.create_sheet("اقدامات | Actions")
set_rtl(ws8)

ws8.merge_cells("A1:D1")
ws8["A1"].value = "اقدامات امروز — Today's Actions"
ws8["A1"].font = make_font(size=14, bold=True, color=WHITE)
ws8["A1"].fill = make_fill(YELLOW)
ws8["A1"].alignment = center_align()

headers_act = ["#", "اقدام", "متقاضی", "اولویت"]
for i, h in enumerate(headers_act, 1):
    ws8.cell(row=3, column=i, value=h)
write_header(ws8, 3, 4, YELLOW)

actions = [
    [1, "ارسال ایمیل Health NZ — ندا (مامایی)", "👩 ندا", "🔴 فوری"],
    [2, "ارسال ایمیل Health NZ — توحید (IT)", "👨 توحید", "🔴 فوری"],
    [3, "ارسال ایمیل RGH Global — ندا", "👩 ندا", "🔴 فوری"],
    [4, "اصلاح لینکدین ندا — عکس + Headline + About", "👩 ندا", "🔴 فوری"],
    [5, "اصلاح لینکدین توحید — Headline + About + Open to Work", "👨 توحید", "🔴 فوری"],
    [6, "فعال‌سازی Open to Work در لینکدین", "هر دو", "🔴 فوری"],
    [7, "ارسال ایمیل Alberta Health Services", "👩 ندا", "🟠 مهم"],
    [8, "ثبت‌نام در SEEK Australia", "هر دو", "🟠 مهم"],
    [9, "بررسی آگهی‌های جدید Job Bank Canada", "هر دو", "🟡 متوسط"],
    [10, "پیگیری ایمیل‌های ارسال شده قبلی", "هر دو", "🟡 متوسط"],
]

for i, row_data in enumerate(actions):
    r = 4 + i
    for j, val in enumerate(row_data):
        ft = "en" if j == 0 else "fa"
        bg = None
        if j == 3:
            if "فوری" in str(val):
                bg = LIGHT_RED
            elif "مهم" in str(val):
                bg = LIGHT_YELLOW
            else:
                bg = LIGHT_GRAY
        write_cell(ws8, r, j + 1, val, ft, 10, bg=bg)

for c in range(1, 5):
    ws8.column_dimensions[get_column_letter(c)].width = [5, 50, 15, 14][c-1]

# ═══════════════════════════════════════════════════
# SAVE
# ═══════════════════════════════════════════════════
os.makedirs("dashboard", exist_ok=True)
fname = f"dashboard/MigrationHunter_Complete_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
wb.save(fname)
print(f"\n✅ {fname}")
print(f"📊 ۸ شیت:")
print(f"   1. داشبورد | Dashboard")
print(f"   2. ایمیل‌ها | Emails")
print(f"   3. متن ایمیل | Email Text")
print(f"   4. نقاط ضعف ندا")
print(f"   5. نقاط ضعف توحید")
print(f"   6. لینک‌های جستجو ({len(search_links)} لینک)")
print(f"   7. مقایسه کشورها")
print(f"   8. اقدامات")
print(f"📧 ۵ ایمیل طبیعی ذخیره شد")
print(f"🔗 {len(search_links)} لینک فعال")
