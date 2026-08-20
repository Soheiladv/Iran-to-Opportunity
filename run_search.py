#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Migration Hunter — جستجوی پارامتری
انتخاب کشور + لینکدین + عناوین دوزبانه
"""
import os
import sys
import json
import argparse
from datetime import datetime
from pathlib import Path

BASE_DIR = Path(__file__).parent
NOW = datetime.now()
DATE_STR = NOW.strftime("%Y-%m-%d %H:%M")

# ════════════════════════════════════════════════════════════
# لیست کشورها
# ════════════════════════════════════════════════════════════

COUNTRIES = {
    # Tier 1
    "nz": {"en": "New Zealand", "fa": "نیوزیلند", "tier": 1},
    "au": {"en": "Australia", "fa": "استرالیا", "tier": 1},
    "de": {"en": "Germany", "fa": "آلمان", "tier": 1},
    "ca": {"en": "Canada", "fa": "کانادا", "tier": 1},
    
    # Tier 2 — اروپا
    "at": {"en": "Austria", "fa": "اتریش", "tier": 2},
    "ie": {"en": "Ireland", "fa": "ایرلند", "tier": 2},
    "nl": {"en": "Netherlands", "fa": "هلند", "tier": 2},
    "se": {"en": "Sweden", "fa": "سوئد", "tier": 2},
    "no": {"en": "Norway", "fa": "نروژ", "tier": 2},
    "dk": {"en": "Denmark", "fa": "دانمارک", "tier": 2},
    "fi": {"en": "Finland", "fa": "فنلاند", "tier": 2},
    "uk": {"en": "United Kingdom", "fa": "بریتانیا", "tier": 2},
    "ch": {"en": "Switzerland", "fa": "سوئیس", "tier": 2},
    "be": {"en": "Belgium", "fa": "بلژیک", "tier": 2},
    "pt": {"en": "Portugal", "fa": "پرتغال", "tier": 2},
    "es": {"en": "Spain", "fa": "اسپانیا", "tier": 2},
    "it": {"en": "Italy", "fa": "ایتالیا", "tier": 2},
    "fr": {"en": "France", "fa": "فرانسه", "tier": 2},
    "pl": {"en": "Poland", "fa": "لهستان", "tier": 2},
    "cz": {"en": "Czech Republic", "fa": "چک", "tier": 2},
    "hu": {"en": "Hungary", "fa": "مجارستان", "tier": 2},
    "ro": {"en": "Romania", "fa": "رومانی", "tier": 2},
    "hr": {"en": "Croatia", "fa": "کرواسی", "tier": 2},
}

# ════════════════════════════════════════════════════════════
# لینکدین‌ها
# ════════════════════════════════════════════════════════════

DEFAULT_LINKEDINS = {
    "neda": {
        "name": "Neda Arjmand",
        "url": "https://www.linkedin.com/in/neda-arjmand",
        "profession": "Midwife",
        "applicant": "NEDA"
    },
    "tohid": {
        "name": "Tohid Arjmand",
        "url": "https://www.linkedin.com/in/tohid-arjmand",
        "profession": "IT Operations Manager",
        "applicant": "TOHID"
    }
}

LINKEDIN_DB = BASE_DIR / "memory" / "LINKEDIN_DB.json"

def load_linkedin_db():
    if LINKEDIN_DB.exists():
        with open(LINKEDIN_DB, 'r', encoding='utf-8') as f:
            return json.load(f)
    data = {"linkedins": list(DEFAULT_LINKEDINS.values())}
    save_linkedin_db(data)
    return data

def save_linkedin_db(data):
    LINKEDIN_DB.parent.mkdir(exist_ok=True)
    with open(LINKEDIN_DB, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

def add_linkedin(name, url, profession, applicant):
    data = load_linkedin_db()
    # Check duplicate
    for li in data["linkedins"]:
        if li["url"] == url:
            print(f"⚠️ لینکدین تکراری: {url}")
            return
    entry = {
        "name": name,
        "url": url,
        "profession": profession,
        "applicant": applicant
    }
    data["linkedins"].append(entry)
    save_linkedin_db(data)
    print(f"✅ لینکدین اضافه شد: {name} ({url})")

def list_linkedin():
    data = load_linkedin_db()
    print("\n📋 لینکدین‌های ثبت شده:")
    print("-" * 50)
    for i, li in enumerate(data["linkedins"], 1):
        print(f"  {i}. {li['name']}")
        print(f"     URL: {li['url']}")
        print(f"     شغل: {li['profession']}")
        print(f"     متقاضی: {li['applicant']}")
        print()

# ════════════════════════════════════════════════════════════
# انتخاب کشور ( تعاملی )
# ════════════════════════════════════════════════════════════

def select_countries_interactive():
    """نمایش لیست کشور و دریافت انتخاب"""
    print("\n" + "=" * 60)
    print("🌍 انتخاب کشورهای جستجو")
    print("=" * 60)
    
    # Tier 1
    print("\n📌 Tier 1 (اولویت بالا):")
    tier1 = {k: v for k, v in COUNTRIES.items() if v["tier"] == 1}
    for code, info in tier1.items():
        print(f"  {code.upper():4s} — {info['fa']:12s} ({info['en']})")
    
    # Tier 2 — Europe
    print("\n📌 Tier 2 (اروپا):")
    tier2 = {k: v for k, v in COUNTRIES.items() if v["tier"] == 2}
    for code, info in tier2.items():
        print(f"  {code.upper():4s} — {info['fa']:12s} ({info['en']})")
    
    print("\n" + "-" * 60)
    print("گزینه‌ها:")
    print("  all     — جستجوی همه کشورها")
    print("  tier1   — فقط Tier 1 (nz, au, de, ca)")
    print("  europe  — فقط اروپا")
    print("  nz,de   — کشورهای انتخابی (با کاما)")
    print("-" * 60)
    
    choice = input("انتخاب شما: ").strip().lower()
    
    if choice == "all":
        selected = list(COUNTRIES.keys())
    elif choice == "tier1":
        selected = [k for k, v in COUNTRIES.items() if v["tier"] == 1]
    elif choice == "europe":
        selected = [k for k, v in COUNTRIES.items() if v["tier"] == 2]
    else:
        selected = [c.strip() for c in choice.split(",") if c.strip() in COUNTRIES]
    
    if not selected:
        print("❌ انتخاب نامعتبر — همه کشورها انتخاب شد")
        selected = list(COUNTRIES.keys())
    
    print(f"\n✅ کشورهای انتخاب شده: {len(selected)}")
    for code in selected:
        print(f"   • {COUNTRIES[code]['fa']} ({COUNTRIES[code]['en']})")
    
    return selected

def select_countries_args(country_arg):
    """انتخاب از آرگومان خط فرمان"""
    if not country_arg:
        return select_countries_interactive()
    
    if country_arg == "all":
        return list(COUNTRIES.keys())
    elif country_arg == "tier1":
        return [k for k, v in COUNTRIES.items() if v["tier"] == 1]
    elif country_arg == "europe":
        return [k for k, v in COUNTRIES.items() if v["tier"] == 2]
    else:
        selected = [c.strip() for c in country_arg.split(",") if c.strip() in COUNTRIES]
        if not selected:
            print("❌ کشور نامعتبر")
            return list(COUNTRIES.keys())
        return selected

# ════════════════════════════════════════════════════════════
# ساخت Excel با عناوین دوزبانه
# ════════════════════════════════════════════════════════════

def build_excel(selected_countries, linkedin_data):
    """ساخت Excel با شیت‌های دوزبانه"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("❌ نیاز به openpyxl: pip install openpyxl")
        return
    
    wb = Workbook()
    
    # فونت فارسی
    FA_FONT = Font(name='B Mitra', size=11)
    FA_TITLE = Font(name='B Mitra', size=14, bold=True)
    HEADER_FONT = Font(name='B Mitra', size=11, bold=True, color='FFFFFF')
    
    # رنگ‌ها
    GREEN = PatternFill(start_color='92D050', end_color='92D050', fill_type='solid')
    YELLOW = PatternFill(start_color='FFD966', end_color='FFD966', fill_type='solid')
    ORANGE = PatternFill(start_color='F4B084', end_color='F4B084', fill_type='solid')
    BLUE = PatternFill(start_color='9BC2E6', end_color='9BC2E6', fill_type='solid')
    HEADER_FILL = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    
    BORDER = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    def style_header(ws, row, cols):
        for col in range(1, cols + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = BORDER
    
    def style_cell(ws, row, col, font=FA_FONT):
        cell = ws.cell(row=row, column=col)
        cell.font = font
        cell.alignment = Alignment(vertical='center', wrap_text=True, horizontal='right')
        cell.border = BORDER
        return cell
    
    # ── شیت ۱: داشبورد / Dashboard ──
    ws1 = wb.active
    ws1.title = "📊 داشبورد | Dashboard"
    ws1.sheet_view.rightToLeft = True
    
    ws1.merge_cells('A1:F1')
    ws1.cell(row=1, column=1, value="📊 داشبورد اصلی | Main Dashboard").font = FA_TITLE
    
    ws1.cell(row=2, column=1, value=f"تاریخ بروزرسانی | Updated: {DATE_STR}").font = FA_FONT
    
    # جدول خلاصه
    headers1 = ["شاخص | Metric", "مقدار | Value"]
    for i, h in enumerate(headers1, 1):
        ws1.cell(row=4, column=i, value=h)
    style_header(ws1, 4, 2)
    
    summary = [
        ("کشورهای انتخاب شده | Selected Countries", len(selected_countries)),
        ("لینکدین‌های ثبت شده | Registered LinkedIn", len(linkedin_data["linkedins"])),
        ("کل منابع | Total Sources", 13),
        ("کل کارفرمایان | Total Employers", 7),
        ("کل فرصت‌ها | Total Jobs", 8),
        ("ایمیل‌های ارسالی | Emails Sent", 5),
        ("یادآوری‌ها | Reminders", 12),
    ]
    for r, (label, value) in enumerate(summary, 5):
        style_cell(ws1, r, 1).value = label
        style_cell(ws1, r, 2).value = value
    
    ws1.column_dimensions['A'].width = 40
    ws1.column_dimensions['B'].width = 20
    
    # ── شیت ۲: کشورهای انتخاب شده / Selected Countries ──
    ws2 = wb.create_sheet("🌍 کشورها | Countries")
    ws2.sheet_view.rightToLeft = True
    
    ws2.merge_cells('A1:D1')
    ws2.cell(row=1, column=1, value="🌍 کشورهای انتخاب شده برای جستجو | Selected Countries").font = FA_TITLE
    
    headers2 = ["کد | Code", "نام فارسی | Persian", "نام انگلیسی | English", "اولویت | Tier"]
    for i, h in enumerate(headers2, 1):
        ws2.cell(row=3, column=i, value=h)
    style_header(ws2, 3, 4)
    
    for r, code in enumerate(selected_countries, 4):
        info = COUNTRIES[code]
        style_cell(ws2, r, 1).value = code.upper()
        style_cell(ws2, r, 2).value = info["fa"]
        style_cell(ws2, r, 3).value = info["en"]
        c = style_cell(ws2, r, 4)
        c.value = f"Tier {info['tier']}"
        c.fill = GREEN if info["tier"] == 1 else BLUE
    
    for col in ['A', 'B', 'C', 'D']:
        ws2.column_dimensions[col].width = 25
    
    # ── شیت ۳: لینکدین / LinkedIn ──
    ws3 = wb.create_sheet("🔗 لینکدین | LinkedIn")
    ws3.sheet_view.rightToLeft = True
    
    ws3.merge_cells('A1:F1')
    ws3.cell(row=1, column=1, value="🔗 لینکدین‌های ثبت شده | Registered LinkedIn Profiles").font = FA_TITLE
    
    headers3 = ["ردیف | #", "نام | Name", "لینک | URL", "شغل | Profession", "متقاضی | Applicant", "وضعیت | Status"]
    for i, h in enumerate(headers3, 1):
        ws3.cell(row=3, column=i, value=h)
    style_header(ws3, 3, 6)
    
    for r, li in enumerate(linkedin_data["linkedins"], 4):
        style_cell(ws3, r, 1).value = r - 3
        style_cell(ws3, r, 2).value = li["name"]
        style_cell(ws3, r, 3).value = li["url"]
        style_cell(ws3, r, 4).value = li["profession"]
        style_cell(ws3, r, 5).value = li["applicant"]
        style_cell(ws3, r, 6).value = "🟢 فعال"
    
    for col in ['A', 'B', 'C', 'D', 'E', 'F']:
        ws3.column_dimensions[col].width = 25
    
    # ── شیت‌های لینکدین (هر لینکدین یک شیت) ──
    for li in linkedin_data["linkedins"]:
        short_name = li["name"].split()[0].lower()
        sheet_name = f"👤 {li['name'][:20]}"
        ws_li = wb.create_sheet(sheet_name)
        ws_li.sheet_view.rightToLeft = True
        
        ws_li.merge_cells('A1:D1')
        ws_li.cell(row=1, column=1, value=f"👤 پروفایل | Profile: {li['name']}").font = FA_TITLE
        
        fields = [
            ("نام | Name", li["name"]),
            ("لینک | URL", li["url"]),
            ("شغل | Profession", li["profession"]),
            ("متقاضی | Applicant", li["applicant"]),
            ("آخرین بررسی | Last Checked", DATE_STR),
            ("وضعیت | Status", "🟢 فعال"),
        ]
        for r, (label, value) in enumerate(fields, 3):
            style_cell(ws_li, r, 1).value = label
            style_cell(ws_li, r, 2).value = value
        
        ws_li.column_dimensions['A'].width = 30
        ws_li.column_dimensions['B'].width = 50
    
    # ── شیت ۴: فرصت‌های ندا / Neda Jobs ──
    ws4 = wb.create_sheet("👩 ندا | Neda Jobs")
    ws4.sheet_view.rightToLeft = True
    
    ws4.merge_cells('A1:H1')
    ws4.cell(row=1, column=1, value="👩 فرصت‌های مامایی ندا | Neda Midwifery Opportunities").font = FA_TITLE
    
    headers4 = ["ردیف | #", "کارفرما | Employer", "کشور | Country", "شغل | Job", "لینک | Link", "حمایت | Sponsor", "امتیاز | Score", "وضعیت | Status"]
    for i, h in enumerate(headers4, 1):
        ws4.cell(row=3, column=i, value=h)
    style_header(ws4, 3, 8)
    
    neda_jobs = [
        (1, "Health New Zealand", "nz", "Midwife", "https://www.healthnz.govt.nz", "✅", 85, "🟢"),
        (2, "RGH Global", "nz", "Midwife Sponsorship", "https://www.rgh-global.com", "✅", 79, "🟢"),
        (3, "Working In Health NZ", "nz", "International Midwife", "https://www.workingin-health.co.nz", "✅", 79, "🟢"),
        (4, "Holalemania GmbH", "de", "Geburtshelfer/in", "https://holalemania.de/en/", "✅", 82, "🟢"),
        (5, "Saskatchewan HA", "ca", "Midwife", "https://www.saskhealthauthority.ca", "⚠️", 78, "🟡"),
        (6, "Hays Healthcare", "au", "Midwife", "https://www.hays.com.au", "⚠️", 72, "🟡"),
        (7, "TalentOrange", "de", "Healthcare", "https://www.talentorange.com", "✅", 80, "🟢"),
    ]
    
    for r, (num, emp, ctry, job, url, sp, score, st) in enumerate(neda_jobs, 4):
        style_cell(ws4, r, 1).value = num
        style_cell(ws4, r, 2).value = emp
        style_cell(ws4, r, 3).value = COUNTRIES.get(ctry, {}).get("fa", ctry)
        style_cell(ws4, r, 4).value = job
        style_cell(ws4, r, 5).value = url
        style_cell(ws4, r, 6).value = sp
        c = style_cell(ws4, r, 7)
        c.value = score
        c.fill = GREEN if score >= 80 else YELLOW if score >= 70 else ORANGE
        style_cell(ws4, r, 8).value = st
    
    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']:
        ws4.column_dimensions[col].width = 22
    
    # ── شیت ۵: فرصت‌های توحید / Tohid Jobs ──
    ws5 = wb.create_sheet("👨 توحید | Tohid Jobs")
    ws5.sheet_view.rightToLeft = True
    
    ws5.merge_cells('A1:H1')
    ws5.cell(row=1, column=1, value="👨 فرصت‌های IT توحید | Tohid IT Opportunities").font = FA_TITLE
    
    headers5 = ["ردیف | #", "کارفرما | Employer", "کشور | Country", "شغل | Job", "لینک | Link", "حمایت | Sponsor", "امتیاز | Score", "وضعیت | Status"]
    for i, h in enumerate(headers5, 1):
        ws5.cell(row=3, column=i, value=h)
    style_header(ws5, 3, 8)
    
    tohid_jobs = [
        (1, "Arbeitnow", "de", "IT Manager", "https://www.arbeitnow.com", "⚠️", 65, "🟡"),
    ]
    
    for r, (num, emp, ctry, job, url, sp, score, st) in enumerate(tohid_jobs, 4):
        style_cell(ws5, r, 1).value = num
        style_cell(ws5, r, 2).value = emp
        style_cell(ws5, r, 3).value = COUNTRIES.get(ctry, {}).get("fa", ctry)
        style_cell(ws5, r, 4).value = job
        style_cell(ws5, r, 5).value = url
        style_cell(ws5, r, 6).value = sp
        c = style_cell(ws5, r, 7)
        c.value = score
        c.fill = GREEN if score >= 80 else YELLOW if score >= 70 else ORANGE
        style_cell(ws5, r, 8).value = st
    
    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']:
        ws5.column_dimensions[col].width = 22
    
    # ── شیت ۶: مقایسه / Comparison ──
    ws6 = wb.create_sheet("📈 مقایسه | Comparison")
    ws6.sheet_view.rightToLeft = True
    
    ws6.merge_cells('A1:E1')
    ws6.cell(row=1, column=1, value="📈 مقایسه کشورها | Country Comparison").font = FA_TITLE
    
    headers6 = ["کشور | Country", "امتیاز ندا | Neda Score", "امتیاز توحید | Tohid Score", "زبان | Language", "ویزا | Visa"]
    for i, h in enumerate(headers6, 1):
        ws6.cell(row=3, column=i, value=h)
    style_header(ws6, 3, 5)
    
    comparison = [
        ("nz", "نیوزیلند", 85, 60, "انگلیسی", "AEWV/Green List"),
        ("au", "استرالیا", 72, 55, "انگلیسی", "482/189/190"),
        ("de", "آلمان", 82, 65, "آلمانی B1-B2", "Work Visa/Blue Card"),
        ("ca", "کانادا", 78, 55, "CLB 7", "Express Entry/PNP"),
    ]
    
    for r, (code, fa_name, neda_s, tohid_s, lang, visa) in enumerate(comparison, 4):
        style_cell(ws6, r, 1).value = f"{fa_name} ({code.upper()})"
        c1 = style_cell(ws6, r, 2)
        c1.value = neda_s
        c1.fill = GREEN if neda_s >= 80 else YELLOW
        c2 = style_cell(ws6, r, 3)
        c2.value = tohid_s
        c2.fill = GREEN if tohid_s >= 80 else YELLOW
        style_cell(ws6, r, 4).value = lang
        style_cell(ws6, r, 5).value = visa
    
    for col in ['A', 'B', 'C', 'D', 'E']:
        ws6.column_dimensions[col].width = 25
    
    # ── ذخیره ──
    dashboard_dir = BASE_DIR / "dashboard"
    dashboard_dir.mkdir(exist_ok=True)
    filename = f"MigrationHunter_{NOW.strftime('%Y%m%d_%H%M')}.xlsx"
    filepath = dashboard_dir / filename
    wb.save(filepath)
    print(f"\n✅ داشبورد ایجاد شد: {filename}")
    print(f"   شیت‌ها: {len(wb.sheetnames)}")
    for s in wb.sheetnames:
        print(f"   • {s}")
    return filepath

# ════════════════════════════════════════════════════════════
# MAIN
# ════════════════════════════════════════════════════════════

def main():
    parser = argparse.ArgumentParser(description='Migration Hunter — جستجوی پارامتری')
    parser.add_argument('--country', type=str, help='کشورها: all, tier1, europe, یا nz,de,ca')
    parser.add_argument('--linkedin-add', action='store_true', help='اضافه کردن لینکدین جدید')
    parser.add_argument('--linkedin-list', action='store_true', help='لیست لینکدین‌ها')
    parser.add_argument('--excel', action='store_true', help='ساخت داشبورد Excel')
    parser.add_argument('--interactive', action='store_true', help='انتخاب تعاملی کشور')
    
    # LinkedIn params
    parser.add_argument('--li-name', type=str, help='نام لینکدین')
    parser.add_argument('--li-url', type=str, help='لینک لینکدین')
    parser.add_argument('--li-profession', type=str, help='شغل')
    parser.add_argument('--li-applicant', type=str, help='متقاضی (NEDA/TOHID)')
    
    args = parser.parse_args()
    
    print("=" * 60)
    print("🔍 Migration Hunter — جستجوی پارامتری")
    print(f"📅 {DATE_STR}")
    print("=" * 60)
    
    # LinkedIn operations
    if args.linkedin_add:
        if not all([args.li_name, args.li_url, args.li_profession, args.li_applicant]):
            print("❌ پارامترهای لینکدین ناقص:")
            print("   --li-name 'نام' --li-url 'URL' --li-profession 'شغل' --li-applicant 'NEDA/TOHID'")
            return
        add_linkedin(args.li_name, args.li_url, args.li_profession, args.li_applicant.upper())
        return
    
    if args.linkedin_list:
        list_linkedin()
        return
    
    # Country selection
    if args.interactive or not args.country:
        selected = select_countries_interactive()
    else:
        selected = select_countries_args(args.country)
    
    # Load LinkedIn data
    linkedin_data = load_linkedin_db()
    
    # Build Excel
    if args.excel or True:  # Always build Excel
        build_excel(selected, linkedin_data)
    
    print("\n" + "=" * 60)
    print("✅ عملیات تکمیل شد!")
    print("=" * 60)

if __name__ == "__main__":
    main()
