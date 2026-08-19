#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
اضافه کردن شیت وضعیت ایمیل و فرم به داشبورد
"""

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
import os

FONT = 'B Mitra'
BD = '1F4E79'
BM = '2E75B6'
GR = 'C6EFCE'
YL = 'FFEB9C'
RD = 'FFC7CE'
OR = 'FFD966'
BL = 'BDD7EE'
WH = 'FFFFFF'
GY = 'F2F2F2'

HF = Font(name=FONT, bold=True, size=11, color=WH)
TF = Font(name=FONT, size=16, bold=True, color=BD)
SF = Font(name=FONT, size=12, bold=True, color=BM)
CF = Font(name=FONT, size=10)
BF = Font(name=FONT, size=10, bold=True)
LF = Font(name=FONT, size=10, color='0563C1', underline='single')
STF = Font(name=FONT, size=9, italic=True, color='666666')

HFI = PatternFill(start_color=BD, end_color=BD, fill_type='solid')
GF = PatternFill(start_color=GR, end_color=GR, fill_type='solid')
YF = PatternFill(start_color=YL, end_color=YL, fill_type='solid')
OF = PatternFill(start_color=OR, end_color=OR, fill_type='solid')
BLF = PatternFill(start_color=BL, end_color=BL, fill_type='solid')
GYF = PatternFill(start_color=GY, end_color=GY, fill_type='solid')
RDF = PatternFill(start_color=RD, end_color=RD, fill_type='solid')

TB = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
CA = Alignment(horizontal='center', vertical='center', wrap_text=True, readingOrder=2)
RA = Alignment(horizontal='right', vertical='center', wrap_text=True, readingOrder=2)

NOW = datetime.now()
NOW_STR = NOW.strftime('%Y-%m-%d')
NOW_TIME = NOW.strftime('%H:%M:%S')
NOW_FULL = f"{NOW_STR} {NOW_TIME}"

def sc(cell, font=None, fill=None, align=None):
    if font is None: font = CF
    if align is None: align = RA
    cell.font = font; cell.alignment = align; cell.border = TB
    if fill: cell.fill = fill

def st(ws, row, end_col, title):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=end_col)
    c = ws.cell(row=row, column=1, value=title); c.font = TF; c.alignment = CA

def sh(ws, row, headers):
    for i, v in enumerate(headers, 1):
        c = ws.cell(row=row, column=i, value=v); sc(c, HF, HFI, CA)

def sd(ws, sr, data, cw=None):
    for r, rd in enumerate(data, sr):
        for ci, v in enumerate(rd, 1):
            c = ws.cell(row=r, column=ci, value=v)
            fl = None; sv = str(v)
            if sv in ['✅','Confirmed','آماده','جدید','P1','🔥 P1','آماده ارسال','تأیید','فوری','READY TO APPLY','ارسال شده','پاسخ دریافت','مصاحبه']:
                fl = GF
            elif sv in ['🟡','Likely','P2','🟢 P2','مهم','آماده نیست','نیاز','نیاز به بررسی','نیاز به ارتقاء','منتظر پاسخ','در انتظار']:
                fl = YF
            elif sv in ['🔵 P3','شناسایی','ممکن','NEW','ایمیل ارسال','فرم تکمیل']:
                fl = BLF
            elif sv in ['❌','REJECTED','EXPIRED','منقضی','رد شده']:
                fl = RDF
            elif sv in ['—','ندارد']:
                fl = GYF
            elif '🔗' in sv:
                fl = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
            sc(c, fill=fl)
    if cw:
        for i, w in enumerate(cw, 1): ws.column_dimensions[get_column_letter(i)].width = w

def main():
    # خواندن فایل فعلی
    base = os.path.dirname(os.path.abspath(__file__))
    excel_path = os.path.join(base, 'dashboard', f'MigrationHunter_Final_v2_{NOW.strftime("%Y%m%d_%H%M")}.xlsx')
    
    # اگر فایل جدید نبود، قدیمی‌ترین را باز کن
    if not os.path.exists(excel_path):
        dashboard_dir = os.path.join(base, 'dashboard')
        files = sorted([f for f in os.listdir(dashboard_dir) if f.startswith('MigrationHunter_Final_v2_') and f.endswith('.xlsx')])
        if files:
            excel_path = os.path.join(dashboard_dir, files[-1])
        else:
            print("❌ فایل Excel یافت نشد!")
            return
    
    print(f"📂 باز کردن: {excel_path}")
    wb = openpyxl.load_workbook(excel_path)
    
    # ========== شیت جدید: وضعیت ایمیل و فرم ==========
    ws_email = wb.create_sheet("📧 وضعیت ایمیل و فرم")
    ws_email.sheet_view.rightToLeft = True
    
    st(ws_email, 1, 14, f"وضعیت ایمیل‌ها و فرم‌ها | {NOW_FULL}")
    
    # ستون‌های وضعیت
    r = 3; sh(ws_email, r, [
        'ردیف', 'متقاضی', 'کارفرما', 'کشور', 'نوع ارسال',
        'تاریخ ارسال', 'ساعت ارسال', 'وضعیت ایمیل',
        'تاریخ پاسخ', 'وضعیت پاسخ', 'تاریخ مصاحبه',
        'وضعیت نهایی', 'یادداشت', '🕐 ثبت'
    ])
    
    sd(ws_email, r+1, [
        [1, 'ندا', 'Working In Health NZ', 'nz', 'ایمیل + فرم آنلاین',
         NOW_STR, NOW_TIME, 'ایمیل ارسال',
         '—', 'منتظر پاسخ', '—',
         'در انتظار', 'تکمیل فرم آنلاین اولویت دارد', NOW_TIME],
        [2, 'ندا', 'Health New Zealand', 'nz', 'ایمیل',
         NOW_STR, NOW_TIME, 'ایمیل ارسال',
         '—', 'منتظر پاسخ', '—',
         'در انتظار', 'سازمان دولتی — پاسخ کندتر', NOW_TIME],
        [3, 'ندا', 'RGH Global', 'nz', 'ایمیل',
         NOW_STR, NOW_TIME, 'ایمیل ارسال',
         '—', 'منتظر پاسخ', '—',
         'در انتظار', 'آژانس متخصص', NOW_TIME],
        [4, 'ندا', 'Holalemania', 'de', 'ایمیل + تلفن',
         NOW_STR, NOW_TIME, 'ایمیل ارسال',
         '—', 'منتظر پاسخ', '—',
         'در انتظار', 'تلفن: +49-40-41496505', NOW_TIME],
        [5, 'ندا', 'TalentOrange', 'de', 'فرم آنلاین',
         '—', '—', 'فرم تکمیل',
         '—', 'منتظر پاسخ', '—',
         'آماده نیست', 'نیاز به ایمیل جداگانه', NOW_TIME],
        [6, 'توحید', 'Arbeitnow', 'de', 'جستجوی وب',
         '—', '—', 'شناسایی',
         '—', '—', '—',
         'شناسایی', 'نیاز به ایمیل درخواست', NOW_TIME],
    ], [8,12,22,8,18,14,12,16,14,16,14,14,25,12])
    
    # ========== خلاصه وضعیت ==========
    r = 12; ws_email.cell(row=r, column=1, value="📊 خلاصه وضعیت").font = SF
    r = 13; sh(ws_email, r, ['وضعیت', 'تعداد', 'درصصد', 'توضیح'])
    sd(ws_email, r+1, [
        ['ایمیل ارسال شده', 4, '67%', 'در انتظار پاسخ'],
        ['فرم تکمیل شده', 1, '17%', 'نیاز به پیگیری'],
        ['شناسایی شده', 1, '17%', 'نیاز به ارسال'],
        ['پاسخ دریافت شده', 0, '0%', '—'],
        ['مصاحبه', 0, '0%', '—'],
        ['Job Offer', 0, '0%', '—'],
    ], [22,12,12,25])
    
    # ========== راهنمای پیگیری ==========
    r = 21; ws_email.cell(row=r, column=1, value="📋 راهنمای پیگیری").font = SF
    r = 22; sh(ws_email, r, ['وضعیت', 'اقدام بعدی', 'زمان پیگیری'])
    sd(ws_email, r+1, [
        ['ایمیل ارسال', 'منتظر ۵-۷ روز کاری', '۷ روز'],
        ['فرم تکمیل', 'پیگیری با ایمیل', '۳ روز'],
        ['شناسایی', 'ارسال ایمیل درخواست', 'فوری'],
        ['پاسخ دریافت', 'بررسی و پاسخ', '۲۴ ساعت'],
        ['مصاحبه', 'آماده شدن', 'بر اساس تاریخ'],
    ], [18,30,14])
    
    # ========== شیت وضعیت فرم‌ها ==========
    ws_form = wb.create_sheet("📋 وضعیت فرم‌ها")
    ws_form.sheet_view.rightToLeft = True
    
    st(ws_form, 1, 10, f"وضعیت فرم‌های آنلاین | {NOW_FULL}")
    
    r = 3; sh(ws_form, r, [
        'ردیف', 'کارفرما', 'نوع فرم', 'لینک فرم',
        'تاریخ تکمیل', 'وضعیت', 'تاریخ پاسخ', 'وضعیت نهایی', 'یادداشت', '🕐 ثبت'
    ])
    
    sd(ws_form, r+1, [
        [1, 'Working In Health NZ', 'فرم ثبت‌نام', '🔗',
         '—', 'آماده تکمیل', '—', 'در انتظار', 'اولویت اول — فرم آنلاین', NOW_TIME],
        [2, 'Health New Zealand', 'فرم درخواست', '🔗',
         '—', 'آماده تکمیل', '—', 'در انتظار', 'از طریق وب‌سایت', NOW_TIME],
        [3, 'Holalemania', 'فرم تماس', '🔗',
         '—', 'آماده تکمیل', '—', 'در انتظار', 'ایمیل + تلفن', NOW_TIME],
    ], [8,22,16,10,14,16,14,14,25,12])
    
    # لینک‌ها
    form_links = [
        'https://www.workingin-health.co.nz/midwifery-jobs/',
        'https://www.healthnz.govt.nz/careers/international',
        'https://holalemania.de/en/midwives/',
    ]
    for i, link in enumerate(form_links):
        c = ws_form.cell(row=4+i, column=4); c.hyperlink = link; c.font = LF; c.value = '🔗'
    
    # ========== ذخیره ==========
    out_path = os.path.join(base, 'dashboard', f'MigrationHunter_Final_v2_{NOW.strftime("%Y%m%d_%H%M")}.xlsx')
    wb.save(out_path)
    print(f"\n✅ شیت‌های جدید اضافه شد!")
    print(f"📊 کل شیت‌ها: {len(wb.sheetnames)}")
    for i, s in enumerate(wb.sheetnames, 1):
        print(f"   {i}. {s}")
    print(f"\n📁 فایل: {out_path}")

if __name__ == '__main__':
    main()
