#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Migration Hunter — DASHBOARD COMPREHENSIVE v2.1
آپدیت MigrationHunter_Final_v2.xlsx
فونت: B Mitra | جهت: RTL | رنگبندی: سبز/زرد/نارنجی/آبی/قرمز/خاکستری
ساعت ثبت: در تمام شیت‌ها
"""

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.utils import get_column_letter
from datetime import datetime

# ==========================================
# STYLES
# ==========================================

FONT = 'B Mitra'
BD = '1F4E79'
BM = '2E75B6'
GR = 'C6EFCE'
YL = 'FFEB9C'
RD = 'FFC7CE'
OR = 'FFD966'
BL = 'BDD7EE'
WH = 'FFFFFF'
GY = 'F2F2F2'
DK = 'D9E1F2'

# Fonts
HF = Font(name=FONT, bold=True, size=11, color=WH)
TF = Font(name=FONT, size=16, bold=True, color=BD)
SF = Font(name=FONT, size=12, bold=True, color=BM)
CF = Font(name=FONT, size=10)
BF = Font(name=FONT, size=10, bold=True)
LF = Font(name=FONT, size=10, color='0563C1', underline='single')
STF = Font(name=FONT, size=9, italic=True, color='666666')

# Fills
HFI = PatternFill(start_color=BD, end_color=BD, fill_type='solid')
GF = PatternFill(start_color=GR, end_color=GR, fill_type='solid')
YF = PatternFill(start_color=YL, end_color=YL, fill_type='solid')
OF = PatternFill(start_color=OR, end_color=OR, fill_type='solid')
BLF = PatternFill(start_color=BL, end_color=BL, fill_type='solid')
GYF = PatternFill(start_color=GY, end_color=GY, fill_type='solid')
DKF = PatternFill(start_color=DK, end_color=DK, fill_type='solid')
RDF = PatternFill(start_color=RD, end_color=RD, fill_type='solid')

# Border & Alignment
TB = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
CA = Alignment(horizontal='center', vertical='center', wrap_text=True, readingOrder=2)
RA = Alignment(horizontal='right', vertical='center', wrap_text=True, readingOrder=2)

NOW = datetime.now()
NOW_STR = NOW.strftime('%Y-%m-%d')
NOW_TIME = NOW.strftime('%H:%M:%S')
NOW_FULL = f"{NOW_STR} {NOW_TIME}"

# ==========================================
# HELPERS
# ==========================================

def sc(cell, font=None, fill=None, align=None):
    if font is None: font = CF
    if align is None: align = RA
    cell.font = font; cell.alignment = align; cell.border = TB
    if fill: cell.fill = fill

def st(ws, row, end_col, title):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=end_col)
    c = ws.cell(row=row, column=1, value=title); c.font = TF; c.alignment = CA

def sh(ws, row, headers):
    for i, v in enumerate(headers, 1):
        c = ws.cell(row=row, column=i, value=v); sc(c, HF, HFI, CA)

def sd(ws, sr, data, cw=None):
    for r, rd in enumerate(data, sr):
        for ci, v in enumerate(rd, 1):
            c = ws.cell(row=r, column=ci, value=v)
            fl = None; sv = str(v)
            if sv in ['✅','Confirmed','آماده','جدید','P1','🔥 P1','آماده ارسال','تأیید','فوری','READY TO APPLY']: fl = GF
            elif sv in ['🟡','Likely','P2','🟢 P2','مهم','آماده نیست','نیاز','نیاز به بررسی','نیاز به ارتقاء']: fl = YF
            elif sv in ['🔵 P3','شناسایی','ممکن','NEW']: fl = BLF
            elif sv in ['❌','REJECTED','EXPIRED','منقضی']: fl = RDF
            elif sv in ['—','ندارد']: fl = GYF
            elif '🔗' in sv: fl = DKF
            sc(c, fill=fl)
    if cw:
        for i, w in enumerate(cw, 1): ws.column_dimensions[get_column_letter(i)].width = w

def add_ts(ws, row, col=1):
    """اضافه کردن ساعت ثبت"""
    c = ws.cell(row=row, column=col, value=f"📅 {NOW_FULL}")
    c.font = STF

def add_chart(ws, title, data_col, data_start, data_end, cat_col, cat_start, cat_end, loc, ctype="bar"):
    ch = BarChart() if ctype == "bar" else PieChart()
    if ctype == "bar": ch.type = "col"; ch.title = title; ch.y_axis.title = "امتیاز"; ch.style = 10
    else: ch.title = title; ch.style = 10
    d = Reference(ws, min_col=data_col, min_row=data_start, max_row=data_end)
    c = Reference(ws, min_col=cat_col, min_row=cat_start, max_row=cat_end)
    ch.add_data(d, titles_from_data=False); ch.set_categories(c)
    ws.add_chart(ch, loc)

# ==========================================
# MAIN
# ==========================================

def main():
    wb = openpyxl.Workbook()
    if 'Sheet' in wb.sheetnames: del wb['Sheet']

    # ========== 1. شیت مستر ==========
    ws = wb.create_sheet("📊 شیت مستر", 0); ws.sheet_view.rightToLeft = True
    st(ws, 1, 14, f"گزارش جامع شکار فرصت شغلی — {NOW_FULL}")
    st(ws, 2, 14, "خانواده آرجمند | انگلیسی A2 + آلمانی A1 | جستجو بدون توقف")
    add_ts(ws, 3, 1)

    r = 5; ws.cell(row=r, column=1, value="📊 خلاصه وضعیت کلی").font = SF
    r = 6; sh(ws, r, ['شاخص', 'ندا', 'توحید', 'جمع'])
    sd(ws, r+1, [
        ['تعداد فرصت‌ها', 5, 5, 10],
        ['Sponsorship تأیید', 5, 0, 5],
        ['Sponsorship احتمالی', 0, 5, 5],
        ['ایمیل آماده', 5, 0, 5],
        ['کاور لیتر آماده', 5, 0, 5],
        ['امتیاز تناسب', '85/100', '75/100', '80/100'],
        ['آخرین جستجو', NOW_STR, NOW_STR, NOW_STR],
    ], [25, 18, 18, 18])

    r = 14; ws.cell(row=r, column=1, value="🌍 مقایسه کشورها").font = SF
    r = 15; sh(ws, r, ['کشور', 'امتیاز', 'ندا', 'توحید', 'Sponsorship', 'زبان', 'سرعت', 'شانس'])
    sd(ws, r+1, [
        ['nz', 85, 'عالی', 'خوب', '۱۰۰%', 'IELTS/OET', 'سریع', 85],
        ['au', 78, 'خوب', 'متوسط', '۹۰%', 'IELTS/OET', 'متوسط', 78],
        ['de', 78, 'عالی', 'خوب', '۹۰%', 'A1-A2', 'متوسط', 78],
        ['ca', 60, 'محدود', 'متوسط', '۶۰%', 'IELTS 6.5', 'طولانی', 60],
    ], [12, 12, 12, 12, 12, 15, 12, 12])

    r = 21; ws.cell(row=r, column=1, value="📈 امتیاز تناسب مسیر").font = SF
    r = 22; sh(ws, r, ['معیار', 'ندا-nz', 'ندا-de', 'توحید-de', 'ندا-au'])
    sd(ws, r+1, [
        ['تناسب حرفه‌ای', 85, 80, 78, 80],
        ['تناسب مهاجرتی', 80, 78, 80, 75],
        ['تناسب زبان', 60, 70, 72, 60],
        ['تناسب حمایت', 95, 85, 75, 85],
        ['⭐ کل', 85, 78, 75, 76],
    ], [22, 15, 15, 15, 15])
    add_chart(ws, "مقایسه امتیاز", 2, 23, 27, 1, 23, 27, "A29")

    # ========== 2. ندا ==========
    ws2 = wb.create_sheet("👩 ندا — فرصت‌ها"); ws2.sheet_view.rightToLeft = True
    st(ws2, 1, 16, f"فرصت‌های ندا — ماما (Midwife) | {NOW_FULL}")
    add_ts(ws2, 2, 1)

    r = 4; ws2.cell(row=r, column=1, value="📋 مشخصات").font = SF
    sd(ws2, r+1, [
        ['نام', 'ندا آرجمند'],['سن', '۳۸ سال'],['حرفه', 'ماما (Midwife)'],
        ['سابقه', 'بیمارستان میلاد تهران'],['زبان', 'A2 English + A1 German'],
        ['LinkedIn', 'linkedin.com/in/neda-arjmand'],
    ], [22, 30])

    r = 12; ws2.cell(row=r, column=1, value="🎯 فرصت‌های شغلی").font = SF
    r = 13; sh(ws2, r, ['اولویت','کارفرما','کشور','عنوان','حقوق','Sponsorship','زبان','ثبت‌نام','تناسب','🔗 لینک','📧 ایمیل','📝 کاور','🕐 ثبت','اقدام'])
    sd(ws2, r+1, [
        ['🔥 P1','Working In Health NZ','nz','ماما','75-106K NZD','✅','IELTS/OET','کمک ثبت‌نام','85/100','🔗','✅','✅',NOW_TIME,'ارسال'],
        ['🔥 P1','Health New Zealand','nz','ماما','75-106K NZD','✅','IELTS/OET','Midwifery Council','82/100','🔗','✅','✅',NOW_TIME,'ارسال'],
        ['🔥 P1','RGH Global','nz','ماما','75-106K NZD','✅','IELTS/OET','Midwifery Council','80/100','🔗','✅','✅',NOW_TIME,'ارسال'],
        ['🔥 P1','Holalemania','de','Hebamme','€35-50K','✅','A1-A2 آلمانی','Anerkennung','78/100','🔗','✅','✅',NOW_TIME,'ارسال'],
        ['🟢 P2','TalentOrange','de','Hebamme','—','✅','بورسیه B2','کمک ثبت‌نام','75/100','🔗','🟡','🟡',NOW_TIME,'بررسی'],
    ], [12,22,8,16,16,12,14,18,10,10,10,10,12,12])

    links_n = ['https://www.workingin-health.co.nz/midwifery-jobs/','https://www.healthnz.govt.nz/careers/international','https://www.rgh-global.com/jobs/midwife-with-sponsorship/','https://holalemania.de/en/midwives/','https://www.talentorange.com/en/']
    for i, l in enumerate(links_n):
        c = ws2.cell(row=14+i, column=10); c.hyperlink = l; c.font = LF; c.value = '🔗'
    add_chart(ws2, "تناسب فرصت‌ها", 9, 9, 18, 2, 14, 18, "A20")

    # ========== 3. توحید ==========
    ws3 = wb.create_sheet("👨 توحید — فرصت‌ها"); ws3.sheet_view.rightToLeft = True
    st(ws3, 1, 16, f"فرصت‌های توحید — مدیر IT | {NOW_FULL}")
    add_ts(ws3, 2, 1)

    r = 4; ws3.cell(row=r, column=1, value="📋 مشخصات").font = SF
    sd(ws3, r+1, [
        ['نام','توحید آرجمند'],['سن','۴۶ سال'],['حرفه','مدیر IT'],
        ['سابقه','۱۹ سال تجربه IT'],['زبان','A2 English + A1 German'],
        ['LinkedIn','linkedin.com/in/tohid-arjmand'],
    ], [22, 30])

    r = 12; ws3.cell(row=r, column=1, value="🎯 فرصت‌های شغلی").font = SF
    r = 13; sh(ws3, r, ['اولویت','کارفرما','کشور','عنوان','حقوق','Sponsorship','زبان','ثبت‌نام','تناسب','🔗 لینک','📧 ایمیل','📝 کاور','🕐 ثبت','اقدام'])
    sd(ws3, r+1, [
        ['🟢 P2','Arbeitnow (IT)','de','IT Manager','€45-80K','🟢','English','—','75/100','🔗','🟡','🟡',NOW_TIME,'جستجو'],
        ['🟢 P2','SAP','de','IT Manager','€55-80K','🟢','English','—','82/100','🔗','🟡','🟡',NOW_TIME,'جستجو'],
        ['🟢 P2','Siemens','de','IT Infrastructure','€60-90K','🟢','English','—','80/100','🔗','🟡','🟡',NOW_TIME,'جستجو'],
        ['🟢 P2','Deutsche Telekom','de','ICT Manager','€50-75K','🟢','English','—','78/100','🔗','🟡','🟡',NOW_TIME,'جستجو'],
        ['🟢 P2','Allianz','de','IT Operations','€55-85K','🟢','English','—','77/100','🔗','🟡','🟡',NOW_TIME,'جستجو'],
    ], [12,22,8,18,14,12,12,10,10,10,10,10,12,12])

    links_t = ['https://www.arbeitnow.com/visa-sponsorship-jobs','https://www.sap.com/about/careers.html','https://www.siemens.com/global/en/company/jobs.html','https://www.telekom.com/en/careers','https://www.allianz.com/en/careers.html']
    for i, l in enumerate(links_t):
        c = ws3.cell(row=14+i, column=10); c.hyperlink = l; c.font = LF; c.value = '🔗'
    add_chart(ws3, "تناسب فرصت‌ها", 9, 9, 18, 2, 14, 18, "A20")

    # ========== 4. مقایسه ==========
    ws4 = wb.create_sheet("🌍 مقایسه کشورها"); ws4.sheet_view.rightToLeft = True
    st(ws4, 1, 10, f"مقایسه جامع کشورها | {NOW_FULL}")
    add_ts(ws4, 2, 1)
    r = 3; sh(ws4, r, ['کشور','امتیاز','ندا','توحید','سرعت','هزینه','زبان','خانواده','شانس'])
    sd(ws4, r+1, [
        ['nz',85,83,55,85,70,70,95,85],['au',78,76,50,70,60,70,90,78],
        ['de',78,78,72,65,75,72,85,78],['ca',60,25,55,55,65,65,85,60],
    ], [12,12,12,12,12,12,12,12,12])
    add_chart(ws4, "مقایسه امتیاز", 9, 9, 12, 1, 4, 7, "A9")
    r = 10; ws4.cell(row=r, column=1, value="📈 سناریوها").font = SF
    r = 11; sh(ws4, r, ['سناریو','توضیح','nz','au','de','ca','امتیاز'])
    sd(ws4, r+1, [
        ['A: ندا اصلی','ندا ماما پیدا کند',83,76,78,25,82],
        ['B: توحید اصلی','توحید IT پیدا کند',55,50,72,55,72],
        ['C: هر دو مستقل','هر دو جداگانه',80,70,75,50,80],
    ], [18,25,12,12,12,12,12])
    add_chart(ws4, "توزیع سناریوها", 2, 2, 5, 1, 12, 14, "A17", "pie")

    # ========== 5. کارفرمایان ==========
    ws5 = wb.create_sheet("🏢 کارفرمایان"); ws5.sheet_view.rightToLeft = True
    st(ws5, 1, 12, f"کارفرمایان شناسایی شده | {NOW_FULL}")
    add_ts(ws5, 2, 1)
    r = 3; sh(ws5, r, ['نام','کشور','نوع','تخصص','Sponsorship','وب‌سایت','适合','وضعیت','ایمیل','تلفن','امتیاز','🕐 ثبت'])
    sd(ws5, r+1, [
        ['Health New Zealand','nz','دولتی','بهداشت','✅','healthnz.govt.nz','ندا','تأیید','—','—',95,NOW_TIME],
        ['Working In Health NZ','nz','آژانس','مامایی','✅','workingin-health.co.nz','ندا','تأیید','—','—',88,NOW_TIME],
        ['RGH Global','nz','آژانس','مامایی','✅','rgh-global.com','ندا','تأیید','—','—',85,NOW_TIME],
        ['Holalemania GmbH','de','آژانس','مامایی','✅','holalemania.de','ندا','تأیید','info@holalemania.de','+49-40-41496505',85,NOW_TIME],
        ['TalentOrange','de','آژانس','بهداشت','✅','talentorange.com','ندا','تأیید','—','—',82,NOW_TIME],
        ['Hassett Group','au','آژانس','بهداشت','✅','hassett.com.au','ندا','تأیید','—','—',78,NOW_TIME],
        ['SAP','de','شرکت IT','IT','🟢','sap.com','توحید','بررسی','—','—',82,NOW_TIME],
        ['Siemens','de','شرکت IT','IT','🟢','siemens.com','توحید','بررسی','—','—',80,NOW_TIME],
    ], [22,8,10,12,12,24,10,10,24,18,10,12])

    # ========== 6. سازمان‌ها ==========
    ws6 = wb.create_sheet("🏛 سازمان‌ها"); ws6.sheet_view.rightToLeft = True
    st(ws6, 1, 8, f"سازمان‌های دولتی | {NOW_FULL}")
    add_ts(ws6, 2, 1)
    r = 3; sh(ws6, r, ['کشور','سازمان','وب‌سایت','مناسب برای','موضوع','اهمیت','🕐 ثبت'])
    sd(ws6, r+1, [
        ['nz','Immigration NZ','immigration.govt.nz','هر دو','ویزا','بالا',NOW_TIME],
        ['nz','Midwifery Council','midwiferycouncil.health.nz','ندا','ثبت‌نام','بالا',NOW_TIME],
        ['nz','Health NZ','healthnz.govt.nz','ندا','استخدام','بالا',NOW_TIME],
        ['au','Home Affairs','homeaffairs.gov.au','هر دو','ویزا','بالا',NOW_TIME],
        ['au','AHPRA','ahpra.gov.au','ندا','ثبت‌نام','بالا',NOW_TIME],
        ['de','Make it in Germany','make-it-in-germany.com','هر دو','مهاجرت','بالا',NOW_TIME],
        ['de','Bundesagentur','arbeitsagentur.de','توحید','بازار کار','متوسط',NOW_TIME],
    ], [8,22,28,15,15,12,12])

    # ========== 7. درخواست‌ها ==========
    ws7 = wb.create_sheet("📤 درخواست‌ها"); ws7.sheet_view.rightToLeft = True
    st(ws7, 1, 12, f"درخواست‌های آماده ارسال | {NOW_FULL}")
    add_ts(ws7, 2, 1)
    r = 3; sh(ws7, r, ['App ID','متقاضی','کارفرما','کشور','CV','کاور','ایمیل','وضعیت','پاسخ','تاریخ ارسال','🕐 ثبت','اقدام'])
    sd(ws7, r+1, [
        ['APP-001','ندا','Working In Health NZ','nz','✅','✅','✅','آماده ارسال','—','—',NOW_TIME,'تأیید کاربر'],
        ['APP-002','ندا','Health New Zealand','nz','✅','✅','✅','آماده ارسال','—','—',NOW_TIME,'تأیید کاربر'],
        ['APP-003','ندا','RGH Global','nz','✅','✅','✅','آماده ارسال','—','—',NOW_TIME,'تأیید کاربر'],
        ['APP-004','ندا','Holalemania','de','✅','✅','✅','آماده ارسال','—','—',NOW_TIME,'تأیید کاربر'],
        ['APP-005','ندا','TalentOrange','de','✅','🟡','🟡','آماده نیست','—','—',NOW_TIME,'تکمیل'],
        ['APP-006','توحید','SAP','de','✅','🟡','🟡','آماده نیست','—','—',NOW_TIME,'تکمیل'],
    ], [12,12,22,8,8,10,10,16,10,14,12,16])

    # ========== 8. ویزا ==========
    ws8 = wb.create_sheet("🛂 ویزا"); ws8.sheet_view.rightToLeft = True
    st(ws8, 1, 10, f"اطلاعات ویزا — ۴ کشور | {NOW_FULL}")
    add_ts(ws8, 2, 1)
    r = 3; sh(ws8, r, ['کشور','نوع ویزا','الزام شغلی','الزام زبان','خانواده','ثبت‌نام','هزینه','زمان','منبع','🕐 ثبت'])
    sd(ws8, r+1, [
        ['nz','AEWV','Job offer','ANZSCO 1-2: None','✅','جداگانه','—','۴-۸ هفته','immigration.govt.nz',NOW_TIME],
        ['au','482 TSS','Job offer','IELTS 5.0','✅','AHPRA','—','۳-۶ ماه','homeaffairs.gov.au',NOW_TIME],
        ['de','EU Blue Card','€45,934+','None for IT','✅','Anerkennung','—','۳-۶ ماه','make-it-in-germany.com',NOW_TIME],
        ['ca','LMIA + WP','Job offer','IELTS 6.5','✅','مختلف','$1,000+','۶-۱۲ ماه','canada.ca',NOW_TIME],
    ], [8,14,18,18,10,14,12,14,24,12])

    # ========== 9. ثبت‌نام ==========
    ws9 = wb.create_sheet("📋 ثبت‌نام"); ws9.sheet_view.rightToLeft = True
    st(ws9, 1, 10, f"ثبت‌نام حرفه‌ای | {NOW_FULL}")
    add_ts(ws9, 2, 1)
    r = 3; sh(ws9, r, ['کشور','متقاضی','حرفه','نهاد','زبان','آزمون','هزینه','زمان','منبع','وضعیت'])
    sd(ws9, r+1, [
        ['nz','ندا','ماما','Midwifery Council','IELTS 7.0 / OET B','ممکن','NZ$485','۱۲ هفته','midwiferycouncil.health.nz','نیاز'],
        ['au','ندا','ماما','AHPRA','IELTS 7.0 / OET B','ممکن','—','—','ahpra.gov.au','نیاز'],
        ['de','ندا','ماما','Anerkennung','B2 آلمانی','ممکن','—','۶-۱۲ ماه','anerkennung-in-deutschland.de','نیاز'],
        ['de','توحید','IT','IHK Chamber','ندارد','—','—','—','make-it-in-germany.com','بررسی'],
    ], [8,10,10,20,18,12,10,12,28,10])

    # ========== 10. زبان ==========
    ws10 = wb.create_sheet("📝 زبان"); ws10.sheet_view.rightToLeft = True
    st(ws10, 1, 10, f"وضعیت زبان — جستجو بدون توقف | {NOW_FULL}")
    add_ts(ws10, 2, 1)
    r = 3; sh(ws10, r, ['متقاضی','English','German','IELTS نیاز','OET نیاز','وضعیت','اقدام','اولویت','زمان','یادداشت'])
    sd(ws10, r+1, [
        ['توحید','A2','A1','نیاز نیست','نیاز نیست','کافی برای IT','ادامه A2','مهم','—','بسیاری شرکت‌ها English کافی است'],
        ['ندا','A2','A1','IELTS Academic 7.0','OET B','نیاز به ارتقاء','ثبت‌نام OET','فوری','۶-۱۲ ماه','فقط برای ثبت‌نام حرفه‌ای'],
    ], [14,14,14,20,14,18,18,12,14,35])

    # ========== 11. تاریخچه ==========
    ws11 = wb.create_sheet("📜 تاریخچه"); ws11.sheet_view.rightToLeft = True
    st(ws11, 1, 12, f"تاریخچه جستجوها | {NOW_FULL}")
    add_ts(ws11, 2, 1)
    r = 3; sh(ws11, r, ['تاریخ','ساعت','متقاضی','کشورها','منابع','جدید','کل','معتبر','درخواست','پاسخ','مصاحبه','Offer'])
    sd(ws11, r+1, [
        [NOW_STR,NOW_TIME,'هر دو','nz,au,de','8','55','55','7','0','0','0','0'],
    ], [14,12,12,14,10,10,10,10,12,10,10,10])

    # ========== 12. خانواده ==========
    ws12 = wb.create_sheet("👨‍👩‍👧‍👦 خانواده"); ws12.sheet_view.rightToLeft = True
    st(ws12, 1, 10, f"استراتژی خروج خانواده | {NOW_FULL}")
    add_ts(ws12, 2, 1)
    r = 3; sh(ws12, r, ['سناریو','متقاضی','کشور','تناسب','سرعت','هزینه','خانواده','امتیاز','توصیه'])
    sd(ws12, r+1, [
        ['A: ندا اصلی','ندا','nz',80,'سریع','متوسط','عالی',82,'✅ پیشنهاد اول'],
        ['B: توحید اصلی','توحید','de',76,'متوسط','کم','خوب',72,'🟢 گزینه دوم'],
        ['C: هر دو مستقل','هر دو','nz+de',78,'سریع','متوسط','عالی',80,'🟢 قوی‌ترین'],
    ], [18,12,12,12,12,12,12,12,18])

    # ========== 13. اقدامات ==========
    ws13 = wb.create_sheet("🎯 اقدامات"); ws13.sheet_view.rightToLeft = True
    st(ws13, 1, 8, f"برنامه اقدام امروز | {NOW_FULL}")
    add_ts(ws13, 2, 1)
    r = 3; sh(ws13, r, ['ردیف','اقدام','متقاضی','کشور','اولویت','وضعیت','🕐 ثبت','یادداشت'])
    sd(ws13, r+1, [
        [1,'تکمیل فرم Working In Health NZ','ندا','nz','🔴 فوری','انجام نشده',NOW_TIME,'بهترین فرصت — خدمات رایگان'],
        [2,'ایمیل Holalemania','ندا','de','🔴 فوری','انجام نشده',NOW_TIME,'آموزش زبان آلمانی ارائه می‌دهد'],
        [3,'ارسال CV Health New Zealand','ندا','nz','🔴 فوری','انجام نشده',NOW_TIME,'سازمان دولتی — معتبر'],
        [4,'ارسال CV RGH Global','ندا','nz','🟠 مهم','انجام نشده',NOW_TIME,'آژانس متخصص مامایی'],
        [5,'جستجوی دقیق IT آلمان','توحید','de','🟡 متوسط','انجام نشده',NOW_TIME,'بازار خوب برای IT'],
    ], [8,32,12,8,15,15,12,30])

    # ========== 14. ایمیل‌ها ==========
    ws14 = wb.create_sheet("📧 ایمیل‌ها"); ws14.sheet_view.rightToLeft = True
    st(ws14, 1, 10, f"ایمیل‌های درخواست | {NOW_FULL}")
    add_ts(ws14, 2, 1)
    r = 3; sh(ws14, r, ['ردیف','متقاضی','کارفرما','موضوع','وضعیت','لینک','تاریخ','🕐 ثبت','اقدام'])
    sd(ws14, r+1, [
        [1,'ندا','Working In Health NZ','Midwife — Full Service','✅ آماده','🔗',NOW_STR,NOW_TIME,'ارسال'],
        [2,'ندا','Health New Zealand','International Midwife','✅ آماده','🔗',NOW_STR,NOW_TIME,'ارسال'],
        [3,'ندا','RGH Global','Midwife with Sponsorship','✅ آماده','🔗',NOW_STR,NOW_TIME,'ارسال'],
        [4,'ندا','Holalemania','Hebamme — Visa Sponsorship','✅ آماده','🔗',NOW_STR,NOW_TIME,'ارسال'],
        [5,'ندا','TalentOrange','Hebamme — B2 Scholarship','✅ آماده','🔗',NOW_STR,NOW_TIME,'ارسال'],
    ], [8,12,22,30,12,10,14,12,12])

    # ========== 15. کاور لیتر ==========
    ws15 = wb.create_sheet("📝 کاور لیتر"); ws15.sheet_view.rightToLeft = True
    st(ws15, 1, 10, f"کاور لیترها | {NOW_FULL}")
    add_ts(ws15, 2, 1)
    r = 3; sh(ws15, r, ['ردیف','متقاضی','کارفرما','عنوان','وضعیت','تاریخ','🕐 ثبت','اقدام'])
    sd(ws15, r+1, [
        [1,'ندا','Working In Health NZ','Cover Letter — Midwife','✅ آماده',NOW_STR,NOW_TIME,'بررسی و ارسال'],
        [2,'ندا','Health New Zealand','Cover Letter — Midwife','✅ آماده',NOW_STR,NOW_TIME,'بررسی و ارسال'],
        [3,'ندا','RGH Global','Cover Letter — Midwife','✅ آماده',NOW_STR,NOW_TIME,'بررسی و ارسال'],
        [4,'ندا','Holalemania','Cover Letter — Hebamme','✅ آماده',NOW_STR,NOW_TIME,'بررسی و ارسال'],
        [5,'ندا','TalentOrange','Cover Letter — Hebamme','✅ آماده',NOW_STR,NOW_TIME,'بررسی و ارسال'],
    ], [8,12,22,30,12,14,12,18])

    # ========== SAVE ==========
    import os
    out = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'dashboard', f'MigrationHunter_Final_v2_{NOW.strftime("%Y%m%d_%H%M")}.xlsx')
    wb.save(out)
    print(f"\n✅ داشبورد ذخیره شد: {out}")
    print(f"📊 {len(wb.sheetnames)} شیت:")
    for i, s in enumerate(wb.sheetnames, 1): print(f"   {i}. {s}")
    print(f"\n🕐 ساعت ثبت: {NOW_FULL}")
    print(f"🔤 فونت: B Mitra | 📐 RTL | 🎨 رنگبندی کامل")

if __name__ == '__main__':
    main()
