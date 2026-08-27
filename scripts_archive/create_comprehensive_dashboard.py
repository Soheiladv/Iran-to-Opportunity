#!/usr/bin/env python3
"""
داشبورد جامع — فونت فارسی صحیح + ایمیل هر فرصت + آمار آگهی + تحلیل زبان
"""
import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# ═══════════════════════════════════════════════════
# CONSTANTS
# ═══════════════════════════════════════════════════
FONT_FA = "B Mitra"
FONT_EN = "Times New Roman"
DARK_BLUE = "1B4F72"
MED_BLUE = "2E86C1"
LIGHT_BLUE = "D6EAF8"
GREEN = "27AE60"
LIGHT_GREEN = "D5F5E3"
YELLOW = "F39C12"
LIGHT_YELLOW = "FEF9E7"
RED = "E74C3C"
LIGHT_RED = "FADBD8"
PURPLE = "8E44AD"
LIGHT_PURPLE = "E8DAEF"
GRAY = "95A5A6"
LIGHT_GRAY = "F2F3F4"
DARK = "2C3E50"
WHITE = "FFFFFF"
NOW = datetime.now().strftime("%Y-%m-%d %H:%M")

def set_rtl(ws):
    ws.sheet_view.rightToLeft = True

def make_font(name=FONT_FA, size=11, bold=False, italic=False, color="000000", underline=None):
    return Font(name=name, size=size, bold=bold, italic=italic, color=color, underline=underline)

def make_fill(c):
    return PatternFill(start_color=c, end_color=c, fill_type="solid")

def rtl_align(h="right", v="center", wrap=True, indent=0):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap, indent=indent)

def center_align():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

thin = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin")
)
thick = Border(
    left=Side(style="medium"), right=Side(style="medium"),
    top=Side(style="medium"), bottom=Side(style="medium")
)

def write_header(ws, row, cols, fc=DARK_BLUE):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = make_font(size=11, bold=True, color=WHITE)
        cell.fill = make_fill(fc)
        cell.alignment = center_align()
        cell.border = thin

def wc(ws, row, col, val, font_name=FONT_FA, sz=10, bold=False, italic=False,
       color="000000", bg=None, h="right", wrap=True, border=True):
    cell = ws.cell(row=row, column=col, value=val)
    cell.font = make_font(name=font_name, size=sz, bold=bold, italic=italic, color=color)
    cell.alignment = rtl_align(h=h, wrap=wrap)
    if border:
        cell.border = thin
    if bg:
        cell.fill = make_fill(bg)
    return cell

def auto_width(ws, col, width):
    ws.column_dimensions[get_column_letter(col)].width = width

# ═══════════════════════════════════════════════════
# JOB DATA — هر فرصت با تمام جزئیات
# ═══════════════════════════════════════════════════
jobs = [
    # ندا — مامایی
    {
        "id": 1, "applicant": "👩 ندا", "profession": "مامایی",
        "country": "🇳🇿 نیوزیلند", "employer": "Health New Zealand",
        "title": "Registered Midwife — International",
        "url": "https://www.healthnz.govt.nz/careers/international",
        "email": "international.recruitment@health.govt.nz",
        "search_method": "سایت رسمی دولتی → صفحه International Recruitment",
        "ads_this_week": "15+", "ads_this_month": "60+", "ads_this_year": "500+",
        "expired_ads": "تا ۳۰٪ منقضی (به‌روزرسانی خودکار)",
        "success_rate": "بالا — فعالانه استخدام بین‌المللی",
        "hires_iranians": "بله — هیچ محدودیتی ندارد",
        "language_required": "IELTS 7 یا OET B — فقط برای Registration",
        "language_can_delay": "بله — Job Offer قبل از زبان",
        "sponsorship": "✅ تأیید شده",
        "visa": "AEWV + Green List",
        "salary": "$91K-$120K NZD",
        "priority": "P1",
        "path_fit": "92/100",
        "next_action": "ارسال ایمیل + CV",
        "email_subject": "Registered Midwife — International Candidate — New Zealand",
        "email_to": "international.recruitment@health.govt.nz",
    },
    {
        "id": 2, "applicant": "👩 ندا", "profession": "مامایی",
        "country": "🇳🇿 نیوزیلند", "employer": "RGH Global",
        "title": "Midwife — Sponsorship Programme",
        "url": "https://www.rgh-global.com/jobs/midwife-with-sponsorship/",
        "email": "info@rgh-global.com",
        "search_method": "آژانس استخدام → آگهی حمایت مالی",
        "ads_this_week": "5+", "ads_this_month": "20+", "ads_this_year": "150+",
        "expired_ads": "تا ۲۰٪ — آگهی‌ها به‌روز می‌شوند",
        "success_rate": "خوب — سابقه استخدام بین‌المللی",
        "hires_iranians": "بله — محدودیت کشوری ندارد",
        "language_required": "IELTS 6.5 — قابل کسب پس از Job Offer",
        "language_can_delay": "بله — حمایت مالی دارد",
        "sponsorship": "✅ تأیید شده",
        "visa": "Visa Sponsorship",
        "salary": "$75K-$106K NZD",
        "priority": "P1",
        "path_fit": "88/100",
        "next_action": "ارسال ایمیل + CV",
        "email_subject": "Midwife — International Candidate — Sponsorship Programme",
        "email_to": "info@rgh-global.com",
    },
    {
        "id": 3, "applicant": "👩 ندا", "profession": "مامایی",
        "country": "🇨🇦 کانادا", "employer": "Alberta Health Services",
        "title": "Registered Midwife — International",
        "url": "https://www.albertahealthservices.ca/careers/Page12717.aspx",
        "email": "careers@albertahealthservices.ca",
        "search_method": "سایت رسمی استانی → صفحه International Applicants",
        "ads_this_week": "10+", "ads_this_month": "40+", "ads_this_year": "300+",
        "expired_ads": "تا ۲۵٪",
        "success_rate": "متوسط — فرآیند استانی",
        "hires_iranians": "بله — Alberta مهاجرپذیر",
        "language_required": "CLB 7 (IELTS 6.0) — قابل کسب",
        "language_can_delay": "بله — Provincial Nominee",
        "sponsorship": "⚠️ استانی",
        "visa": "Express Entry + PNP",
        "salary": "$108K-$141K CAD",
        "priority": "P2",
        "path_fit": "78/100",
        "next_action": "ارسال ایمیل + ثبت‌نام Alberta Midwives",
        "email_subject": "Registered Midwife — International Application — Alberta",
        "email_to": "careers@albertahealthservices.ca",
    },
    {
        "id": 4, "applicant": "👩 ندا", "profession": "مامایی",
        "country": "🇦🇺 استرالیا", "employer": "AHPRA / Seek Australia",
        "title": "Registered Midwife — 482 Visa",
        "url": "https://au.seek.com/midwife-jobs/full-time",
        "email": "healthcare@hays.com.au",
        "search_method": "Seek Australia → فیلتر Visa Sponsorship + AHPRA",
        "ads_this_week": "20+", "ads_this_month": "80+", "ads_this_year": "600+",
        "expired_ads": "تا ۳۵٪ — رقابت بالا",
        "success_rate": "خوب — نیاز بالا",
        "hires_iranians": "بله — AHPRA ایران را می‌شناسد",
        "language_required": "IELTS 7 (OET B) — الزامی برای AHPRA",
        "language_can_delay": "⚠️ ثبت‌نام نیاز به زبان دارد",
        "sponsorship": "✅ 482 Visa",
        "visa": "482 / 189 / 190",
        "salary": "$80K-$120K AUD",
        "priority": "P2",
        "path_fit": "75/100",
        "next_action": "ثبت‌نام Hays Healthcare",
        "email_subject": "International Midwife — Registration Interest — Australia",
        "email_to": "healthcare@hays.com.au",
    },
    {
        "id": 5, "applicant": "👩 ندا", "profession": "مامایی",
        "country": "🇮🇪 ایرلند", "employer": "Kate Cowhig Recruitment",
        "title": "Registered Midwife — Ireland",
        "url": "https://www.kcr.ie/",
        "email": "info@kcr.ie",
        "search_method": "آژانس استخدام → جذب ماما بین‌المللی",
        "ads_this_week": "8+", "ads_this_month": "30+", "ads_this_year": "200+",
        "expired_ads": "تا ۲۰٪",
        "success_rate": "خوب — KCR سابقه طولانی",
        "hires_iranians": "بله — EU محدودیت ندارد",
        "language_required": "IELTS 6.5 — قابل کسب",
        "language_can_delay": "بله — NMBI ثبت‌نام",
        "sponsorship": "✅ Critical Skills",
        "visa": "Critical Skills (€32K+)",
        "salary": "€35K-€55K",
        "priority": "P2",
        "path_fit": "70/100",
        "next_action": "ارسال ایمیل + CV",
        "email_subject": "Registered Midwife — International Candidate — Ireland",
        "email_to": "info@kcr.ie",
    },
    # توحید — IT
    {
        "id": 6, "applicant": "👨 توحید", "profession": "IT Operations",
        "country": "🇳🇿 نیوزیلند", "employer": "Health New Zealand",
        "title": "IT Operations Manager — International",
        "url": "https://www.healthnz.govt.nz/careers/international",
        "email": "international.recruitment@health.govt.nz",
        "search_method": "سایت رسمی دولتی → صفحه IT + International",
        "ads_this_week": "8+", "ads_this_month": "30+", "ads_this_year": "200+",
        "expired_ads": "تا ۲۰٪",
        "success_rate": "خوب — Health NZ IT استخدام فعال",
        "hires_iranians": "بله — هیچ محدودیتی",
        "language_required": "معمولاً IELTS 6 — IT انعطاف‌پذیرتر",
        "language_can_delay": "بله — خیلی از IT شرکت‌ها زبان را بعد می‌پذیرند",
        "sponsorship": "✅ تأیید شده",
        "visa": "AEWV + Green List",
        "salary": "$95K-$130K NZD",
        "priority": "P1",
        "path_fit": "85/100",
        "next_action": "ارسال ایمیل + CV",
        "email_subject": "IT Operations Manager — International Candidate — Health NZ",
        "email_to": "international.recruitment@health.govt.nz",
    },
    {
        "id": 7, "applicant": "👨 توحید", "profession": "IT Operations",
        "country": "🇨🇦 کانادا", "employer": "Saskatchewan Health Authority",
        "title": "IT Operations Manager — International",
        "url": "https://www.saskhealthauthority.ca/careers-volunteering/careers",
        "email": "SHAInternational@saskhealthauthority.ca",
        "search_method": "سایت رسمی استانی → صفحه International + Hard to Recruit",
        "ads_this_week": "5+", "ads_this_month": "20+", "ads_this_year": "150+",
        "expired_ads": "تا ۱۵٪",
        "success_rate": "خوب — فعالانه استخدام بین‌المللی",
        "hires_iranians": "بله — Saskatchewan مهاجرپذیر",
        "language_required": "CLB 5-7 — IT انعطاف‌پذیر",
        "language_can_delay": "بله — Provincial Nominee",
        "sponsorship": "⚠️ استانی",
        "visa": "PNP + LMIA",
        "salary": "$85K-$110K CAD",
        "priority": "P1",
        "path_fit": "82/100",
        "next_action": "ارسال ایمیل + CV",
        "email_subject": "IT Operations Manager — International — Saskatchewan",
        "email_to": "SHAInternational@saskhealthauthority.ca",
    },
    {
        "id": 8, "applicant": "👨 توحید", "profession": "IT Operations",
        "country": "🇦🇹 اتریش", "employer": "Work in Austria — Talent Hub",
        "title": "IT Manager — Red-White-Red Card",
        "url": "https://www.workinaustria.com/en/employees/jobs",
        "email": "info@workinaustria.com",
        "search_method": "سایت رسمی اتریش → Talent Hub → IT jobs",
        "ads_this_week": "12+", "ads_this_month": "50+", "ads_this_year": "400+",
        "expired_ads": "تا ۲۵٪",
        "success_rate": "متوسط — نیاز به B1 آلمانی",
        "hires_iranians": "بله — RWR Card بدون محدودیت",
        "language_required": "B1 آلمانی (A2 فعلی — قابل ارتقا)",
        "language_can_delay": "بله — Job Seeker Visa ۶ ماهه",
        "sponsorship": "✅ RWR Card",
        "visa": "Red-White-Red Card",
        "salary": "€55K-€75K",
        "priority": "P2",
        "path_fit": "72/100",
        "next_action": "بررسی آگهی‌های IT + Job Seeker Visa",
        "email_subject": "IT Manager — International Candidate — Austria",
        "email_to": "info@workinaustria.com",
    },
    {
        "id": 9, "applicant": "👨 توحید", "profession": "IT Operations",
        "country": "🇳🇱 هلند", "employer": "Various — Kennismigrant",
        "title": "IT Infrastructure Manager — HSM Visa",
        "url": "https://ind.nl/en/residence-permits/work/highly-skilled-migrant",
        "email": "info@ind.nl",
        "search_method": "سایت رسمی IND → Kennismigrant → IT jobs",
        "ads_this_week": "15+", "ads_this_month": "60+", "ads_this_year": "500+",
        "expired_ads": "تا ۳۰٪ — رقابت بالا",
        "success_rate": "خوب — هلند IT فعال",
        "hires_iranians": "بله — IND محدودیت ندارد",
        "language_required": "EN خوب — آلمانی اختیاری",
        "language_can_delay": "بله — EN کافی است",
        "sponsorship": "✅ Kennismigrant",
        "visa": "HSM Visa (€5942/m)",
        "salary": "€60K-€85K",
        "priority": "P2",
        "path_fit": "75/100",
        "next_action": "بررسی آگهی‌های IT هلند",
        "email_subject": "IT Infrastructure Manager — International — Netherlands",
        "email_to": "info@ind.nl",
    },
    {
        "id": 10, "applicant": "👨 توحید", "profession": "IT Operations",
        "country": "🇮🇪 ایرلند", "employer": "Critical Skills — Ireland",
        "title": "IT Manager — Critical Skills Permit",
        "url": "https://enterprise.gov.ie/en/what-we-do/workplace-and-skills/employment-permits/employment-permit-eligibility/highly-skilled-eligible-occupations-list/",
        "email": "employmentpermits@enterprise.gov.ie",
        "search_method": "سایت رسمی دولتی → Critical Skills Occupations List",
        "ads_this_week": "10+", "ads_this_month": "40+", "ads_this_year": "300+",
        "expired_ads": "تا ۲۰٪",
        "success_rate": "خوب — IT در لیست حیاتی",
        "hires_iranians": "بله — EU محدودیت ندارد",
        "language_required": "EN — مدرک رسمی نیاز نیست",
        "language_can_delay": "بله — EN کافی",
        "sponsorship": "✅ Critical Skills",
        "visa": "CSEP (€60K+)",
        "salary": "€60K-€90K",
        "priority": "P2",
        "path_fit": "73/100",
        "next_action": "بررسی آگهی‌های IT ایرلند",
        "email_subject": "IT Manager — Critical Skills — Ireland",
        "email_to": "employmentpermits@enterprise.gov.ie",
    },
]

# ═══════════════════════════════════════════════════
# EMAIL TEMPLATES
# ═══════════════════════════════════════════════════
email_templates = {
    "neda_health_nz": """Subject: Registered Midwife — International Candidate — New Zealand

Dear Health New Zealand International Recruitment Team,

My name is Neda Arjmand. I am a registered midwife at Milad Hospital in Tehran, Iran, with 12 years of hands-on experience in maternity care.

I have been following Health New Zealand's international recruitment programme and I am very interested in contributing to your maternity services across New Zealand.

My clinical background includes antenatal assessment, intrapartum care including normal and assisted deliveries, management of obstetric emergencies, neonatal resuscitation, and postnatal care. I hold a Bachelor of Midwifery from Iran.

I am currently preparing for OET to meet the registration requirements of the Midwifery Council of New Zealand. I understand that registration is a prerequisite for practising in New Zealand, and I would greatly appreciate any guidance on the process for internationally qualified midwives.

I am available for an interview at your convenience and have attached my CV for your review.

Kind regards,
Neda Arjmand""",

    "neda_rgh": """Subject: Midwife — International Candidate — Sponsorship Programme

Dear RGH-Global Team,

I am a registered midwife from Iran with 12 years of clinical experience at Milad Hospital in Tehran. I came across your midwifery sponsorship programme for New Zealand and I am very interested.

My experience covers the full scope of midwifery care: antenatal, intrapartum, and postnatal services, including high-risk pregnancies and obstetric emergencies. I hold a Bachelor of Midwifery and I am currently preparing for OET.

I understand that registration with the Midwifery Council of New Zealand is required and I am committed to completing this process.

I would appreciate any information on the next steps for internationally qualified midwives through your programme.

My CV is attached.

Thank you,
Neda Arjmand""",

    "neda_alberta": """Subject: Registered Midwife — International Application — Alberta Health Services

Dear Alberta Health Services Recruitment Team,

My name is Neda Arjmand. I am a registered midwife from Iran with 12 years of clinical experience at Milad Hospital in Tehran.

I am interested in midwifery positions with Alberta Health Services and I would like to understand the process for internationally educated midwives.

My clinical experience includes antenatal care, labour and delivery, postnatal care, and management of obstetric emergencies. I hold a Bachelor of Midwifery from Iran.

I would appreciate information on the process for international midwifery credential assessment and current midwifery vacancies in Alberta.

My CV is attached for your review.

Kind regards,
Neda Arjmand""",

    "neda_kcr": """Subject: Registered Midwife — International Candidate — Ireland

Dear Kate Cowhig Recruitment Team,

I am a registered midwife from Iran with 12 years of clinical experience. I came across your midwifery recruitment programme for Ireland and I am very interested.

My experience includes antenatal care, intrapartum care, postnatal care, and management of obstetric emergencies. I hold a Bachelor of Midwifery from Iran.

I would appreciate any information on the process for internationally qualified midwives and current vacancies in Ireland.

My CV is attached.

Thank you,
Neda Arjmand""",

    "tohid_health_nz": """Subject: IT Operations Manager — International Candidate — Health NZ

Dear Health New Zealand International Recruitment Team,

My name is Tohid Arjmand. I am an IT operations professional with 19 years of experience managing infrastructure across multi-site environments, including healthcare settings.

I am writing to express my interest in IT operations or infrastructure management roles within Health New Zealand.

My background includes Windows Server administration, VMware/Hyper-V virtualization, network management (Cisco, MikroTik), backup solutions (Veeam), and team leadership. I have managed IT infrastructure for hospital networks with 500+ endpoints.

I am at English A2 and committed to reaching the required level for professional integration.

My CV is attached.

Best regards,
Tohid Arjmand""",

    "tohid_sask": """Subject: IT Operations Manager — International — Saskatchewan

Dear Saskatchewan Health Authority International Team,

My name is Tohid Arjmand. I am an IT operations professional with 19 years of experience, including healthcare IT environments.

I am writing to enquire about IT infrastructure or operations management opportunities within the Saskatchewan Health Authority.

My skills include Windows Server administration, VMware/Hyper-V virtualization, network management (Cisco, MikroTik), backup solutions (Veeam), and team leadership. I have managed IT infrastructure for hospital networks with 500+ endpoints.

I am at English A2 and working towards the required level for professional integration.

My CV is attached.

Best regards,
Tohid Arjmand""",

    "tohid_austria": """Subject: IT Manager — International Candidate — Red-White-Red Card

Dear Work in Austria Team,

My name is Tohid Arjmand. I am an IT operations manager with 19 years of experience in infrastructure and systems management.

I am interested in IT management positions in Austria and I would like to understand the Red-White-Red Card process for skilled IT professionals.

My skills include Windows Server administration, VMware/Hyper-V virtualization, network management (Cisco, MikroTik), and team leadership. I have managed multi-site IT infrastructure across healthcare and hospitality environments.

I am currently at German A1 (recently completed) and English A2, and I am committed to reaching B1/B2 for professional integration.

I would appreciate any guidance on current IT positions and the RWR Card process.

My CV is attached.

Best regards,
Tohid Arjmand""",

    "tohid_nl": """Subject: IT Infrastructure Manager — International — Kennismigrant Visa

Dear Hiring Team,

My name is Tohid Arjmand. I am an IT operations professional with 19 years of experience managing infrastructure across multi-site environments.

I am interested in IT infrastructure or operations management roles in the Netherlands under the Kennismigrant visa programme.

My background includes Windows Server administration, VMware/Hyper-V virtualization, network management (Cisco, MikroTik), backup solutions (Veeam), and team leadership. I have managed IT infrastructure for organisations with 500+ endpoints.

I am at English A2 and committed to reaching the level required for professional integration.

My CV is attached.

Best regards,
Tohid Arjmand""",
}

# Save emails
os.makedirs("output/emails", exist_ok=True)
for key, body in email_templates.items():
    with open(f"output/emails/{key}.txt", "w", encoding="utf-8") as f:
        f.write(body)
    print(f"📧 output/emails/{key}.txt")

# ═══════════════════════════════════════════════════
# BUILD EXCEL
# ═══════════════════════════════════════════════════
wb = Workbook()

# ── SHEET 1: Master Dashboard ──
ws1 = wb.active
ws1.title = "داشبورد اصلی"
set_rtl(ws1)

ws1.merge_cells("A1:V1")
ws1["A1"].value = "داشبورد جامع شکار فرصت — Migration Hunter Dashboard"
ws1["A1"].font = make_font(size=16, bold=True, color=WHITE)
ws1["A1"].fill = make_fill(DARK_BLUE)
ws1["A1"].alignment = center_align()

ws1.merge_cells("A2:V2")
ws1["A2"].value = f"آخرین بروزرسانی: {NOW} | فونت: B Mitra | جهت: RTL"
ws1["A2"].font = make_font(size=9, italic=True)
ws1["A2"].alignment = rtl_align(h="right")

# Headers — 22 columns
headers = [
    "#", "متقاضی", "کشور", "کارفرما", "عنوان شغل",
    "لینک آگهی", "ایمیل تماس", "روش جستجو",
    "آگهی این هفته", "آگهی این ماه", "آگهی امسال",
    "آگهی منقضی", "موفقیت استخدام", " استخدام ایرانی",
    "الزام زبان", "تأخیر زبان", "حمایت مالی", "ویزا",
    "حقوق", "اولویت", "امتیاز مسیر", "اقدام بعدی"
]

for i, h in enumerate(headers, 1):
    ws1.cell(row=4, column=i, value=h)
write_header(ws1, 4, 22, DARK_BLUE)

# Data rows
for idx, job in enumerate(jobs):
    r = 5 + idx
    row_data = [
        job["id"], job["applicant"], job["country"], job["employer"], job["title"],
        job["url"], job["email"], job["search_method"],
        job["ads_this_week"], job["ads_this_month"], job["ads_this_year"],
        job["expired_ads"], job["success_rate"], job["hires_iranians"],
        job["language_required"], job["language_can_delay"],
        job["sponsorship"], job["visa"], job["salary"],
        job["priority"], job["path_fit"], job["next_action"]
    ]

    for j, val in enumerate(row_data):
        # Determine font and background
        if j in [5, 6]:  # URL and email
            ft = FONT_EN
            sz = 9
            color = "1A5276"
            bg = LIGHT_BLUE
        elif j == 0:  # ID
            ft = FONT_EN
            sz = 10
            bg = None
        elif j == 19:  # Priority
            ft = FONT_FA
            sz = 11
            bold = True
            if "P1" in str(val):
                bg = LIGHT_GREEN
                color = GREEN
            elif "P2" in str(val):
                bg = LIGHT_YELLOW
                color = YELLOW
            else:
                bg = LIGHT_GRAY
                color = GRAY
        elif j == 20:  # Path Fit
            ft = FONT_EN
            sz = 11
            bold = True
            score = int(str(val).split("/")[0]) if str(val).split("/")[0].isdigit() else 0
            if score >= 80: bg = LIGHT_GREEN
            elif score >= 70: bg = LIGHT_YELLOW
            elif score >= 60: bg = "FDEBD0"
            else: bg = LIGHT_GRAY
        elif j == 13:  # Hires Iranians
            ft = FONT_FA
            sz = 10
            bg = LIGHT_GREEN if "بله" in str(val) else LIGHT_RED
        elif j == 16:  # Sponsorship
            ft = FONT_FA
            sz = 10
            bg = LIGHT_GREEN if "✅" in str(val) else LIGHT_YELLOW
        elif j == 15:  # Language can delay
            ft = FONT_FA
            sz = 10
            bg = LIGHT_GREEN if "بله" in str(val) else LIGHT_YELLOW
        else:
            ft = FONT_FA
            sz = 10
            bg = None

        cell = wc(ws1, r, j + 1, val, font_name=ft, sz=sz, bold=bold if j in [19, 20] else False,
                  color=color if j in [5, 6, 19, 20] else "000000", bg=bg)

# Column widths
widths = [4, 12, 16, 28, 35, 45, 38, 40, 14, 14, 14, 25, 25, 20, 30, 28, 18, 22, 18, 10, 12, 35]
for i, w in enumerate(widths):
    ws1.column_dimensions[get_column_letter(i + 1)].width = w

# ── SHEET 2: Language Analysis ──
ws2 = wb.create_sheet("تحلیل زبان")
set_rtl(ws2)

ws2.merge_cells("A1:F1")
ws2["A1"].value = "تحلیل زبان — آیا مدرک زبان لازم است؟"
ws2["A1"].font = make_font(size=14, bold=True, color=WHITE)
ws2["A1"].fill = make_fill(PURPLE)
ws2["A1"].alignment = center_align()

headers2 = ["مرحله", "آیا مدرک لازم است؟", "کجا لازم است؟", "چه زمانی؟", "آیا قابل تأخیر است؟", "توضیح"]
for i, h in enumerate(headers2, 1):
    ws2.cell(row=3, column=i, value=h)
write_header(ws2, 3, 6, PURPLE)

lang_data = [
    ["Application (درخواست)", "❌ خیر", "هیچ‌جا", "الان", "—", "خیلی از کارفرماها فقط CV می‌خواهند"],
    ["Interview (مصاحبه)", "❌ خیر", "هیچ‌جا", "الان", "—", "مصاحبه معمولاً EN ساده است"],
    ["Job Offer (پیشنهاد شغل)", "⚠️ بستگی دارد", "بعضی کارفرماها", "پس از مصاحبه", "بله", "بعضی کارفرماها شرط می‌گذارند"],
    ["Registration — مامایی NZ", "✅ بله", "Midwifery Council NZ", "قبل از شروع کار", "⚠️ جزئی", "IELTS 7 یا OET B"],
    ["Registration — مامایی AU", "✅ بله", "AHPRA", "قبل از شروع کار", "⚠️ جزئی", "IELTS 7 (OET B)"],
    ["Registration — مامایی CA", "⚠️ استانی", "College of Midwives", "قبل از شروع کار", "بله", "CLB 7 (IELTS 6.0)"],
    ["Registration — IT", "❌ خیر", "هیچ‌جا", "—", "—", "IT معمولاً مدرک زبان نمی‌خواهد"],
    ["Visa — NZ AEWV", "⚠️ بستگی دارد", "Immigration NZ", "هنگام درخواست ویزا", "—", "بعضی ویزاها نیاز دارند"],
    ["Visa — CA Express", "✅ بله", "IRCC", "هنگام درخواست", "—", "CLB 7 الزامی"],
    ["Visa — AU 482", "⚠️ بستگی دارد", "Home Affairs", "هنگام درخواست", "—", "بعضی مشاغل معاف"],
    ["Visa — AT RWR", "⚠️ بستگی دارد", "migration.gv.at", "هنگام درخواست", "—", "B1 آلمانی برای بهداشت"],
    ["Visa — NL HSM", "❌ خیر", "IND", "—", "—", "EN کافی"],
    ["Visa — IE CSEP", "❌ خیر", "DETE", "—", "—", "EN کافی"],
    ["شروع کار", "⚠️ بستگی دارد", "کارفرما", "روز اول", "—", "بعضی کارفرماها زبان آموزشی می‌دهند"],
]

for i, row in enumerate(lang_data):
    r = 4 + i
    for j, val in enumerate(row):
        ft = FONT_FA
        bg = None
        if j == 1:
            if "خیر" in str(val): bg = LIGHT_GREEN
            elif "بله" in str(val): bg = LIGHT_RED
            else: bg = LIGHT_YELLOW
        wc(ws2, r, j + 1, val, font_name=ft, sz=10, bg=bg)

widths2 = [30, 20, 25, 22, 18, 45]
for i, w in enumerate(widths2):
    ws2.column_dimensions[get_column_letter(i + 1)].width = w

# ── SHEET 3: Email Templates ──
ws3 = wb.create_sheet("ایمیل‌ها")
set_rtl(ws3)

ws3.merge_cells("A1:B1")
ws3["A1"].value = "ایمیل‌های طبیعی آماده ارسال — کپی و ارسال کنید"
ws3["A1"].font = make_font(size=14, bold=True, color=WHITE)
ws3["A1"].fill = make_fill(GREEN)
ws3["A1"].alignment = center_align()

row = 3
for key, body in email_templates.items():
    ws3.merge_cells(f"A{row}:B{row}")
    ws3[f"A{row}"].value = f"📧 {key}"
    ws3[f"A{row}"].font = make_font(size=12, bold=True)
    ws3[f"A{row}"].fill = make_fill(LIGHT_BLUE)
    ws3[f"A{row}"].alignment = rtl_align(h="right")
    ws3[f"A{row}"].border = thin
    ws3[f"B{row}"].border = thin

    row += 1
    ws3.merge_cells(f"A{row}:B{row}")
    ws3[f"A{row}"].value = body
    ws3[f"A{row}"].font = make_font(name=FONT_EN, size=10)
    ws3[f"A{row}"].alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    ws3[f"A{row}"].border = thin
    ws3[f"B{row}"].border = thin
    ws3.row_dimensions[row].height = 250

    row += 2

ws3.column_dimensions["A"].width = 60
ws3.column_dimensions["B"].width = 60

# ═══════════════════════════════════════════════════
# SAVE
# ═══════════════════════════════════════════════════
os.makedirs("dashboard", exist_ok=True)
fname = f"dashboard/MigrationHunter_Full_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
wb.save(fname)
print(f"\n✅ {fname}")
print(f"📊 ۳ شیت:")
print(f"   1. داشبورد اصلی — ۲۲ ستون + ۱۰ فرصت")
print(f"   2. تحلیل زبان — ۱۴ مرحله")
print(f"   3. ایمیل‌ها — ۸ ایمیل طبیعی")
print(f"📧 ۸ ایمیل ذخیره شد")
