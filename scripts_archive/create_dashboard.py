#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Migration Hunter Dashboard - Excel Generator
"""

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

FONT_NAME = 'B Mitra'
HEADER_FONT = Font(name=FONT_NAME, bold=True, size=14, color='FFFFFF')
HEADER_FILL = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
GREEN_FILL = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
YELLOW_FILL = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
RED_FILL = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
CELL_FONT = Font(name=FONT_NAME, size=10)
BOLD_FONT = Font(name=FONT_NAME, size=10, bold=True)
TITLE_FONT = Font(name=FONT_NAME, size=14, bold=True, color='1F4E79')
THIN_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
CENTER_ALIGN = Alignment(horizontal='center', vertical='center', wrap_text=True)
RTL_ALIGN = Alignment(horizontal='right', vertical='center', wrap_text=True, readingOrder=2)

def style_cell(cell, font=CELL_FONT, fill=None, alignment=RTL_ALIGN):
    cell.font = font
    cell.alignment = alignment
    cell.border = THIN_BORDER
    if fill:
        cell.fill = fill

def create_sheet(wb, name, headers, data, widths=None):
    ws = wb.create_sheet(name)
    ws.sheet_view.rightToLeft = True
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        style_cell(cell, HEADER_FONT, HEADER_FILL, CENTER_ALIGN)
    for r, row_data in enumerate(data, 2):
        for c, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c, value=val)
            style_cell(cell)
    if widths:
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w
    return ws

def main():
    wb = openpyxl.Workbook()
    if 'Sheet' in wb.sheetnames:
        del wb['Sheet']
    
    # Dashboard
    ws = wb.create_sheet('Dashboard', 0)
    ws.sheet_view.rightToLeft = True
    ws.merge_cells('A1:F1')
    ws['A1'] = 'Migration Hunter Dashboard'
    ws['A1'].font = TITLE_FONT
    ws['A1'].alignment = CENTER_ALIGN
    
    headers = ['指标', '数值']
    data = [
        ['总机会数', 4],
        ['新机会', 4],
        ['已确认机会', 0],
        ['TOHID机会', 1],
        ['NEDA机会', 3],
        ['已确认雇主', 5],
        ['确认Sponsorship', 4],
        ['准备提交申请', 4],
        ['已发送申请', 0],
        ['回复数', 0],
        ['面试数', 0],
        ['Offer数', 0],
    ]
    create_sheet(wb, 'Summary', headers, data, [30, 15])
    
    # Tohid Jobs
    headers = ['Job ID', 'Employer', 'Country', 'Title', 'Sponsorship', 'Path Fit', 'Status']
    data = [
        ['JOB-004', 'IT Companies DE', 'Germany', 'IT Manager', 'Likely', '75/100', 'NEW'],
    ]
    create_sheet(wb, 'Tohid Jobs', headers, data, [12, 20, 12, 18, 12, 12, 12])
    
    # Neda Jobs
    headers = ['Job ID', 'Employer', 'Country', 'Title', 'Sponsorship', 'Path Fit', 'Status']
    data = [
        ['JOB-001', 'Health New Zealand', 'New Zealand', 'Midwife', 'Confirmed', '80/100', 'NEW'],
        ['JOB-002', 'RGH Global', 'New Zealand', 'Midwife', 'Confirmed', '78/100', 'NEW'],
        ['JOB-003', 'Hassett Group', 'Australia', 'Registered Midwife', 'Confirmed', '78/100', 'NEW'],
    ]
    create_sheet(wb, 'Neda Jobs', headers, data, [12, 20, 14, 18, 12, 12, 12])
    
    # Employers
    headers = ['Employer', 'Country', 'Sponsorship', 'Score', 'Applicant']
    data = [
        ['Health New Zealand', 'NZ', 'Confirmed', 95, 'NEDA'],
        ['RGH Global', 'NZ', 'Confirmed', 85, 'NEDA'],
        ["St Vincent's Health", 'AU', 'Confirmed', 85, 'NEDA'],
        ['Hassett Group', 'AU', 'Confirmed', 80, 'NEDA'],
        ['IT Companies DE', 'DE', 'Likely', 75, 'TOHID'],
    ]
    create_sheet(wb, 'Employers', headers, data, [22, 10, 14, 10, 12])
    
    # Sources
    headers = ['Source', 'Type', 'Country', 'Trust', 'TOHID', 'NEDA']
    data = [
        ['Health NZ', 'Government', 'NZ', 95, 70, 98],
        ['RGH Global', 'Recruiter', 'NZ', 85, 30, 92],
        ['SEEK Australia', 'Job Board', 'AU', 90, 60, 80],
        ['Arbeitnow', 'Job Board', 'DE', 85, 90, 40],
        ['LinkedIn', 'Social', 'Global', 80, 75, 75],
    ]
    create_sheet(wb, 'Sources', headers, data, [18, 14, 10, 10, 10, 10])
    
    # Applications
    headers = ['App ID', 'Applicant', 'Employer', 'Status', 'Next Action']
    data = [
        ['APP-001', 'NEDA', 'Health NZ', 'READY TO SEND', 'User approval'],
        ['APP-002', 'NEDA', 'RGH Global', 'READY TO SEND', 'User approval'],
        ['APP-003', 'NEDA', 'Hassett Group', 'READY TO SEND', 'User approval'],
        ['APP-004', 'TOHID', 'IT Companies DE', 'READY TO SEND', 'User approval'],
    ]
    create_sheet(wb, 'Applications', headers, data, [10, 12, 18, 16, 14])
    
    # Visa
    headers = ['Country', 'Visa', 'Language', 'Family', 'Registration']
    data = [
        ['New Zealand', 'AEWV', 'ANZSCO 1-2: None', 'Yes', 'Separate'],
        ['Australia', '482 TSS', 'IELTS 5.0', 'Yes', 'AHPRA'],
        ['Germany', 'EU Blue Card', 'None (IT)', 'Yes', 'Anerkennung'],
    ]
    create_sheet(wb, 'Visa', headers, data, [16, 14, 18, 10, 14])
    
    # Registration
    headers = ['Country', 'Applicant', 'Regulator', 'English Req', 'Status']
    data = [
        ['New Zealand', 'NEDA', 'Midwifery Council', 'IELTS 7.0/OET', 'Required'],
        ['Australia', 'NEDA', 'AHPRA', 'IELTS 7.0/OET', 'Required'],
        ['Germany', 'TOHID', 'Chamber', 'None for visa', 'Check'],
    ]
    create_sheet(wb, 'Registration', headers, data, [16, 12, 18, 16, 12])
    
    # Language
    headers = ['Applicant', 'English', 'German', 'IELTS', 'OET']
    data = [
        ['TOHID', 'A2', 'A1', 'Not required', 'Not required'],
        ['NEDA', 'A2', 'A1', 'Needed for registration', 'Alternative'],
    ]
    create_sheet(wb, 'Language', headers, data, [14, 10, 10, 22, 14])
    
    # Search History
    headers = ['Date', 'Countries', 'Sources', 'Jobs', 'Applications']
    data = [
        ['2026-08-18', 'NZ, AU, DE', '5', '4', '4 prepared'],
    ]
    create_sheet(wb, 'Search History', headers, data, [14, 16, 12, 10, 16])
    
    filename = 'MigrationHunter/dashboard/MigrationHunter_Dashboard.xlsx'
    wb.save(filename)
    print(f"Dashboard created: {filename}")

if __name__ == '__main__':
    main()
