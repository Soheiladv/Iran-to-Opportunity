#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Migration Hunter Dashboard v2 - Complete RTL with B Mitra Font
"""

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, numbers
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.utils import get_column_letter
from datetime import datetime

# Persian Font
FONT_NAME = 'B Mitra'
FONT_SIZE = 11
FONT_SIZE_SM = 10
FONT_SIZE_LG = 14
FONT_SIZE_XL = 16

# Colors
BLUE_DARK = '1F4E79'
BLUE_MED = '2E75B6'
BLUE_LIGHT = 'BDD7EE'
GREEN = 'C6EFCE'
YELLOW = 'FFEB9C'
RED = 'FFC7CE'
ORANGE = 'FFD966'
WHITE = 'FFFFFF'
GRAY = 'F2F2F2'

# Styles
HEADER_FONT = Font(name=FONT_NAME, bold=True, size=FONT_SIZE, color=WHITE)
HEADER_FILL = PatternFill(start_color=BLUE_DARK, end_color=BLUE_DARK, fill_type='solid')
SUBHEADER_FILL = PatternFill(start_color=BLUE_MED, end_color=BLUE_MED, fill_type='solid')
GREEN_FILL = PatternFill(start_color=GREEN, end_color=GREEN, fill_type='solid')
YELLOW_FILL = PatternFill(start_color=YELLOW, end_color=YELLOW, fill_type='solid')
RED_FILL = PatternFill(start_color=RED, end_color=RED, fill_type='solid')
ORANGE_FILL = PatternFill(start_color=ORANGE, end_color=ORANGE, fill_type='solid')
GRAY_FILL = PatternFill(start_color=GRAY, end_color=GRAY, fill_type='solid')
BLUE_FILL = PatternFill(start_color=BLUE_LIGHT, end_color=BLUE_LIGHT, fill_type='solid')

CELL_FONT = Font(name=FONT_NAME, size=FONT_SIZE_SM)
BOLD_FONT = Font(name=FONT_NAME, size=FONT_SIZE_SM, bold=True)
TITLE_FONT = Font(name=FONT_NAME, size=FONT_SIZE_XL, bold=True, color=BLUE_DARK)
SUBTITLE_FONT = Font(name=FONT_NAME, size=FONT_SIZE_LG, bold=True, color=BLUE_MED)

THIN_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True, readingOrder=2)
RIGHT = Alignment(horizontal='right', vertical='center', wrap_text=True, readingOrder=2)

def sc(cell, font=CELL_FONT, fill=None, align=RIGHT):
    cell.font = font
    cell.alignment = align
    cell.border = THIN_BORDER
    if fill:
        cell.fill = fill

def add_title(ws, row, col_end, title):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=col_end)
    c = ws.cell(row=row, column=1, value=title)
    c.font = TITLE_FONT
    c.alignment = CENTER

def add_subtitle(ws, row, col_end, title):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=col_end)
    c = ws.cell(row=row, column=1, value=title)
    c.font = SUBTITLE_FONT
    c.alignment = CENTER

def add_headers(ws, row, headers):
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=i, value=h)
        sc(c, HEADER_FONT, HEADER_FILL, CENTER)

def add_data(ws, start_row, data, col_widths=None):
    for r, row_data in enumerate(data, start_row):
        for c_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c_idx, value=val)
            score_fill = None
            if isinstance(val, str):
                if val in ['✅', 'Confirmed', 'READY TO SEND', 'NEW', 'Approved']:
                    score_fill = GREEN_FILL
                elif val in ['🟡', 'Likely', 'P2', 'مهم']:
                    score_fill = YELLOW_FILL
                elif val in ['❌', 'REJECTED', 'EXPIRED']:
                    score_fill = RED_FILL
                elif '90' in str(val) or '95' in str(val) or '98' in str(val):
                    score_fill = GREEN_FILL
                elif '80' in str(val) or '85' in str(val):
                    score_fill = YELLOW_FILL
                elif '75' in str(val) or '78' in str(val):
                    score_fill = ORANGE_FILL
            sc(cell, fill=score_fill)
    if col_widths:
        for i, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

def main():
    wb = openpyxl.Workbook()
    if 'Sheet' in wb.sheetnames:
        del wb['Sheet']
    
    # ============ DASHBOARD (Sheet 1) ============
    ws = wb.create_sheet('داشبورد', 0)
    ws.sheet_view.rightToLeft = True
    
    add_title(ws, 1, 10, 'داشبورد شکار فرصت شغلی — Migration Hunter')
    add_subtitle(ws, 2, 10, f'آخرین بروزرسانی: {datetime.now().strftime("%Y-%m-%d")}')
    
    # Summary boxes
    row = 4
    add_subtitle(ws, row, 5, '📊 خلاصه وضعیت')
    
    row = 5
    add_headers(ws, row, ['شاخص', 'تعداد', 'شاخص', 'تعداد'])
    
    summary = [
        ['کل فرصت‌ها', 4, 'فرصت‌های جدید', 4],
        ['فرصت‌های TOHID', 1, 'فرصت‌های NEDA', 3],
        ['کارفرمایان تأیید شده', 5, 'Sponsorship تأیید', 4],
        ['درخواست‌های آماده', 4, 'درخواست‌های ارسال شده', 0],
        ['پاسخ دریافتی', 0, 'مصاحبه', 0],
        ['Job Offer', 0, 'ویزای در حال بررسی', 0],
    ]
    add_data(ws, row+1, summary, [25, 12, 25, 12])
    
    # Language Status
    row = 13
    add_subtitle(ws, row, 5, '📝 وضعیت زبان')
    row = 14
    add_headers(ws, row, ['متقاضی', 'English', 'German', 'IELTS', 'OET'])
    lang_data = [
        ['توحید', 'A2', 'A1', 'نیاز نیست', 'نیاز نیست'],
        ['ندا', 'A2', 'A1', 'نیاز برای ثبت‌نام', 'جایگزین IELTS'],
    ]
    add_data(ws, row+1, lang_data, [14, 12, 12, 22, 18])
    
    # Top Opportunities
    row = 18
    add_subtitle(ws, row, 10, '🔥 فرصت‌های برتر')
    row = 19
    add_headers(ws, row, ['ردیف', 'متقاضی', 'کشور', 'کارفرما', 'عنوان', 'Sponsorship', 'زبان', 'تناسب', 'وضعیت', 'اقدام بعدی'])
    top_jobs = [
        [1, 'ندا', 'nz', 'Health New Zealand', 'ماما', '✅', 'IELTS/OET', '80/100', 'جدید', 'ارسال درخواست'],
        [2, 'ندا', 'nz', 'RGH Global', 'ماما', '✅', 'IELTS/OET', '78/100', 'جدید', 'ارسال درخواست'],
        [3, 'ندا', 'au', 'Hassett Group', 'ماما', '✅', 'IELTS/OET', '78/100', 'جدید', 'ارسال درخواست'],
        [4, 'توحید', 'de', 'شرکت‌های IT', 'مدیر IT', '🟡', 'A2 کافی', '75/100', 'جدید', 'جستجو و ارسال CV'],
    ]
    add_data(ws, row+1, top_jobs, [6, 10, 8, 18, 12, 12, 14, 10, 10, 18])
    
    ws.column_dimensions['A'].width = 6
    for i in range(2, 11):
        ws.column_dimensions[get_column_letter(i)].width = 18
    
    # ============ TOHID JOBS ============
    ws2 = wb.create_sheet('فرصت‌های توحید')
    ws2.sheet_view.rightToLeft = True
    
    add_title(ws2, 1, 12, 'فرصت‌های شغلی توحید آرجمند')
    add_subtitle(ws2, 2, 12, 'IT Operations / Infrastructure / Management')
    
    row = 4
    add_headers(ws2, row, ['Job ID', 'کارفرما', 'کشور', 'شهر', 'عنوان', 'حقوق', 'Sponsorship', 'زبان', 'ثبت‌نام', 'تناسب', 'وضعیت', 'اقدام'])
    
    tohid_jobs = [
        ['JOB-004', 'SAP', 'آلمان', 'والدورف', 'مدیر IT', '€55-80K', '🟢', 'انگلیسی', '—', '82/100', 'جدید', 'جستجو'],
        ['JOB-005', 'Siemens', 'آلمان', 'مونیخ', 'مدیر IT', '€60-90K', '🟢', 'انگلیسی', '—', '80/100', 'جدید', 'جستجو'],
        ['JOB-006', 'Deutsche Telekom', 'آلمان', 'بون', 'مدیر IT', '€50-75K', '🟢', 'انگلیسی', '—', '78/100', 'جدید', 'جستجو'],
        ['JOB-007', 'Allianz', 'آلمان', 'مونیخ', 'مدیر IT', '€55-85K', '🟢', 'انگلیسی', '—', '77/100', 'جدید', 'جستجو'],
        ['JOB-008', 'BMW', 'آلمان', 'مونیخ', 'مدیر IT', '€60-90K', '🟢', 'انگلیسی', '—', '76/100', 'جدید', 'جستجو'],
    ]
    add_data(ws2, row+1, tohid_jobs, [12, 18, 10, 12, 12, 14, 12, 12, 10, 10, 10, 14])
    
    # Chart
    chart = BarChart()
    chart.type = "col"
    chart.title = "تناسب فرصت‌های توحید"
    chart.y_axis.title = "امتیاز"
    chart.style = 10
    data_ref = Reference(ws2, min_col=10, min_row=4, max_row=9)
    cats = Reference(ws2, min_col=2, min_row=5, max_row=9)
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats)
    ws2.add_chart(chart, "A12")
    
    # ============ NEDA JOBS ============
    ws3 = wb.create_sheet('فرصت‌های ندا')
    ws3.sheet_view.rightToLeft = True
    
    add_title(ws3, 1, 12, 'فرصت‌های شغلی ندا آرجمند')
    add_subtitle(ws3, 2, 12, 'Midwife / Registered Midwife / Clinical Midwife')
    
    row = 4
    add_headers(ws3, row, ['Job ID', 'کارفرما', 'کشور', 'شهر', 'عنوان', 'حقوق', 'Sponsorship', 'زبان', 'ثبت‌نام', 'تناسب', 'وضعیت', 'اقدام'])
    
    neda_jobs = [
        ['JOB-001', 'Health New Zealand', 'nz', 'سراسر کشور', 'ماما', '75-106K NZD', '✅', 'IELTS/OET', 'Midwifery Council', '80/100', 'جدید', 'ارسال'],
        ['JOB-002', 'RGH Global', 'nz', 'سراسر کشور', 'ماما', '75-106K NZD', '✅', 'IELTS/OET', 'Midwifery Council', '78/100', 'جدید', 'ارسال'],
        ['JOB-003', 'Hassett Group', 'au', 'Melbourne', 'ماما', '80-120K AUD', '✅', 'IELTS/OET', 'AHPRA', '78/100', 'جدید', 'ارسال'],
        ['JOB-009', 'Talent Angels', 'au', 'Victoria', 'ماما', '80-120K AUD', '✅', 'IELTS/OET', 'AHPRA', '76/100', 'جدید', 'ارسال'],
        ['JOB-010', 'HealthX', 'au', 'سراسر کشور', 'ماما', '67-95K AUD', '✅', 'IELTS/OET', 'AHPRA', '74/100', 'جدید', 'ارسال'],
    ]
    add_data(ws3, row+1, neda_jobs, [12, 18, 8, 14, 12, 16, 10, 14, 18, 10, 10, 12])
    
    # Chart
    chart2 = BarChart()
    chart2.type = "col"
    chart2.title = "تناسب فرصت‌های ندا"
    chart2.y_axis.title = "امتیاز"
    chart2.style = 10
    data_ref2 = Reference(ws3, min_col=10, min_row=4, max_row=9)
    cats2 = Reference(ws3, min_col=2, min_row=5, max_row=9)
    chart2.add_data(data_ref2, titles_from_data=True)
    chart2.set_categories(cats2)
    ws3.add_chart(chart2, "A12")
    
    # ============ EMPLOYERS ============
    ws4 = wb.create_sheet('کارفرمایان')
    ws4.sheet_view.rightToLeft = True
    
    add_title(ws4, 1, 10, 'بانک کارفرمایان')
    
    row = 3
    add_headers(ws4, row, ['کارفرما', 'کشور', 'صنعت', 'وب‌سایت', ' استخدام بین‌المللی', 'Sponsorship', 'امتیاز', 'مناسب برای', 'تاریخ بررسی', 'یادداشت'])
    
    employers = [
        ['Health New Zealand', 'nz', 'بهداشت', 'healthnz.govt.nz', '✅', '✅', '95/100', 'ندا', '2026-08-18', 'دولتی'],
        ['RGH Global', 'nz', 'بهداشت', 'rgh-global.com', '✅', '✅', '85/100', 'ندا', '2026-08-18', 'آژانس'],
        ["St Vincent's Health", 'au', 'بهداشت', 'svha.org.au', '✅', '✅', '85/100', 'ندا', '2026-08-18', 'بیمارستان'],
        ['Hassett Group', 'au', 'بهداشت', 'hassett.com.au', '✅', '✅', '80/100', 'ندا', '2026-08-18', 'آژانس'],
        ['HealthX', 'au', 'بهداشت', 'healthx.com.au', '✅', '✅', '82/100', 'ندا', '2026-08-18', 'آژانس'],
        ['SAP', 'de', 'IT', 'sap.com', '✅', '🟡', '80/100', 'توحید', '2026-08-18', 'شرکت بزرگ'],
        ['Siemens', 'de', 'IT', 'siemens.com', '✅', '🟡', '78/100', 'توحید', '2026-08-18', 'شرکت بزرگ'],
    ]
    add_data(ws4, row+1, employers, [20, 8, 10, 20, 16, 12, 12, 12, 14, 12])
    
    # ============ SOURCES ============
    ws5 = wb.create_sheet('منابع')
    ws5.sheet_view.rightToLeft = True
    
    add_title(ws5, 1, 10, 'بانک منابع')
    
    row = 3
    add_headers(ws5, row, ['منبع', 'نوع', 'کشور', 'اعتماد', 'کیفیت', 'Sponsorship', 'TOHID', 'NEDA', 'آخرین بررسی', 'یادداشت'])
    
    sources = [
        ['Health NZ', 'دولتی', 'nz', '95', '90', '95', '70', '98', '2026-08-18', 'فعال'],
        ['RGH Global', 'کاریابی', 'nz', '85', '85', '90', '30', '92', '2026-08-18', 'متخصص مامایی'],
        ['SEEK Australia', 'هیئت شغلی', 'au', '90', '80', '70', '60', '80', '2026-08-18', 'نیاز به فیلتر'],
        ['Arbeitnow', 'هیئت شغلی', 'de', '85', '85', '80', '90', '40', '2026-08-18', 'IT focused'],
        ['LinkedIn', 'اجتماعی', 'global', '80', '75', '60', '75', '75', '2026-08-18', 'ترکیبی'],
    ]
    add_data(ws5, row+1, sources, [16, 12, 10, 10, 10, 12, 10, 10, 14, 16])
    
    # ============ APPLICATIONS ============
    ws6 = wb.create_sheet('درخواست‌ها')
    ws6.sheet_view.rightToLeft = True
    
    add_title(ws6, 1, 10, 'بانک درخواست‌ها')
    
    row = 3
    add_headers(ws6, row, ['App ID', 'متقاضی', 'کارفرما', 'کشور', 'تاریخ', 'CV', 'Cover Letter', 'وضعیت', 'پاسخ', 'اقدام بعدی'])
    
    apps = [
        ['APP-001', 'ندا', 'Health New Zealand', 'nz', '2026-08-18', '✅', '✅', 'آماده ارسال', '—', 'تأیید کاربر'],
        ['APP-002', 'ندا', 'RGH Global', 'nz', '2026-08-18', '✅', '✅', 'آماده ارسال', '—', 'تأیید کاربر'],
        ['APP-003', 'ندا', 'Hassett Group', 'au', '2026-08-18', '✅', '✅', 'آماده ارسال', '—', 'تأیید کاربر'],
        ['APP-004', 'توحید', 'IT Companies DE', 'de', '2026-08-18', '✅', '✅', 'آماده ارسال', '—', 'تأیید کاربر'],
    ]
    add_data(ws6, row+1, apps, [12, 12, 18, 8, 14, 8, 14, 16, 10, 16])
    
    # ============ VISA ============
    ws7 = wb.create_sheet('ویزا')
    ws7.sheet_view.rightToLeft = True
    
    add_title(ws7, 1, 8, 'بانک اطلاعات ویزا')
    
    row = 3
    add_headers(ws7, row, ['کشور', 'نوع ویزا', 'الزام شغلی', 'الزام زبان', 'خانواده', 'ثبت‌نام', 'منبع رسمی', 'آخرین بررسی'])
    
    visa_data = [
        ['nz', 'AEWV', 'Job offer', 'ANZSCO 1-2: None', '✅', 'جداگانه', 'immigration.govt.nz', '2026-08-18'],
        ['au', '482 TSS', 'Job offer', 'IELTS 5.0', '✅', 'AHPRA', 'homeaffairs.gov.au', '2026-08-18'],
        ['de', 'EU Blue Card', '€45,934+', 'None (IT)', '✅', 'Anerkennung', 'make-it-in-germany.com', '2026-08-18'],
        ['ca', 'LMIA', 'Job offer', 'IELTS 6.5', '✅', 'مختلف', 'canada.ca', '2026-08-18'],
    ]
    add_data(ws7, row+1, visa_data, [8, 14, 14, 18, 10, 14, 22, 14])
    
    # ============ REGISTRATION ============
    ws8 = wb.create_sheet('ثبت‌نام')
    ws8.sheet_view.rightToLeft = True
    
    add_title(ws8, 1, 10, 'بانک ثبت‌نام حرفه‌ای')
    
    row = 3
    add_headers(ws8, row, ['کشور', 'متقاضی', 'حرفه', 'نهاد', 'مسیر', 'زبان', 'آزمون', 'هزینه', 'زمان', 'منبع'])
    
    reg_data = [
        ['nz', 'ندا', 'ماما', 'Midwifery Council', 'Overseas Midwife', 'IELTS 7.0/OET', 'ممکن است', 'NZ$485', '۱۲ هفته', 'midwiferycouncil.health.nz'],
        ['au', 'ندا', 'ماما', 'AHPRA', 'International', 'IELTS 7.0/OET', 'ممکن است', '—', '—', 'ahpra.gov.au'],
        ['de', 'توحید', 'IT', 'Chamber', 'Anerkennung', 'None for visa', '—', '—', '—', 'anerkennung-in-deutschland.de'],
    ]
    add_data(ws8, row+1, reg_data, [8, 12, 10, 18, 16, 16, 12, 12, 12, 28])
    
    # ============ LANGUAGE ============
    ws9 = wb.create_sheet('زبان')
    ws9.sheet_view.rightToLeft = True
    
    add_title(ws9, 1, 8, 'وضعیت زبان')
    
    row = 3
    add_headers(ws9, row, ['متقاضی', 'English', 'German', 'IELTS نیاز', 'OET نیاز', 'وضعیت', 'اقدام', 'اولویت'])
    
    lang_full = [
        ['توحید', 'A2', 'A1', 'نیاز نیست', 'نیاز نیست', 'کافی برای IT', 'ادامه A2', 'مهم'],
        ['ندا', 'A2', 'A1', 'IELTS Academic 7.0', 'OET B', 'نیاز به ارتقاء', 'ثبت‌نام OET', 'فوری'],
    ]
    add_data(ws9, row+1, lang_full, [14, 10, 10, 20, 14, 18, 16, 12])
    
    # ============ SEARCH HISTORY ============
    ws10 = wb.create_sheet('تاریخچه')
    ws10.sheet_view.rightToLeft = True
    
    add_title(ws10, 1, 10, 'تاریخچه جستجو')
    
    row = 3
    add_headers(ws10, row, ['تاریخ', 'متقاضی', 'کشورها', 'منابع', 'منابع جدید', 'فرصت‌ها', 'معتبر', 'درخواست‌ها', 'پاسخ', 'Offer'])
    
    history = [
        ['2026-08-18', 'هر دو', 'nz,au,de', '5', '5', '4', '4', '4', '0', '0'],
    ]
    add_data(ws10, row+1, history, [14, 12, 14, 10, 12, 10, 10, 12, 10, 10])
    
    # Save
    filename = 'MigrationHunter/dashboard/MigrationHunter_Dashboard.xlsx'
    wb.save(filename)
    print(f"✅ Dashboard v2 created: {filename}")
    print(f"📊 {len(wb.sheetnames)} sheets:")
    for s in wb.sheetnames:
        print(f"   - {s}")

if __name__ == '__main__':
    main()
