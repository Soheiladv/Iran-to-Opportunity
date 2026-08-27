#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Migration Hunter FINAL Excel - Complete with Charts, Colors, Email Features
"""

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from datetime import datetime

FONT = 'B Mitra'
BD = '1F4E79'
BM = '2E75B6'
GR = 'C6EFCE'
YL = 'FFEB9C'
RD = 'FFC7CE'
OR = 'FFD966'
BL = 'BDD7EE'
WH = 'FFFFFF'
GY = 'F2F2F2'

HF = Font(name=FONT, bold=True, size=11, color=WH)
HFI = PatternFill(start_color=BD, end_color=BD, fill_type='solid')
SF = PatternFill(start_color=BM, end_color=BM, fill_type='solid')
GF = PatternFill(start_color=GR, end_color=GR, fill_type='solid')
YF = PatternFill(start_color=YL, end_color=YL, fill_type='solid')
RF = PatternFill(start_color=RD, end_color=RD, fill_type='solid')
OF = PatternFill(start_color=OR, end_color=OR, fill_type='solid')
BF = PatternFill(start_color=BL, end_color=BL, fill_type='solid')
GYF = PatternFill(start_color=GY, end_color=GY, fill_type='solid')
CF = Font(name=FONT, size=10)
BF2 = Font(name=FONT, size=10, bold=True)
TF = Font(name=FONT, size=14, bold=True, color=BD)
SF2 = Font(name=FONT, size=12, bold=True, color=BM)
LF = Font(name=FONT, size=10, color='0563C1', underline='single')
TB = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
CA = Alignment(horizontal='center', vertical='center', wrap_text=True, readingOrder=2)
RA = Alignment(horizontal='right', vertical='center', wrap_text=True, readingOrder=2)

def sc(c, f=CF, fl=None, a=RA):
    c.font=f; c.alignment=a; c.border=TB
    if fl: c.fill=fl

def st(ws, r, ce, t):
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ce)
    c=ws.cell(row=r, column=1, value=t); c.font=TF; c.alignment=CA

def sh(ws, r, h):
    for i,v in enumerate(h,1):
        c=ws.cell(row=r, column=i, value=v); sc(c, HF, HFI, CA)

def sd(ws, sr, d, cw=None):
    for r, rd in enumerate(d, sr):
        for ci, v in enumerate(rd, 1):
            c=ws.cell(row=r, column=ci, value=v)
            fl=None
            sv=str(v)
            if sv in ['✅','Confirmed','آماده','جدید','P1','🔥 P1','READY TO SEND','آماده ارسال']: fl=GF
            elif sv in ['🟡','Likely','P2','🟢 P2','مهم','آماده نیست','نیاز']: fl=YF
            elif sv in ['🔵 P3','شناسایی']: fl=BF
            elif sv in ['❌','REJECTED','EXPIRED']: fl=RF
            elif '🔗' in sv: fl=BF
            sc(c, fl=fl)
    if cw:
        for i,w in enumerate(cw,1): ws.column_dimensions[get_column_letter(i)].width=w

def add_chart(ws, title, data_start, data_end, cat_col, cat_start, cat_end, location):
    ch = BarChart()
    ch.type = "col"
    ch.title = title
    ch.y_axis.title = "امتیاز"
    ch.style = 10
    d = Reference(ws, min_col=data_start, min_row=cat_start-1, max_col=data_end, max_row=cat_end)
    c = Reference(ws, min_col=cat_col, min_row=cat_start, max_row=cat_end)
    ch.add_data(d, titles_from_data=True)
    ch.set_categories(c)
    ws.add_chart(ch, location)

def main():
    wb = openpyxl.Workbook()
    if 'Sheet' in wb.sheetnames: del wb['Sheet']
    
    # ========== 1. شیت مستر ==========
    ws = wb.create_sheet("📊 شیت مستر", 0)
    ws.sheet_view.rightToLeft = True
    
    st(ws, 1, 14, "گزارش جامع شکار فرصت شغلی — Migration Hunter v6.0")
    st(ws, 2, 14, "خانواده آرجمند — توحید و ندا | انگلیسی A2 + آلمانی A1")
    
    r = 4; ws.cell(row=r, column=1, value="📊 خلاصه وضعیت کلی").font = SF2
    r = 5; sh(ws, r, ['شاخص', 'ندا', 'توحید', 'جمع'])
    sd(ws, r+1, [
        ['تعداد فرصت‌ها', 5, 5, 10],
        ['Sponsorship تأیید', 5, 0, 5],
        ['Sponsorship احتمالی', 0, 5, 5],
        ['ایمیل آماده', 4, 0, 4],
        ['کاور لیتر آماده', 4, 0, 4],
        ['امتیاز تناسب', '80/100', '76/100', '78/100'],
    ], [25, 18, 18, 18])
    
    r = 12; ws.cell(row=r, column=1, value="🌍 مقایسه کشورها").font = SF2
    r = 13; sh(ws, r, ['کشور', 'امتیاز', 'ندا', 'توحید', 'Sponsorship', 'زبان', 'سرعت', 'شانس'])
    sd(ws, r+1, [
        ['nz', 85, 'عالی', 'خوب', '۱۰۰%', 'IELTS/OET', 'سریع', 85],
        ['au', 78, 'خوب', 'متوسط', '۹۰%', 'IELTS/OET', 'متوسط', 78],
        ['de', 65, 'محدود', 'خوب', '۷۰%', 'A1 کافی(IT)', 'متوسط', 65],
        ['ca', 60, 'محدود', 'متوسط', '۶۰%', 'IELTS 6.5', 'طولانی', 60],
    ], [12, 12, 12, 12, 12, 15, 12, 12])
    
    r = 19; ws.cell(row=r, column=1, value="📈 امتیاز تناسب مسیر").font = SF2
    r = 20; sh(ws, r, ['معیار', 'ندا-nz', 'ندا-au', 'توحید-de', 'توحید-nz'])
    sd(ws, r+1, [
        ['تناسب حرفه‌ای', 85, 80, 75, 70],
        ['تناسب مهاجرتی', 80, 75, 80, 75],
        ['تناسب زبان', 60, 60, 70, 65],
        ['تناسب حمایت', 90, 85, 70, 75],
        ['⭐ کل', 80, 76, 76, 76],
    ], [22, 15, 15, 15, 15])
    
    # Chart
    add_chart(ws, "مقایسه امتیاز تناسب", 2, 5, 1, 21, 25, "A26")
    
    # ========== 2. ندا ==========
    ws2 = wb.create_sheet("👩 ندا — فرصت‌ها")
    ws2.sheet_view.rightToLeft = True
    
    st(ws2, 1, 16, "فرصت‌های ندا آرجمند — ماما (Midwife)")
    st(ws2, 2, 16, "⚠️ لینک شغل + ایمیل + کاور لیتر برای هر فرصت")
    
    r = 4; ws2.cell(row=r, column=1, value="📋 مشخصات").font = SF2
    sd(ws2, r+1, [
        ['نام', 'ندا آرجمند'],
        ['سن', '۳۸ سال'],
        ['حرفه', 'ماما'],
        ['سابقه', 'بیمارستان میلاد تهران'],
        ['زبان', 'A2 English + A1 German'],
        ['LinkedIn', 'linkedin.com/in/neda-arjmand'],
    ], [22, 30])
    
    r = 12; ws2.cell(row=r, column=1, value="🎯 فرصت‌های شغلی با ایمیل").font = SF2
    r = 13; sh(ws2, r, ['اولویت', 'کارفرما', 'کشور', 'عنوان', 'حقوق', 'Sponsorship', 'تناسب', '🔗 لینک', '📧 ایمیل', '📝 کاور', 'اقدام'])
    sd(ws2, r+1, [
        ['🔥 P1', 'Health New Zealand', 'nz', 'ماما', '75-106K NZD', '✅', '80/100', '🔗', '✅', '✅', 'ارسال'],
        ['🔥 P1', 'RGH Global', 'nz', 'ماما', '75-106K NZD', '✅', '78/100', '🔗', '✅', '✅', 'ارسال'],
        ['🟢 P2', 'Hassett Group', 'au', 'ماما', '80-120K AUD', '✅', '78/100', '🔗', '✅', '✅', 'ارسال'],
        ['🟢 P2', 'Talent Angels', 'au', 'ماما', '80-120K AUD', '✅', '76/100', '🔗', '🟡', '🟡', 'بررسی'],
        ['🟢 P2', 'HealthX', 'au', 'ماما', '67-95K AUD', '✅', '74/100', '🔗', '✅', '✅', 'ارسال'],
    ], [12, 20, 8, 12, 16, 12, 10, 10, 12, 12, 12])
    
    # Add hyperlinks
    links = [
        'https://www.healthnz.govt.nz/careers/international',
        'https://www.rgh-global.com/jobs/midwife-with-sponsorship/',
        'https://www.hassett.com.au/job-details/registered-midwife-visa-sponsorship-pathway-to-pr-in-healthcare-medicine-jobs-1475247',
        'https://au.linkedin.com/jobs/view/registered-midwife-at-talent-angels-4231144020',
        'https://www.healthx.com.au/international-candidates/visa-sponsorship',
    ]
    for i, link in enumerate(links):
        cell = ws2.cell(row=r+1+i, column=8)
        cell.hyperlink = link
        cell.font = LF
        cell.value = '🔗'
    
    add_chart(ws2, "تناسب فرصت‌های ندا", 7, 7, 2, 14, 18, "A20")
    
    # ========== 3. توحید ==========
    ws3 = wb.create_sheet("👨 توحید — فرصت‌ها")
    ws3.sheet_view.rightToLeft = True
    
    st(ws3, 1, 16, "فرصت‌های توحید آرجمند — مدیر IT")
    st(ws3, 2, 16, "⚠️ لینک شغل + ایمیل + کاور لیتر برای هر فرصت")
    
    r = 4; ws3.cell(row=r, column=1, value="📋 مشخصات").font = SF2
    sd(ws3, r+1, [
        ['نام', 'توحید آرجمند'],
        ['سن', '۴۶ سال'],
        ['حرفه', 'مدیر IT'],
        ['سابقه', '۱۹ سال تجربه IT'],
        ['زبان', 'A2 English + A1 German'],
        ['LinkedIn', 'linkedin.com/in/tohid-arjmand'],
    ], [22, 30])
    
    r = 12; ws3.cell(row=r, column=1, value="🎯 فرصت‌های شغلی با ایمیل").font = SF2
    r = 13; sh(ws3, r, ['اولویت', 'کارفرما', 'کشور', 'عنوان', 'حقوق', 'Sponsorship', 'تناسب', '🔗 لینک', '📧 ایمیل', '📝 کاور', 'اقدام'])
    sd(ws3, r+1, [
        ['🟢 P2', 'SAP', 'de', 'مدیر IT', '€55-80K', '🟢', '82/100', '🔗', '🟡', '🟡', 'جستجو'],
        ['🟢 P2', 'Siemens', 'de', 'مدیر IT', '€60-90K', '🟢', '80/100', '🔗', '🟡', '🟡', 'جستجو'],
        ['🟢 P2', 'Deutsche Telekom', 'de', 'مدیر IT', '€50-75K', '🟢', '78/100', '🔗', '🟡', '🟡', 'جستجو'],
        ['🟢 P2', 'Allianz', 'de', 'مدیر IT', '€55-85K', '🟢', '77/100', '🔗', '🟡', '🟡', 'جستجو'],
        ['🟢 P2', 'BMW', 'de', 'مدیر IT', '€60-90K', '🟢', '76/100', '🔗', '🟡', '🟡', 'جستجو'],
    ], [12, 20, 8, 12, 14, 12, 10, 10, 12, 12, 12])
    
    # Add hyperlinks
    links3 = [
        'https://www.sap.com/about/careers.html',
        'https://www.siemens.com/global/en/company/jobs.html',
        'https://www.telekom.com/en/careers',
        'https://www.allianz.com/en/careers.html',
        'https://www.bmwgroup.com/en/careers.html',
    ]
    for i, link in enumerate(links3):
        cell = ws3.cell(row=r+1+i, column=8)
        cell.hyperlink = link
        cell.font = LF
        cell.value = '🔗'
    
    add_chart(ws3, "تناسب فرصت‌های توحید", 7, 7, 2, 14, 18, "A20")
    
    # ========== 4. مقایسه ==========
    ws4 = wb.create_sheet("🌍 مقایسه")
    ws4.sheet_view.rightToLeft = True
    
    st(ws4, 1, 10, "مقایسه جامع کشورها")
    r = 3; sh(ws4, r, ['کشور', 'امتیاز', 'ندا', 'توحید', 'سرعت', 'هزینه', 'زبان', 'خانواده'])
    sd(ws4, r+1, [
        ['nz', 85, 83, 55, 85, 70, 70, 95],
        ['au', 78, 76, 50, 70, 60, 70, 90],
        ['de', 65, 30, 72, 65, 75, 50, 85],
        ['ca', 60, 25, 55, 55, 65, 65, 85],
    ], [12, 12, 12, 12, 12, 12, 12, 12])
    
    add_chart(ws4, "مقایسه امتیاز کلی", 2, 2, 1, 4, 7, "A9")
    
    r = 15; ws4.cell(row=r, column=1, value="📈 سناریوها").font = SF2
    r = 16; sh(ws4, r, ['سناریو', 'nz', 'au', 'de', 'ca'])
    sd(ws4, r+1, [
        ['ندا اصلی', 83, 76, 30, 25],
        ['توحید اصلی', 55, 50, 72, 55],
        ['هر دو مستقل', 80, 70, 65, 50],
    ], [18, 12, 12, 12, 12])
    
    # Pie Chart
    pie = PieChart()
    pie.title = "توزیع سناریوها"
    pie.style = 10
    d = Reference(ws4, min_col=2, min_row=16, max_col=5, max_row=19)
    c = Reference(ws4, min_col=1, min_row=17, max_row=19)
    pie.add_data(d, titles_from_data=True)
    pie.set_categories(c)
    ws4.add_chart(pie, "A21")
    
    # ========== 5. کارفرمایان ==========
    ws5 = wb.create_sheet("🏢 کارفرمایان")
    ws5.sheet_view.rightToLeft = True
    
    st(ws5, 1, 10, "کارفرمایان شناسایی شده")
    r = 3; sh(ws5, r, ['نام', 'کشور', 'نوع', 'تخصص', 'Sponsorship', 'وب‌سایت', '适合', 'وضعیت'])
    sd(ws5, r+1, [
        ['Health New Zealand', 'nz', 'دولتی', 'بهداشت', '✅', 'healthnz.govt.nz', 'ندا', 'تأیید'],
        ['RGH Global', 'nz', 'آژانس', 'مامایی', '✅', 'rgh-global.com', 'ندا', 'تأیید'],
        ["St Vincent's Health", 'au', 'بیمارستان', 'بهداشت', '✅', 'svha.org.au', 'ندا', 'تأیید'],
        ['Hassett Group', 'au', 'آژانس', 'بهداشت', '✅', 'hassett.com.au', 'ندا', 'تأیید'],
        ['HealthX', 'au', 'آژانس', 'بهداشت', '✅', 'healthx.com.au', 'ندا', 'تأیید'],
        ['SAP', 'de', 'شرکت IT', 'IT', '🟡', 'sap.com', 'توحید', 'بررسی'],
        ['Siemens', 'de', 'شرکت IT', 'IT', '🟡', 'siemens.com', 'توحید', 'بررسی'],
    ], [22, 10, 12, 12, 12, 22, 12, 12])
    
    # ========== 6. سازمان‌ها ==========
    ws6 = wb.create_sheet("🏛 سازمان‌ها")
    ws6.sheet_view.rightToLeft = True
    
    st(ws6, 1, 8, "سازمان‌های دولتی")
    r = 3; sh(ws6, r, ['کشور', 'سازمان', 'وب‌سایت', 'مناسب برای', 'موضوع'])
    sd(ws6, r+1, [
        ['nz', 'Immigration NZ', 'immigration.govt.nz', 'هر دو', 'ویزا'],
        ['nz', 'Midwifery Council', 'midwiferycouncil.health.nz', 'ندا', 'ثبت‌نام'],
        ['nz', 'Health NZ', 'healthnz.govt.nz', 'ندا', 'استخدام'],
        ['au', 'Home Affairs', 'homeaffairs.gov.au', 'هر دو', 'ویزا'],
        ['au', 'AHPRA', 'ahpra.gov.au', 'ندا', 'ثبت‌نام'],
        ['de', 'Make it in Germany', 'make-it-in-germany.com', 'هر دو', 'مهاجرت'],
        ['de', 'Bundesagentur', 'arbeitsagentur.de', 'توحید', 'بازار کار'],
    ], [12, 22, 28, 15, 15])
    
    # ========== 7. درخواست‌ها ==========
    ws7 = wb.create_sheet("📤 درخواست‌ها")
    ws7.sheet_view.rightToLeft = True
    
    st(ws7, 1, 12, "درخواست‌های آماده ارسال")
    r = 3; sh(ws7, r, ['App ID', 'متقاضی', 'کارفرما', 'کشور', 'CV', 'کاور', 'ایمیل', 'وضعیت', 'پاسخ', 'اقدام بعدی'])
    sd(ws7, r+1, [
        ['APP-001', 'ندا', 'Health NZ', 'nz', '✅', '✅', '✅', 'آماده ارسال', '—', 'تأیید کاربر'],
        ['APP-002', 'ندا', 'RGH Global', 'nz', '✅', '✅', '✅', 'آماده ارسال', '—', 'تأیید کاربر'],
        ['APP-003', 'ندا', 'Hassett Group', 'au', '✅', '✅', '✅', 'آماده ارسال', '—', 'تأیید کاربر'],
        ['APP-004', 'توحید', 'SAP', 'de', '✅', '🟡', '🟡', 'آماده نیست', '—', 'تکمیل'],
    ], [12, 12, 18, 8, 8, 10, 10, 16, 10, 16])
    
    # ========== 8. ویزا ==========
    ws8 = wb.create_sheet("🛂 ویزا")
    ws8.sheet_view.rightToLeft = True
    
    st(ws8, 1, 8, "اطلاعات ویزا")
    r = 3; sh(ws8, r, ['کشور', 'نوع ویزا', 'الزام شغلی', 'زبان', 'خانواده', 'ثبت‌نام', 'منبع', 'تاریخ'])
    sd(ws8, r+1, [
        ['nz', 'AEWV', 'Job offer', 'ANZSCO 1-2: None', '✅', 'جداگانه', 'immigration.govt.nz', '2026-08-18'],
        ['au', '482 TSS', 'Job offer', 'IELTS 5.0', '✅', 'AHPRA', 'homeaffairs.gov.au', '2026-08-18'],
        ['de', 'EU Blue Card', '€45,934+', 'None (IT)', '✅', 'Anerkennung', 'make-it-in-germany.com', '2026-08-18'],
        ['ca', 'LMIA', 'Job offer', 'IELTS 6.5', '✅', 'مختلف', 'canada.ca', '2026-08-18'],
    ], [8, 14, 14, 18, 10, 14, 22, 14])
    
    # ========== 9. ثبت‌نام ==========
    ws9 = wb.create_sheet("📋 ثبت‌نام")
    ws9.sheet_view.rightToLeft = True
    
    st(ws9, 1, 10, "ثبت‌نام حرفه‌ای")
    r = 3; sh(ws9, r, ['کشور', 'متقاضی', 'حرفه', 'نهاد', 'زبان', 'آزمون', 'هزینه', 'زمان', 'منبع', 'وضعیت'])
    sd(ws9, r+1, [
        ['nz', 'ندا', 'ماما', 'Midwifery Council', 'IELTS 7.0/OET', 'ممکن است', 'NZ$485', '۱۲ هفته', 'midwiferycouncil.health.nz', 'نیاز'],
        ['au', 'ندا', 'ماما', 'AHPRA', 'IELTS 7.0/OET', 'ممکن است', '—', '—', 'ahpra.gov.au', 'نیاز'],
        ['de', 'توحید', 'IT', 'Chamber', 'None for visa', '—', '—', '—', 'anerkennung-in-deutschland.de', 'بررسی'],
    ], [8, 12, 10, 18, 16, 12, 12, 12, 28, 10])
    
    # ========== 10. زبان ==========
    ws10 = wb.create_sheet("📝 زبان")
    ws10.sheet_view.rightToLeft = True
    
    st(ws10, 1, 8, "وضعیت زبان")
    r = 3; sh(ws10, r, ['متقاضی', 'English', 'German', 'IELTS نیاز', 'OET نیاز', 'وضعیت', 'اقدام', 'اولویت'])
    sd(ws10, r+1, [
        ['توحید', 'A2', 'A1', 'نیاز نیست', 'نیاز نیست', 'کافی برای IT', 'ادامه A2', 'مهم'],
        ['ندا', 'A2', 'A1', 'IELTS Academic 7.0', 'OET B', 'نیاز به ارتقاء', 'ثبت‌نام OET', 'فوری'],
    ], [14, 10, 10, 20, 14, 18, 16, 12])
    
    # ========== 11. تاریخچه ==========
    ws11 = wb.create_sheet("📜 تاریخچه")
    ws11.sheet_view.rightToLeft = True
    
    st(ws11, 1, 10, "تاریخچه جستجو")
    r = 3; sh(ws11, r, ['تاریخ', 'متقاضی', 'کشورها', 'منابع', 'جدید', 'فرصت‌ها', 'معتبر', 'درخواست‌ها', 'پاسخ', 'Offer'])
    sd(ws11, r+1, [['2026-08-18', 'هر دو', 'nz,au,de', '5', '5', '10', '10', '4', '0', '0']], [14, 12, 14, 10, 10, 10, 10, 12, 10, 10])
    
    # ========== 12. خانواده ==========
    ws12 = wb.create_sheet("👨‍👩‍👧‍👦 خانواده")
    ws12.sheet_view.rightToLeft = True
    
    st(ws12, 1, 10, "استراتژی خروج خانواده")
    r = 3; sh(ws12, r, ['سناریو', 'متقاضی', 'کشور', 'تناسب', 'سرعت', 'هزینه', 'خانواده', 'امتیاز'])
    sd(ws12, r+1, [
        ['A: ندا اصلی', 'ندا', 'nz', 80, 'سریع', 'متوسط', 'عالی', 82],
        ['B: توحید اصلی', 'توحید', 'de', 76, 'متوسط', 'کم', 'خوب', 72],
        ['C: هر دو مستقل', 'هر دو', 'nz+au', 78, 'سریع', 'بالا', 'عالی', 80],
    ], [18, 12, 12, 12, 12, 12, 12, 12])
    
    r = 8; ws12.cell(row=r, column=1, value="🎯 پیشنهاد: سناریو C").font = BF2
    
    # ========== 13. اقدامات ==========
    ws13 = wb.create_sheet("🎯 اقدامات")
    ws13.sheet_view.rightToLeft = True
    
    st(ws13, 1, 8, "برنامه اقدام امروز")
    r = 3; sh(ws13, r, ['ردیف', 'اقدام', 'متقاضی', 'اولویت', 'وضعیت', 'یادداشت'])
    sd(ws13, r+1, [
        [1, 'ارسال Health NZ', 'ندا', '🔴 فوری', 'انجام نشده', 'بهترین فرصت'],
        [2, 'ارسال Hassett Group', 'ندا', '🔴 فوری', 'انجام نشده', 'حمایت ویزا'],
        [3, 'ارسال RGH Global', 'ندا', '🔴 فوری', 'انجام نشده', 'آژانس متخصص'],
        [4, 'جستجو IT آلمان', 'توحید', '🟡 متوسط', 'انجام نشده', 'بازار خوب'],
        [5, 'ثبت‌نام OET', 'ندا', '🟠 مهم', 'انجام نشده', 'پیش‌نیاز ثبت‌نام'],
    ], [8, 25, 12, 15, 15, 20])
    
    # ========== 14. ایمیل‌ها ==========
    ws14 = wb.create_sheet("📧 ایمیل‌ها")
    ws14.sheet_view.rightToLeft = True
    
    st(ws14, 1, 10, "ایمیل‌های درخواست — آماده ارسال")
    r = 3; sh(ws14, r, ['ردیف', 'متقاضی', 'کارفرما', 'موضوع', 'وضعیت', 'لینک', 'تاریخ', 'اقدام'])
    sd(ws14, r+1, [
        [1, 'ندا', 'Health New Zealand', 'International Midwife', '✅ آماده', '🔗', '2026-08-18', 'ارسال'],
        [2, 'ندا', 'RGH Global', 'Midwife Application', '✅ آماده', '🔗', '2026-08-18', 'ارسال'],
        [3, 'ندا', 'Hassett Group', 'Visa Sponsorship Inquiry', '✅ آماده', '🔗', '2026-08-18', 'ارسال'],
    ], [8, 12, 20, 25, 12, 10, 14, 12])
    
    # ========== 15. کاور لیتر ==========
    ws15 = wb.create_sheet("📝 کاور لیتر")
    ws15.sheet_view.rightToLeft = True
    
    st(ws15, 1, 10, "کاور لیترها — آماده ارسال")
    r = 3; sh(ws15, r, ['ردیف', 'متقاضی', 'کارفرما', 'عنوان', 'وضعیت', 'تاریخ', 'اقدام'])
    sd(ws15, r+1, [
        [1, 'ندا', 'Health New Zealand', 'Cover Letter - Midwife', '✅ آماده', '2026-08-18', 'بررسی و ارسال'],
        [2, 'ندا', 'RGH Global', 'Cover Letter - Midwife', '✅ آماده', '2026-08-18', 'بررسی و ارسال'],
        [3, 'ندا', 'Hassett Group', 'Cover Letter - Midwife', '✅ آماده', '2026-08-18', 'بررسی و ارسال'],
    ], [8, 12, 20, 25, 12, 14, 18])
    
    fn = 'MigrationHunter/dashboard/MigrationHunter_Dashboard.xlsx'
    wb.save(fn)
    print(f"✅ {fn}")
    print(f"📊 {len(wb.sheetnames)} sheets:")
    for s in wb.sheetnames:
        print(f"   - {s}")

if __name__ == '__main__':
    main()
