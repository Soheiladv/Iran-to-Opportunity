#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Excel جستجوی Saskatchewan Health Authority
"""
import os
from datetime import datetime
from pathlib import Path

BASE_DIR = Path(__file__).parent
DASHBOARD_DIR = BASE_DIR / "dashboard"
ARCHIVE_DIR = BASE_DIR / "archive"
NOW = datetime.now()
DATE_STR = NOW.strftime("%Y-%m-%d %H:%M")

def build():
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    wb = Workbook()

    FA = Font(name='B Mitra', size=11)
    FA_T = Font(name='B Mitra', size=14, bold=True)
    HD = Font(name='B Mitra', size=11, bold=True, color='FFFFFF')
    GF = PatternFill(start_color='92D050', end_color='92D050', fill_type='solid')
    YF = PatternFill(start_color='FFD966', end_color='FFD966', fill_type='solid')
    BF = PatternFill(start_color='9BC2E6', end_color='9BC2E6', fill_type='solid')
    HF = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    GR = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
    BD = Border(left=Side('thin'), right=Side('thin'), top=Side('thin'), bottom=Side('thin'))

    def sh(ws, r, c):
        cell = ws.cell(row=r, column=c)
        cell.font = HD; cell.fill = HF
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = BD

    def sc(ws, r, c, v=None, f=FA, fill=None):
        cell = ws.cell(row=r, column=c)
        if v is not None: cell.value = v
        cell.font = f
        cell.alignment = Alignment(vertical='center', wrap_text=True, horizontal='right')
        cell.border = BD
        if fill: cell.fill = fill
        return cell

    # ── شیت ۱: خلاصه Saskatchewan ──
    ws1 = wb.active
    ws1.title = " Saskatchewan | خلاصه"
    ws1.sheet_view.rightToLeft = True
    ws1.merge_cells('A1:D1')
    ws1.cell(1, 1, f"🇨🇦 Saskatchewan Health Authority — {DATE_STR}").font = FA_T

    h1 = ["فیلد | Field", "مقدار | Value", "توضیح | Notes", "وضعیت | Status"]
    for i, x in enumerate(h1, 1): ws1.cell(3, i, x)
    sh(ws1, 3, 4)

    info = [
        ("نام سازمان", "Saskatchewan Health Authority", "سازمان دولتی بهداشت", "✅"),
        ("کشور", "🇨🇦 کانادا", "استان Saskatchewan", "✅"),
        ("ایمیل اصلی", "SHAInternational@saskhealthauthority.ca", "استخدام بین‌المللی", "✅"),
        ("ایمیل عمومی", "careers@saskhealthauthority.ca", "مشاغل عمومی", "✅"),
        ("لینک شغلی", "https://www.saskhealthauthority.ca/careers", "صفحه مشاغل", "✅"),
        ("لینک بین‌المللی", "saskhealthauthority.ca/.../internationally-trained", "متخصصان بین‌المللی", "✅"),
        ("سیستم EOI", "Government of Saskatchewan website", "Expression of Interest", "✅"),
        ("ثبت‌نام ماما", "College of Midwives of Saskatchewan", "نیاز به ثبت‌نام", "⚠️"),
        ("ثبت‌نام IT", "نیاز ندارد", "IT نیاز به ثبت حرفه‌ای ندارد", "✅"),
        ("مسکن", "متوسط $289,600", "ارزان‌ترین در کانادا", "✅"),
        ("آفتاب", "۲,۰۰۰+ ساعت", "آفتابی‌ترین استان", "✅"),
        ("HHR Plan", "HHR@health.gov.sk.ca", "منابع انسانی بهداشت", "✅"),
    ]
    for r, (f, v, n, s) in enumerate(info, 4):
        sc(ws1, r, 1, f)
        c = sc(ws1, r, 2, v)
        if "✅" in s: c.fill = GR
        sc(ws1, r, 3, n)
        sc(ws1, r, 4, s)

    for col in 'ABCD': ws1.column_dimensions[col].width = 35

    # ── شیت ۲: ایمیل ندا ──
    ws2 = wb.create_sheet(" ندا — ایمیل | Neda Email")
    ws2.sheet_view.rightToLeft = True
    ws2.merge_cells('A1:B1')
    ws2.cell(1, 1, "👩 ایمیل ندا — Saskatchewan Health Authority").font = FA_T

    email_neda = [
        ("گیرنده", "SHAInternational@saskhealthauthority.ca"),
        ("موضوع", "International Midwife — Expression of Interest"),
        ("معرفی", "ماما با ۱۲+ سال تجربه بالینی"),
        ("محل کار", "بیمارستان میلاد تهران"),
        ("تخصص", "آنتناتال، لیبر، پست‌ناتال، بارداری پرخطر"),
        ("زبان", "English A2 — در حال آمادگی"),
        ("دلیل Saskatchewan", "مسکن ارزان، آفتاب، محیط حمایتی"),
        ("درخواست", "راهنمایی فرآیند EOI + ثبت‌نام"),
        ("پیوست", "CV"),
    ]
    for r, (f, v) in enumerate(email_neda, 3):
        sc(ws2, r, 1, f).fill = BF
        sc(ws2, r, 2, v)
    ws2.column_dimensions['A'].width = 25
    ws2.column_dimensions['B'].width = 50

    # ── شیت ۳: ایمیل توحید ──
    ws3 = wb.create_sheet(" توحید — ایمیل | Tohid Email")
    ws3.sheet_view.rightToLeft = True
    ws3.merge_cells('A1:B1')
    ws3.cell(1, 1, "👨 ایمیل توحید — Saskatchewan Health Authority").font = FA_T

    email_tohid = [
        ("گیرنده", "SHAInternational@saskhealthauthority.ca"),
        ("موضوع", "IT Operations Manager — International Candidate"),
        ("معرفی", "مدیر عملیات IT با ۱۹+ سال تجربه"),
        ("تخصص", "IT Infrastructure, VMware, Cisco, Windows Server"),
        ("ارتباط با IT بهداشت", "سیستم‌های High Availability، امنیت داده"),
        ("زبان", "English A2"),
        ("دلیل Saskatchewan", "بخش IT رو به رشد، مسکن ارزان"),
        ("پیوست", "CV"),
    ]
    for r, (f, v) in enumerate(email_tohid, 3):
        sc(ws3, r, 1, f).fill = BF
        sc(ws3, r, 2, v)
    ws3.column_dimensions['A'].width = 25
    ws3.column_dimensions['B'].width = 50

    # ── شیت ۴: ایمیل‌های مفید ──
    ws4 = wb.create_sheet(" ایمیل‌ها | Emails")
    ws4.sheet_view.rightToLeft = True
    ws4.merge_cells('A1:D1')
    ws4.cell(1, 1, "📧 ایمیل‌های مفید Saskatchewan").font = FA_T

    h4 = ["سازمان | Organization", "ایمیل | Email", "حوزه | Field", "وضعیت | Status"]
    for i, x in enumerate(h4, 1): ws4.cell(3, i, x)
    sh(ws4, 3, 4)

    emails = [
        ("SHA International", "SHAInternational@saskhealthauthority.ca", "استخدام بین‌المللی", "🟢"),
        ("SHA Careers", "careers@saskhealthauthority.ca", "مشاغل عمومی", "🟢"),
        ("HHR Saskatchewan", "HHR@health.gov.sk.ca", "منابع انسانی", "🟢"),
        ("eHealth Saskatchewan", "careers@ehealthsask.ca", "IT بهداشت", "🟢"),
        ("Government of Saskatchewan", "https://www.saskatchewan.ca", "سیستم EOI", "🟢"),
    ]
    for r, (org, email, field, status) in enumerate(emails, 4):
        sc(ws4, r, 1, org)
        c = sc(ws4, r, 2, email)
        if "@" in email: c.fill = GR
        sc(ws4, r, 3, field)
        sc(ws4, r, 4, status)

    for col in 'ABCD': ws4.column_dimensions[col].width = 35

    # ── شیت ۵: فرآیند ثبت‌نام ──
    ws5 = wb.create_sheet(" ثبت‌نام | Registration")
    ws5.sheet_view.rightToLeft = True
    ws5.merge_cells('A1:C1')
    ws5.cell(1, 1, "📋 فرآیند ثبت‌نام حرفه‌ای — Saskatchewan").font = FA_T

    h5 = ["مرحله | Step", "توضیح | Description", "وضعیت | Status"]
    for i, x in enumerate(h5, 1): ws5.cell(3, i, x)
    sh(ws5, 3, 3)

    steps = [
        ("۱. EOI ثبت‌نام", "ایجاد پروفایل Expression of Interest", "🔵 در انتظار"),
        ("۲. بررسی مدارک", "ارزیابی مدارک تحصیلی", "🔵 در انتظار"),
        ("۳. ثبت‌نام حرفه‌ای", "ثبت‌نام در College of Midwives", "🔵 در انتظار"),
        ("۴. آزمون زبان", "IELTS Academic / OET", "🔵 در انتظار"),
        ("۵. پیشنهاد شغل", "دریافت Job Offer", "🔵 در انتظار"),
        ("۶. LMIA", "Labour Market Impact Assessment", "🔵 در انتظار"),
        ("۷. ویزا", "Apply for Work Permit", "🔵 در انتظار"),
    ]
    for r, (step, desc, status) in enumerate(steps, 4):
        sc(ws5, r, 1, step).fill = BF
        sc(ws5, r, 2, desc)
        sc(ws5, r, 3, status)

    for col in 'ABC': ws5.column_dimensions[col].width = 35

    # ── شیت ۶: مقایسه Saskatchewan vs سایر ──
    ws6 = wb.create_sheet(" مقایسه | Comparison")
    ws6.sheet_view.rightToLeft = True
    ws6.merge_cells('A1:E1')
    ws6.cell(1, 1, "📈 مقایسه Saskatchewan با سایر استان‌ها").font = FA_T

    h6 = ["استان | Province", "هزینه مسکن | Housing", "آفتاب | Sunshine", "استخدام بین‌المللی | Intl", "امتیاز | Score"]
    for i, x in enumerate(h6, 1): ws6.cell(3, i, x)
    sh(ws6, 3, 5)

    provinces = [
        ("Saskatchewan", "$289,600", "۲,۰۰۰+ ساعت", "✅ فعال", 90),
        ("Ontario", "$1,052,920", "متوسط", "✅", 70),
        ("British Columbia", "$1,089,600", "متوسط", "✅", 65),
        ("Alberta", "$450,000", "خوب", "✅", 75),
        ("Manitoba", "$350,000", "خوب", "✅", 80),
    ]
    for r, (prov, house, sun, intl, score) in enumerate(provinces, 4):
        sc(ws6, r, 1, prov)
        c = sc(ws6, r, 2, house)
        if "$289" in house: c.fill = GR
        sc(ws6, r, 3, sun)
        sc(ws6, r, 4, intl)
        c = sc(ws6, r, 5, score)
        c.fill = GF if score >= 80 else YF

    for col in 'ABCDE': ws6.column_dimensions[col].width = 28

    # ── ذخیره ──
    DASHBOARD_DIR.mkdir(exist_ok=True)
    fn = f"Saskatchewan_Search_{NOW.strftime('%Y%m%d_%H%M')}.xlsx"
    fp = DASHBOARD_DIR / fn
    wb.save(fp)
    print(f"✅ Excel Saskatchewan: {fn}")
    print(f"   شیت‌ها: {len(wb.sheetnames)}")
    for s in wb.sheetnames:
        print(f"   • {s}")

if __name__ == "__main__":
    build()
