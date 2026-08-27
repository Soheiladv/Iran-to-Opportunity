# -*- coding: utf-8 -*-
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.chart import BarChart, Reference
from openpyxl.utils import get_column_letter
import os

# Create workbook
wb = openpyxl.Workbook()

# Styles
font_b_mitra = Font(name='B Mitra', size=14)
font_b_mitra_bold = Font(name='B Mitra', size=14, bold=True)
font_b_mitra_white = Font(name='B Mitra', size=14, bold=True, color='FFFFFF')
font_b_mitra_small = Font(name='B Mitra', size=11)

rtl_alignment = Alignment(horizontal='right', vertical='center', wrap_text=True)
center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
left_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Colors
green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
yellow_fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
orange_fill = PatternFill(start_color='FFD7B5', end_color='FFD7B5', fill_type='solid')
blue_fill = PatternFill(start_color='BDD7EE', end_color='BDD7EE', fill_type='solid')
header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
dark_fill = PatternFill(start_color='2F5597', end_color='2F5597', fill_type='solid')
light_gray_fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')

def set_rtl(ws):
    ws.sheet_properties.rightToLeft = True

def style_header_row(ws, row, cols):
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = font_b_mitra_white
        cell.fill = header_fill
        cell.alignment = center_alignment
        cell.border = thin_border

def style_data_row(ws, row, cols, fill=None):
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = font_b_mitra
        cell.alignment = rtl_alignment
        cell.border = thin_border
        if fill:
            cell.fill = fill

def add_score_color(ws, row, col, score):
    cell = ws.cell(row=row, column=col)
    cell.font = font_b_mitra_bold
    cell.alignment = center_alignment
    cell.border = thin_border
    if score >= 80:
        cell.fill = green_fill
    elif score >= 70:
        cell.fill = yellow_fill
    else:
        cell.fill = orange_fill

# ==========================================
# SHEET 1: شیت مستر (Dashboard)
# ==========================================
ws1 = wb.active
ws1.title = "داشبورد"
set_rtl(ws1)

# Title
ws1.merge_cells('A1:H1')
title_cell = ws1['A1']
title_cell.value = "داشبورد شکار فرصت شغلی — Migration Hunter"
title_cell.font = font_b_mitra_bold
title_cell.fill = dark_fill
title_cell.font = Font(name='B Mitra', size=18, bold=True, color='FFFFFF')
title_cell.alignment = center_alignment

# Date
ws1.merge_cells('A2:H2')
ws1['A2'].value = "آخرین بروزرسانی: ۱۹ آگوست ۲۰۲۶ — چرخه ۳"
ws1['A2'].font = font_b_mitra
ws1['A2'].alignment = center_alignment

# Summary
ws1['A4'].value = "خلاصه وضعیت"
ws1['A4'].font = font_b_mitra_bold

summary_data = [
    ("شاخص", "مقدار"),
    (" فرصت‌های شناسایی شده", "۶"),
    ("فرصت‌های آماده اعمال", "۴"),
    ("کارفرمایان جدید", "۳"),
    ("منابع جدید", "۳"),
    ("درخواست‌های آماده", "۴"),
    ("فرصت ندا", "۵"),
    ("فرصت توحید", "۱"),
]

for i, (label, value) in enumerate(summary_data):
    row = 5 + i
    ws1.cell(row=row, column=1, value=label)
    ws1.cell(row=row, column=2, value=value)
    if i == 0:
        style_header_row(ws1, row, 2)
    else:
        style_data_row(ws1, row, 2)
        ws1.cell(row=row, column=2).alignment = center_alignment

# Top Opportunities
ws1['A14'].value = "فرصت‌های برتر"
ws1['A14'].font = font_b_mitra_bold

headers = ["متقاضی", "کشور", "کارفرما", "عنوان", "Sponsorship", "زبان", "تناسب", "وضعیت"]
for col, h in enumerate(headers, 1):
    ws1.cell(row=15, column=col, value=h)
style_header_row(ws1, 15, len(headers))

jobs = [
    ("ندا", "nz", "Working In NZ", "Midwife", "Confirmed", "A2", 85, "🟢 READY"),
    ("ندا", "nz", "Health NZ", "Midwife", "Confirmed", "A2", 82, "🟢 READY"),
    ("ندا", "nz", "RGH Global", "Midwife", "Confirmed", "A2", 80, "🟢 READY"),
    ("ندا", "de", "Holalemania", "Hebamme", "Confirmed", "A1", 78, "🟢 READY"),
    ("ندا", "de", "TalentOrange", "Hebamme", "Confirmed", "A1", 75, "🟡 NEEDS"),
    ("توحید", "de", "IT Companies", "IT Manager", "Likely", "A2", 75, "🟡 NEEDS"),
]

for i, job in enumerate(jobs):
    row = 16 + i
    for col, val in enumerate(job, 1):
        ws1.cell(row=row, column=col, value=val)
    style_data_row(ws1, row, len(headers))
    if job[6] >= 80:
        ws1.cell(row=row, column=7).fill = green_fill
    elif job[6] >= 70:
        ws1.cell(row=row, column=7).fill = yellow_fill

# Chart
chart1 = BarChart()
chart1.type = "col"
chart1.title = "امتیاز تناسب فرصت‌ها"
chart1.y_axis.title = "امتیاز"
chart1.x_axis.title = "فرصت"
chart1.style = 10

data = Reference(ws1, min_col=7, min_row=15, max_row=21)
cats = Reference(ws1, min_col=3, min_row=16, max_row=21)
chart1.add_data(data, titles_from_data=True)
chart1.set_categories(cats)
chart1.shape = 4
ws1.add_chart(chart1, "A23")

# Column widths
ws1.column_dimensions['A'].width = 20
ws1.column_dimensions['B'].width = 15
ws1.column_dimensions['C'].width = 20
ws1.column_dimensions['D'].width = 25
ws1.column_dimensions['E'].width = 18
ws1.column_dimensions['F'].width = 12
ws1.column_dimensions['G'].width = 12
ws1.column_dimensions['H'].width = 15

# ==========================================
# SHEET 2: ندا — فرصت‌ها + ایمیل + کاور لیتر
# ==========================================
ws2 = wb.create_sheet("ندا — فرصت‌ها + ایمیل")
set_rtl(ws2)

ws2.merge_cells('A1:L1')
ws2['A1'].value = "فرصت‌های ندا — Midwife"
ws2['A1'].font = Font(name='B Mitra', size=16, bold=True, color='FFFFFF')
ws2['A1'].fill = PatternFill(start_color='E91E63', end_color='E91E63', fill_type='solid')
ws2['A1'].alignment = center_alignment

headers_neda = ["ردیف", "کارفرما", "کشور", "شهر", "عنوان", "حقوق", "Sponsorship", "زبان", "ثبت‌نام", "تناسب", "لینک شغل", "ایمیل درخواست", "کاور لیتر", "اقدام بعدی"]
for col, h in enumerate(headers_neda, 1):
    ws2.cell(row=2, column=col, value=h)
style_header_row(ws2, 2, len(headers_neda))

neda_jobs = [
    (1, "Working In Health NZ", "nz", "Auckland", "Midwife", "RAIGAN", "Confirmed", "A2", "Required", 85, "https://www.workingin-health.co.nz/midwifery-jobs/", "🟢 آماده", "🟢 آماده", "تکمیل فرم آنلاین"),
    (2, "Health New Zealand", "nz", "سراسر NZ", "Midwife", "75K-106K NZD", "Confirmed", "A2", "Required", 82, "https://www.healthnz.govt.nz/careers/international", "🟢 آماده", "🟢 آماده", "ارسال CV"),
    (3, "RGH Global", "nz", "سراسر NZ", "Midwife", "75K-106K NZD", "Confirmed", "A2", "Required", 80, "https://www.rgh-global.com/jobs/midwife-with-sponsorship/", "🟢 آماده", "🟢 آماده", "ارسال CV"),
    (4, "Holalemania", "de", "Hamburg", "Hebamme", "UNKNOWN", "Confirmed", "A1-A2", "Anerkennung", 78, "https://holalemania.de/en/", "🟢 آماده", "🟢 آماده", "ایمیل به info@holalemania.de"),
    (5, "TalentOrange", "de", "سراسر DE", "Hebamme", "UNKNOWN", "Confirmed", "B2 (بورسیه)", "کمک", 75, "https://www.talentorange.com/en/", "🟡 نیاز", "🟡 نیاز", "بررسی وبسایت"),
]

for i, job in enumerate(neda_jobs):
    row = 3 + i
    for col, val in enumerate(job, 1):
        ws2.cell(row=row, column=col, value=val)
    style_data_row(ws2, row, len(headers_neda))
    # Score color
    score = job[9]
    if score >= 80:
        ws2.cell(row=row, column=10).fill = green_fill
    elif score >= 70:
        ws2.cell(row=row, column=10).fill = yellow_fill
    # Link color
    link_cell = ws2.cell(row=row, column=11)
    link_cell.font = Font(name='B Mitra', size=11, color='0563C1', underline='single')
    link_cell.hyperlink = job[10]

# Chart
chart2 = BarChart()
chart2.title = "امتیاز تناسب — ندا"
chart2.y_axis.title = "امتیاز"
chart2.style = 10
data2 = Reference(ws2, min_col=10, min_row=2, max_row=7)
cats2 = Reference(ws2, min_col=2, min_row=3, max_row=7)
chart2.add_data(data2, titles_from_data=True)
chart2.set_categories(cats2)
ws2.add_chart(chart2, "A9")

for col_letter in ['A','B','C','D','E','F','G','H','I','J','K','L','M','N']:
    ws2.column_dimensions[col_letter].width = 18

# ==========================================
# SHEET 3: توحید — فرصت‌ها + ایمیل + کاور لیتر
# ==========================================
ws3 = wb.create_sheet("توحید — فرصت‌ها + ایمیل")
set_rtl(ws3)

ws3.merge_cells('A1:L1')
ws3['A1'].value = "فرصت‌های توحید — IT Manager"
ws3['A1'].font = Font(name='B Mitra', size=16, bold=True, color='FFFFFF')
ws3['A1'].fill = PatternFill(start_color='1565C0', end_color='1565C0', fill_type='solid')
ws3['A1'].alignment = center_alignment

headers_tohid = ["ردیف", "کارفرما", "کشور", "شهر", "عنوان", "حقوق", "Sponsorship", "زبان", "ثبت‌نام", "تناسب", "لینک شغل", "ایمیل درخواست", "کاور لیتر", "اقدام بعدی"]
for col, h in enumerate(headers_tohid, 1):
    ws3.cell(row=2, column=col, value=h)
style_header_row(ws3, 2, len(headers_tohid))

tohid_jobs = [
    (1, "SAP", "de", "Waldorf", "IT Infrastructure Manager", "55K-80K EUR", "Likely", "English", "N/A", 78, "https://jobs.sap.com/", "🟡 نیاز", "🟡 نیاز", "بررسی careers page"),
    (2, "Siemens", "de", "Munich", "IT Operations Manager", "60K-90K EUR", "Likely", "English", "N/A", 76, "https://jobs.siemens.com/", "🟡 نیاز", "🟡 نیاز", "بررسی careers page"),
    (3, "Deutsche Telekom", "de", "Bonn", "IT Manager", "50K-75K EUR", "Likely", "English", "N/A", 75, "https://jobs.telekom.com/", "🟡 نیاز", "🟡 نیاز", "بررسی careers page"),
    (4, "Allianz", "de", "Munich", "IT Operations Manager", "55K-85K EUR", "Likely", "English", "N/A", 74, "https://jobs.allianz.com/", "🟡 نیاز", "🟡 نیاز", "بررسی careers page"),
    (5, "BMW", "de", "Munich", "IT Infrastructure Engineer", "60K-90K EUR", "Likely", "English", "N/A", 73, "https://www.bmwgroup.jobs/", "🟡 نیاز", "🟡 نیاز", "بررسی careers page"),
]

for i, job in enumerate(tohid_jobs):
    row = 3 + i
    for col, val in enumerate(job, 1):
        ws3.cell(row=row, column=col, value=val)
    style_data_row(ws3, row, len(headers_tohid))
    score = job[9]
    if score >= 80:
        ws3.cell(row=row, column=10).fill = green_fill
    elif score >= 70:
        ws3.cell(row=row, column=10).fill = yellow_fill
    link_cell = ws3.cell(row=row, column=11)
    link_cell.font = Font(name='B Mitra', size=11, color='0563C1', underline='single')
    link_cell.hyperlink = job[10]

chart3 = BarChart()
chart3.title = "امتیاز تناسب — توحید"
chart3.y_axis.title = "امتیاز"
chart3.style = 10
data3 = Reference(ws3, min_col=10, min_row=2, max_row=7)
cats3 = Reference(ws3, min_col=2, min_row=3, max_row=7)
chart3.add_data(data3, titles_from_data=True)
chart3.set_categories(cats3)
ws3.add_chart(chart3, "A9")

for col_letter in ['A','B','C','D','E','F','G','H','I','J','K','L','M','N']:
    ws3.column_dimensions[col_letter].width = 18

# ==========================================
# SHEET 4: مقایسه کشورها
# ==========================================
ws4 = wb.create_sheet("مقایسه کشورها")
set_rtl(ws4)

ws4.merge_cells('A1:F1')
ws4['A1'].value = "مقایسه کشورها"
ws4['A1'].font = Font(name='B Mitra', size=16, bold=True, color='FFFFFF')
ws4['A1'].fill = dark_fill
ws4['A1'].alignment = center_alignment

headers_country = ["کشور", "ندا (امتیاز)", "توحید (امتیاز)", "زبان مورد نیاز", "سرعت", "توضیح"]
for col, h in enumerate(headers_country, 1):
    ws4.cell(row=2, column=col, value=h)
style_header_row(ws4, 2, len(headers_country))

countries = [
    ("nz نیوزیلند", 85, 70, "A2 (ویزا) + IELTS/OET (ثبت‌نام)", "متوسط", "Green List = اقامت فوری"),
    ("de آلمان", 78, 78, "A1-A2 آلمانی", "متوسط", "IT انگلیسی. ماما آلمانی."),
    ("au استرالیا", 75, 65, "A2 (ویза) + AHPRA (ثبت‌نام)", "کند", "حمایت ویزا محدود"),
    ("ca کانادا", 70, 65, "A2 (ویза)", "کند", "LMIA پیچیده"),
]

for i, country in enumerate(countries):
    row = 3 + i
    for col, val in enumerate(country, 1):
        ws4.cell(row=row, column=col, value=val)
    style_data_row(ws4, row, len(headers_country))
    # Color for NEDA score
    if country[1] >= 80:
        ws4.cell(row=row, column=2).fill = green_fill
    elif country[1] >= 70:
        ws4.cell(row=row, column=2).fill = yellow_fill
    # Color for TOHID score
    if country[2] >= 80:
        ws4.cell(row=row, column=3).fill = green_fill
    elif country[2] >= 70:
        ws4.cell(row=row, column=3).fill = yellow_fill

chart4 = BarChart()
chart4.title = "مقایسه امتیاز کشورها"
chart4.y_axis.title = "امتیاز"
chart4.style = 10
data4 = Reference(ws4, min_col=2, max_col=3, min_row=2, max_row=6)
cats4 = Reference(ws4, min_col=1, min_row=3, max_row=6)
chart4.add_data(data4, titles_from_data=True)
chart4.set_categories(cats4)
ws4.add_chart(chart4, "A8")

for col_letter in ['A','B','C','D','E','F']:
    ws4.column_dimensions[col_letter].width = 25

# ==========================================
# SHEET 5: ایمیل‌ها و کاور لیتر
# ==========================================
ws5 = wb.create_sheet("ایمیل‌ها و کاور لیتر")
set_rtl(ws5)

ws5.merge_cells('A1:F1')
ws5['A1'].value = "ایمیل‌ها و کاور لیتر آماده"
ws5['A1'].font = Font(name='B Mitra', size=16, bold=True, color='FFFFFF')
ws5['A1'].fill = PatternFill(start_color='4CAF50', end_color='4CAF50', fill_type='solid')
ws5['A1'].alignment = center_alignment

headers_email = ["ردیف", "کارفرما", "متقاضی", "موضوع ایمیل", "وضعیت", "لینک"]
for col, h in enumerate(headers_email, 1):
    ws5.cell(row=2, column=col, value=h)
style_header_row(ws5, 2, len(headers_email))

emails = [
    (1, "Working In Health NZ", "ندا", "International Midwife — Registration Interest", "🟢 آماده", "https://www.workingin-health.co.nz/midwifery-jobs/"),
    (2, "Health New Zealand", "ندا", "International Midwife — Clinical Experience", "🟢 آماده", "https://www.healthnz.govt.nz/careers/international"),
    (3, "RGH Global", "ندا", "Midwife Application — Visa Sponsorship", "🟢 آماده", "https://www.rgh-global.com/jobs/midwife-with-sponsorship/"),
    (4, "Holalemania", "ندا", "International Midwife — Anerkennung Inquiry", "🟢 آماده", "https://holalemania.de/en/"),
    (5, "TalentOrange", "ندا", "Midwife Application — Program Inquiry", "🟡 نیاز", "https://www.talentorange.com/en/"),
]

for i, email in enumerate(emails):
    row = 3 + i
    for col, val in enumerate(email, 1):
        ws5.cell(row=row, column=col, value=val)
    style_data_row(ws5, row, len(headers_email))
    link_cell = ws5.cell(row=row, column=6)
    link_cell.font = Font(name='B Mitra', size=11, color='0563C1', underline='single')
    link_cell.hyperlink = email[5]

for col_letter in ['A','B','C','D','E','F']:
    ws5.column_dimensions[col_letter].width = 25

# ==========================================
# SHEET 6: زبان
# ==========================================
ws6 = wb.create_sheet("زبان")
set_rtl(ws6)

ws6.merge_cells('A1:E1')
ws6['A1'].value = "وضعیت زبان — جستجو بدون توقف"
ws6['A1'].font = Font(name='B Mitra', size=16, bold=True, color='FFFFFF')
ws6['A1'].fill = PatternFill(start_color='FF9800', end_color='FF9800', fill_type='solid')
ws6['A1'].alignment = center_alignment

headers_lang = ["متقاضی", "زبان فعلی", "وضعیت", "نیاز بعدی", "زمان تقریبی"]
for col, h in enumerate(headers_lang, 1):
    ws6.cell(row=2, column=col, value=h)
style_header_row(ws6, 2, len(headers_lang))

lang_data = [
    ("توحید", "English A2", "✅ تأیید شده", "IELTS 6.5 (اختیاری)", "۶-۱۲ ماه"),
    ("توحید", "German A1", "✅ تازه تکمیل", "A2 آلمانی", "۳-۶ ماه"),
    ("ندا", "English A2", "✅ تأیید شده", "IELTS Academic 7.0 / OET", "۶-۱۲ ماه"),
    ("ندا", "German A1", "✅ تازه تکمیل", "A2 آلمانی", "۳-۶ ماه"),
]

for i, lang in enumerate(lang_data):
    row = 3 + i
    for col, val in enumerate(lang, 1):
        ws6.cell(row=row, column=col, value=val)
    style_data_row(ws6, row, len(headers_lang))
    ws6.cell(row=row, column=3).fill = green_fill

# Important note
ws6.merge_cells('A8:E8')
ws6['A8'].value = "⚠️ زبان = مانع جستجو نیست. جستجو و زبان همزمان ادامه دارد."
ws6['A8'].font = Font(name='B Mitra', size=14, bold=True, color='FF0000')
ws6['A8'].alignment = center_alignment

for col_letter in ['A','B','C','D','E']:
    ws6.column_dimensions[col_letter].width = 25

# ==========================================
# SHEET 7: ویزا
# ==========================================
ws7 = wb.create_sheet("ویزا")
set_rtl(ws7)

ws7.merge_cells('A1:F1')
ws7['A1'].value = "اطلاعات ویزا — ۴ کشور"
ws7['A1'].font = Font(name='B Mitra', size=16, bold=True, color='FFFFFF')
ws7['A1'].fill = dark_fill
ws7['A1'].alignment = center_alignment

headers_visa = ["کشور", "ویزا", "الزام شغلی", "الزام زبان", "خانواده", "منبع رسمی"]
for col, h in enumerate(headers_visa, 1):
    ws7.cell(row=2, column=col, value=h)
style_header_row(ws7, 2, len(headers_visa))

visa_data = [
    ("nz نیوزیلند", "AEWV", "Job Offer + Accredited Employer", "ANZSCO 1-2: No", "همراه", "immigration.govt.nz"),
    ("nz نیوزیلند", "Green List", "Green List occupation", "No", "PR فوری", "immigration.govt.nz"),
    ("de آلمان", "EU Blue Card", "Job Offer > €45,934", "No", "همراه", "bamf.de"),
    ("de آلمان", "Work Visa", "Job Offer", "No", "همquate", "bamf.de"),
    ("au استرالیا", "482 TSS", "Job Offer + Sponsor", "IELTS 5.0", "همراه", "homeaffairs.gov.au"),
    ("ca کانادا", "LMIA", "Job Offer + LMIA", "CLB 5+", "همراه", "canada.ca"),
]

for i, visa in enumerate(visa_data):
    row = 3 + i
    for col, val in enumerate(visa, 1):
        ws7.cell(row=row, column=col, value=val)
    style_data_row(ws7, row, len(headers_visa))

for col_letter in ['A','B','C','D','E','F']:
    ws7.column_dimensions[col_letter].width = 25

# ==========================================
# SHEET 8: ثبت‌نام حرفه‌ای
# ==========================================
ws8 = wb.create_sheet("ثبت‌نام")
set_rtl(ws8)

ws8.merge_cells('A1:F1')
ws8['A1'].value = "ثبت‌نام حرفه‌ای — مامایی"
ws8['A1'].font = Font(name='B Mitra', size=16, bold=True, color='FFFFFF')
ws8['A1'].fill = PatternFill(start_color='9C27B0', end_color='9C27B0', fill_type='solid')
ws8['A1'].alignment = center_alignment

headers_reg = ["کشور", "نهاد", "مسیر ثبت‌نام", "الزام زبان", "زمان تقریبی", "هزینه"]
for col, h in enumerate(headers_reg, 1):
    ws8.cell(row=2, column=col, value=h)
style_header_row(ws8, 2, len(headers_reg))

reg_data = [
    ("nz نیوزیلند", "Midwifery Council NZ", "ارزیابی مدارک + آزمون", "IELTS Academic 7.0 / OET", "۶-۱۲ ماه", "NZ$485"),
    ("de آلمان", "Anerkennungsstelle", "ارزیابی مدارک + B2 آلمانی", "B2 آلمانی", "۱۲-۱۸ ماه", "متغیر"),
    ("au استرالیا", "AHPRA", "ارزیابی مدارک + IELTS", "IELTS 7.0", "۶-۱۲ ماه", "AUD$500+"),
]

for i, reg in enumerate(reg_data):
    row = 3 + i
    for col, val in enumerate(reg, 1):
        ws8.cell(row=row, column=col, value=val)
    style_data_row(ws8, row, len(headers_reg))

for col_letter in ['A','B','C','D','E','F']:
    ws8.column_dimensions[col_letter].width = 25

# ==========================================
# SHEET 9: اقدامات
# ==========================================
ws9 = wb.create_sheet("اقدامات")
set_rtl(ws9)

ws9.merge_cells('A1:E1')
ws9['A1'].value = "۵ اقدام برتر امروز"
ws9['A1'].font = Font(name='B Mitra', size=16, bold=True, color='FFFFFF')
ws9['A1'].fill = PatternFill(start_color='F44336', end_color='F44336', fill_type='solid')
ws9['A1'].alignment = center_alignment

headers_action = ["ردیف", "متقاضی", "اقدام", "کشور", "اولویت"]
for col, h in enumerate(headers_action, 1):
    ws9.cell(row=2, column=col, value=h)
style_header_row(ws9, 2, len(headers_action))

actions = [
    (1, "ندا", "تکمیل فرم Working In Health NZ", "nz", "🔴 فوری"),
    (2, "ندا", "ایمیل Holalemania", "de", "🔴 فوری"),
    (3, "ندا", "ارسال CV RGH Global", "nz", "🔴 فوری"),
    (4, "ندا", "ارسال CV Health New Zealand", "nz", "🟠 مهم"),
    (5, "توحید", "جستجوی دقیق IT آلمان", "de", "🟠 مهم"),
]

for i, action in enumerate(actions):
    row = 3 + i
    for col, val in enumerate(action, 1):
        ws9.cell(row=row, column=col, value=val)
    style_data_row(ws9, row, len(headers_action))
    if "فوری" in str(action[4]):
        ws9.cell(row=row, column=5).fill = PatternFill(start_color='FF6B6B', end_color='FF6B6B', fill_type='solid')
    else:
        ws9.cell(row=row, column=5).fill = orange_fill

for col_letter in ['A','B','C','D','E']:
    ws9.column_dimensions[col_letter].width = 25

# Save
output_dir = os.path.join(os.path.dirname(__file__), 'dashboard')
os.makedirs(output_dir, exist_ok=True)
output_path = os.path.join(output_dir, 'MigrationHunter_Dashboard.xlsx')
wb.save(output_path)
print(f"Dashboard saved: {output_path}")
print(f"Sheets: {wb.sheetnames}")
