#!/usr/bin/env python3
"""
جستجوی کامل اروپا — اتریش، اسکاندیناوی، هلند، ایرلند، دانمارک، فنلاند
همراه با مقایسه کشورها و لینک‌های جستجو
"""
import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

FONT_FA = "B Mitra"
FONT_EN = "Times New Roman"
DARK_BLUE = "1B4F72"
MED_BLUE = "2E86C1"
GREEN = "27AE60"
LIGHT_GREEN = "D5F5E3"
YELLOW = "F39C12"
LIGHT_YELLOW = "FEF9E7"
RED = "E74C3C"
LIGHT_RED = "FADBD8"
PURPLE = "8E44AD"
LIGHT_PURPLE = "E8DAEF"
GRAY = "BDC3C7"
LIGHT_GRAY = "F2F3F4"
DARK = "2C3E50"
WHITE = "FFFFFF"
NOW = datetime.now().strftime("%Y-%m-%d %H:%M")

def set_rtl(ws):
    ws.sheet_view.rightToLeft = True

def mf(name=FONT_FA, size=11, bold=False, color="000000"):
    return Font(name=name, size=size, bold=bold, color=color)

def mfill(c):
    return PatternFill(start_color=c, end_color=c, fill_type="solid")

def ra(h="right", v="center", wrap=True):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def ca():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

thin = Border(left=Side(style="thin"), right=Side(style="thin"),
              top=Side(style="thin"), bottom=Side(style="thin"))

def write_h(ws, row, cols, fc=DARK_BLUE):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = mf(size=11, bold=True, color=WHITE)
        cell.fill = mfill(fc)
        cell.alignment = ca()
        cell.border = thin

def wc(ws, row, col, val, ft="fa", sz=10, bold=False, color="000000", bg=None, h="right"):
    cell = ws.cell(row=row, column=col, value=val)
    fn = FONT_FA if ft == "fa" else FONT_EN
    cell.font = mf(name=fn, size=sz, bold=bold, color=color)
    cell.alignment = ra(h=h)
    cell.border = thin
    if bg:
        cell.fill = mfill(bg)
    return cell

# ═══════════════════════════════════════════════════
# ALL EUROPEAN SEARCH LINKS
# ═══════════════════════════════════════════════════
all_links = [
    # ── AUSTRIA 🇦🇹 ──
    ["🇦🇹 اتریش", "Work in Austria — Talent Hub", "government", "https://www.workinaustria.com/en/employees/jobs", "پلتفرم رسمی کار اتریش", "✅ فعال"],
    ["🇦🇹 اتریش", "Red-White-Red Card", "government", "https://www.migration.gv.at/en/types-of-immigration/permanent-immigration/very-highly-qualified-workers/", "ویزای کار ماهر اتریش", "✅ فعال"],
    ["🇦🇹 اتریش", "RWR Card — WorkInAustria", "government", "https://www.workinaustria.com/en/residence-employment/red-white-red-card", "راهنمای کامل کارت قرمز-سفید-قرمز", "✅ فعال"],
    ["🇦🇹 اتریش", "Red-White-Red Card Nurses 2026", "government", "https://www.movetoaustria.at/en/blog/red-white-red-card-nurses-austria-2026", "پرستاران در لیست کمبود 2026", "✅ فعال"],
    ["🇦🇹 اتریش", "Faruse — Healthcare Austria", "recruiter", "https://v2.faruse.com/healthcare-jobs-in-austria-for-foreigners-with-visa-sponsorship", "مشاغل بهداشت با حمایت ویزا", "✅ فعال"],
    ["🇦🇹 اتریش", "Indeed Austria — Visa", "jobboard", "https://sg.indeed.com/q-visa-sponsorship-austria-jobs.html", "665 آگهی حمایت ویزا", "✅ فعال"],
    ["🇦🇹 اتریش", "LinkedIn Vienna — Visa", "linkedin", "https://www.linkedin.com/jobs/visa-sponsorship-jobs-vienna-va", "1000+ آگهی وین", "⚠️ لاگین"],
    ["🇦🇹 اتریش", "OeAD — RWR Card", "government", "https://oead.at/en/to-austria/entry-and-residence/researcher-without-hosting-agreement/red-white-red-card", "سازمان تبادل دانشگاهی", "✅ فعال"],
    ["🇦🇹 اتریش", "Y-Axis — Austria IT", "recruiter", "https://www.y-axis.com/blog/visa-sponsorship-jobs-in-austria/", "589K شغل IT در Q2 2026", "✅ فعال"],
    ["🇦🇹 اتریش", "IFMOSA — Europe", "recruiter", "https://ifmosawork.com/europe-visa-sponsored-jobs/", "ایمیل و تلفن مستقیم کارفرماها", "✅ فعال"],

    # ── SWEDEN 🇸🇪 ──
    ["🇸🇪 سوئد", "Vårdförbundet — Work in Sweden", "association", "https://www.vardforbundet.se/in-english/work-in-sweden/", "اتحادیه پرستاری سوئد — ماما هم پذیرفته", "✅ فعال"],
    ["🇸🇪 سوئد", "Nordic Cooperation — Healthcare", "government", "https://www.norden.org/en/info-norden/work-doctor-or-nurse-sweden", "راهنمای کار در سوئد", "✅ فعال"],
    ["🇸🇪 سوئد", "Taj HR — Sweden Nurses", "recruiter", "https://tajhrservices.com/jobs/nurses-sweden", "کمبود 10,000+ پرستار — حقوق SEK 32-45K", "✅ فعال"],
    ["🇸🇪 سوئد", "Expomedics — Sweden", "recruiter", "https://expomedics.com/jobs/nursing-position-in-sweden-free-language-course-for-all-family", "دوره رایگان زبان برای کل خانواده", "✅ فعال"],
    ["🇸🇪 سوئد", "MediCarrera — Scandinavia", "recruiter", "https://medicarrera.com/jobs/", "مشاغل بهداشت اسکاندیناوی", "✅ فعال"],
    ["🇸🇪 سوئد", "Faruse — Sweden Healthcare", "recruiter", "https://v2.faruse.com/english-speaking-healthcare-jobs-in-sweden", "مامایی بین‌المللی", "✅ فعال"],
    ["🇸🇪 سوئد", "Indeed Sweden — Healthcare", "jobboard", "https://www.indeed.com/q-healthcare-sweden-jobs.html", "آگهی‌های بهداشت سوئد", "✅ فعال"],

    # ── NORWAY 🇳🇴 ──
    ["🇳🇴 نروژ", "EP Advisory — Norway Jobs", "recruiter", "https://ep-advisory.com/blog/jobs-in-norway-for-foreigners/", "حقوق از €5,000/ماه — بیکاری 4.3%", "✅ فعال"],
    ["🇳🇴 نروژ", "Tech Careers Norway", "specialist", "https://www.tech-careers-no.com/navigating-the-job-market-it-jobs-for-foreigners-in-norway/", "راهنمای IT نروژ برای خارجی‌ها", "✅ فعال"],
    ["🇳🇴 نروژ", "Deel — Norway Work Visa", "government", "https://www.deel.com/blog/how-to-get-a-visa-and-work-permit-in-norway/", "راهنمای ویزای کار نروژ 2026", "✅ فعال"],
    ["🇳🇴 نروژ", "Glassdoor Norway — Visa", "jobboard", "https://www.glassdoor.com/Job/norway-visa-jobs-SRCH_IL.0,6_IN180_KO7,11.htm", "118 آگهی ویزا", "✅ فعال"],
    ["🇳🇴 نروژ", "AtoZ — Norway Employment", "recruiter", "https://www.atozserwisplus.com/jobs-europe/norway-employment-trends-2026-high-paying-jobs-salary-stats-work-permit-info", "نروژ به 100K نیروی ماهر نیاز دارد", "✅ فعال"],
    ["🇳🇴 نروژ", "Indeed Norway — IT Sponsorship", "jobboard", "https://au.indeed.com/q-norway-visa-sponsorship,-it-security-jobs.html", "IT infrastructure + VPN + firewall", "✅ فعال"],

    # ── DENMARK 🇩🇰 ──
    ["🇩🇰 دانمارک", "WorkInDenmark", "government", "https://www.workindenmark.dk/", "سایت رسمی کار دانمارک", "✅ فعال"],
    ["🇩🇰 دانمارک", "Danish Healthcare Sector", "government", "https://www.workindenmark.dk/working-in-denmark/sectors-with-high-demand/healthcare", "بهداشت دانمارک — نیاز به مجوز", "✅ فعال"],
    ["🇩🇰 دانمارک", "Positive List — Skilled Work", "government", "https://www.nyidanmark.dk/pl-PL/You-want-to-apply/Work/The-Positive-Lists/Positive-List-for-skilled-work", "لیست مشاغل کمبود", "✅ فعال"],
    ["🇩🇰 دانمارک", "STPS — Midwife Authorization", "government", "https://en.stps.dk/health-professionals-and-authorities/registration-of-healthcare-professionals/midwife/non-eu-countries/application-and-approval-of-qualifications", "ثبت‌نام ماما — غیرEU", "✅ فعال"],
    ["🇩🇰 دانمارک", "Taj HR — Denmark Nurses", "recruiter", "https://tajhrservices.com/jobs/nurses-denmark", "حقوق DKK 30-42K/ماه", "✅ فعال"],
    ["🇩🇰 دانمارک", "Medicolink", "recruiter", "https://medicolink.com/our-services/", "جذب پزشک و پرستار دانمارک", "✅ فعال"],
    ["🇩🇰 دانمارک", "JobsInEnglish.dk", "jobboard", "https://jobsinenglish.dk/jobs-in-denmark-with-visa-sponsorship/", "204 آگهی با حمایت ویزا", "✅ فعال"],
    ["🇩🇰 دانمارک", "Denmark 2026 Immigration", "government", "https://corporateimmigrationpartners.com/denmark-new-2026-immigration-scheme-targets-global-talent/", "طرح جدید مهاجرت 2026", "✅ فعال"],

    # ── FINLAND 🇫🇮 ──
    ["🇫🇮 فنلاند", "Finncare — International", "recruiter", "https://finncare.fi/intlrecruit/", "جذب پرستار و مراقب بین‌المللی", "✅ فعال"],
    ["🇫🇮 فنلاند", "Job Market Finland", "government", "https://tyomarkkinatori.fi/en/personal-customers/professional-information/profession-fields/health-care", "بازار کار فنلاند — بهداشت", "✅ فعال"],
    ["🇫🇮 فنланد", "Dynamic Health — Finland", "recruiter", "https://www.dynamichealthstaff.com/nursing-jobs-in-finland", "راهنمای LVV + ثبت‌نام", "✅ فعال"],
    ["🇫🇮 فنلاند", "AYK Global — Finland", "recruiter", "https://ayk.global/hire-healthcare-workers-for-finland", "جذب نیروی بهداشت فنلاند", "✅ فعال"],
    ["🇫🇮 فنلاند", "Faruse — Finland Healthcare", "recruiter", "https://v2.faruse.com/healthcare-jobs-in-finland-for-foreigners", "مشاغل بهداشت خارجی‌ها", "✅ فعال"],

    # ── NETHERLANDS 🇳🇱 ──
    ["🇳🇱 هلند", "IND — Kennismigrant", "government", "https://ind.nl/en/residence-permits/work/highly-skilled-migrant", "ویزای مهاجر ماهر هلند", "✅ فعال"],
    ["🇳🇱 هلند", "MediCarrera — Netherlands", "recruiter", "https://medicarrera.com/jobs-netherlands/", "مشاغل بهداشت هلند", "✅ فعال"],
    ["🇳🇱 هلند", "OTTO Healthcare — NL", "recruiter", "https://ottohealthcare.eu/en/information-for-candidates/our-program-outside-europe", "برنامه ۵ ساله بهداشت", "✅ فعال"],
    ["🇳🇱 هلند", "Vitae Professionals", "recruiter", "https://www.vitaeprofessionals.com/jobs", "مشاغل پرستاری هلند", "✅ فعال"],
    ["🇳🇱 هلند", "Faruse — NL Healthcare", "recruiter", "https://v2.faruse.com/healthcare-jobs-in-netherlands-with-visa-sponsorship", "حمایت ویزا بهداشت", "✅ فعال"],
    ["🇳🇱 هلند", "NextLevelJobs — NL IT", "specialist", "https://nextleveljobs.eu/blog/visa-sponsorship/nl", "آستانه €5942/ماه — 15+ شرکت فعال", "✅ فعال"],
    ["🇳🇱 هلند", "Jaabz — NL Visa Sponsorship", "jobboard", "https://jaabz.com/jobs/in/netherlands/visasponsorship", "لیست به‌روز کارفرمایان", "✅ فعال"],
    ["🇳🇱 هلند", "Jobbatical — NL Guide", "recruiter", "https://www.jobbatical.com/blog/netherlands-international-employees-visa-guide", "راهنمای ویزا 2026", "✅ فعال"],
    ["🇳🇱 هلند", "DutchReview — Sponsored Job", "specialist", "https://dutchreview.com/expat/sponsored-job-netherlands/", "راهنمای عملی کار حمایتی", "✅ فعال"],
    ["🇳🇱 هلند", "Glassdoor NL — Visa", "jobboard", "https://www.glassdoor.com/Job/nederland-visa-sponsorship-jobs-SRCH_IL.0,9_IC3057018_KO10,26.htm", "433 آگهی حمایت ویزا", "✅ فعال"],

    # ── IRELAND 🇮🇪 ──
    ["🇮🇪 ایرلند", "DETE — Critical Skills", "government", "https://enterprise.gov.ie/en/what-we-do/workplace-and-skills/employment-permits/employment-permit-eligibility/highly-skilled-eligible-occupations-list/", "لیست مشاغل حیاتی", "✅ فعال"],
    ["🇮🇪 ایرلند", "Workpermit.com — Ireland", "government", "https://workpermit.com/immigration/ireland/irish-critical-skills-occupations-list", "حداقل €60K/سال", "✅ فعال"],
    ["🇮🇪 ایرلند", "Recruitroo — CSEP 2026", "recruiter", "https://www.recruitroo.com/blog/critical-skills-employment-permit-csep-ireland-2026", "راهنمای کارفرما", "✅ فعال"],
    ["🇮🇪 ایرلند", "Kate Cowhig — Midwives", "recruiter", "https://www.kcr.ie/", "جذب ماما برای بیمارستان‌های ایرلند", "✅ فعال"],
    ["🇮🇪 ایرلند", "CPL Healthcare — Midwife", "recruiter", "https://www.cplhealthcare.com/job/midwife-jobs-ireland/", "حمایت ویزا Non-EU", "✅ فعال"],
    ["🇮🇪 ایرلند", "Servisource — Overseas", "recruiter", "https://servisource.ie/nursing-jobs-ireland-overseas-nurses/", "مشاغل پرستاری خارجی‌ها", "✅ فعال"],
    ["🇮🇪 ایرلند", "Med Connect", "recruiter", "https://medconnect.ie/candidates/nurses", "کمک ثبت‌نام NMBI", "✅ فعال"],
    ["🇮🇪 ایرلند", "IrishJobs — Overseas Midwife", "jobboard", "https://www.irishjobs.ie/jobs/overseas-midwife", "121 آگهی مامای بین‌المللی", "✅ فعال"],
    ["🇮🇪 ایرلند", "Indeed Ireland — Midwife", "jobboard", "https://ie.indeed.com/q-visa-sponsorship,-midwife-jobs.html", "34 آگهی ماما با حمایت ویزا", "✅ فعال"],
    ["🇮🇪 ایرلند", "LinkedIn Ireland — Midwife", "linkedin", "https://ie.linkedin.com/jobs/midwife-jobs", "541 آگهی ماما", "⚠️ لاگین"],

    # ── EXISTING (NZ, CA, AU, DE) ──
    ["🇳🇿 نیوزیلند", "Health NZ International", "government", "https://www.healthnz.govt.nz/careers/international", "استخدام رسمی بین‌المللی", "✅ فعال"],
    ["🇳🇿 نیوزیلند", "RGH Global — Midwife", "recruiter", "https://www.rgh-global.com/jobs/midwife-with-sponsorship/", "حقوق 75-106K NZD", "✅ فعال"],
    ["🇳🇿 نیوزیلند", "Seek NZ — Midwife", "jobboard", "https://nz.seek.com/midwife-jobs", "207 آگهی ماما", "✅ فعال"],
    ["🇳🇿 نیوزیلند", "Seek NZ — IT Hotel", "jobboard", "https://nz.seek.com/it-manager-hotel-jobs", "170 آگهی IT هتل", "✅ فعال"],
    ["🇨🇦 کانادا", "Alberta HA — International", "government", "https://www.albertahealthservices.ca/careers/Page12717.aspx", "استخدام بین‌المللی Albertا", "✅ فعال"],
    ["🇨🇦 کانادا", "Alberta Midwives", "association", "https://www.alberta-midwives.ca/job-board", "تخت‌های مامایی Albertا", "✅ فعال"],
    ["🇨🇦 کانادا", "Saskatchewan HA", "government", "https://www.saskhealthauthority.ca/careers-volunteering/careers", "استخدام بین‌المللی", "✅ فعال"],
    ["🇨🇦 کانادا", "Indeed Canada — IT", "jobboard", "https://ca.indeed.com/q-it-infrastructure-manager-visa-sponsorship-jobs.html", "607 آگهی IT", "✅ فعال"],
    ["🇦🇺 استرالیا", "AHPRA — IQNM", "government", "https://www.nursingmidwiferyboard.gov.au/Accreditation/IQNM/Before-you-apply", "ثبت‌نام مامای بین‌المللی", "✅ فعال"],
    ["🇦🇺 استرالیا", "Seek Australia — Midwife", "jobboard", "https://au.seek.com/midwife-jobs/full-time", "262 آگهی م تمام‌وقت", "✅ فعال"],
    ["🇦🇺 استرالیا", "482Jobs.com", "specialist", "https://482jobs.com/", "سایت تخصصی 482", "✅ فعال"],
    ["🇩🇪 آلمان", "Make it in Germany", "government", "https://www.make-it-in-germany.com/en/", "سایت رسمی کار آلمان", "✅ فعال"],
    ["🇩🇪 آلمان", "StepStone Germany", "jobboard", "https://www.stepstone.de/", "بزرگترین سایت کار آلمان", "✅ فعال"],
]

# ═══════════════════════════════════════════════════
# COUNTRY COMPARISON (expanded)
# ═══════════════════════════════════════════════════
country_data = [
    ["🇳🇿 نیوزیلند", "90", "65", "✅ تأیید", "Green List + AEWV", "IELTS 6.5-7", "85", "بهترین مسیر مامایی — حمایت مالی فعال"],
    ["🇦🇺 استرالیا", "85", "70", "✅ AHPRA", "482 / 189 / 190", "IELTS 7", "80", "ثبت‌نام AHPRA — نیاز بالا"],
    ["🇨🇦 کانادا", "80", "75", "⚠️ استانی", "Express Entry + PNP", "CLB 7", "78", "Alberta + Saskatchewan فعال"],
    ["🇩🇪 آلمان", "40", "80", "⚠️ بررسی", "EU Blue Card", "B2 آلمانی", "65", "IT عالی — مامایی سخت"],
    ["🇦🇹 اتریش", "50", "75", "✅ RWR Card", "Red-White-Red Card", "B2 آلمانی", "70", "589K شغل IT — ماما در لیست"],
    ["🇮🇪 ایرلند", "70", "70", "✅ Critical Skills", "CSEP (€60K+)", "IELTS 6.5", "72", "541 آگهی ماما — IT در لیست حیاتی"],
    ["🇳🇱 هلند", "55", "80", "✅ Kennismigrant", "HSM Visa (€5942/m)", "EN خوب", "70", "15+ شرکت IT فعال"],
    ["🇸🇪 سوئد", "60", "65", "⚠️ ثبت‌نام", "Work Permit", "سوئدی/B2", "62", "کمبود 10K+ پرستار"],
    ["🇳🇴 نروژ", "55", "70", "⚠️ ارزیابی", "Skilled Worker", "نروژی/B2", "63", "حقوق €5K+/ماه — IT خوب"],
    ["🇩🇰 دانمارک", "50", "70", "✅ Positive List", "Pay Limit / Positive List", "دانمارکی/B2", "60", "طرح جدید 2026 — IT در لیست"],
    ["🇫🇮 فنلاند", "45", "60", "⚠️ LVV", "Residence Permit", "فنلاندی/B2", "55", "Finncare فعال"],
]

# ═══════════════════════════════════════════════════
# BUILD EXCEL
# ═══════════════════════════════════════════════════
wb = Workbook()

# ── SHEET 1: Country Comparison ──
ws1 = wb.active
ws1.title = "مقایسه کشورها"
set_rtl(ws1)

ws1.merge_cells("A1:H1")
ws1["A1"].value = "مقایسه کشورها — رتبه‌بندی شانس مهاجرت"
ws1["A1"].font = mf(size=14, bold=True, color=WHITE)
ws1["A1"].fill = mfill(PURPLE)
ws1["A1"].alignment = ca()

headers1 = ["کشور", "مامایی\n(0-100)", "IT\n(0-100)", "حمایت مالی", "مسیر ویزا", "زبان", "امتیاز\nکل", "توضیح"]
for i, h in enumerate(headers1, 1):
    ws1.cell(row=3, column=i, value=h)
write_h(ws1, 3, 8, PURPLE)

for i, row in enumerate(country_data):
    r = 4 + i
    for j, val in enumerate(row):
        ft = "en" if j in [1, 2, 6] else "fa"
        bg = None
        if j == 6:
            score = int(val) if val.isdigit() else 0
            if score >= 80: bg = LIGHT_GREEN
            elif score >= 70: bg = LIGHT_YELLOW
            elif score >= 60: bg = "FDEBD0"
            else: bg = LIGHT_GRAY
        wc(ws1, r, j + 1, val, ft, 10, bg=bg)

widths1 = [18, 12, 10, 16, 28, 16, 10, 40]
for i, w in enumerate(widths1):
    ws1.column_dimensions[get_column_letter(i + 1)].width = w

# ── SHEET 2: All Search Links ──
ws2 = wb.create_sheet("لینک‌های جستجو")
set_rtl(ws2)

ws2.merge_cells("A1:G1")
ws2["A1"].value = f"لینک‌های جستجو — {len(all_links)} لینک فعال"
ws2["A1"].font = mf(size=14, bold=True, color=WHITE)
ws2["A1"].fill = mfill(MED_BLUE)
ws2["A1"].alignment = ca()

headers2 = ["#", "کشور", "منبع", "نوع", "لینک", "توضیح", "وضعیت"]
for i, h in enumerate(headers2, 1):
    ws2.cell(row=3, column=i, value=h)
write_h(ws2, 3, 7, MED_BLUE)

for i, row in enumerate(all_links):
    r = 4 + i
    wc(ws2, r, 1, i + 1, "en", 9)
    for j, val in enumerate(row):
        ft = "en" if j in [3, 4] else "fa"
        bg = None
        if j == 6:
            bg = LIGHT_GREEN if "فعال" in str(val) else LIGHT_YELLOW
        wc(ws2, r, j + 2, val, ft, 9, bg=bg)

widths2 = [5, 18, 35, 14, 60, 40, 12]
for i, w in enumerate(widths2):
    ws2.column_dimensions[get_column_letter(i + 1)].width = w

# ── SHEET 3: Austria Detail ──
ws3 = wb.create_sheet("اتریش | Austria")
set_rtl(ws3)

ws3.merge_cells("A1:D1")
ws3["A1"].value = "🇦🇹 اتریش — راهنمای جامع مهاجرت"
ws3["A1"].font = mf(size=14, bold=True, color=WHITE)
ws3["A1"].fill = mfill(RED)
ws3["A1"].alignment = ca()

austria_info = [
    ["ویزا", "Red-White-Red Card", "✅"],
    ["آستانه حقوق", "€55,678/سال (2026)", "—"],
    ["سیستم امتیاز", "70+ امتیاز", "—"],
    ["Job Seeker Visa", "۶ ماهه — بدون پیشنهاد شغل", "✅"],
    ["ثبت‌نام ماما", "B2 آلمانی + ارزیابی مدرک", "⚠️"],
    ["IT", "589K شغل IT در Q2 2026", "✅"],
    ["سایت رسمی", "workinaustria.com", "✅"],
    ["migration.gv.at", "قوانین مهاجرت", "✅"],
    ["آلمانی مورد نیاز", "B2 برای بهداشت — B1 برای IT", "—"],
    ["PR پس از", "۵ سال اقامت", "—"],
]

headers3 = ["مورد", "توضیح", "وضعیت"]
for i, h in enumerate(headers3, 1):
    ws3.cell(row=3, column=i, value=h)
write_h(ws3, 3, 3, RED)

for i, row in enumerate(austria_info):
    r = 4 + i
    for j, val in enumerate(row):
        ft = "fa"
        bg = None
        if j == 2:
            if "فعال" in str(val) or val == "✅": bg = LIGHT_GREEN
            elif "⚠️" in str(val): bg = LIGHT_YELLOW
        wc(ws3, r, j + 1, val, ft, 10, bg=bg)

ws3.column_dimensions["A"].width = 25
ws3.column_dimensions["B"].width = 45
ws3.column_dimensions["C"].width = 12

# ── SHEET 4: Scandinavia Detail ──
ws4 = wb.create_sheet("اسکاندیناوی")
set_rtl(ws4)

ws4.merge_cells("A1:E1")
ws4["A1"].value = "🇸🇪🇳🇴🇩🇰🇫🇮 اسکاندیناوی — مقایسه"
ws4["A1"].font = mf(size=14, bold=True, color=WHITE)
ws4["A1"].fill = mfill(DARK_BLUE)
ws4["A1"].alignment = ca()

scandi_headers = ["کشور", "مامایی", "IT", "زبان مورد نیاز", "ویزا"]
for i, h in enumerate(scandi_headers, 1):
    ws4.cell(row=3, column=i, value=h)
write_h(ws4, 3, 5, DARK_BLUE)

scandi_data = [
    ["🇸🇪 سوئد", "60/100 — کمبود 10K+", "65/100", "سوئدی B2", "Work Permit"],
    ["🇳🇴 نروژ", "55/100 — نیاز متوسط", "70/100 — حقوق بالا", "نروژی B2", "Skilled Worker"],
    ["🇩🇰 دانمارک", "50/100 — نیاز به مجوز", "70/100 — Positive List", "دانمارکی B2", "Pay Limit / Positive List"],
    ["🇫🇮 فنلاند", "45/100 — Finncare فعال", "60/100", "فنلاندی B2", "Residence Permit"],
]

for i, row in enumerate(scandi_data):
    r = 4 + i
    for j, val in enumerate(row):
        wc(ws4, r, j + 1, val, "en" if j == 4 else "fa", 10)

ws4.column_dimensions["A"].width = 18
ws4.column_dimensions["B"].width = 25
ws4.column_dimensions["C"].width = 25
ws4.column_dimensions["D"].width = 22
ws4.column_dimensions["E"].width = 25

# ── SHEET 5: Netherlands + Ireland ──
ws5 = wb.create_sheet("هلند+ایرلند")
set_rtl(ws5)

ws5.merge_cells("A1:E1")
ws5["A1"].value = "🇳🇱 هلند + 🇮🇪 ایرلند — مقایسه"
ws5["A1"].font = mf(size=14, bold=True, color=WHITE)
ws5["A1"].fill = mfill(GREEN)
ws5["A1"].alignment = ca()

headers5 = ["کشور", "مامایی", "IT", "زبان", "ویزا"]
for i, h in enumerate(headers5, 1):
    ws5.cell(row=3, column=i, value=h)
write_h(ws5, 3, 5, GREEN)

ni_data = [
    ["🇳🇱 هلند", "55/100 — نیاز متوسط", "80/100 — عالی", "EN خوب + هلندی", "Kennismigrant (€5942/m)"],
    ["🇮🇪 ایرلند", "70/100 — 541 آگهی", "70/100 — Critical Skills", "EN", "CSEP (€60K+)"],
]

for i, row in enumerate(ni_data):
    r = 4 + i
    for j, val in enumerate(row):
        wc(ws5, r, j + 1, val, "en" if j == 4 else "fa", 10)

ws5.column_dimensions["A"].width = 18
ws5.column_dimensions["B"].width = 28
ws5.column_dimensions["C"].width = 28
ws5.column_dimensions["D"].width = 22
ws5.column_dimensions["E"].width = 28

# ═══════════════════════════════════════════════════
# SAVE
# ═══════════════════════════════════════════════════
os.makedirs("dashboard", exist_ok=True)
fname = f"dashboard/Europe_Complete_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
wb.save(fname)
print(f"\n✅ {fname}")
print(f"📊 ۵ شیت:")
print(f"   1. مقایسه کشورها — ۱۱ کشور")
print(f"   2. لینک‌های جستجو — {len(all_links)} لینک")
print(f"   3. اتریش — راهنمای جامع")
print(f"   4. اسکاندیناوی — ۴ کشور")
print(f"   5. هلند + ایرلند")
