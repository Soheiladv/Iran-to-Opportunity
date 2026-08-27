#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Migration Hunter — Email Tracker & Follow-up System
سیستم پیگیری ایمیل‌ها و نوتیفیکیشن خودکار

نحوه اجرا:
    python email_tracker.py                    # نمایش وضعیت
    python email_tracker.py --add              # اضافه کردن ایمیل جدید
    python email_tracker.py --update           # بروزرسانی وضعیت
    python email_tracker.py --check            # بررسی نوتیفیکیشن
    python email_tracker.py --export           # خروجی به Excel
"""

import os
import json
from datetime import datetime, timedelta
from pathlib import Path

BASE_DIR = Path(__file__).parent
TRACKER_FILE = BASE_DIR / "memory" / "EMAIL_TRACKER.json"
NOTIFICATION_FILE = BASE_DIR / "output" / "NOTIFICATIONS.md"

# ==========================================
# EMAIL TRACKER
# ==========================================

class EmailTracker:
    """سیستم پیگیری ایمیل‌ها"""
    
    def __init__(self):
        self.data = self._load()
    
    def _load(self):
        """بارگذاری اطلاعات"""
        if TRACKER_FILE.exists():
            with open(TRACKER_FILE, 'r', encoding='utf-8') as f:
                return json.load(f)
        return {"emails": [], "follow_ups": [], "notifications": []}
    
    def _save(self):
        """ذخیره اطلاعات"""
        TRACKER_FILE.parent.mkdir(exist_ok=True)
        with open(TRACKER_FILE, 'w', encoding='utf-8') as f:
            json.dump(self.data, f, ensure_ascii=False, indent=2)
    
    def add_email(self, applicant, employer, country, email_type, subject, recipient_email="", notes=""):
        """اضافه کردن ایمیل جدید"""
        now = datetime.now()
        email = {
            "id": f"EMAIL-{len(self.data['emails'])+1:03d}",
            "applicant": applicant,
            "employer": employer,
            "country": country,
            "type": email_type,  # application, follow_up, inquiry
            "subject": subject,
            "recipient_email": recipient_email,
            "status": "sent",  # sent, pending, replied, bounced
            "sent_date": now.strftime("%Y-%m-%d"),
            "sent_time": now.strftime("%H:%M:%S"),
            "follow_up_date": (now + timedelta(days=7)).strftime("%Y-%m-%d"),
            "follow_up_status": "pending",  # pending, sent, completed
            "reply_date": None,
            "reply_status": None,  # positive, negative, no_reply
            "interview_date": None,
            "notes": notes,
            "created_at": now.isoformat()
        }
        
        self.data["emails"].append(email)
        self._save()
        
        # اضافه کردن نوتیفیکیشن پیگیری
        self._add_notification(email)
        
        print(f"✅ ایمیل ثبت شد: {email['id']}")
        print(f"   متقاضی: {applicant}")
        print(f"   کارفرما: {employer}")
        print(f"   تاریخ ارسال: {email['sent_date']}")
        print(f"   تاریخ پیگیری: {email['follow_up_date']}")
        
        return email
    
    def update_status(self, email_id, status, reply_date=None, reply_status=None, notes=None):
        """بروزرسانی وضعیت ایمیل"""
        for email in self.data["emails"]:
            if email["id"] == email_id:
                email["status"] = status
                if reply_date:
                    email["reply_date"] = reply_date
                if reply_status:
                    email["reply_status"] = reply_status
                if notes:
                    email["notes"] = notes
                self._save()
                print(f"✅ وضعیت بروزرسانی شد: {email_id} → {status}")
                return email
        
        print(f"❌ ایمیل یافت نشد: {email_id}")
        return None
    
    def check_follow_ups(self):
        """بررسی پیگیری‌های ضروری"""
        today = datetime.now().strftime("%Y-%m-%d")
        pending = []
        
        for email in self.data["emails"]:
            if email["follow_up_date"] and email["follow_up_date"] <= today:
                if email["follow_up_status"] == "pending":
                    pending.append(email)
        
        return pending
    
    def _add_notification(self, email):
        """اضافه کردن نوتیفیکیشن"""
        notification = {
            "email_id": email["id"],
            "type": "follow_up",
            "date": email["follow_up_date"],
            "message": f"پیگیری ایمیل به {email['employer']}",
            "status": "pending"
        }
        self.data["notifications"].append(notification)
        self._save()
    
    def get_all_emails(self):
        """دریافت تمام ایمیل‌ها"""
        return self.data["emails"]
    
    def get_email_by_id(self, email_id):
        """دریافت ایمیل با ID"""
        for email in self.data["emails"]:
            if email["id"] == email_id:
                return email
        return None
    
    def get_statistics(self):
        """آمار کلی"""
        emails = self.data["emails"]
        return {
            "total": len(emails),
            "sent": len([e for e in emails if e["status"] == "sent"]),
            "replied": len([e for e in emails if e["status"] == "replied"]),
            "pending_follow_up": len(self.check_follow_ups()),
            "positive_replies": len([e for e in emails if e.get("reply_status") == "positive"]),
            "interviews": len([e for e in emails if e.get("interview_date")])
        }
    
    def generate_notification_report(self):
        """تولید گزارش نوتیفیکیشن"""
        pending = self.check_follow_ups()
        today = datetime.now().strftime("%Y-%m-%d")
        
        report = [
            f"# نوتیفیکیشن‌های پیگیری — {today}",
            "",
            "---",
            "",
        ]
        
        if pending:
            report.append("## ⚠️ پیگیری‌های ضروری")
            report.append("")
            for email in pending:
                days_passed = (datetime.now() - datetime.fromisoformat(email["created_at"])).days
                report.append(f"### {email['id']}: {email['employer']}")
                report.append(f"| فیلد | مقدار |")
                report.append(f"|------|-------|")
                report.append(f"| متقاضی | {email['applicant']} |")
                report.append(f"| کارفرما | {email['employer']} |")
                report.append(f"| کشور | {email['country']} |")
                report.append(f"| تاریخ ارسال | {email['sent_date']} |")
                report.append(f"| روزهای گذشته | {days_passed} |")
                report.append(f"| تاریخ پیگیری | {email['follow_up_date']} |")
                report.append(f"| وضعیت | ⚠️ نیاز به پیگیری |")
                report.append("")
        else:
            report.append("## ✅ هیچ پیگیری ضروری نیست")
            report.append("")
        
        # آمار کلی
        stats = self.get_statistics()
        report.append("## 📊 آمار کلی")
        report.append("")
        report.append("| شاخص | مقدار |")
        report.append("|------|-------|")
        report.append(f"| کل ایمیل‌ها | {stats['total']} |")
        report.append(f"| ارسال شده | {stats['sent']} |")
        report.append(f"| پاسخ دریافت | {stats['replied']} |")
        report.append(f"| پیگیری ضروری | {stats['pending_follow_up']} |")
        report.append(f"| پاسخ مثبت | {stats['positive_replies']} |")
        report.append(f"| مصاحبه | {stats['interviews']} |")
        report.append("")
        
        content = "\n".join(report)
        NOTIFICATION_FILE.write_text(content, encoding='utf-8')
        
        return content

# ==========================================
# EXCEL EXPORTER
# ==========================================

class ExcelExporter:
    """خروجی به Excel"""
    
    def export(self, tracker):
        """خروجی وضعیت ایمیل‌ها به Excel"""
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        except ImportError:
            print("❌ openpyxl نصب نیست")
            return
        
        wb = openpyxl.Workbook()
        if 'Sheet' in wb.sheetnames: del wb['Sheet']
        
        FONT = 'B Mitra'
        BD = '1F4E79'
        BM = '2E75B6'
        GR = 'C6EFCE'
        YL = 'FFEB9C'
        RD = 'FFC7CE'
        WH = 'FFFFFF'
        
        HF = Font(name=FONT, bold=True, size=11, color=WH)
        TF = Font(name=FONT, size=14, bold=True, color=BD)
        CF = Font(name=FONT, size=10)
        SF = Font(name=FONT, size=12, bold=True, color=BM)
        
        HFI = PatternFill(start_color=BD, end_color=BD, fill_type='solid')
        GF = PatternFill(start_color=GR, end_color=GR, fill_type='solid')
        YF = PatternFill(start_color=YL, end_color=YL, fill_type='solid')
        RDF = PatternFill(start_color=RD, end_color=RD, fill_type='solid')
        
        TB = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        CA = Alignment(horizontal='center', vertical='center', wrap_text=True, readingOrder=2)
        RA = Alignment(horizontal='right', vertical='center', wrap_text=True, readingOrder=2)
        
        def sc(cell, font=CF, fill=None, align=RA):
            cell.font = font; cell.alignment = align; cell.border = TB
            if fill: cell.fill = fill
        
        # شیت ۱: وضعیت ایمیل‌ها
        ws = wb.create_sheet("📧 وضعیت ایمیل‌ها", 0)
        ws.sheet_view.rightToLeft = True
        
        now_str = datetime.now().strftime('%Y-%m-%d %H:%M')
        ws.merge_cells('A1:N1')
        c = ws.cell(row=1, column=1, value=f"وضعیت ایمیل‌ها — {now_str}")
        c.font = TF; c.alignment = CA
        
        headers = ['ID', 'متقاضی', 'کارفرما', 'کشور', 'نوع', 'موضوع', 'تاریخ ارسال', 'ساعت', 'وضعیت', 'تاریخ پاسخ', 'وضعیت پاسخ', 'تاریخ پیگیری', 'وضعیت پیگیری', 'یادداشت']
        for i, h in enumerate(headers, 1):
            c = ws.cell(row=3, column=i, value=h); sc(c, HF, HFI, CA)
        
        for row_idx, email in enumerate(tracker.get_all_emails(), 4):
            data = [
                email['id'], email['applicant'], email['employer'], email['country'],
                email['type'], email['subject'], email['sent_date'], email['sent_time'],
                email['status'], email.get('reply_date', '—') or '—',
                email.get('reply_status', '—') or '—',
                email['follow_up_date'], email['follow_up_status'], email['notes']
            ]
            for ci, val in enumerate(data, 1):
                c = ws.cell(row=row_idx, column=ci, value=val)
                fl = None
                sv = str(val)
                if sv in ['sent', 'پاسخ مثبت']: fl = GF
                elif sv in ['pending', 'در انتظار']: fl = YF
                elif sv in ['negative', 'رد شده']: fl = RDF
                sc(c, fill=fl)
        
        for i in range(1, len(headers)+1):
            ws.column_dimensions[chr(64+i) if i <= 26 else 'A'].width = 15
        
        # شیت ۲: پیگیری‌ها
        ws2 = wb.create_sheet("📋 پیگیری‌ها")
        ws2.sheet_view.rightToLeft = True
        
        ws2.merge_cells('A1:G1')
        c = ws2.cell(row=1, column=1, value=f"پیگیری‌ها — {now_str}")
        c.font = TF; c.alignment = CA
        
        headers2 = ['ID', 'متقاضی', 'کارفرما', 'تاریخ ارسال', 'تاریخ پیگیری', 'وضعیت', 'روزهای گذشته']
        for i, h in enumerate(headers2, 1):
            c = ws2.cell(row=3, column=i, value=h); sc(c, HF, HFI, CA)
        
        row_idx = 4
        for email in tracker.get_all_emails():
            days = (datetime.now() - datetime.fromisoformat(email['created_at'])).days
            data = [
                email['id'], email['applicant'], email['employer'],
                email['sent_date'], email['follow_up_date'],
                email['follow_up_status'], days
            ]
            for ci, val in enumerate(data, 1):
                c = ws2.cell(row=row_idx, column=ci, value=val)
                fl = None
                if str(val) == 'pending' and days >= 7: fl = RDF
                elif str(val) == 'pending': fl = YF
                elif str(val) == 'sent': fl = GF
                sc(c, fill=fl)
            row_idx += 1
        
        for i in range(1, len(headers2)+1):
            ws2.column_dimensions[chr(64+i)].width = 18
        
        # شیت ۳: آمار
        ws3 = wb.create_sheet("📊 آمار")
        ws3.sheet_view.rightToLeft = True
        
        ws3.merge_cells('A1:D1')
        c = ws3.cell(row=1, column=1, value="آمار کلی")
        c.font = TF; c.alignment = CA
        
        stats = tracker.get_statistics()
        stats_data = [
            ['کل ایمیل‌ها', stats['total']],
            ['ارسال شده', stats['sent']],
            ['پاسخ دریافت', stats['replied']],
            ['پیگیری ضروری', stats['pending_follow_up']],
            ['پاسخ مثبت', stats['positive_replies']],
            ['مصاحبه', stats['interviews']],
        ]
        
        for i, (label, val) in enumerate(stats_data, 3):
            ws3.cell(row=i, column=1, value=label).font = SF
            ws3.cell(row=i, column=2, value=val).font = CF
            ws3.cell(row=i, column=1).alignment = RA
            ws3.cell(row=i, column=2).alignment = CA
        
        ws3.column_dimensions['A'].width = 20
        ws3.column_dimensions['B'].width = 15
        
        # ذخیره
        out = BASE_DIR / "dashboard" / f"EMAIL_STATUS_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        wb.save(out)
        print(f"✅ خروجی Excel: {out}")
        return out

# ==========================================
# MAIN
# ==========================================

def main():
    import argparse
    
    parser = argparse.ArgumentParser(description='Email Tracker & Follow-up System')
    parser.add_argument('--add', action='store_true', help='Add new email')
    parser.add_argument('--update', action='store_true', help='Update email status')
    parser.add_argument('--check', action='store_true', help='Check follow-ups')
    parser.add_argument('--export', action='store_true', help='Export to Excel')
    parser.add_argument('--stats', action='store_true', help='Show statistics')
    parser.add_argument('--report', action='store_true', help='Generate notification report')
    
    # Email details for --add
    parser.add_argument('--applicant', type=str, help='Applicant name')
    parser.add_argument('--employer', type=str, help='Employer name')
    parser.add_argument('--country', type=str, help='Country')
    parser.add_argument('--type', type=str, default='application', help='Email type')
    parser.add_argument('--subject', type=str, help='Email subject')
    parser.add_argument('--email', type=str, default='', help='Recipient email')
    parser.add_argument('--notes', type=str, default='', help='Notes')
    
    # Update details
    parser.add_argument('--id', type=str, help='Email ID for update')
    parser.add_argument('--status', type=str, help='New status')
    parser.add_argument('--reply', type=str, help='Reply status (positive/negative/no_reply)')
    
    args = parser.parse_args()
    
    tracker = EmailTracker()
    
    if args.add:
        if not all([args.applicant, args.employer, args.country, args.subject]):
            print("❌ لطفاً تمام فیلدها را پر کنید")
            print("   --applicant NAME --employer NAME --country CODE --subject SUBJECT")
            return
        tracker.add_email(args.applicant, args.employer, args.country, args.type, args.subject, args.email, args.notes)
    
    elif args.update:
        if not all([args.id, args.status]):
            print("❌ لطفاً ID و وضعیت را مشخص کنید")
            print("   --id EMAIL-001 --status replied --reply positive")
            return
        tracker.update_status(args.id, args.status, reply_status=args.reply)
    
    elif args.check:
        pending = tracker.check_follow_ups()
        if pending:
            print(f"⚠️ {len(pending)} پیگیری ضروری:")
            for email in pending:
                days = (datetime.now() - datetime.fromisoformat(email['created_at'])).days
                print(f"   {email['id']}: {email['employer']} ({days} روز)")
        else:
            print("✅ هیچ پیگیری ضروری نیست")
    
    elif args.export:
        exporter = ExcelExporter()
        exporter.export(tracker)
    
    elif args.stats:
        stats = tracker.get_statistics()
        print("\n📊 آمار کلی:")
        print(f"   کل ایمیل‌ها: {stats['total']}")
        print(f"   ارسال شده: {stats['sent']}")
        print(f"   پاسخ دریافت: {stats['replied']}")
        print(f"   پیگیری ضروری: {stats['pending_follow_up']}")
        print(f"   پاسخ مثبت: {stats['positive_replies']}")
        print(f"   مصاحبه: {stats['interviews']}")
    
    elif args.report:
        report = tracker.generate_notification_report()
        print(report)
    
    else:
        # نمایش وضعیت
        emails = tracker.get_all_emails()
        stats = tracker.get_statistics()
        
        print("\n" + "=" * 60)
        print("📧 Email Tracker — وضعیت ایمیل‌ها")
        print("=" * 60)
        
        if emails:
            for email in emails:
                days = (datetime.now() - datetime.fromisoformat(email['created_at'])).days
                print(f"\n  {email['id']}: {email['employer']} ({email['country']})")
                print(f"    متقاضی: {email['applicant']}")
                print(f"    تاریخ ارسال: {email['sent_date']} {email['sent_time']}")
                print(f"    وضعیت: {email['status']}")
                print(f"    روزهای گذشته: {days}")
                print(f"    پیگیری: {email['follow_up_date']} ({email['follow_up_status']})")
        else:
            print("  هیچ ایمیلی ثبت نشده")
        
        print(f"\n📊 آمار: {stats['total']} ایمیل | {stats['pending_follow_up']} پیگیری ضروری")
        print("=" * 60)

if __name__ == "__main__":
    main()
