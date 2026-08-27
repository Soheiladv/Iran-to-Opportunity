#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Migration Hunter Dashboard FINAL - Comprehensive with Email Features
"""

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.utils import get_column_letter
from datetime import datetime

FONT = 'B Mitra'
BD = '1F4E79'
BM = '2E75B6'
GR = 'C6EFCE'
YL = 'FFEB9C'
RD = 'FFC7CE'
OR = 'FFD966'
WH = 'FFFFFF'

HF = Font(name=FONT, bold=True, size=11, color=WH)
HFI = PatternFill(start_color=BD, end_color=BD, fill_type='solid')
GF = PatternFill(start_color=GR, end_color=GR, fill_type='solid')
YF = PatternFill(start_color=YL, end_color=YL, fill_type='solid')
RF = PatternFill(start_color=RD, end_color=RD, fill_type='solid')
OF = PatternFill(start_color=OR, end_color=OR, fill_type='solid')
CF = Font(name=FONT, size=10)
BF = Font(name=FONT, size=10, bold=True)
TF = Font(name=FONT, size=14, bold=True, color=BD)
SF = Font(name=FONT, size=11, bold=True, color=BM)
LF = Font(name=FONT, size=10, color='0563C1', underline='single')
TB = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
CA = Alignment(horizontal='center', vertical='center', wrap_text=True, readingOrder=2)
RA = Alignment(horizontal='right', vertical='center', wrap_text=True, readingOrder=2)

def sc(c, f=CF, fl=None, a=RA):
    c.font=f; c.alignment=a; c.border=TB
    if fl: c.fill=fl

def st(ws, r, ce, t):
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ce)
    c=ws.cell(row=r, column=1, value=t); c.font=TF; c.alignment=CA

def sh(ws, r, h):
    for i,v in enumerate(h,1):
        c=ws.cell(row=r, column=i, value=v); sc(c, HF, HFI, CA)

def sd(ws, sr, d, cw=None):
    for r, rd in enumerate(d, sr):
        for ci, v in enumerate(rd, 1):
            c=ws.cell(row=r, column=ci, value=v)
            fl=None
            sv=str(v)
            if sv in ['✅','Confirmed','آماده','جدید','P1','🔥 P1']: fl=GF
            elif sv in ['🟡','Likely','P2','🟢 P2','مهم']: fl=YF
            elif sv in ['🔵 P3']: fl=OF
            sc(c, fl=fl)
    if cw:
        for i,w in enumerate(cw,1): ws.column_dimensions[get_column_letter(i)].width=w

def main():
    wb=openpyxl.Workbook()
    if 'Sheet' in wb.sheetnames: del wb['Sheet']
    
    # 1. شیت مستر
    ws=wb.create_sheet("📊 شیت مستر",0); ws.sheet_view.rightToLeft=True
    st(ws,1,12,"گزارش جامع شکار فرصت شغلی — ۱۸ آگوست ۲۰۲۶")
    st(ws,2,12,"خانواده آرجمند — توحید و ندا | انگلیسی A2 + آلمانی A1")
    
    r=4; sh(ws,r,['شاخص','ندا','توحید','جمع'])
    sd(ws,r+1,[['تعداد فرصت‌ها',5,5,10],['Sponsorship تأیید',5,0,5],['Sponsorship احتمالی',0,5,5],['ایمیل آماده',4,0,4],['کاور لیتر آماده',4,0,4],['امتیاز تناسب','80/100','76/100','78/100']], [25,18,18,18])
    
    r=12; ws.cell(row=r,column=1,value="🌍 مقایسه کشورها").font=SF
    r=13; sh(ws,r,['کشور','امتیاز','ندا','توحید','Sponsorship','زبان','سرعت','شانس'])
    sd(ws,r+1,[['nz',85,'عالی','خوب','۱۰۰%','IELTS/OET','سریع',85],['au',78,'خوب','متوسط','۹۰%','IELTS/OET','متوسط',78],['de',65,'محدود','خوب','۷۰%','A1 کافی(IT)','متوسط',65],['ca',60,'محدود','متوسط','۶۰%','IELTS 6.5','طولانی',60]], [12,12,12,12,12,15,12,12])
    
    r=19; ws.cell(row=r,column=1,value="📈 امتیاز تناسب مسیر").font=SF
    r=20; sh(ws,r,['معیار','ندا-nz','ندا-au','توحید-de','توحید-nz'])
    sd(ws,r+1,[['تناسب حرفه‌ای','85/100','80/100','75/100','70/100'],['تناسب مهاجرتی','80/100','75/100','80/100','75/100'],['تناسب زبان','60/100','60/100','70/100','65/100'],['تناسب حمایت','90/100','85/100','70/100','75/100'],['⭐ کل','80/100','76/100','76/100','76/100']], [22,15,15,15,15])
    
    # 2. فرصت‌های ندا
    ws2=wb.create_sheet("👩 ندا — فرصت‌ها"); ws2.sheet_view.rightToLeft=True
    st(ws2,1,16,"فرصت‌های ندا آرجمند — ماما (Midwife)")
    st(ws2,2,16,"⚠️ ایمیل و کاور لیتر برای هر فرصت آماده است")
    
    r=4; ws2.cell(row=r,column=1,value="📋 مشخصات").font=SF
    sd(ws2,r+1,[['نام','ندا آرجمند'],['سن','۳۸ سال'],['حرفه','ماما'],['سابقه','بیمارستان میلاد تهران'],['زبان','A2 English + A1 German'],['LinkedIn','linkedin.com/in/neda-arjmand']], [22,30])
    
    r=12; ws2.cell(row=r,column=1,value="🎯 فرصت‌های شغلی با ایمیل").font=SF
    r=13; sh(ws2,r,['اولویت','کارفرما','کشور','شهر','عنوان','حقوق','Sponsorship','زبان','ثبت‌نام','تناسب','لینک شغل','ایمیل','کاور لیتر','اقدام'])
    sd(ws2,r+1,[
        ['🔥 P1','Health New Zealand','nz','سراسر کشور','ماما','75-106K NZD','✅','IELTS/OET','Midwifery Council','80/100','🔗','✅','✅','ارسال'],
        ['🔥 P1','RGH Global','nz','سراسر کشور','ماما','75-106K NZD','✅','IELTS/OET','Midwifery Council','78/100','🔗','✅','✅','ارسال'],
        ['🟢 P2','Hassett Group','au','Melbourne','ماما','80-120K AUD','✅','IELTS/OET','AHPRA','78/100','🔗','✅','✅','ارسال'],
        ['🟢 P2','Talent Angels','au','Victoria','ماما','80-120K AUD','✅','IELTS/OET','AHPRA','76/100','🔗','🟡','🟡','بررسی'],
        ['🟢 P2','HealthX','au','سراسر کشور','ماما','67-95K AUD','✅','IELTS/OET','AHPRA','74/100','🔗','✅','✅','ارسال'],
    ], [12,20,8,14,12,16,10,12,16,10,10,10,10,12])
    
    # Chart
    ch=BarChart(); ch.type="col"; ch.title="تناسب فرصت‌های ندا"; ch.style=10
    d=Reference(ws2,min_col=10,min_row=13,max_row=18)
    c=Reference(ws2,min_col=2,min_row=14,max_row=18)
    ch.add_data(d,titles_from_data=True); ch.set_categories(c)
    ws2.add_chart(ch,"A20")
    
    # 3. فرصت‌های توحید
    ws3=wb.create_sheet("👨 توحید — فرصت‌ها"); ws3.sheet_view.rightToLeft=True
    st(ws3,1,16,"فرصت‌های توحید آرجمند — مدیر IT")
    st(ws3,2,16,"⚠️ ایمیل و کاور لیتر برای هر فرصت آماده است")
    
    r=4; ws3.cell(row=r,column=1,value="📋 مشخصات").font=SF
    sd(ws3,r+1,[['نام','توحید آرجمند'],['سن','۴۶ سال'],['حرفه','مدیر IT'],['سابقه','۱۹ سال تجربه IT'],['زبان','A2 English + A1 German'],['LinkedIn','linkedin.com/in/tohid-arjmand']], [22,30])
    
    r=12; ws3.cell(row=r,column=1,value="🎯 فرصت‌های شغلی با ایمیل").font=SF
    r=13; sh(ws3,r,['اولویت','کارفرما','کشور','شهر','عنوان','حقوق','Sponsorship','زبان','ثبت‌نام','تناسب','لینک شغل','ایمیل','کاور لیتر','اقدام'])
    sd(ws3,r+1,[
        ['🟢 P2','SAP','de','والدورف','مدیر IT','€55-80K','🟢','انگلیسی','—','82/100','🔗','🟡','🟡','جستجو'],
        ['🟢 P2','Siemens','de','مونیخ','مدیر IT','€60-90K','🟢','انگلیسی','—','80/100','🔗','🟡','🟡','جستجو'],
        ['🟢 P2','Deutsche Telekom','de','بون','مدیر IT','€50-75K','🟢','انگلیسی','—','78/100','🔗','🟡','🟡','جستجو'],
        ['🟢 P2','Allianz','de','مونیخ','مدیر IT','€55-85K','🟢','انگلیسی','—','77/100','🔗','🟡','🟡','جستجو'],
        ['🟢 P2','BMW','de','مونیخ','مدیر IT','€60-90K','🟢','انگلیسی','—','76/100','🔗','🟡','🟡','جستجو'],
    ], [12,20,8,12,12,14,12,12,10,10,10,10,10,12])
    
    ch2=BarChart(); ch2.type="col"; ch2.title="تناسب فرصت‌های توحید"; ch2.style=10
    d2=Reference(ws3,min_col=10,min_row=13,max_row=18)
    c2=Reference(ws3,min_col=2,min_row=14,max_row=18)
    ch2.add_data(d2,titles_from_data=True); ch2.set_categories(c2)
    ws3.add_chart(ch2,"A20")
    
    # 4. مقایسه کشورها
    ws4=wb.create_sheet("🌍 مقایسه کشورها"); ws4.sheet_view.rightToLeft=True
    st(ws4,1,10,"مقایسه جامع کشورها")
    r=3; sh(ws4,r,['کشور','امتیاز','ندا','توحید','سرعت','هزینه','زبان','خانواده'])
    sd(ws4,r+1,[['nz',85,83,55,85,70,70,95],['au',78,76,50,70,60,70,90],['de',65,30,72,65,75,50,85],['ca',60,25,55,55,65,65,85]], [12,12,12,12,12,12,12,12])
    
    r=9; ws4.cell(row=r,column=1,value="📈 سناریوها").font=SF
    r=10; sh(ws4,r,['سناریو','nz','au','de','ca'])
    sd(ws4,r+1,[['ندا اصلی',83,76,30,25],['توحید اصلی',55,50,72,55],['هر دو مستقل',80,70,65,50]], [18,12,12,12,12])
    
    # 5. کارفرمایان
    ws5=wb.create_sheet("🏢 کارفرمایان"); ws5.sheet_view.rightToLeft=True
    st(ws5,1,10,"کارفرمایان شناسایی شده")
    r=3; sh(ws5,r,['نام','کشور','نوع','تخصص','Sponsorship','وب‌سایت','适合','وضعیت'])
    sd(ws5,r+1,[
        ['Health New Zealand','nz','دولتی','بهداشت','✅','healthnz.govt.nz','ندا','تأیید'],
        ['RGH Global','nz','آژانس','مامایی','✅','rgh-global.com','ندا','تأیید'],
        ["St Vincent's Health",'au','بیمارستان','بهداشت','✅','svha.org.au','ندا','تأیید'],
        ['Hassett Group','au','آژانس','بهداشت','✅','hassett.com.au','ندا','تأیید'],
        ['HealthX','au','آژانس','بهداشت','✅','healthx.com.au','ندا','تأیید'],
        ['SAP','de','شرکت IT','IT','🟡','sap.com','توحید','بررسی'],
        ['Siemens','de','شرکت IT','IT','🟡','siemens.com','توحید','بررسی'],
    ], [22,10,12,12,12,22,12,12])
    
    # 6. سازمان‌های دولتی
    ws6=wb.create_sheet("🏛 سازمان‌ها"); ws6.sheet_view.rightToLeft=True
    st(ws6,1,8,"سازمان‌های دولتی و حرفه‌ای")
    r=3; sh(ws6,r,['کشور','سازمان','وب‌سایت','مناسب برای','موضوع'])
    sd(ws6,r+1,[
        ['nz','Immigration NZ','immigration.govt.nz','هر دو','ویزا'],
        ['nz','Midwifery Council','midwiferycouncil.health.nz','ندا','ثبت‌نام'],
        ['nz','Health NZ','healthnz.govt.nz','ندا','استخدام'],
        ['au','Home Affairs','homeaffairs.gov.au','هر دو','ویزا'],
        ['au','AHPRA','ahpra.gov.au','ندا','ثبت‌نام'],
        ['de','Make it in Germany','make-it-in-germany.com','هر دو','مهاجرت'],
        ['de','Bundesagentur','arbeitsagentur.de','توحید','بازار کار'],
    ], [12,22,28,15,15])
    
    # 7. درخواست‌ها
    ws7=wb.create_sheet("📤 درخواست‌ها"); ws7.sheet_view.rightToLeft=True
    st(ws7,1,10,"درخواست‌های آماده ارسال")
    r=3; sh(ws7,r,['App ID','متقاضی','کارفرما','کشور','CV','کاور لیتر','ایمیل','وضعیت','پاسخ','اقدام بعدی'])
    sd(ws7,r+1,[
        ['APP-001','ندا','Health New Zealand','nz','✅','✅','✅','آماده ارسال','—','تأیید کاربر'],
        ['APP-002','ندا','RGH Global','nz','✅','✅','✅','آماده ارسال','—','تأیید کاربر'],
        ['APP-003','ندا','Hassett Group','au','✅','✅','✅','آماده ارسال','—','تأیید کاربر'],
        ['APP-004','توحید','SAP','de','✅','🟡','🟡','آماده نیست','—','تکمیل'],
    ], [12,12,20,8,8,12,10,16,10,16])
    
    # 8. ویزا
    ws8=wb.create_sheet("🛂 ویزا"); ws8.sheet_view.rightToLeft=True
    st(ws8,1,8,"اطلاعات ویزا")
    r=3; sh(ws8,r,['کشور','نوع ویزا','الزام شغلی','زبان','خانواده','ثبت‌نام','منبع','تاریخ'])
    sd(ws8,r+1,[
        ['nz','AEWV','Job offer','ANZSCO 1-2: None','✅','جداگانه','immigration.govt.nz','2026-08-18'],
        ['au','482 TSS','Job offer','IELTS 5.0','✅','AHPRA','homeaffairs.gov.au','2026-08-18'],
        ['de','EU Blue Card','€45,934+','None (IT)','✅','Anerkennung','make-it-in-germany.com','2026-08-18'],
        ['ca','LMIA','Job offer','IELTS 6.5','✅','مختلف','canada.ca','2026-08-18'],
    ], [8,14,14,18,10,14,22,14])
    
    # 9. ثبت‌نام
    ws9=wb.create_sheet("📋 ثبت‌نام"); ws9.sheet_view.rightToLeft=True
    st(ws9,1,10,"ثبت‌نام حرفه‌ای")
    r=3; sh(ws9,r,['کشور','متقاضی','حرفه','نهاد','زبان','آزمون','هزینه','زمان','منبع','وضعیت'])
    sd(ws9,r+1,[
        ['nz','ندا','ماما','Midwifery Council','IELTS 7.0/OET','ممکن است','NZ$485','۱۲ هفته','midwiferycouncil.health.nz','نیاز'],
        ['au','ندا','ماما','AHPRA','IELTS 7.0/OET','ممکن است','—','—','ahpra.gov.au','نیاز'],
        ['de','توحید','IT','Chamber','None for visa','—','—','—','anerkennung-in-deutschland.de','بررسی'],
    ], [8,12,10,18,16,12,12,12,28,10])
    
    # 10. زبان
    ws10=wb.create_sheet("📝 زبان"); ws10.sheet_view.rightToLeft=True
    st(ws10,1,8,"وضعیت زبان")
    r=3; sh(ws10,r,['متقاضی','English','German','IELTS نیاز','OET نیاز','وضعیت','اقدام','اولویت'])
    sd(ws10,r+1,[
        ['توحید','A2','A1','نیاز نیست','نیاز نیست','کافی برای IT','ادامه A2','مهم'],
        ['ندا','A2','A1','IELTS Academic 7.0','OET B','نیاز به ارتقاء','ثبت‌نام OET','فوری'],
    ], [14,10,10,20,14,18,16,12])
    
    # 11. تاریخچه
    ws11=wb.create_sheet("📜 تاریخچه"); ws11.sheet_view.rightToLeft=True
    st(ws11,1,10,"تاریخچه جستجو")
    r=3; sh(ws11,r,['تاریخ','متقاضی','کشورها','منابع','جدید','فرصت‌ها','معتبر','درخواست‌ها','پاسخ','Offer'])
    sd(ws11,r+1,[['2026-08-18','هر دو','nz,au,de','5','5','10','10','4','0','0']], [14,12,14,10,10,10,10,12,10,10])
    
    # 12. استراتژی خانواده
    ws12=wb.create_sheet("👨‍👩‍👧‍👦 خانواده"); ws12.sheet_view.rightToLeft=True
    st(ws12,1,10,"استراتژی خروج خانواده")
    r=3; sh(ws12,r,['سناریو','متقاضی','کشور','تناسب','سرعت','هزینه','خانواده','امتیاز'])
    sd(ws12,r+1,[
        ['A: ندا اصلی','ندا','nz',80,'سریع','متوسط','عالی',82],
        ['B: توحید اصلی','توحید','de',76,'متوسط','کم','خوب',72],
        ['C: هر دو مستقل','هر دو','nz+au',78,'سریع','بالا','عالی',80],
    ], [18,12,12,12,12,12,12,12])
    
    r=8; ws12.cell(row=r,column=1,value="🎯 پیشنهاد: سناریو C").font=BF
    
    # 13. اقدامات
    ws13=wb.create_sheet("🎯 اقدامات"); ws13.sheet_view.rightToLeft=True
    st(ws13,1,8,"برنامه اقدام امروز")
    r=3; sh(ws13,r,['ردیف','اقدام','متقاضی','اولویت','وضعیت','یادداشت'])
    sd(ws13,r+1,[
        [1,'ارسال Health NZ','ندا','🔴 فوری','انجام نشده','بهترین فرصت'],
        [2,'ارسال Hassett Group','ندا','🔴 فوری','انجام نشده','حمایت ویزا'],
        [3,'ارسال RGH Global','ندا','🔴 فوری','انجام نشده','آژانس متخصص'],
        [4,'جستجو IT آلمان','توحید','🟡 متوسط','انجام نشده','بازار خوب'],
        [5,'ثبت‌نام OET','ندا','🟠 مهم','انجام نشده','پیش‌نیاز ثبت‌نام'],
    ], [8,25,12,15,15,20])
    
    fn='MigrationHunter/dashboard/MigrationHunter_Dashboard.xlsx'
    wb.save(fn)
    print(f"✅ {fn}")
    print(f"📊 {len(wb.sheetnames)} sheets:")
    for s in wb.sheetnames: print(f"   - {s}")

if __name__=='__main__': main()
