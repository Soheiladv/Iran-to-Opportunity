#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Migration Hunter Dashboard v3 - With Email/Cover Letter Options
"""

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.chart import BarChart, Reference
from openpyxl.utils import get_column_letter
from datetime import datetime

FONT_NAME = 'B Mitra'
BLUE_DARK = '1F4E79'
BLUE_MED = '2E75B6'
GREEN = 'C6EFCE'
YELLOW = 'FFEB9C'
RED = 'FFC7CE'
ORANGE = 'FFD966'
WHITE = 'FFFFFF'
GRAY = 'F2F2F2'

HEADER_FONT = Font(name=FONT_NAME, bold=True, size=11, color=WHITE)
HEADER_FILL = PatternFill(start_color=BLUE_DARK, end_color=BLUE_DARK, fill_type='solid')
GREEN_FILL = PatternFill(start_color=GREEN, end_color=GREEN, fill_type='solid')
YELLOW_FILL = PatternFill(start_color=YELLOW, end_color=YELLOW, fill_type='solid')
RED_FILL = PatternFill(start_color=RED, end_color=RED, fill_type='solid')
ORANGE_FILL = PatternFill(start_color=ORANGE, end_color=ORANGE, fill_type='solid')
GRAY_FILL = PatternFill(start_color=GRAY, end_color=GRAY, fill_type='solid')

CELL_FONT = Font(name=FONT_NAME, size=10)
BOLD_FONT = Font(name=FONT_NAME, size=10, bold=True)
TITLE_FONT = Font(name=FONT_NAME, size=14, bold=True, color=BLUE_DARK)
LINK_FONT = Font(name=FONT_NAME, size=10, color='0563C1', underline='single')

THIN_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True, readingOrder=2)
RIGHT = Alignment(horizontal='right', vertical='center', wrap_text=True, readingOrder=2)

def sc(cell, font=CELL_FONT, fill=None, align=RIGHT):
    cell.font = font
    cell.alignment = align
    cell.border = THIN_BORDER
    if fill:
        cell.fill = fill

def add_title(ws, row, col_end, title):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=col_end)
    c = ws.cell(row=row, column=1, value=title)
    c.font = TITLE_FONT
    c.alignment = CENTER

def add_headers(ws, row, headers):
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=i, value=h)
        sc(c, HEADER_FONT, HEADER_FILL, CENTER)

def main():
    wb = openpyxl.Workbook()
    if 'Sheet' in wb.sheetnames:
        del wb['Sheet']
    
    # ============ ندا — فرصت‌ها با ایمیل ============
    ws = wb.create_sheet('ندا — فرصت‌ها + ایمیل', 0)
    ws.sheet_view.rightToLeft = True
    
    add_title(ws, 1, 15, 'فرصت‌های ندا آرجمند — با گزینه ایمیل و کاور لیتر')
    
    row = 3
    add_headers(ws, row, [
        'ردیف', 'کارفرما', 'کشور', 'شهر', 'عنوان', 'حقوق',
        'Sponsorship', 'زبان', 'ثبت‌نام', 'تناسب',
        'وضعیت', 'لینک شغل', 'ایمیل درخواست', 'کاور لیتر', 'اقدام بعدی'
    ])
    
    neda_jobs = [
        [1, 'Health New Zealand', 'nz', 'سراسر کشور', 'ماما', '75-106K NZD',
         '✅', 'IELTS/OET', 'Midwifery Council', '80/100',
         'جدید', 'healthnz.govt.nz/careers', '✅ آماده', '✅ آماده', 'ارسال'],
        [2, 'RGH Global', 'nz', 'سراسر کشور', 'ماما', '75-106K NZD',
         '✅', 'IELTS/OET', 'Midwifery Council', '78/100',
         'جدید', 'rgh-global.com/jobs', '✅ آماده', '✅ آماده', 'ارسال'],
        [3, 'Hassett Group', 'au', 'Melbourne', 'ماما', '80-120K AUD',
         '✅', 'IELTS/OET', 'AHPRA', '78/100',
         'جدید', 'hassett.com.au', '✅ آماده', '✅ آماده', 'ارسال'],
        [4, 'Talent Angels', 'au', 'Victoria', 'ماما', '80-120K AUD',
         '✅', 'IELTS/OET', 'AHPRA', '76/100',
         'جدید', 'LinkedIn', '🟡 نیاز', '🟡 نیاز', 'بررسی'],
        [5, 'HealthX', 'au', 'سراسر کشور', 'ماما', '67-95K AUD',
         '✅', 'IELTS/OET', 'AHPRA', '74/100',
         'جدید', 'healthx.com.au', '✅ آماده', '✅ آماده', 'ارسال'],
    ]
    
    for r, row_data in enumerate(neda_jobs, row+1):
        for c_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c_idx, value=val)
            fill = None
            if str(val) == '✅ آماده':
                fill = GREEN_FILL
            elif str(val) == '🟡 نیاز':
                fill = YELLOW_FILL
            sc(cell, fill=fill)
    
    widths = [6, 20, 8, 14, 12, 16, 10, 12, 16, 10, 10, 22, 14, 14, 14]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    
    # ============ توحید — فرصت‌ها با ایمیل ============
    ws2 = wb.create_sheet('توحید — فرصت‌ها + ایمیل')
    ws2.sheet_view.rightToLeft = True
    
    add_title(ws2, 1, 15, 'فرصت‌های توحید آرجمند — با گزینه ایمیل و کاور لیتر')
    
    row = 3
    add_headers(ws2, row, [
        'ردیف', 'کارفرما', 'کشور', 'شهر', 'عنوان', 'حقوق',
        'Sponsorship', 'زبان', 'ثبت‌نام', 'تناسب',
        'وضعیت', 'لینک شغل', 'ایمیل درخواست', 'کاور لیتر', 'اقدام بعدی'
    ])
    
    tohid_jobs = [
        [1, 'SAP', 'de', 'والدورف', 'مدیر IT', '€55-80K',
         '🟢', 'انگلیسی', '—', '82/100',
         'جدید', 'sap.com/careers', '🟡 نیاز', '🟡 نیاز', 'جستجو'],
        [2, 'Siemens', 'de', 'مونیخ', 'مدیر IT', '€60-90K',
         '🟢', 'انگلیسی', '—', '80/100',
         'جدید', 'siemens.com/careers', '🟡 نیاز', '🟡 نیاز', 'جستجو'],
        [3, 'Deutsche Telekom', 'de', 'بون', 'مدیر IT', '€50-75K',
         '🟢', 'انگلیسی', '—', '78/100',
         'جدید', 'telekom.com/careers', '🟡 نیاز', '🟡 نیاز', 'جستجو'],
        [4, 'Allianz', 'de', 'مونیخ', 'مدیر IT', '€55-85K',
         '🟢', 'انگلیسی', '—', '77/100',
         'جدید', 'allianz.com/careers', '🟡 نیاز', '🟡 نیاز', 'جستجو'],
        [5, 'BMW', 'de', 'مونیخ', 'مدیر IT', '€60-90K',
         '🟢', 'انگلیسی', '—', '76/100',
         'جدید', 'bmw.com/careers', '🟡 نیاز', '🟡 نیاز', 'جستجو'],
    ]
    
    for r, row_data in enumerate(tohid_jobs, row+1):
        for c_idx, val in enumerate(row_data, 1):
            cell = ws2.cell(row=r, column=c_idx, value=val)
            fill = None
            if str(val) == '✅ آماده':
                fill = GREEN_FILL
            elif str(val) == '🟡 نیاز':
                fill = YELLOW_FILL
            sc(cell, fill=fill)
    
    for i, w in enumerate(widths, 1):
        ws2.column_dimensions[get_column_letter(i)].width = w
    
    # ============ داشبورد ============
    ws3 = wb.create_sheet('داشبورد')
    ws3.sheet_view.rightToLeft = True
    
    add_title(ws3, 1, 8, 'داشبورد شکار فرصت شغلی')
    
    row = 3
    add_headers(ws3, row, ['شاخص', 'تعداد'])
    summary = [
        ['کل فرصت‌ها', 10],
        ['فرصت‌های ندا', 5],
        ['فرصت‌های توحید', 5],
        ['ایمیل آماده', 6],
        ['کاور لیتر آماده', 6],
        ['لینک شغل', 10],
        ['Sponsorship تأیید', 8],
    ]
    for r, row_data in enumerate(summary, row+1):
        for c_idx, val in enumerate(row_data, 1):
            cell = ws3.cell(row=r, column=c_idx, value=val)
            sc(cell)
    
    ws3.column_dimensions['A'].width = 25
    ws3.column_dimensions['B'].width = 15
    
    # Save
    filename = 'MigrationHunter/dashboard/MigrationHunter_Dashboard.xlsx'
    wb.save(filename)
    print(f"✅ Dashboard v3 created: {filename}")
    print(f"📊 {len(wb.sheetnames)} sheets:")
    for s in wb.sheetnames:
        print(f"   - {s}")

if __name__ == '__main__':
    main()
