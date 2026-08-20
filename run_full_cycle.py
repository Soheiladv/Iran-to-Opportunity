#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Migration Hunter — اجرای کامل
اروپا + نقاط ضعف لینکدین + داشبورد نهایی
"""
import os
import json
from datetime import datetime
from pathlib import Path

BASE_DIR = Path(__file__).parent
NOW = datetime.now()
DATE_STR = NOW.strftime("%Y-%m-%d %H:%M")

# ════════════════════════════════════════════════════════════
# کشورهای اروپا
# ════════════════════════════════════════════════════════════

EUROPE_COUNTRIES = {
    "at": {"en": "Austria", "fa": "اتریش", "visa": "Red-White-Red Card", "lang": "German B1-B2"},
    "ie": {"en": "Ireland", "fa": "ایرلند", "visa": "Critical Skills", "lang": "English"},
    "nl": {"en": "Netherlands", "fa": "هلند", "visa": "Kennismigrant", "lang": "English/Dutch"},
    "se": {"en": "Sweden", "fa": "سوئد", "visa": "Work Permit", "lang": "Swedish/English"},
    "no": {"en": "Norway", "fa": "نروژ", "visa": "Skilled Worker", "lang": "Norwegian/English"},
    "dk": {"en": "Denmark", "fa": "دانمارک", "visa": "Pay Limit/Positive List", "lang": "Danish/English"},
    "fi": {"en": "Finland", "fa": "فنلاند", "visa": "Residence Permit", "lang": "Finnish/English"},
    "uk": {"en": "United Kingdom", "fa": "بریتانیا", "visa": "Skilled Worker", "lang": "English"},
    "ch": {"en": "Switzerland", "fa": "سوئیس", "visa": "Work Permit", "lang": "German/French/Italian"},
    "be": {"en": "Belgium", "fa": "بلژیک", "visa": "EU Blue Card", "lang": "Dutch/French/English"},
    "pt": {"en": "Portugal", "fa": "پرتغال", "visa": "D1 Visa", "lang": "Portuguese/English"},
    "es": {"en": "Spain", "fa": "اسپانیا", "visa": "EU Blue Card", "lang": "Spanish/English"},
    "it": {"en": "Italy", "fa": "ایتالیا", "visa": "Work Permit", "lang": "Italian/English"},
    "fr": {"en": "France", "fa": "فرانسه", "visa": "Passeport Talent", "lang": "French/English"},
    "pl": {"en": "Poland", "fa": "لهستان", "visa": "Work Permit", "lang": "Polish/English"},
    "cz": {"en": "Czech Republic", "fa": "چک", "visa": "EU Blue Card", "lang": "Czech/English"},
    "hu": {"en": "Hungary", "fa": "مجارستان", "visa": "White Card", "lang": "Hungarian/English"},
    "ro": {"en": "Romania", "fa": "رومانی", "visa": "Work Permit", "lang": "Romanian/English"},
    "hr": {"en": "Croatia", "fa": "کرواسی", "visa": "Work Permit", "lang": "Croatian/English"},
}

# ════════════════════════════════════════════════════════════
# نقاط ضعف لینکدین
# ════════════════════════════════════════════════════════════

LINKEDIN_WEAKNESSES = {
    "neda": {
        "name": "Neda Arjmand",
        "profession": "Midwife",
        "weaknesses": [
            {
                "category": "Profile Photo",
                "issue": "عدم وجود عکس حرفه‌ای",
                "impact": "بالا",
                "fix": "عکس پرتره حرفه‌ای با پس‌زمینه ساده بگذارید",
                "priority": "🔴 فوری"
            },
            {
                "category": "Headline",
                "issue": "عنوان ساده و غیرجذاب",
                "impact": "بالا",
                "fix": "عنوان توصیفی: 'Registered Midwife | Maternal Healthcare | Open to International Opportunities'",
                "priority": "🔴 فوری"
            },
            {
                "category": "About Section",
                "issue": "ندارد یا خیلی کوتاه",
                "impact": "بالا",
                "fix": "نوشتن ۳-۵ پاراگراف درباره تجربه بالینی، مهارت‌ها و هدف شغلی",
                "priority": "🔴 فوری"
            },
            {
                "category": "Experience",
                "issue": "توضیحات شغلی ناقص",
                "impact": "متوسط",
                "fix": "افزودن دستاوردها و مسئولیت‌ها با اعداد و آمار",
                "priority": "🟠 مهم"
            },
            {
                "category": "Skills",
                "issue": "مهارت‌های کمتر از ۵",
                "impact": "متوسط",
                "fix": "افزودن ۱۰-۱۵ مهارت تخصصی مامایی",
                "priority": "🟠 مهم"
            },
            {
                "category": "Recommendations",
                "issue": "ندارد",
                "impact": "متوسط",
                "fix": "درخواست توصیه‌نامه از همکاران و مدیران",
                "priority": "🟡 متوسط"
            },
            {
                "category": "Certifications",
                "issue": "ندارد",
                "impact": "متوسط",
                "fix": "افزودن گواهینامه‌های مامایی و دوره‌ها",
                "priority": "🟠 مهم"
            },
            {
                "category": "Languages",
                "issue": "فقط فارسی",
                "impact": "بالا",
                "fix": "افزودن English (A2) و German (A1) با صداقت",
                "priority": "🔴 فوری"
            },
            {
                "category": "Location",
                "issue": "تهران — محدودکننده",
                "impact": "متوسط",
                "fix": "تغییر به 'Tehran, Iran • Open to Remote/International'",
                "priority": "🟡 متوسط"
            },
            {
                "category": "Activity",
                "issue": "فعالیت کم",
                "impact": "متوسط",
                "fix": "لایک و کامنت روی پست‌های حوزه مامایی بین‌المللی",
                "priority": "🟡 متوسط"
            },
        ]
    },
    "tohid": {
        "name": "Tohid Arjmand",
        "profession": "IT Operations Manager",
        "weaknesses": [
            {
                "category": "Profile Photo",
                "issue": "عدم وجود عکس حرفه‌ای",
                "impact": "بالا",
                "fix": "عکس پرتره حرفه‌ای با پس‌زمینه ساده",
                "priority": "🔴 فوری"
            },
            {
                "category": "Headline",
                "issue": "عنوان غیراستاندارد",
                "impact": "بالا",
                "fix": "عنوان: 'IT Operations Manager | Infrastructure | Network Administration | Open to International Roles'",
                "priority": "🔴 فوری"
            },
            {
                "category": "About Section",
                "issue": "ندارد یا فنی و خشک",
                "impact": "بالا",
                "fix": "نوشتن خلاصه حرفه‌ای با تمرکز بر مدیریت و زیرساخت",
                "priority": "🔴 فوری"
            },
            {
                "category": "Experience",
                "issue": "توضیحات فنی بدون نتیجه",
                "impact": "متوسط",
                "fix": "افزودن دستاوردها با معیارهای کمّی (تعداد سرور، کاهش downtime و...)",
                "priority": "🟠 مهم"
            },
            {
                "category": "Skills",
                "issue": "مهارت‌های زیاد ولی پراکنده",
                "impact": "متوسط",
                "fix": "تمرکز بر ۱۰ مهارت کلیدی + endorsements",
                "priority": "🟠 مهم"
            },
            {
                "category": "Recommendations",
                "issue": "ندارد",
                "impact": "بالا",
                "fix": "درخواست توصیه‌نامه از مدیران و همکاران",
                "priority": "🔴 فوری"
            },
            {
                "category": "Certifications",
                "issue": "ندارد",
                "impact": "متوسط",
                "fix": "افزودن گواهینامه‌های Microsoft/Cisco/VMware",
                "priority": "🟠 مهم"
            },
            {
                "category": "Languages",
                "issue": "فقط فارسی",
                "impact": "بالا",
                "fix": "افزودن English (A2) و German (A1)",
                "priority": "🔴 فوری"
            },
            {
                "category": "Location",
                "issue": "تهران",
                "impact": "متوسط",
                "fix": "تغییر به 'Tehran, Iran • Open to Remote/International'",
                "priority": "🟡 متوسط"
            },
            {
                "category": "Open to Work",
                "issue": "فعال نیست",
                "impact": "بالا",
                "fix": "فعال کردن 'Open to Work' badge برای کارفرمایان",
                "priority": "🔴 فوری"
            },
        ]
    }
}

# ════════════════════════════════════════════════════════════
# ساخت Excel نهایی
# ════════════════════════════════════════════════════════════

def build_final_excel():
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    except ImportError:
        print("❌ pip install openpyxl")
        return

    wb = Workbook()

    # Styles
    FA = Font(name='B Mitra', size=11)
    FA_T = Font(name='B Mitra', size=14, bold=True)
    HD = Font(name='B Mitra', size=11, bold=True, color='FFFFFF')
    GF = PatternFill(start_color='92D050', end_color='92D050', fill_type='solid')
    YF = PatternFill(start_color='FFD966', end_color='FFD966', fill_type='solid')
    OF = PatternFill(start_color='F4B084', end_color='F4B084', fill_type='solid')
    BF = PatternFill(start_color='9BC2E6', end_color='9BC2E6', fill_type='solid')
    RF = PatternFill(start_color='FF6B6B', end_color='FF6B6B', fill_type='solid')
    HF = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    BD = Border(left=Side('thin'), right=Side('thin'), top=Side('thin'), bottom=Side('thin'))

    def sh(ws, r, c):
        cell = ws.cell(row=r, column=c)
        cell.font = HD; cell.fill = HF
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = BD

    def sc(ws, r, c, v=None, f=FA, fill=None):
        cell = ws.cell(row=r, column=c)
        if v: cell.value = v
        cell.font = f
        cell.alignment = Alignment(vertical='center', wrap_text=True, horizontal='right')
        cell.border = BD
        if fill: cell.fill = fill
        return cell

    # ── شیت ۱: داشبورد ──
    ws1 = wb.active
    ws1.title = "📊 Dashboard | داشبورد"
    ws1.sheet_view.rightToLeft = True
    ws1.merge_cells('A1:F1')
    ws1.cell(1, 1, f"📊 داشبورد جامع | {DATE_STR}").font = FA_T

    h = ["شاخص | Metric", "مقدار | Value"]
    for i, x in enumerate(h, 1): ws1.cell(3, i, x)
    sh(ws1, 3, 2)

    data = [
        ("کشورهای اروپا | Europe", 19),
        ("لینکدین‌ها | LinkedIn", 2),
        ("کل منابع | Sources", 13),
        ("کل کارفرمایان | Employers", 7),
        ("کل فرصت‌ها | Jobs", 8),
        ("نقاط ضعف شناسایی شده | Weaknesses", 20),
    ]
    for r, (l, v) in enumerate(data, 4):
        sc(ws1, r, 1, l); sc(ws1, r, 2, v)
    ws1.column_dimensions['A'].width = 40
    ws1.column_dimensions['B'].width = 15

    # ── شیت ۲: کشورهای اروپا ──
    ws2 = wb.create_sheet("🌍 Europe | اروپا")
    ws2.sheet_view.rightToLeft = True
    ws2.merge_cells('A1:E1')
    ws2.cell(1, 1, "🌍 کشورهای اروپا برای جستجو | European Countries").font = FA_T

    h2 = ["کد | Code", "نام | Country", "ویزا | Visa", "زبان | Language", "اولویت | Priority"]
    for i, x in enumerate(h2, 1): ws2.cell(3, i, x)
    sh(ws2, 3, 5)

    for r, (code, info) in enumerate(EUROPE_COUNTRIES.items(), 4):
        sc(ws2, r, 1, code.upper())
        sc(ws2, r, 2, f"{info['fa']} ({info['en']})")
        sc(ws2, r, 3, info['visa'])
        sc(ws2, r, 4, info['lang'])
        c = sc(ws2, r, 5, "🟢")
        c.fill = GF

    for col in 'ABCDE': ws2.column_dimensions[col].width = 28

    # ── شیت ۳: نقاط ضعف ندا ──
    ws3 = wb.create_sheet("⚠️ Neda Weaknesses | نقاط ضعف ندا")
    ws3.sheet_view.rightToLeft = True
    ws3.merge_cells('A1:E1')
    ws3.cell(1, 1, f"⚠️ تحلیل نقاط ضعف لینکدین — {LINKEDIN_WEAKNESSES['neda']['name']}").font = FA_T

    h3 = ["دسته | Category", "مشکل | Issue", "تأثیر | Impact", "راه‌حل | Fix", "اولویت | Priority"]
    for i, x in enumerate(h3, 1): ws3.cell(3, i, x)
    sh(ws3, 3, 5)

    for r, w in enumerate(LINKEDIN_WEAKNESSES['neda']['weaknesses'], 4):
        sc(ws3, r, 1, w['category'])
        sc(ws3, r, 2, w['issue'])
        imp = sc(ws3, r, 3, w['impact'])
        imp.fill = RF if w['impact'] == 'بالا' else OF if w['impact'] == 'متوسط' else YF
        sc(ws3, r, 4, w['fix'])
        pri = sc(ws3, r, 5, w['priority'])

    for col in 'ABCDE': ws3.column_dimensions[col].width = 35

    # ── شیت ۴: نقاط ضعف توحید ──
    ws4 = wb.create_sheet("⚠️ Tohid Weaknesses | نقاط ضعف توحید")
    ws4.sheet_view.rightToLeft = True
    ws4.merge_cells('A1:E1')
    ws4.cell(1, 1, f"⚠️ تحلیل نقاط ضعف لینکدین — {LINKEDIN_WEAKNESSES['tohid']['name']}").font = FA_T

    for i, x in enumerate(h3, 1): ws4.cell(3, i, x)
    sh(ws4, 3, 5)

    for r, w in enumerate(LINKEDIN_WEAKNESSES['tohid']['weaknesses'], 4):
        sc(ws4, r, 1, w['category'])
        sc(ws4, r, 2, w['issue'])
        imp = sc(ws4, r, 3, w['impact'])
        imp.fill = RF if w['impact'] == 'بالا' else OF if w['impact'] == 'متوسط' else YF
        sc(ws4, r, 4, w['fix'])
        pri = sc(ws4, r, 5, w['priority'])

    for col in 'ABCDE': ws4.column_dimensions[col].width = 35

    # ── شیت ۵: فرصت‌های ندا ──
    ws5 = wb.create_sheet("👩 Neda Jobs | فرصت‌های ندا")
    ws5.sheet_view.rightToLeft = True
    ws5.merge_cells('A1:H1')
    ws5.cell(1, 1, "👩 فرصت‌های مامایی ندا | Neda Midwifery Opportunities").font = FA_T

    h5 = ["ردیف | #", "کارفرما | Employer", "کشور | Country", "شغل | Job", "لینک | Link", "حمایت | Sponsor", "امتیاز | Score", "وضعیت | Status"]
    for i, x in enumerate(h5, 1): ws5.cell(3, i, x)
    sh(ws5, 3, 8)

    jobs_neda = [
        (1, "Health New Zealand", "nz", "Midwife", "healthnz.govt.nz", "✅", 85, "🟢"),
        (2, "RGH Global", "nz", "Midwife", "rgh-global.com", "✅", 79, "🟢"),
        (3, "Working In Health NZ", "nz", "Midwife", "workingin-health.co.nz", "✅", 79, "🟢"),
        (4, "Holalemania GmbH", "de", "Geburtshelfer", "holalemania.de", "✅", 82, "🟢"),
        (5, "Saskatchewan HA", "ca", "Midwife", "saskhealthauthority.ca", "⚠️", 78, "🟡"),
        (6, "Hays Healthcare", "au", "Midwife", "hays.com.au", "⚠️", 72, "🟡"),
        (7, "TalentOrange", "de", "Healthcare", "talentorange.com", "✅", 80, "🟢"),
    ]
    for r, (n, emp, c, j, u, s, sc2, st) in enumerate(jobs_neda, 4):
        sc(ws5, r, 1, n); sc(ws5, r, 2, emp)
        sc(ws5, r, 3, EUROPE_COUNTRIES.get(c, {}).get('fa', c) if c in EUROPE_COUNTRIES else {'nz':'نیوزیلند','au':'استرالیا','de':'آلمان','ca':'کانادا'}.get(c, c))
        sc(ws5, r, 4, j); sc(ws5, r, 5, u); sc(ws5, r, 6, s)
        cell_sc = sc(ws5, r, 7, sc2)
        cell_sc.fill = GF if sc2 >= 80 else YF if sc2 >= 70 else OF
        sc(ws5, r, 8, st)

    for col in 'ABCDEFGH': ws5.column_dimensions[col].width = 22

    # ── شیت ۶: فرصت‌های توحید ──
    ws6 = wb.create_sheet("👨 Tohid Jobs | فرصت‌های توحید")
    ws6.sheet_view.rightToLeft = True
    ws6.merge_cells('A1:H1')
    ws6.cell(1, 1, "👨 فرصت‌های IT توحید | Tohid IT Opportunities").font = FA_T
    for i, x in enumerate(h5, 1): ws6.cell(3, i, x)
    sh(ws6, 3, 8)

    jobs_tohid = [
        (1, "Arbeitnow", "de", "IT Manager", "arbeitnow.com", "⚠️", 65, "🟡"),
    ]
    for r, (n, emp, c, j, u, s, sc2, st) in enumerate(jobs_tohid, 4):
        sc(ws6, r, 1, n); sc(ws6, r, 2, emp)
        sc(ws6, r, 3, {'de':'آلمان'}.get(c, c))
        sc(ws6, r, 4, j); sc(ws6, r, 5, u); sc(ws6, r, 6, s)
        cell_sc = sc(ws6, r, 7, sc2)
        cell_sc.fill = GF if sc2 >= 80 else YF if sc2 >= 70 else OF
        sc(ws6, r, 8, st)

    for col in 'ABCDEFGH': ws6.column_dimensions[col].width = 22

    # ── شیت ۷: مقایسه ──
    ws7 = wb.create_sheet("📈 Comparison | مقایسه")
    ws7.sheet_view.rightToLeft = True
    ws7.merge_cells('A1:E1')
    ws7.cell(1, 1, "📈 مقایسه کشورها | Country Comparison").font = FA_T

    h7 = ["کشور | Country", "ندا | Neda", "توحید | Tohid", "زبان | Language", "ویزا | Visa"]
    for i, x in enumerate(h7, 1): ws7.cell(3, i, x)
    sh(ws7, 3, 5)

    comp = [
        ("nz", 85, 60, "English", "AEWV/Green List"),
        ("au", 72, 55, "English", "482/189/190"),
        ("de", 82, 65, "German B1-B2", "Blue Card"),
        ("ca", 78, 55, "CLB 7", "Express Entry/PNP"),
        ("uk", 70, 50, "English", "Skilled Worker"),
        ("ie", 68, 48, "English", "Critical Skills"),
    ]
    for r, (c, n, t, l, v) in enumerate(comp, 4):
        sc(ws7, r, 1, {'nz':'نیوزیلند','au':'استرالیا','de':'آلمان','ca':'کانادا','uk':'بریتانیا','ie':'ایرلند'}.get(c, c))
        c1 = sc(ws7, r, 2, n); c1.fill = GF if n >= 80 else YF
        c2 = sc(ws7, r, 3, t); c2.fill = GF if t >= 80 else YF
        sc(ws7, r, 4, l); sc(ws7, r, 5, v)

    for col in 'ABCDE': ws7.column_dimensions[col].width = 25

    # ── شیت ۸: اقدامات ──
    ws8 = wb.create_sheet("🎯 Actions | اقدامات")
    ws8.sheet_view.rightToLeft = True
    ws8.merge_cells('A1:D1')
    ws8.cell(1, 1, "🎯 ۵ اقدام برتر امروز | Top 5 Actions").font = FA_T

    h8 = ["ردیف | #", "اقدام | Action", "کشور | Country", "اولویت | Priority"]
    for i, x in enumerate(h8, 1): ws8.cell(3, i, x)
    sh(ws8, 3, 4)

    actions = [
        (1, "ارسال CV به Saskatchewan Health Authority", "🇨🇦 کانادا", "🔴 فوری"),
        (2, "ثبت‌نام در Hays Healthcare Australia", "🇦🇺 استرالیا", "🔴 فوری"),
        (3, "پیگیری Working In Health NZ", "🇳🇿 نیوزیلند", "🟠 مهم"),
        (4, "شروع آمادگی OET", "🌍 بین‌المللی", "🟠 مهم"),
        (5, "جستجوی IT آلمان/اتریش", "🇩🇪🇦🇹 اروپا", "🟠 مهم"),
    ]
    for r, (n, a, c, p) in enumerate(actions, 4):
        sc(ws8, r, 1, n); sc(ws8, r, 2, a)
        sc(ws8, r, 3, c); sc(ws8, r, 4, p)

    for col in 'ABCD': ws8.column_dimensions[col].width = 35

    # ── ذخیره ──
    dash = BASE_DIR / "dashboard"
    dash.mkdir(exist_ok=True)
    fn = f"MigrationHunter_Full_{NOW.strftime('%Y%m%d_%H%M')}.xlsx"
    fp = dash / fn
    wb.save(fp)
    print(f"\n✅ داشبورد نهایی: {fn}")
    print(f"   شیت‌ها: {len(wb.sheetnames)}")
    for s in wb.sheetnames:
        print(f"   • {s}")
    return fp

# ── اجرا ──
if __name__ == "__main__":
    print("=" * 60)
    print("🔄 اجرای کامل — اروپا + نقاط ضعف لینکدین")
    print(f"📅 {DATE_STR}")
    print("=" * 60)
    build_final_excel()
    print("\n✅ تمام!")
