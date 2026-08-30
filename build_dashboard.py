#!/usr/bin/env python3
"""
MigrationHunter — Unified Dashboard Builder
خواندن از memory/*.md → تولید Excel 12 شیتی

اجرا: python build_dashboard.py
"""
import os, sys, re, glob, io
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import (Font, Alignment, PatternFill, Border, Side,
                              numbers)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, PieChart, Reference
from config_loader import get_applicant_label, get_all_applicant_labels, get_applicants

# Fix Windows console encoding for emoji support
if sys.platform == 'win32':
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8', errors='replace')
    os.environ['PYTHONIOENCODING'] = 'utf-8'

# ═══════════════════════════════════════════════════
# CONSTANTS
# ═══════════════════════════════════════════════════
BASE = os.path.dirname(os.path.abspath(__file__))
MEM = os.path.join(BASE, "memory")
PRO = os.path.join(BASE, "profiles")
DASH = os.path.join(BASE, "dashboard")
OUT = os.path.join(BASE, "output")
ARCH = os.path.join(DASH, "archive")

FONT_FA = "B Mitra"
FONT_EN = "Times New Roman"
NOW = datetime.now()
DATE_STR = NOW.strftime("%Y-%m-%d %H:%M")
FILE_DATE = NOW.strftime("%Y%m%d_%H%M")

# Colors
C_DARK = "1B4F72"
C_MED = "2E86C1"
C_LIGHT = "D6EAF8"
C_GREEN = "27AE60"
C_LGREEN = "D5F5E3"
C_YELLOW = "F39C12"
C_LYELLOW = "FEF9E7"
C_RED = "E74C3C"
C_LRED = "FADBD8"
C_PURPLE = "8E44AD"
C_LPURPLE = "E8DAEF"
C_GRAY = "95A5A6"
C_LGRAY = "F2F3F4"
C_WHITE = "FFFFFF"
C_DARK2 = "2C3E50"

# ═══════════════════════════════════════════════════
# STYLE HELPERS
# ═══════════════════════════════════════════════════
thin = Border(left=Side("thin"), right=Side("thin"),
              top=Side("thin"), bottom=Side("thin"))
thick = Border(left=Side("medium"), right=Side("medium"),
               top=Side("medium"), bottom=Side("medium"))

def rtl(ws):
    ws.sheet_view.rightToLeft = True

def fa(name=FONT_FA, sz=10, bold=False, color="000000", italic=False):
    return Font(name=name, size=sz, bold=bold, color=color, italic=italic)

def en(name=FONT_EN, sz=10, bold=False, color="000000"):
    return Font(name=name, size=sz, bold=bold, color=color)

def fill(c):
    return PatternFill(start_color=c, end_color=c, fill_type="solid")

def ac(h="right", v="center", wrap=True):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def center():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def wc(ws, r, c, val, font=None, bg=None, align=None, border=True):
    cell = ws.cell(row=r, column=c, value=val)
    if font: cell.font = font
    if bg: cell.fill = fill(bg)
    if align: cell.alignment = align
    else: cell.alignment = ac()
    if border: cell.border = thin
    return cell

def header_row(ws, row, cols, bg=C_DARK):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = fa(sz=10, bold=True, color=C_WHITE)
        cell.fill = fill(bg)
        cell.alignment = center()
        cell.border = thin

def auto_width(ws, col, w):
    ws.column_dimensions[get_column_letter(col)].width = w

def freeze(ws, cell="A2"):
    ws.freeze_panes = cell

# ═══════════════════════════════════════════════════
# PARSE MARKDOWN TABLES
# ═══════════════════════════════════════════════════
def parse_md_tables(filepath):
    """Parse all markdown tables from a file. Returns list of list-of-dicts."""
    if not os.path.exists(filepath):
        return []
    with open(filepath, "r", encoding="utf-8") as f:
        content = f.read()
    
    tables = []
    lines = content.split("\n")
    i = 0
    while i < len(lines):
        line = lines[i].strip()
        # Detect table header row (contains |)
        if "|" in line and i + 1 < len(lines) and re.match(r'^[\s|:-]+$', lines[i+1].strip()):
            headers = [h.strip() for h in line.split("|") if h.strip()]
            i += 2  # skip separator
            rows = []
            while i < len(lines) and "|" in lines[i]:
                cells = [c.strip() for c in lines[i].split("|") if c.strip()]
                if len(cells) >= len(headers):
                    rows.append(dict(zip(headers, cells[:len(headers)])))
                elif cells:
                    row = {}
                    for j, h in enumerate(headers):
                        row[h] = cells[j] if j < len(cells) else ""
                    rows.append(row)
                i += 1
            if rows:
                tables.append({"headers": headers, "rows": rows})
        i += 1
    return tables

def parse_md_sections(filepath):
    """Parse sections with headers. Returns dict of section_name -> content."""
    if not os.path.exists(filepath):
        return {}
    with open(filepath, "r", encoding="utf-8") as f:
        content = f.read()
    sections = {}
    current = "__top__"
    buf = []
    for line in content.split("\n"):
        if line.startswith("## "):
            sections[current] = "\n".join(buf)
            current = line[3:].strip()
            buf = []
        else:
            buf.append(line)
    sections[current] = "\n".join(buf)
    return sections

# ═══════════════════════════════════════════════════
# LOAD ALL DATA FROM MEMORY
# ═══════════════════════════════════════════════════
def load_all_data():
    data = {}
    
    # Sources
    tables = parse_md_tables(os.path.join(MEM, "SOURCE_BANK.md"))
    data["sources"] = tables[0]["rows"] if tables else []
    
    # Employers
    tables = parse_md_tables(os.path.join(MEM, "EMPLOYER_BANK.md"))
    data["employers"] = tables[0]["rows"] if tables else []
    
    # Jobs
    tables = parse_md_tables(os.path.join(MEM, "JOB_BANK.md"))
    data["jobs"] = tables[0]["rows"] if tables else []
    
    # Evidence
    sections = parse_md_sections(os.path.join(MEM, "EVIDENCE_REGISTRY.md"))
    data["evidence_sections"] = sections
    
    # Application Bank
    app_tables = parse_md_tables(os.path.join(MEM, "APPLICATION_BANK.md"))
    data["applications"] = app_tables
    
    # Visa
    visa_tables = parse_md_tables(os.path.join(MEM, "VISA_BANK.md"))
    data["visa"] = []
    for t in visa_tables:
        data["visa"].extend(t["rows"])
    
    # Registration
    reg_tables = parse_md_tables(os.path.join(MEM, "REGISTRATION_BANK.md"))
    data["registration"] = []
    for t in reg_tables:
        data["registration"].extend(t["rows"])
    
    # Search History
    hist_tables = parse_md_tables(os.path.join(MEM, "SEARCH_HISTORY.md"))
    data["search_history"] = []
    for t in hist_tables:
        data["search_history"].extend(t["rows"])
    
    # Profiles
    data["tohid"] = parse_md_sections(os.path.join(PRO, "TOHID_PROFILE.md"))
    data["neda"] = parse_md_sections(os.path.join(PRO, "NEDA_PROFILE.md"))
    
    return data

# ═══════════════════════════════════════════════════
# ENRICH OPPORTUNITIES (merge jobs + evidence + apps)
# ═══════════════════════════════════════════════════
def build_opportunities(data):
    """Build enriched opportunity list from all memory banks."""
    opps = []
    
    # Evidence opportunities (most detailed)
    evidence_map = {}
    for key, content in data.get("evidence_sections", {}).items():
        if key.startswith("MH-2026-"):
            # Extract ID
            opp_id = key.split("—")[0].strip() if "—" in key else key.split(" ")[0].strip()
            # Extract info from key
            parts = key.split("—")
            applicant = parts[1].strip() if len(parts) > 1 else ""
            employer = parts[2].strip() if len(parts) > 2 else ""
            job = parts[3].strip() if len(parts) > 3 else ""
            
            # Extract scores from content
            evidence_score = 0
            final_score = 0
            decision = ""
            for line in content.split("\n"):
                if "Evidence Score:" in line:
                    m = re.search(r'(\d+)/100', line)
                    if m: evidence_score = int(m.group(1))
                if "Final Score:" in line:
                    m = re.search(r'(\d+)/100', line)
                    if m: final_score = int(m.group(1))
                if "Decision:" in line:
                    decision = line.split("Decision:")[1].strip().split("\n")[0].strip()
            
            evidence_map[opp_id] = {
                "id": opp_id,
                "applicant": applicant,
                "employer": employer,
                "job": job,
                "evidence_score": evidence_score,
                "final_score": final_score,
                "decision": decision,
            }
    
    # Enrich from JOB_BANK
    for j in data.get("jobs", []):
        opp = {
            "id": "",
            "applicant": j.get("متقاضی", ""),
            "country": j.get("کشور", ""),
            "employer": j.get("کارفرما", ""),
            "job": j.get("شغل", ""),
            "evidence_score": int(j.get("امتیاز", "0") or "0"),
            "final_score": 0,
            "status": j.get("وضعیت", ""),
            "decision": "",
            "email": "",
            "url": "",
            "sponsorship": "",
            "visa": "",
            "language": "",
            "registration": "",
            "salary": "",
            "source": "",
            "search_method": "",
        }
        
        # Find matching evidence
        for eid, ev in evidence_map.items():
            if ev["employer"].lower() in opp["employer"].lower() or opp["employer"].lower() in ev["employer"].lower():
                opp["id"] = eid
                opp["evidence_score"] = ev["evidence_score"]
                opp["final_score"] = ev["final_score"]
                opp["decision"] = ev["decision"]
                break
        
        if not opp["id"]:
            opp["id"] = f"MH-2026-{len(opps)+1:03d}"
        
        opps.append(opp)
    
    # Add evidence-only opportunities not in JOB_BANK
    for eid, ev in evidence_map.items():
        found = any(o["id"] == eid for o in opps)
        if not found:
            opps.append({
                "id": eid,
                "applicant": ev["applicant"],
                "country": "",
                "employer": ev["employer"],
                "job": ev["job"],
                "evidence_score": ev["evidence_score"],
                "final_score": ev["final_score"],
                "status": "",
                "decision": ev["decision"],
                "email": "",
                "url": "",
                "sponsorship": "",
                "visa": "",
                "language": "",
                "registration": "",
                "salary": "",
                "source": "",
                "search_method": "",
            })
    
    return opps

# ═══════════════════════════════════════════════════
# VERIFIED EMAILS (from previous research)
# ═══════════════════════════════════════════════════
VERIFIED_EMAILS = {
    "Health New Zealand": "international.recruitment@health.govt.nz",
    "RGH Global": "info@rgh-global.com",
    "Alberta Health Services": "careers@albertahealthservices.ca",
    "Saskatchewan HA": "SHAInternational@saskhealthauthority.ca",
    "Kate Cowhig Ireland": "info@kcr.ie",
    "CPL Healthcare Ireland": "info@cplhealthcare.com",
    "Holalemania GmbH": "info@holalemania.de",
    "TalentOrange": "info@talentorange.de",
    "Make it in Germany": "info@make-it-in-germany.com",
    "Work in Austria": "info@workinaustria.com",
    "IND Netherlands": "info@ind.nl",
    "Hays Healthcare": "healthcare@hays.com.au",
    "MediCarrera NL": "info@medicarrera.com",
    "Finncare Finland": "info@finncare.fi",
    "WorkInDenmark": "info@workindenmark.dk",
    "Vårdförbundet Sweden": "info@vardforbundet.se",
    "Job Bank Canada": "info@jobbank.gc.ca",
    "Working In Health NZ": "info@workinginhealth.co.nz",
}

EMAIL_URLS = {
    "Health New Zealand": "https://www.healthnz.govt.nz/careers/international",
    "RGH Global": "https://www.rgh-global.com/jobs/midwife-with-sponsorship/",
    "Saskatchewan HA": "https://www.saskhealthauthority.ca/careers-volunteering/careers/internationally-trained-health-care-professionals",
    "Alberta Health Services": "https://www.albertahealthservices.ca/careers/",
    "Holalemania GmbH": "https://www.holalemania.de",
    "TalentOrange": "https://www.talentorange.com",
    "Make it in Germany": "https://www.make-it-in-germany.com",
    "Work in Austria": "https://www.workinaustria.com",
    "IND Netherlands": "https://www.ind.nl",
    "Hays Healthcare": "https://www.hays.com.au",
    "Kate Cowhig Ireland": "https://www.kcr.ie",
    "CPL Healthcare Ireland": "https://www.cplhealthcare.com",
}

COUNTRY_FLAGS = {
    "NZ": "🇳🇿", "AU": "🇦🇺", "CA": "🇨🇦", "DE": "🇩🇪",
    "IE": "🇮🇪", "AT": "🇦🇹", "NL": "🇳🇱", "SE": "🇸🇪",
    "NO": "🇳🇴", "DK": "🇩🇰", "FI": "🇫🇮",
}

APPLICANT_EMOJI = get_all_applicant_labels()

# ═══════════════════════════════════════════════════
# SHEET BUILDERS
# ═══════════════════════════════════════════════════

def build_sheet_01_dashboard(wb, data, opps):
    """داشبورد اصلی با KPI Cards"""
    ws = wb.create_sheet("01 داشبورد")
    rtl(ws)
    
    # Title
    wc(ws, 1, 1, f"داشبورد شکار فرصت — {DATE_STR}",
       font=fa(sz=16, bold=True, color=C_DARK))
    ws.merge_cells("A1:H1")
    
    # KPI Cards
    kpis = [
        ("کل فرصت‌ها", len(opps), C_MED),
        ("تأیید شده", sum(1 for o in opps if "APPLY" in str(o.get("decision",""))), C_GREEN),
        ("نیاز به بررسی", sum(1 for o in opps if "REVIEW" in str(o.get("decision","")) or "NEEDS" in str(o.get("decision",""))), C_YELLOW),
        ("ایمیل ارسال شده", 0, C_RED),
    ]
    
    row = 3
    for i, (label, val, color) in enumerate(kpis):
        col = i * 2 + 1
        wc(ws, row, col, label, font=fa(sz=9, bold=True, color=C_WHITE), bg=color, align=center())
        wc(ws, row + 1, col, val, font=fa(sz=20, bold=True, color=color), align=center())
    
    # Applicant breakdown
    row = 7
    wc(ws, row, 1, "تعداد فرصت‌ها بر اساس متقاضی", font=fa(sz=12, bold=True, color=C_DARK))
    ws.merge_cells(f"A{row}:H{row}")
    row += 1
    # Dynamic applicant counts from config
    applicants = get_applicants()
    for i, a in enumerate(applicants):
        app_id = a["id"].upper()
        app_label = get_applicant_label(a["id"])
        profession = a.get("profession", "")
        cnt = sum(1 for o in opps if app_id in str(o.get("applicant", "")).upper() or a.get("name_fa", "") in str(o.get("applicant", "")))
        col = i * 2 + 1
        wc(ws, row, col, f"{app_label} — {profession}", font=fa(sz=10, bold=True), bg=C_LPURPLE if i == 0 else C_LIGHT, align=center())
        wc(ws, row, col + 1, cnt, font=fa(sz=14, bold=True, color=C_PURPLE if i == 0 else C_MED), align=center())
    
    # Country breakdown
    row += 2
    wc(ws, row, 1, "توزیع فرصت‌ها بر اساس کشور", font=fa(sz=12, bold=True, color=C_DARK))
    ws.merge_cells(f"A{row}:H{row}")
    row += 1
    countries = {}
    for o in opps:
        c = o.get("country", "نامشخص")
        countries[c] = countries.get(c, 0) + 1
    for c, count in sorted(countries.items(), key=lambda x: -x[1]):
        wc(ws, row, 1, c, font=fa(sz=10), align=center())
        wc(ws, row, 2, count, font=fa(sz=10, bold=True), align=center())
        row += 1
    
    # TOP 5
    row += 1
    wc(ws, row, 1, "۵ فرصت برتر", font=fa(sz=12, bold=True, color=C_DARK))
    ws.merge_cells(f"A{row}:H{row}")
    row += 1
    top_headers = ["#", "متقاضی", "کارفرما", "کشور", "شغل", "Evidence", "Final", "تصمیم"]
    for i, h in enumerate(top_headers):
        wc(ws, row, i + 1, h, font=fa(sz=9, bold=True, color=C_WHITE), bg=C_DARK, align=center())
    row += 1
    
    sorted_opps = sorted(opps, key=lambda x: x.get("final_score", 0) or x.get("evidence_score", 0), reverse=True)
    for idx, o in enumerate(sorted_opps[:5], 1):
        applicant_label = APPLICANT_EMOJI.get(str(o.get("applicant","")).upper(), o.get("applicant",""))
        country = o.get("country", "")
        for code, flag in COUNTRY_FLAGS.items():
            if code in str(country):
                country = f"{flag} {country}"
                break
        
        wc(ws, row, 1, idx, align=center())
        wc(ws, row, 2, applicant_label, align=center())
        wc(ws, row, 3, o.get("employer", ""), font=fa(sz=9))
        wc(ws, row, 4, country, align=center())
        wc(ws, row, 5, o.get("job", ""), font=fa(sz=9))
        wc(ws, row, 6, o.get("evidence_score", ""), align=center())
        wc(ws, row, 7, o.get("final_score", "") or "—", align=center())
        decision = o.get("decision", "")
        bg = C_LGREEN if "APPLY" in decision else C_LYELLOW if "REVIEW" in decision or "NEEDS" in decision else C_LRED
        wc(ws, row, 8, decision, font=fa(sz=9, bold=True), bg=bg, align=center())
        row += 1
    
    # Widths
    widths = [6, 14, 25, 14, 25, 10, 10, 25]
    for i, w in enumerate(widths):
        auto_width(ws, i + 1, w)
    
    freeze(ws, "A2")

def build_sheet_02_opportunities(wb, opps):
    """فرصت‌های کاری — جدول کامل"""
    ws = wb.create_sheet("02 فرصت‌ها")
    rtl(ws)
    
    headers = ["ID", "متقاضی", "کشور", "کارفرما", "شغل", "لینک آگهی",
               "ایمیل", "Evidence", "Final", "حمایت", "ویزا", "زبان",
               "ثبت‌نام", "وضعیت", "تصمیم", "تاریخ"]
    
    for i, h in enumerate(headers):
        wc(ws, 1, i + 1, h, font=fa(sz=9, bold=True, color=C_WHITE), bg=C_DARK, align=center())
    
    for idx, o in enumerate(opps, 2):
        applicant_label = APPLICANT_EMOJI.get(str(o.get("applicant","")).upper(), o.get("applicant",""))
        country = o.get("country", "")
        for code, flag in COUNTRY_FLAGS.items():
            if code in str(country):
                country = f"{flag} {country}"
                break
        
        email = o.get("email", "")
        if not email:
            for emp_name, emp_email in VERIFIED_EMAILS.items():
                if emp_name.lower() in o.get("employer", "").lower():
                    email = emp_email
                    break
        
        url = o.get("url", "")
        if not url:
            for emp_name, emp_url in EMAIL_URLS.items():
                if emp_name.lower() in o.get("employer", "").lower():
                    url = emp_url
                    break
        
        vals = [
            o.get("id", ""),
            applicant_label,
            country,
            o.get("employer", ""),
            o.get("job", ""),
            url,
            email,
            o.get("evidence_score", ""),
            o.get("final_score", "") or "—",
            o.get("sponsorship", ""),
            o.get("visa", ""),
            o.get("language", ""),
            o.get("registration", ""),
            o.get("status", "NEW"),
            o.get("decision", ""),
            DATE_STR,
        ]
        
        for ci, v in enumerate(vals):
            font = fa(sz=9)
            if ci == 5 and v:  # URL
                font = en(sz=9, color="0563C1", bold=True)
            elif ci == 6 and v:  # Email
                font = en(sz=9, color="0563C1")
            
            bg = None
            if ci == 14:  # Decision column
                if "APPLY" in str(v): bg = C_LGREEN
                elif "REVIEW" in str(v) or "NEEDS" in str(v): bg = C_LYELLOW
                elif "REJECT" in str(v): bg = C_LRED
            
            wc(ws, idx, ci + 1, v, font=font, bg=bg)
    
    # Auto widths
    widths = [14, 12, 14, 22, 22, 35, 30, 10, 10, 10, 12, 12, 12, 12, 22, 18]
    for i, w in enumerate(widths):
        auto_width(ws, i + 1, w)
    
    freeze(ws, "A2")
    # Auto filter
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(opps)+1}"

def build_sheet_applicant(wb, opps, app_info, sheet_num):
    """شیت فرصت‌های هر متقاضی — داینامیک از config"""
    app_id = app_info["id"].upper()
    name_fa = app_info.get("name_fa", app_id)
    profession = app_info.get("profession", "")
    emoji = app_info.get("emoji", "?")
    color = C_PURPLE if app_info.get("gender") == "female" else C_MED
    
    sheet_name = f"{sheet_num:02d} {name_fa} — {profession}"
    ws = wb.create_sheet(sheet_name)
    rtl(ws)
    
    app_opps = [o for o in opps if app_id in str(o.get("applicant","")).upper() or name_fa in str(o.get("applicant",""))]
    
    wc(ws, 1, 1, f"{emoji} فرصت‌های {profession} {name_fa} — {len(app_opps)} فرصت", font=fa(sz=14, bold=True, color=color))
    ws.merge_cells("A1:L1")
    
    headers = ["#", "کارفرما", "کشور", "شغل", "Evidence", "Final", "ایمیل", "حمایت", "ویزا", "زبان", "ثبت‌نام", "تصمیم"]
    for i, h in enumerate(headers):
        wc(ws, 3, i + 1, h, font=fa(sz=9, bold=True, color=C_WHITE), bg=color, align=center())
    
    for idx, o in enumerate(app_opps, 1):
        country = o.get("country", "")
        for code, flag in COUNTRY_FLAGS.items():
            if code in str(country): country = f"{flag} {country}"; break
        
        email = o.get("email", "")
        if not email:
            for emp_name, emp_email in VERIFIED_EMAILS.items():
                if emp_name.lower() in o.get("employer", "").lower():
                    email = emp_email; break
        
        vals = [idx, o.get("employer",""), country, o.get("job",""),
                o.get("evidence_score",""), o.get("final_score","") or "—",
                email, o.get("sponsorship",""), o.get("visa",""),
                o.get("language",""), o.get("registration",""), o.get("decision","")]
        
        for ci, v in enumerate(vals):
            bg = None
            if ci == 11:
                if "APPLY" in str(v): bg = C_LGREEN
                elif "REVIEW" in str(v) or "NEEDS" in str(v): bg = C_LYELLOW
            wc(ws, idx + 3, ci + 1, v, font=fa(sz=9), bg=bg)
    
    widths = [5, 22, 16, 22, 10, 10, 32, 12, 14, 12, 14, 22]
    for i, w in enumerate(widths): auto_width(ws, i + 1, w)
    freeze(ws, "A4")

def build_sheet_05_employers(wb, data):
    """بانک کارفرمایان"""
    ws = wb.create_sheet("05 کارفرمایان")
    rtl(ws)
    
    wc(ws, 1, 1, "بانک کارفرمایان — Employer Bank", font=fa(sz=14, bold=True, color=C_DARK))
    ws.merge_cells("A1:H1")
    
    headers = ["#", "نام", "کشور", "نوع", "حمایت", "امتیاز", "ایمیل", "آخرین بررسی"]
    for i, h in enumerate(headers):
        wc(ws, 3, i + 1, h, font=fa(sz=9, bold=True, color=C_WHITE), bg=C_DARK, align=center())
    
    for idx, e in enumerate(data.get("employers", []), 1):
        email = VERIFIED_EMAILS.get(e.get("نام", ""), "—")
        vals = [idx, e.get("نام",""), e.get("کشور",""), e.get("نوع",""),
                e.get("حمایت",""), e.get("امتیاز",""), email, DATE_STR]
        for ci, v in enumerate(vals):
            wc(ws, idx + 3, ci + 1, v, font=fa(sz=9))
    
    auto_width(ws, 1, 5)
    auto_width(ws, 2, 25)
    auto_width(ws, 3, 10)
    auto_width(ws, 4, 14)
    auto_width(ws, 5, 10)
    auto_width(ws, 6, 10)
    auto_width(ws, 7, 35)
    auto_width(ws, 8, 18)
    freeze(ws, "A4")

def build_sheet_06_emails(wb, opps):
    """ایمیل‌های آماده"""
    ws = wb.create_sheet("06 ایمیل‌ها")
    rtl(ws)
    
    wc(ws, 1, 1, "ایمیل‌های معتبر و آماده ارسال", font=fa(sz=14, bold=True, color=C_GREEN))
    ws.merge_cells("A1:F1")
    
    headers = ["#", "متقاضی", "کارفرما", "ایمیل", "لینک آگهی", "وضعیت"]
    for i, h in enumerate(headers):
        wc(ws, 3, i + 1, h, font=fa(sz=9, bold=True, color=C_WHITE), bg=C_GREEN, align=center())
    
    row = 4
    idx = 1
    for o in opps:
        email = o.get("email", "")
        if not email:
            for emp_name, emp_email in VERIFIED_EMAILS.items():
                if emp_name.lower() in o.get("employer", "").lower():
                    email = emp_email; break
        
        url = o.get("url", "")
        if not url:
            for emp_name, emp_url in EMAIL_URLS.items():
                if emp_name.lower() in o.get("employer", "").lower():
                    url = emp_url; break
        
        if email and email != "—":
            applicant_label = APPLICANT_EMOJI.get(str(o.get("applicant","")).upper(), o.get("applicant",""))
            wc(ws, row, 1, idx, align=center())
            wc(ws, row, 2, applicant_label, align=center())
            wc(ws, row, 3, o.get("employer",""), font=fa(sz=9))
            wc(ws, row, 4, email, font=en(sz=9, color="0563C1"))
            wc(ws, row, 5, url, font=en(sz=9, color="0563C1"))
            wc(ws, row, 6, "آماده ارسال", font=fa(sz=9, color=C_GREEN), bg=C_LGREEN, align=center())
            row += 1
            idx += 1
    
    widths = [5, 14, 25, 35, 40, 16]
    for i, w in enumerate(widths): auto_width(ws, i + 1, w)
    freeze(ws, "A4")

def build_sheet_07_applications(wb, data):
    """Application Pipeline"""
    ws = wb.create_sheet("07 درخواست‌ها")
    rtl(ws)
    
    wc(ws, 1, 1, "Application Pipeline — فرآیند درخواست", font=fa(sz=14, bold=True, color=C_DARK))
    ws.merge_cells("A1:K1")
    
    # Pipeline stages
    stages = ["DISCOVERED", "VERIFIED", "SHORTLISTED", "READY", "SENT",
              "FOLLOW_UP", "RESPONSE", "INTERVIEW", "OFFER", "REJECTED", "CLOSED"]
    
    wc(ws, 2, 1, "مراحل:", font=fa(sz=10, bold=True))
    for i, s in enumerate(stages):
        bg = C_LGREEN if s in ["SENT", "RESPONSE", "INTERVIEW", "OFFER"] else \
             C_LYELLOW if s in ["VERIFIED", "SHORTLISTED", "READY"] else C_LGRAY
        wc(ws, 2, i + 2, s, font=fa(sz=8, bold=True), bg=bg, align=center())
    
    headers = ["ID", "متقاضی", "کارفرما", "شغل", "کشور", "ایمیل",
               "تاریخ ارسال", "وضعیت", "پاسخ", "مصاحبه", "یادداشت"]
    row = 4
    for i, h in enumerate(headers):
        wc(ws, row, i + 1, h, font=fa(sz=9, bold=True, color=C_WHITE), bg=C_DARK, align=center())
    
    row = 5
    for app_table in data.get("applications", []):
        for r in app_table.get("rows", []):
            vals = [
                r.get("ID", ""),
                r.get("Applicant", r.get("متقاضی", "")),
                r.get("Employer", r.get("کارفرما", "")),
                r.get("Job", r.get("شغل", "")),
                r.get("Country", r.get("کشور", "")),
                r.get("Email", r.get("ایمیل", "")),
                r.get("Applied At", r.get("تاریخ", "")),
                r.get("Status", r.get("وضعیت", "DISCOVERED")),
                r.get("Response", r.get("پاسخ", "")),
                r.get("Interview", r.get("مصاحبه", "")),
                r.get("Next Action", r.get("یادداشت", "")),
            ]
            for ci, v in enumerate(vals):
                bg = None
                status = str(vals[7]).upper()
                if "SENT" in status: bg = C_LGREEN
                elif "APPROVED" in status or "READY" in status: bg = C_LYELLOW
                elif "INTERVIEW" in status: bg = C_LPURPLE
                elif "OFFER" in status: bg = C_LGREEN
                elif "REJECT" in status: bg = C_LRED
                wc(ws, row, ci + 1, v, font=fa(sz=9), bg=bg)
            row += 1
    
    # Empty rows for manual entry
    for r in range(row, row + 10):
        wc(ws, r, 1, "", bg=C_LGRAY)
        for ci in range(2, 12):
            wc(ws, r, ci, "", bg=C_LGRAY)
    
    widths = [14, 14, 22, 22, 14, 32, 16, 16, 12, 12, 20]
    for i, w in enumerate(widths): auto_width(ws, i + 1, w)
    freeze(ws, "A5")

def build_sheet_08_followup(wb, opps):
    """Follow-up Tracker"""
    ws = wb.create_sheet("08 پیگیری")
    rtl(ws)
    
    wc(ws, 1, 1, "پیگیری درخواست‌ها — Follow-up Tracker", font=fa(sz=14, bold=True, color=C_YELLOW))
    ws.merge_cells("A1:I1")
    
    headers = ["#", "متقاضی", "کارفرما", "ایمیل", "تاریخ ارسال",
               "روزهای سپری‌شده", "یادآوری پیگیری", "وضعیت پاسخ", "اقدام بعدی"]
    for i, h in enumerate(headers):
        wc(ws, 3, i + 1, h, font=fa(sz=9, bold=True, color=C_WHITE), bg=C_YELLOW, align=center())
    
    # Default follow-up timeline
    row = 4
    for idx, o in enumerate(opps, 1):
        email = o.get("email", "")
        if not email:
            for emp_name, emp_email in VERIFIED_EMAILS.items():
                if emp_name.lower() in o.get("employer", "").lower():
                    email = emp_email; break
        
        applicant_label = APPLICANT_EMOJI.get(str(o.get("applicant","")).upper(), o.get("applicant",""))
        followup_date = (NOW + timedelta(days=7)).strftime("%Y-%m-%d")
        
        vals = [idx, applicant_label, o.get("employer",""), email,
                "منتظر ارسال", "—", followup_date, "بدون پاسخ", "ارسال اولیه"]
        for ci, v in enumerate(vals):
            wc(ws, row, ci + 1, v, font=fa(sz=9))
        row += 1
    
    # Empty rows for manual entry
    for r in range(row, row + 15):
        for ci in range(1, 10):
            wc(ws, r, ci, "", bg=C_LGRAY)
    
    widths = [5, 14, 22, 32, 16, 16, 16, 16, 18]
    for i, w in enumerate(widths): auto_width(ws, i + 1, w)
    freeze(ws, "A4")

def build_sheet_09_visa(wb, data):
    """اطلاعات ویزا"""
    ws = wb.create_sheet("09 ویزا")
    rtl(ws)
    
    wc(ws, 1, 1, "اطلاعات ویزا — Visa Bank", font=fa(sz=14, bold=True, color=C_DARK))
    ws.merge_cells("A1:H1")
    
    headers = ["کشور", "نوع ویزا", "شرط شغل", "شرط حقوق", "زبان ویزا", "خانواده", "ثبت‌نام", "منبع رسمی"]
    for i, h in enumerate(headers):
        wc(ws, 3, i + 1, h, font=fa(sz=9, bold=True, color=C_WHITE), bg=C_DARK, align=center())
    
    for idx, v in enumerate(data.get("visa", []), 4):
        vals = [v.get("Country",""), v.get("Visa Type",""), v.get("Job Requirement",""),
                v.get("Salary Threshold",""), v.get("Language (Visa)","") or v.get("Language",""),
                v.get("Family Rules",""), v.get("Registration",""), v.get("Official Source","")]
        for ci, val in enumerate(vals):
            wc(ws, idx, ci + 1, val, font=fa(sz=9))
    
    widths = [14, 22, 22, 18, 18, 18, 14, 35]
    for i, w in enumerate(widths): auto_width(ws, i + 1, w)
    freeze(ws, "A4")

def build_sheet_10_registration(wb, data):
    """ثبت‌نام حرفه‌ای"""
    ws = wb.create_sheet("10 ثبت‌نام")
    rtl(ws)
    
    wc(ws, 1, 1, "ثبت‌نام حرفه‌ای — Registration", font=fa(sz=14, bold=True, color=C_PURPLE))
    ws.merge_cells("A1:I1")
    
    headers = ["کشور", "متقاضی", "حرفه", "نهاد ثبت", "مسیر ثبت", "زبان مورد نیاز", "آزمون", "هزینه", "منبع"]
    for i, h in enumerate(headers):
        wc(ws, 3, i + 1, h, font=fa(sz=9, bold=True, color=C_WHITE), bg=C_PURPLE, align=center())
    
    row = 4
    for r in data.get("registration", []):
        vals = [r.get("Country",""), r.get("Applicant", r.get("Profession","")),
                r.get("Profession",""), r.get("Regulator",""), r.get("Registration Route",""),
                r.get("English Requirement", r.get("Required Score","")),
                r.get("Exam",""), r.get("Fees",""), r.get("Official URL", r.get("Official Source",""))]
        for ci, v in enumerate(vals):
            wc(ws, row, ci + 1, v, font=fa(sz=9))
        row += 1
    
    widths = [14, 14, 14, 22, 22, 20, 12, 12, 35]
    for i, w in enumerate(widths): auto_width(ws, i + 1, w)
    freeze(ws, "A4")

def build_sheet_11_evidence(wb, data):
    """Evidence Matrix"""
    ws = wb.create_sheet("11 Evidence")
    rtl(ws)
    
    wc(ws, 1, 1, "ارزیابی Evidence — Evidence Matrix", font=fa(sz=14, bold=True, color=C_DARK))
    ws.merge_cells("A1:G1")
    
    headers = ["ID", "متقاضی", "کارفرما", "Evidence Score", "Final Score", "Decision", "تاریخ بررسی"]
    for i, h in enumerate(headers):
        wc(ws, 3, i + 1, h, font=fa(sz=9, bold=True, color=C_WHITE), bg=C_DARK, align=center())
    
    row = 4
    for key, content in data.get("evidence_sections", {}).items():
        if not key.startswith("MH-2026-"):
            continue
        
        parts = key.split("—")
        opp_id = parts[0].strip() if parts else key
        applicant = parts[1].strip() if len(parts) > 1 else ""
        employer = parts[2].strip() if len(parts) > 2 else ""
        
        evidence_score = ""
        final_score = ""
        decision = ""
        for line in content.split("\n"):
            if "Evidence Score:" in line:
                m = re.search(r'(\d+)/100', line)
                if m: evidence_score = int(m.group(1))
            if "Final Score:" in line:
                m = re.search(r'(\d+)/100', line)
                if m: final_score = int(m.group(1))
            if "Decision:" in line:
                decision = line.split("Decision:")[1].strip().split("\n")[0].strip()
        
        vals = [opp_id, applicant, employer, evidence_score,
                final_score or "—", decision, DATE_STR]
        
        for ci, v in enumerate(vals):
            bg = None
            if ci == 5:
                if "APPLY" in str(v): bg = C_LGREEN
                elif "REVIEW" in str(v) or "NEEDS" in str(v): bg = C_LYELLOW
                elif "REJECT" in str(v): bg = C_LRED
            wc(ws, row, ci + 1, v, font=fa(sz=9), bg=bg)
        row += 1
    
    widths = [14, 14, 25, 14, 14, 28, 18]
    for i, w in enumerate(widths): auto_width(ws, i + 1, w)
    freeze(ws, "A4")

def build_sheet_12_history(wb, data):
    """تاریخچه جستجو"""
    ws = wb.create_sheet("12 تاریخچه")
    rtl(ws)
    
    wc(ws, 1, 1, "تاریخچه جستجو — Search History", font=fa(sz=14, bold=True, color=C_GRAY))
    ws.merge_cells("A1:G1")
    
    headers = ["تاریخ", "متقاضی", "کشورها", "منابع", "مشاغل یافت‌شده", "مشاغل معتبر", "توضیحات"]
    for i, h in enumerate(headers):
        wc(ws, 3, i + 1, h, font=fa(sz=9, bold=True, color=C_WHITE), bg=C_GRAY, align=center())
    
    for idx, s in enumerate(data.get("search_history", []), 4):
        vals = [s.get("Date",""), s.get("Applicant",""), s.get("Countries",""),
                s.get("Sources Searched",""), s.get("Jobs Found",""),
                s.get("Valid Jobs",""), s.get("New Sources","")]
        for ci, v in enumerate(vals):
            wc(ws, idx, ci + 1, v, font=fa(sz=9))
    
    widths = [14, 14, 18, 22, 16, 16, 30]
    for i, w in enumerate(widths): auto_width(ws, i + 1, w)
    freeze(ws, "A4")

# ═══════════════════════════════════════════════════
# ARCHIVE OLD EXCELS
# ═══════════════════════════════════════════════════
def archive_old():
    os.makedirs(ARCH, exist_ok=True)
    archived = 0
    for f in glob.glob(os.path.join(DASH, "*.xlsx")):
        basename = os.path.basename(f)
        if "build_dashboard" not in basename:
            dest = os.path.join(ARCH, basename)
            if not os.path.exists(dest):
                try:
                    os.replace(f, dest)
                    archived += 1
                    print(f"  📦 آرشیو: {basename}")
                except:
                    pass
    # Also clean nested dashboard/dashboard/
    nested = os.path.join(DASH, "dashboard")
    if os.path.isdir(nested):
        import shutil
        shutil.rmtree(nested, ignore_errors=True)
        print("  🗑️ حذف dashboard/dashboard/ تکراری")
    # Clean lock files
    for f in glob.glob(os.path.join(DASH, "~$*.xlsx")):
        try:
            os.remove(f)
            print(f"  🗑️ حذف lock file: {os.path.basename(f)}")
        except:
            pass
    print(f"📦 {archived} فایل آرشیو شد")

# ═══════════════════════════════════════════════════
# MAIN
# ═══════════════════════════════════════════════════
def build_sheet_13_email_analysis(wb, data):
    """تحلیل ایمیل شغلی با آمار درصدی"""
    ws = wb.create_sheet("13 تحلیل ایمیل")
    rtl(ws)
    
    # Load email analysis
    import json
    email_file = os.path.join(MEM, "EMAIL_ANALYSIS.json")
    if not os.path.exists(email_file):
        wc(ws, 1, 1, "تحلیل ایمیل — داده‌ای موجود نیست", font=fa(sz=14, bold=True, color=C_GRAY))
        return
    
    with open(email_file, "r", encoding="utf-8") as f:
        email_data = json.load(f)
    
    emails = email_data.get("emails", [])
    total = email_data.get("total_emails", 0)
    job_related = email_data.get("job_related", 0)
    
    # Title
    wc(ws, 1, 1, f"تحلیل ایمیل شغلی — {job_related} ایمیل از {total} کل", font=fa(sz=14, bold=True, color=C_DARK))
    ws.merge_cells("A1:H1")
    
    # KPI
    wc(ws, 3, 1, "کل", font=fa(sz=8, bold=True, color=C_WHITE), bg=C_DARK, align=center())
    wc(ws, 3, 2, total, font=fa(sz=14, bold=True), align=center())
    pct = round(job_related / total * 100) if total else 0
    wc(ws, 3, 3, "مرتبط با کار", font=fa(sz=8, bold=True, color=C_WHITE), bg=C_GREEN, align=center())
    wc(ws, 3, 4, f"{job_related} ({pct}%)", font=fa(sz=14, bold=True, color=C_GREEN), align=center())
    
    # Category breakdown
    by_cat = {}
    for e in emails:
        cat = e.get("category", "unknown")
        by_cat[cat] = by_cat.get(cat, 0) + 1
    
    row = 5
    wc(ws, row, 1, "دسته", font=fa(sz=9, bold=True, color=C_WHITE), bg=C_DARK, align=center())
    wc(ws, row, 2, "تعداد", font=fa(sz=9, bold=True, color=C_WHITE), bg=C_DARK, align=center())
    wc(ws, row, 3, "درصد", font=fa(sz=9, bold=True, color=C_WHITE), bg=C_DARK, align=center())
    row += 1
    
    cat_info = {
        "interview": ("🗣️ مصاحبه", C_GREEN, C_LGREEN),
        "offer": ("🎉 پیشنهاد", C_PURPLE, C_LPURPLE),
        "rejection": ("❌ رد شده", C_RED, C_LRED),
        "follow_up": ("⏰ پیگیری", C_YELLOW, C_LYELLOW),
        "inquiry": ("💬 استعلام", C_MED, C_LIGHT),
    }
    
    for cat_key in ["interview", "offer", "rejection", "follow_up", "inquiry"]:
        label, color, bg = cat_info.get(cat_key, (cat_key, C_GRAY, C_LGRAY))
        count = by_cat.get(cat_key, 0)
        p = round(count / job_related * 100) if job_related else 0
        wc(ws, row, 1, label, font=fa(sz=9, bold=True), bg=bg, align=center())
        wc(ws, row, 2, count, font=fa(sz=10, bold=True), align=center())
        wc(ws, row, 3, f"{p}%", font=fa(sz=10, bold=True, color=color), align=center())
        row += 1
    
    # By applicant
    row += 1
    wc(ws, row, 1, "متقاضی", font=fa(sz=9, bold=True, color=C_WHITE), bg=C_DARK, align=center())
    wc(ws, row, 2, "تعداد", font=fa(sz=9, bold=True, color=C_WHITE), bg=C_DARK, align=center())
    wc(ws, row, 3, "درصد", font=fa(sz=9, bold=True, color=C_WHITE), bg=C_DARK, align=center())
    row += 1
    
    by_app = {"TOHID": 0, "NEDA": 0}
    for e in emails:
        app = e.get("applicant", "UNKNOWN")
        if app in by_app: by_app[app] += 1
    
    for a in get_applicants():
        app_key = a["id"].upper()
        app_label = get_applicant_label(a["id"])
        count = by_app.get(app_key, 0)
        p = round(count / job_related * 100) if job_related else 0
        wc(ws, row, 1, app_label, font=fa(sz=9, bold=True), align=center())
        wc(ws, row, 2, count, font=fa(sz=10, bold=True), align=center())
        wc(ws, row, 3, f"{p}%", font=fa(sz=10, bold=True, color=C_MED), align=center())
        row += 1
    
    # Top senders
    row += 1
    wc(ws, row, 1, "فرستنده‌ها", font=fa(sz=9, bold=True, color=C_WHITE), bg=C_DARK, align=center())
    wc(ws, row, 2, "تعداد", font=fa(sz=9, bold=True, color=C_WHITE), bg=C_DARK, align=center())
    wc(ws, row, 3, "درصد", font=fa(sz=9, bold=True, color=C_WHITE), bg=C_DARK, align=center())
    row += 1
    
    senders = {}
    for e in emails:
        sender = e.get("from", "").split("<")[0].strip().strip('"')[:40]
        if sender: senders[sender] = senders.get(sender, 0) + 1
    
    for sender, count in sorted(senders.items(), key=lambda x: -x[1])[:8]:
        p = round(count / job_related * 100) if job_related else 0
        wc(ws, row, 1, sender, font=fa(sz=8))
        wc(ws, row, 2, count, font=fa(sz=9, bold=True), align=center())
        wc(ws, row, 3, f"{p}%", font=fa(sz=9, color=C_MED), align=center())
        row += 1
    
    # Recent emails (last 10)
    row += 1
    wc(ws, row, 1, "آخرین ایمیل‌ها", font=fa(sz=9, bold=True, color=C_WHITE), bg=C_DARK, align=center())
    ws.merge_cells(f"A{row}:E{row}")
    row += 1
    
    for e in sorted(emails, key=lambda x: x.get("date", ""), reverse=True)[:10]:
        cat = e.get("category", "unknown")
        label, color, bg = cat_info.get(cat, (cat_key, C_GRAY, C_LGRAY))
        wc(ws, row, 1, e.get("date", "")[:10], font=fa(sz=8), align=center())
        wc(ws, row, 2, e.get("from", "")[:35], font=fa(sz=8))
        wc(ws, row, 3, e.get("subject", "")[:50], font=fa(sz=8))
        wc(ws, row, 4, label, font=fa(sz=8, bold=True), bg=bg, align=center())
        row += 1
    
    # Widths
    for i, w in enumerate([16, 35, 50, 14, 10, 10, 10, 10]):
        auto_width(ws, i + 1, w)
    freeze(ws, "A2")


def main():
    print("=" * 60)
    print("MigrationHunter — Build Dashboard")
    print(f"📅 {DATE_STR}")
    print("=" * 60)
    
    # Step 1: Archive old files
    print("\n📦 Phase 1: آرشیو فایل‌های قدیمی...")
    archive_old()
    
    # Step 2: Load data from memory
    print("\n📂 Phase 2: خواندن حافظه‌ها...")
    data = load_all_data()
    print(f"  ✅ Sources: {len(data.get('sources', []))}")
    print(f"  ✅ Employers: {len(data.get('employers', []))}")
    print(f"  ✅ Jobs: {len(data.get('jobs', []))}")
    print(f"  ✅ Evidence sections: {len(data.get('evidence_sections', {}))}")
    print(f"  ✅ Visa entries: {len(data.get('visa', []))}")
    print(f"  ✅ Registration entries: {len(data.get('registration', []))}")
    
    # Step 3: Build opportunities
    print("\n🎯 Phase 3: ساخت لیست فرصت‌ها...")
    opps = build_opportunities(data)
    print(f"  ✅ {len(opps)} فرصت شناسایی شد")
    
    # Step 4: Create workbook
    print("\n📊 Phase 4: ساخت Excel...")
    wb = Workbook()
    # Remove default sheet
    wb.remove(wb.active)
    
    # Build all 12 sheets
    print("  📝 Sheet 01: داشبورد...")
    build_sheet_01_dashboard(wb, data, opps)
    
    print("  📝 Sheet 02: فرصت‌ها...")
    build_sheet_02_opportunities(wb, opps)
    
    # Dynamic applicant sheets from config
    for i, app_info in enumerate(get_applicants()):
        sheet_num = 3 + i
        print(f"  📝 Sheet {sheet_num:02d}: {app_info.get('name_fa', app_info['id'])}...")
        build_sheet_applicant(wb, opps, app_info, sheet_num)
    
    print("  📝 Sheet 05: کارفرمایان...")
    build_sheet_05_employers(wb, data)
    
    print("  📝 Sheet 06: ایمیل‌ها...")
    build_sheet_06_emails(wb, opps)
    
    print("  📝 Sheet 07: درخواست‌ها...")
    build_sheet_07_applications(wb, data)
    
    print("  📝 Sheet 08: پیگیری...")
    build_sheet_08_followup(wb, opps)
    
    print("  📝 Sheet 09: ویزا...")
    build_sheet_09_visa(wb, data)
    
    print("  📝 Sheet 10: ثبت‌نام...")
    build_sheet_10_registration(wb, data)
    
    print("  📝 Sheet 11: Evidence...")
    build_sheet_11_evidence(wb, data)
    
    print("  📝 Sheet 12: تاریخچه...")
    build_sheet_12_history(wb, data)
    
    print("  📝 Sheet 13: تحلیل ایمیل...")
    build_sheet_13_email_analysis(wb, data)
    
    # Step 5: Save
    os.makedirs(DASH, exist_ok=True)
    filename = f"MigrationHunter_Dashboard_{FILE_DATE}.xlsx"
    filepath = os.path.join(DASH, filename)
    wb.save(filepath)
    print(f"\n✅ ذخیره شد: {filepath}")
    
    # Summary
    print("\n" + "=" * 60)
    print("📊 خلاصه")
    print("=" * 60)
    print(f"  فایل: {filename}")
    print(f"  شیت‌ها: 13")
    print(f"  فرصت‌ها: {len(opps)}")
    for a in get_applicants():
        cnt = sum(1 for o in opps if a['id'].upper() in str(o.get('applicant','')).upper() or a.get('name_fa','') in str(o.get('applicant','')))
        print(f"  {a.get('emoji','?')} {a.get('name_fa', a['id'])}: {cnt}")
    print(f"  RTL: ✅ همه شیت‌ها")
    print(f"  فونت: B Mitra + Times New Roman")
    print(f"  Freeze: ✅")
    print(f"  Auto Filter: ✅")
    print(f"  Application Pipeline: ✅")
    print(f"  Follow-up Tracker: ✅")
    print("=" * 60)

if __name__ == "__main__":
    main()
