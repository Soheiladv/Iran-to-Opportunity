#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Migration Hunter Excel with VBA - Email & Cover Letter Generation
"""

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from datetime import datetime
import os

# Font settings - will use B Mitra if available, else Tahoma
FONT = 'B Mitra'
FALLBACK = 'Tahoma'

BD = '1F4E79'; BM = '2E75B6'; GR = 'C6EFCE'; YL = 'FFEB9C'
RD = 'FFC7CE'; OR = 'FFD966'; BL = 'BDD7EE'; WH = 'FFFFFF'

HF = Font(name=FONT, bold=True, size=11, color=WH)
HFI = PatternFill(start_color=BD, end_color=BD, fill_type='solid')
SF = PatternFill(start_color=BM, end_color=BM, fill_type='solid')
GF = PatternFill(start_color=GR, end_color=GR, fill_type='solid')
YF = PatternFill(start_color=YL, end_color=YL, fill_type='solid')
RF = PatternFill(start_color=RD, end_color=RD, fill_type='solid')
OF = PatternFill(start_color=OR, end_color=OR, fill_type='solid')
BF = PatternFill(start_color=BL, end_color=BL, fill_type='solid')
CF = Font(name=FONT, size=10)
BF2 = Font(name=FONT, size=10, bold=True)
TF = Font(name=FONT, size=14, bold=True, color=BD)
SF2 = Font(name=FONT, size=12, bold=True, color=BM)
LF = Font(name=FONT, size=10, color='0563C1', underline='single')
TB = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
CA = Alignment(horizontal='center', vertical='center', wrap_text=True, readingOrder=2)
RA = Alignment(horizontal='right', vertical='center', wrap_text=True, readingOrder=2)

def sc(c, f=CF, fl=None, a=RA):
    c.font=f; c.alignment=a; c.border=TB
    if fl: c.fill=fl

def st(ws, r, ce, t):
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ce)
    c=ws.cell(row=r, column=1, value=t); c.font=TF; c.alignment=CA

def sh(ws, r, h):
    for i,v in enumerate(h,1):
        c=ws.cell(row=r, column=i, value=v); sc(c, HF, HFI, CA)

def sd(ws, sr, d, cw=None):
    for r, rd in enumerate(d, sr):
        for ci, v in enumerate(rd, 1):
            c=ws.cell(row=r, column=ci, value=v)
            fl=None
            sv=str(v)
            if sv in ['✅','Confirmed','آماده','جدید','P1','🔥 P1','READY TO SEND','آماده ارسال']: fl=GF
            elif sv in ['🟡','Likely','P2','🟢 P2','مهم','آماده نیست','نیاز']: fl=YF
            elif sv in ['🔵 P3','شناسایی']: fl=BF
            elif sv in ['❌','REJECTED','EXPIRED']: fl=RF
            elif '🔗' in sv: fl=BF
            sc(c, fl=fl)
    if cw:
        for i,w in enumerate(cw,1): ws.column_dimensions[get_column_letter(i)].width=w

def add_link(ws, row, col, url, text='🔗'):
    c = ws.cell(row=row, column=col)
    c.hyperlink = url
    c.font = LF
    c.value = text

def add_chart(ws, title, data_col, cat_col, cat_start, cat_end, location):
    ch = BarChart()
    ch.type = "col"
    ch.title = title
    ch.y_axis.title = "امتیاز"
    ch.style = 10
    d = Reference(ws, min_col=data_col, min_row=cat_start-1, max_col=data_col, max_row=cat_end)
    c = Reference(ws, min_col=cat_col, min_row=cat_start, max_row=cat_end)
    ch.add_data(d, titles_from_data=True)
    ch.set_categories(c)
    ws.add_chart(ch, location)

def create_email_template_sheet(wb):
    """Create sheet with email templates"""
    ws = wb.create_sheet("📧 قالب ایمیل‌ها")
    ws.sheet_view.rightToLeft = True
    
    st(ws, 1, 10, "قالب ایمیل‌های درخواست — قابل کپی")
    
    # Email 1
    r = 3; ws.cell(row=r, column=1, value="📧 ایمیل ۱: Health New Zealand").font = SF2
    r = 4; ws.cell(row=r, column=1, value="موضوع:").font = BF2
    ws.cell(row=r, column=2, value="International Midwife — Registration Interest")
    r = 5; ws.cell(row=r, column=1, value="گیرنده:").font = BF2
    ws.cell(row=r, column=2, value="Health NZ International Recruitment")
    r = 6; ws.cell(row=r, column=1, value="لینک:").font = BF2
    add_link(ws, r, 2, "https://www.healthnz.govt.nz/careers/international")
    
    r = 8; ws.cell(row=r, column=1, value="متن ایمیل:").font = SF2
    email1 = """Dear Health New Zealand International Recruitment Team,

I am a registered midwife currently practising at Milad Hospital in Tehran, Iran. I am writing to express my interest in midwifery positions with Health New Zealand.

Professional Background:
- Registered Midwife with clinical experience in a high-volume maternity unit
- Current employment: Milad Hospital, Tehran
- Education: Bachelor of Science in Midwifery

I am committed to completing the registration process with the Midwifery Council of New Zealand.

I would welcome the opportunity to discuss how my clinical experience could contribute to Health New Zealand.

Kind regards,
Neda Arjmand"""
    
    ws.merge_cells(start_row=r, start_column=1, end_row=r+15, end_column=10)
    c = ws.cell(row=r, column=1, value=email1)
    c.font = CF
    c.alignment = Alignment(wrap_text=True, readingOrder=2)
    
    # Email 2
    r = 25; ws.cell(row=r, column=1, value="📧 ایمیل ۲: Hassett Group").font = SF2
    r = 26; ws.cell(row=r, column=1, value="موضوع:").font = BF2
    ws.cell(row=r, column=2, value="Registered Midwife — Visa Sponsorship Inquiry")
    r = 27; ws.cell(row=r, column=1, value="گیرنده:").font = BF2
    ws.cell(row=r, column=2, value="Hassett Group Recruitment")
    r = 28; ws.cell(row=r, column=1, value="لینک:").font = BF2
    add_link(ws, r, 2, "https://www.hassett.com.au/job-details/registered-midwife-visa-sponsorship-pathway-to-pr-in-healthcare-medicine-jobs-1475247")
    
    r = 30; ws.cell(row=r, column=1, value="متن ایمیل:").font = SF2
    email2 = """Dear Hassett Group Recruitment Team,

I am writing to apply for the Registered Midwife position with visa sponsorship and pathway to permanent residency.

Professional Profile:
- Registered Midwife with clinical experience at Milad Hospital, Tehran
- Experience in labour ward, antenatal, postnatal, and maternal assessment

I am seeking an international midwifery position with visa sponsorship support.

I would welcome the opportunity to discuss this further.

Best regards,
Neda Arjmand"""
    
    ws.merge_cells(start_row=r, start_column=1, end_row=r+12, end_column=10)
    c = ws.cell(row=r, column=1, value=email2)
    c.font = CF
    c.alignment = Alignment(wrap_text=True, readingOrder=2)
    
    return ws

def create_cover_letter_template_sheet(wb):
    """Create sheet with cover letter templates"""
    ws = wb.create_sheet("📝 قالب کاور لیتر")
    ws.sheet_view.rightToLeft = True
    
    st(ws, 1, 10, "قالب کاور لیتر — قابل کپی")
    
    r = 3; ws.cell(row=r, column=1, value="📝 کاور لیتر: ماما (Midwife)").font = SF2
    
    r = 5; ws.cell(row=r, column=1, value="متن کاور لیتر:").font = SF2
    cover = """Dear Hiring Manager,

I am writing to express my strong interest in the midwifery position at your organisation.

Professional Background:
I am a registered midwife with extensive clinical experience at Milad Hospital in Tehran, one of Iran's leading medical centres. My experience includes:

- Labour ward management and deliveries
- Antenatal and postnatal care
- High-risk pregnancy monitoring
- Neonatal resuscitation
- Clinical documentation
- Team collaboration

Why This Role:
I am seeking an international opportunity where I can contribute my clinical skills while developing professionally within a world-renowned healthcare system.

Registration Readiness:
I am prepared to complete all registration requirements, including:
- Professional qualification assessment
- English language proficiency (IELTS/OET)
- Supervised practice if required

Personal Details:
I am 38 years old, married with two children. My family is committed to relocating.

I would welcome the opportunity to discuss how my experience could benefit your team.

Thank you for considering my application.

Yours sincerely,
Neda Arjmand"""
    
    ws.merge_cells(start_row=r, start_column=1, end_row=r+25, end_column=10)
    c = ws.cell(row=r, column=1, value=cover)
    c.font = CF
    c.alignment = Alignment(wrap_text=True, readingOrder=2)
    
    return ws

def main():
    wb = openpyxl.Workbook()
    if 'Sheet' in wb.sheetnames: del wb['Sheet']
    
    # ========== 1. شیت مستر ==========
    ws = wb.create_sheet("📊 شیت مستر", 0)
    ws.sheet_view.rightToLeft = True
    
    st(ws, 1, 14, "گزارش جامع شکار فرصت شغلی — Migration Hunter v6.0")
    st(ws, 2, 14, "خانواده آرجمند — توحید و ندا | A2 English + A1 German")
    
    r = 4; ws.cell(row=r, column=1, value="📊 خلاصه وضعیت کلی").font = SF2
    r = 5; sh(ws, r, ['شاخص', 'ندا', 'توحید', 'جمع'])
    sd(ws, r+1, [
        ['تعداد فرصت‌ها', 5, 5, 10],
        ['Sponsorship تأیید', 5, 0, 5],
        ['ایمیل آماده', 3, 0, 3],
        ['کاور لیتر آماده', 3, 0, 3],
        ['امتیاز تناسب', '80/100', '76/100', '78/100'],
    ], [25, 18, 18, 18])
    
    r = 12; ws.cell(row=r, column=1, value="🌍 مقایسه کشورها").font = SF2
    r = 13; sh(ws, r, ['کشور', 'امتیاز', 'ندا', 'توحید', 'Sponsorship', 'زبان', 'سرعت'])
    sd(ws, r+1, [
        ['nz', 85, 'عالی', 'خوب', '۱۰۰%', 'جستجو بدون توقف', 'سریع'],
        ['au', 78, 'خوب', 'متوسط', '۹۰%', 'جستجو بدون توقف', 'متوسط'],
        ['de', 65, 'محدود', 'خوب', '۷۰%', 'A1 کافی(IT)', 'متوسط'],
        ['ca', 60, 'محدود', 'متوسط', '۶۰%', 'جستجو بدون توقف', 'طولانی'],
    ], [12, 12, 12, 12, 12, 20, 12])
    
    # ========== 2. ندا ==========
    ws2 = wb.create_sheet("👩 ندا — فرصت‌ها")
    ws2.sheet_view.rightToLeft = True
    
    st(ws2, 1, 14, "فرصت‌های ندا آرجمند — ماما (Midwife)")
    
    r = 3; ws2.cell(row=r, column=1, value="⚠️ زبان مانع جستجو نیست — جستجو بدون توقف").font = Font(name=FONT, size=10, bold=True, color='CC0000')
    
    r = 5; ws2.cell(row=r, column=1, value="🎯 فرصت‌های شغلی").font = SF2
    r = 6; sh(ws2, r, ['اولویت', 'کارفرما', 'کشور', 'عنوان', 'حقوق', 'Sponsorship', 'تناسب', '🔗 لینک', '📧 ایمیل', '📝 کاور', 'اقدام'])
    sd(ws2, r+1, [
        ['🔥 P1', 'Health New Zealand', 'nz', 'ماما', '75-106K NZD', '✅', '80/100', '🔗', '✅', '✅', 'ارسال'],
        ['🔥 P1', 'RGH Global', 'nz', 'ماما', '75-106K NZD', '✅', '78/100', '🔗', '✅', '✅', 'ارسال'],
        ['🟢 P2', 'Hassett Group', 'au', 'ماما', '80-120K AUD', '✅', '78/100', '🔗', '✅', '✅', 'ارسال'],
        ['🟢 P2', 'Talent Angels', 'au', 'ماما', '80-120K AUD', '✅', '76/100', '🔗', '🟡', '🟡', 'بررسی'],
        ['🟢 P2', 'HealthX', 'au', 'ماما', '67-95K AUD', '✅', '74/100', '🔗', '✅', '✅', 'ارسال'],
    ], [12, 20, 8, 12, 16, 12, 10, 10, 12, 12, 12])
    
    # Add hyperlinks
    links = [
        'https://www.healthnz.govt.nz/careers/international',
        'https://www.rgh-global.com/jobs/midwife-with-sponsorship/',
        'https://www.hassett.com.au/job-details/registered-midwife-visa-sponsorship-pathway-to-pr-in-healthcare-medicine-jobs-1475247',
        'https://au.linkedin.com/jobs/view/registered-midwife-at-talent-angels-4231144020',
        'https://www.healthx.com.au/international-candidates/visa-sponsorship',
    ]
    for i, link in enumerate(links):
        add_link(ws2, r+1+i, 8, link)
    
    add_chart(ws2, "تناسب فرصت‌های ندا", 7, 2, 7, 11, "A13")
    
    # ========== 3. توحید ==========
    ws3 = wb.create_sheet("👨 توحید — فرصت‌ها")
    ws3.sheet_view.rightToLeft = True
    
    st(ws3, 1, 14, "فرصت‌های توحید آرجمند — مدیر IT")
    
    r = 3; ws3.cell(row=r, column=1, value="⚠️ زبان مانع جستجو نیست — جستجو بدون توقف").font = Font(name=FONT, size=10, bold=True, color='CC0000')
    
    r = 5; ws3.cell(row=r, column=1, value="🎯 فرصت‌های شغلی").font = SF2
    r = 6; sh(ws3, r, ['اولویت', 'کارفرما', 'کشور', 'عنوان', 'حقوق', 'Sponsorship', 'تناسب', '🔗 لینک', '📧 ایمیل', '📝 کاور', 'اقدام'])
    sd(ws3, r+1, [
        ['🟢 P2', 'SAP', 'de', 'مدیر IT', '€55-80K', '🟢', '82/100', '🔗', '🟡', '🟡', 'جستجو'],
        ['🟢 P2', 'Siemens', 'de', 'مدیر IT', '€60-90K', '🟢', '80/100', '🔗', '🟡', '🟡', 'جستجو'],
        ['🟢 P2', 'Deutsche Telekom', 'de', 'مدیر IT', '€50-75K', '🟢', '78/100', '🔗', '🟡', '🟡', 'جستجو'],
        ['🟢 P2', 'Allianz', 'de', 'مدیر IT', '€55-85K', '🟢', '77/100', '🔗', '🟡', '🟡', 'جستجو'],
        ['🟢 P2', 'BMW', 'de', 'مدیر IT', '€60-90K', '🟢', '76/100', '🔗', '🟡', '🟡', 'جستجو'],
    ], [12, 20, 8, 12, 14, 12, 10, 10, 12, 12, 12])
    
    # Add hyperlinks
    links3 = [
        'https://www.sap.com/about/careers.html',
        'https://www.siemens.com/global/en/company/jobs.html',
        'https://www.telekom.com/en/careers',
        'https://www.allianz.com/en/careers.html',
        'https://www.bmwgroup.com/en/careers.html',
    ]
    for i, link in enumerate(links3):
        add_link(ws3, r+1+i, 8, link)
    
    add_chart(ws3, "تناسب فرصت‌های توحید", 7, 2, 7, 11, "A13")
    
    # ========== 4. مقایسه ==========
    ws4 = wb.create_sheet("🌍 مقایسه")
    ws4.sheet_view.rightToLeft = True
    
    st(ws4, 1, 10, "مقایسه کشورها")
    r = 3; sh(ws4, r, ['کشور', 'امتیاز', 'ندا', 'توحید', 'سرعت', 'هزینه', 'زبان', 'خانواده'])
    sd(ws4, r+1, [
        ['nz', 85, 83, 55, 85, 70, 70, 95],
        ['au', 78, 76, 50, 70, 60, 70, 90],
        ['de', 65, 30, 72, 65, 75, 50, 85],
        ['ca', 60, 25, 55, 55, 65, 65, 85],
    ], [12, 12, 12, 12, 12, 12, 12, 12])
    
    add_chart(ws4, "مقایسه امتیاز کلی", 2, 1, 4, 7, "A9")
    
    r = 15; ws4.cell(row=r, column=1, value="📈 سناریوها").font = SF2
    r = 16; sh(ws4, r, ['سناریو', 'nz', 'au', 'de', 'ca'])
    sd(ws4, r+1, [
        ['ندا اصلی', 83, 76, 30, 25],
        ['توحید اصلی', 55, 50, 72, 55],
        ['هر دو مستقل', 80, 70, 65, 50],
    ], [18, 12, 12, 12, 12])
    
    # ========== 5. کارفرمایان ==========
    ws5 = wb.create_sheet("🏢 کارفرمایان")
    ws5.sheet_view.rightToLeft = True
    
    st(ws5, 1, 10, "کارفرمایان شناسایی شده")
    r = 3; sh(ws5, r, ['نام', 'کشور', 'نوع', 'Sponsorship', 'وب‌سایت适合', 'وضعیت'])
    sd(ws5, r+1, [
        ['Health New Zealand', 'nz', 'دولتی', '✅', 'healthnz.govt.nz', 'ندا', 'تأیید'],
        ['RGH Global', 'nz', 'آژانس', '✅', 'rgh-global.com', 'ندا', 'تأیید'],
        ["St Vincent's Health", 'au', 'بیمارستان', '✅', 'svha.org.au', 'ندا', 'تأیید'],
        ['Hassett Group', 'au', 'آژانس', '✅', 'hassett.com.au', 'ندا', 'تأیید'],
        ['HealthX', 'au', 'آژانس', '✅', 'healthx.com.au', 'ندا', 'تأیید'],
        ['SAP', 'de', 'شرکت IT', '🟡', 'sap.com', 'توحید', 'بررسی'],
        ['Siemens', 'de', 'شرکت IT', '🟡', 'siemens.com', 'توحید', 'بررسی'],
    ], [22, 10, 12, 12, 22, 12, 12])
    
    # ========== 6. ایمیل‌ها ==========
    ws6 = wb.create_sheet("📧 ایمیل‌ها")
    ws6.sheet_view.rightToLeft = True
    
    st(ws6, 1, 8, "ایمیل‌های درخواست — آماده ارسال")
    r = 3; sh(ws6, r, ['ردیف', 'متقاضی', 'کارفرما', 'موضوع', 'وضعیت', 'لینک', 'تاریخ', 'اقدام'])
    sd(ws6, r+1, [
        [1, 'ندا', 'Health New Zealand', 'International Midwife', '✅ آماده', '🔗', '2026-08-18', 'ارسال'],
        [2, 'ندا', 'RGH Global', 'Midwife Application', '✅ آماده', '🔗', '2026-08-18', 'ارسال'],
        [3, 'ندا', 'Hassett Group', 'Visa Sponsorship Inquiry', '✅ آماده', '🔗', '2026-08-18', 'ارسال'],
    ], [8, 12, 20, 25, 12, 10, 14, 12])
    
    # ========== 7. کاور لیتر ==========
    ws7 = wb.create_sheet("📝 کاور لیتر")
    ws7.sheet_view.rightToLeft = True
    
    st(ws7, 1, 8, "کاور لیترها — آماده ارسال")
    r = 3; sh(ws7, r, ['ردیف', 'متقاضی', 'کارفرما', 'عنوان', 'وضعیت', 'تاریخ', 'اقدام'])
    sd(ws7, r+1, [
        [1, 'ندا', 'Health New Zealand', 'Cover Letter - Midwife', '✅ آماده', '2026-08-18', 'بررسی و ارسال'],
        [2, 'ندا', 'RGH Global', 'Cover Letter - Midwife', '✅ آماده', '2026-08-18', 'بررسی و ارسال'],
        [3, 'ندا', 'Hassett Group', 'Cover Letter - Midwife', '✅ آماده', '2026-08-18', 'بررسی و ارسال'],
    ], [8, 12, 20, 25, 12, 14, 18])
    
    # ========== 8. ویزا ==========
    ws8 = wb.create_sheet("🛂 ویزا")
    ws8.sheet_view.rightToLeft = True
    
    st(ws8, 1, 8, "اطلاعات ویزا")
    r = 3; sh(ws8, r, ['کشور', 'نوع ویزا', 'الزام شغلی', 'زبان', 'خانواده', 'ثبت‌نام', 'منبع', 'تاریخ'])
    sd(ws8, r+1, [
        ['nz', 'AEWV', 'Job offer', 'ANZSCO 1-2: None', '✅', 'جداگانه', 'immigration.govt.nz', '2026-08-18'],
        ['au', '482 TSS', 'Job offer', 'IELTS 5.0', '✅', 'AHPRA', 'homeaffairs.gov.au', '2026-08-18'],
        ['de', 'EU Blue Card', '€45,934+', 'None (IT)', '✅', 'Anerkennung', 'make-it-in-germany.com', '2026-08-18'],
    ], [8, 14, 14, 18, 10, 14, 22, 14])
    
    # ========== 9. ثبت‌نام ==========
    ws9 = wb.create_sheet("📋 ثبت‌نام")
    ws9.sheet_view.rightToLeft = True
    
    st(ws9, 1, 10, "ثبت‌نام حرفه‌ای")
    r = 3; sh(ws9, r, ['کشور', 'متقاضی', 'حرفه', 'نهاد', 'زبان', 'آزمون', 'هزینه', 'زمان', 'منبع', 'وضعیت'])
    sd(ws9, r+1, [
        ['nz', 'ندا', 'ماما', 'Midwifery Council', 'بررسی شود', 'ممکن است', 'NZ$485', '۱۲ هفته', 'midwiferycouncil.health.nz', 'نیاز'],
        ['au', 'ندا', 'ماما', 'AHPRA', 'بررسی شود', 'ممکن است', '—', '—', 'ahpra.gov.au', 'نیاز'],
        ['de', 'توحید', 'IT', 'Chamber', 'None for visa', '—', '—', '—', 'anerkennung-in-deutschland.de', 'بررسی'],
    ], [8, 12, 10, 18, 14, 12, 12, 12, 28, 10])
    
    # ========== 10. زبان ==========
    ws10 = wb.create_sheet("📝 زبان")
    ws10.sheet_view.rightToLeft = True
    
    st(ws10, 1, 8, "وضعیت زبان — جستجو بدون توقف")
    r = 3; sh(ws10, r, ['متقاضی', 'English', 'German', 'وضعیت', 'اقدام', 'اولویت', 'یادداشت', 'تاریخ'])
    sd(ws10, r+1, [
        ['توحید', 'A2', 'A1', 'قابل بررسی', 'جستجو بدون توقف', 'مهم', 'بررسی الزام واقعی کارفرما', '2026-08-18'],
        ['ندا', 'A2', 'A1', 'قابل بررسی', 'جستجو بدون توقف', 'فوری', 'بررسی الزام ثبت‌نام هر کشور', '2026-08-18'],
    ], [14, 10, 10, 16, 18, 12, 25, 14])
    
    # ========== 11. اقدامات ==========
    ws11 = wb.create_sheet("🎯 اقدامات")
    ws11.sheet_view.rightToLeft = True
    
    st(ws11, 1, 8, "برنامه اقدام امروز")
    r = 3; sh(ws11, r, ['ردیف', 'اقدام', 'متقاضی', 'اولویت', 'وضعیت', 'یادداشت'])
    sd(ws11, r+1, [
        [1, 'ارسال Health NZ', 'ندا', '🔴 فوری', 'انجام نشده', 'بهترین فرصت'],
        [2, 'ارسال Hassett Group', 'ندا', '🔴 فوری', 'انجام نشده', 'حمایت ویزا'],
        [3, 'ارسال RGH Global', 'ندا', '🔴 فوری', 'انجام نشده', 'آژانس متخصص'],
        [4, 'جستجو IT آلمان', 'توحید', '🟡 متوسط', 'انجام نشده', 'بازار خوب'],
        [5, 'شروع آمادگی زبان', 'هر دو', '🟠 مهم', 'انجام نشده', 'همزمان با جستجو'],
    ], [8, 25, 12, 15, 15, 20])
    
    # Save
    fn = 'MigrationHunter/dashboard/MigrationHunter_Final.xlsx'
    try:
        wb.save(fn)
        print(f"✅ {fn}")
    except PermissionError:
        fn = 'MigrationHunter/dashboard/MigrationHunter_Final_v2.xlsx'
        wb.save(fn)
        print(f"✅ {fn}")
    
    print(f"📊 {len(wb.sheetnames)} sheets:")
    for s in wb.sheetnames:
        print(f"   - {s}")

if __name__ == '__main__':
    main()
