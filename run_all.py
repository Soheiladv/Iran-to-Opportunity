#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Migration Hunter — اجرای کامل تمام آپدیت‌ها
آرشیو + داشبورد v2 + Saskatchewan + حافظه + گزارش
"""
import os
import json
import shutil
from datetime import datetime
from pathlib import Path

BASE_DIR = Path(__file__).parent
DASHBOARD_DIR = BASE_DIR / "dashboard"
ARCHIVE_DIR = BASE_DIR / "archive"
MEMORY_DIR = BASE_DIR / "memory"
OUTPUT_DIR = BASE_DIR / "output"
PROFILES_DIR = BASE_DIR / "profiles"
NOW = datetime.now()
DATE_STR = NOW.strftime("%Y-%m-%d %H:%M")

# ════════════════════════════════════════════════════════════
# ۱. آرشیو خودکار Excel قدیمی
# ════════════════════════════════════════════════════════════

def archive_old_excels():
    ARCHIVE_DIR.mkdir(exist_ok=True)
    archived = 0
    for f in DASHBOARD_DIR.glob("*.xlsx"):
        age_hours = (NOW.timestamp() - f.stat().st_mtime) / 3600
        if age_hours > 1:
            dest = ARCHIVE_DIR / f.name
            if not dest.exists():
                shutil.move(str(f), str(dest))
                archived += 1
                print(f"  📦 {f.name}")
    print(f"📦 {archived} فایل آرشیو شد")

# ════════════════════════════════════════════════════════════
# ۲. بروزرسانی حافظه‌ها
# ════════════════════════════════════════════════════════════

def update_memory():
    MEMORY_DIR.mkdir(exist_ok=True)
    
    # SOURCE_BANK
    src = MEMORY_DIR / "SOURCE_BANK.md"
    src.write_text(f"""# SOURCE_BANK — بانک منابع
آخرین بروزرسانی: {DATE_STR}

---

| ID | نام | نوع | کشور | امتیاز | وضعیت |
|----|-----|-----|------|--------|-------|
| SRC-001 | Health New Zealand | Government | NZ | 95 | ✅ |
| SRC-002 | RGH Global | Recruiter | NZ | 85 | ✅ |
| SRC-003 | Working In Health NZ | Recruiter | NZ | 88 | ✅ |
| SRC-004 | Holalemania GmbH | Recruiter | DE | 85 | ✅ |
| SRC-005 | TalentOrange | Recruiter | DE | 82 | ✅ |
| SRC-006 | Saskatchewan HA | Government | CA | 90 | ✅ |
| SRC-007 | Hays Healthcare | Recruiter | AU | 82 | ✅ |
| SRC-008 | ANMF | Association | AU | 88 | ✅ |
| SRC-009 | CAM | Association | CA | 88 | ✅ |
| SRC-010 | Job Bank Canada | Government | CA | 90 | ✅ |
| SRC-011 | SEEK Australia | Job Board | AU | 85 | ✅ |
| SRC-012 | Arbeitnow | Job Board | DE | 80 | ✅ |
| SRC-013 | LinkedIn | LinkedIn | Global | 80 | ✅ |

---
 آخرین بروزرسانی: {DATE_STR}
""", encoding='utf-8')
    print("  ✅ SOURCE_BANK")
    
    # EMPLOYER_BANK
    emp = MEMORY_DIR / "EMPLOYER_BANK.md"
    emp.write_text(f"""# EMPLOYER_BANK — بانک کارفرمایان
آخرین بروزرسانی: {DATE_STR}

---

| نام | کشور | نوع | حمایت | امتیاز | ایمیل |
|-----|------|-----|--------|--------|-------|
| Health New Zealand | NZ | Government | ✅ | 85 | careers@healthnz.govt.nz |
| RGH Global | NZ | Recruiter | ✅ | 79 | info@rgh-global.com |
| Working In Health NZ | NZ | Recruiter | ✅ | 79 | — |
| Holalemania GmbH | DE | Recruiter | ✅ | 82 | info@holalemania.de |
| TalentOrange | DE | Recruiter | ✅ | 80 | — |
| Saskatchewan HA | CA | Government | ⚠️ | 78 | SHAInternational@... |
| Hays Healthcare | AU | Recruiter | ⚠️ | 72 | — |

---
""", encoding='utf-8')
    print("  ✅ EMPLOYER_BANK")
    
    # JOB_BANK
    job = MEMORY_DIR / "JOB_BANK.md"
    job.write_text(f"""# JOB_BANK — بانک مشاغل
آخرین بروزرسانی: {DATE_STR}

---

| # | متقاضی | کارفرما | کشور | شغل | امتیاز | وضعیت |
|---|--------|---------|------|------|--------|-------|
| 1 | NEDA | Health New Zealand | NZ | Midwife | 85 | 🟢 |
| 2 | NEDA | RGH Global | NZ | Midwife | 79 | 🟢 |
| 3 | NEDA | Working In Health NZ | NZ | Midwife | 79 | 🟢 |
| 4 | NEDA | Holalemania | DE | Midwife | 82 | 🟢 |
| 5 | NEDA | Saskatchewan HA | CA | Midwife | 78 | 🟡 |
| 6 | NEDA | Hays Healthcare | AU | Midwife | 72 | 🟡 |
| 7 | NEDA | TalentOrange | DE | Healthcare | 80 | 🟢 |
| 8 | TOHID | Arbeitnow | DE | IT Manager | 65 | 🟡 |

---
""", encoding='utf-8')
    print("  ✅ JOB_BANK")
    
    # LINKEDIN_DB
    li_db = MEMORY_DIR / "LINKEDIN_DB.json"
    li_data = {
        "linkedins": [
            {"name": "Neda Arjmand", "url": "https://www.linkedin.com/in/neda-arjmand", "profession": "Midwife", "applicant": "NEDA"},
            {"name": "Tohid Arjmand", "url": "https://www.linkedin.com/in/tohid-arjmand", "profession": "IT Operations Manager", "applicant": "TOHID"}
        ]
    }
    with open(li_db, 'w', encoding='utf-8') as f:
        json.dump(li_data, f, ensure_ascii=False, indent=2)
    print("  ✅ LINKEDIN_DB")

# ════════════════════════════════════════════════════════════
# ۳. ساخت داشبورد v2
# ════════════════════════════════════════════════════════════

def build_main_dashboard():
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    wb = Workbook()
    FA = Font(name='B Mitra', size=11)
    FA_T = Font(name='B Mitra', size=14, bold=True)
    HD = Font(name='B Mitra', size=11, bold=True, color='FFFFFF')
    GF = PatternFill(start_color='92D050', end_color='92D050', fill_type='solid')
    YF = PatternFill(start_color='FFD966', end_color='FFD966', fill_type='solid')
    OF = PatternFill(start_color='F4B084', end_color='F4B084', fill_type='solid')
    RF = PatternFill(start_color='FF6B6B', end_color='FF6B6B', fill_type='solid')
    HF = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    GR = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
    BF = PatternFill(start_color='9BC2E6', end_color='9BC2E6', fill_type='solid')
    BD = Border(left=Side('thin'), right=Side('thin'), top=Side('thin'), bottom=Side('thin'))

    def sh(ws, r, c):
        cell = ws.cell(row=r, column=c)
        cell.font = HD; cell.fill = HF
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = BD

    def sc(ws, r, c, v=None, f=FA, fill=None):
        cell = ws.cell(row=r, column=c)
        if v is not None: cell.value = v
        cell.font = f
        cell.alignment = Alignment(vertical='center', wrap_text=True, horizontal='right')
        cell.border = BD
        if fill: cell.fill = fill
        return cell

    ALL_JOBS = [
        {"id":1,"app":"NEDA","emp":"Health New Zealand","ctry":"nz","ctry_f":"نیوزیلند","title":"Midwife","url":"healthnz.govt.nz","email":"careers@healthnz.govt.nz","phone":None,"form":"healthnz.govt.nz/careers/international","sponsor":"Confirmed","intl":True,"free":False,"reloc":False,"lang_train":False,"url_ok":True,"has_email":True,"has_form":True,"has_phone":False,"fresh":True,"green":True,"lang_req":"IELTS/OET","visa":"AEWV + Green List","notes":"Green List"},
        {"id":2,"app":"NEDA","emp":"RGH Global","ctry":"nz","ctry_f":"نیوزیلند","title":"Midwife","url":"rgh-global.com","email":"info@rgh-global.com","phone":None,"form":None,"sponsor":"Confirmed","intl":True,"free":False,"reloc":True,"lang_train":False,"url_ok":True,"has_email":True,"has_form":False,"has_phone":False,"fresh":True,"green":False,"lang_req":"IELTS/OET","visa":"AEWV","notes":"متخصص ماما"},
        {"id":3,"app":"NEDA","emp":"Working In Health NZ","ctry":"nz","ctry_f":"نیوزیلند","title":"Midwife","url":"workingin-health.co.nz","email":None,"phone":None,"form":"workingin-health.co.nz/midwifery-jobs","sponsor":"Confirmed","intl":True,"free":True,"reloc":True,"lang_train":False,"url_ok":True,"has_email":False,"has_form":True,"has_phone":False,"fresh":True,"green":True,"lang_req":"Registration","visa":"AEWV + Green List","notes":"رایگان"},
        {"id":4,"app":"NEDA","emp":"Holalemania GmbH","ctry":"de","ctry_f":"آلمان","title":"Geburtshelfer/in","url":"holalemania.de","email":"info@holalemania.de","phone":"+49-40-41 49 65 05","form":"holalemania.de/en/application","sponsor":"Confirmed","intl":True,"free":True,"reloc":True,"lang_train":True,"url_ok":True,"has_email":True,"has_form":True,"has_phone":True,"fresh":True,"green":False,"lang_req":"German B1-B2","visa":"Work Visa + Recognition","notes":"آموزش زبان"},
        {"id":5,"app":"NEDA","emp":"TalentOrange","ctry":"de","ctry_f":"آلمان","title":"Healthcare","url":"talentorange.com","email":None,"phone":None,"form":"talentorange.com/en/for-candidates","sponsor":"Confirmed","intl":True,"free":True,"reloc":True,"lang_train":True,"url_ok":True,"has_email":False,"has_form":True,"has_phone":False,"fresh":True,"green":False,"lang_req":"German B2","visa":"Work Visa","notes":"بورسیه زبان"},
        {"id":6,"app":"NEDA","emp":"Saskatchewan HA","ctry":"ca","ctry_f":"کانادا","title":"Midwife","url":"saskhealthauthority.ca","email":"SHAInternational@...","phone":None,"form":"saskhealthauthority.ca/.../internationally-trained","sponsor":"Likely","intl":True,"free":False,"reloc":True,"lang_train":False,"url_ok":True,"has_email":True,"has_form":True,"has_phone":False,"fresh":True,"green":False,"lang_req":"CLB 7","visa":"Provincial Nominee","notes":"EOI system"},
        {"id":7,"app":"NEDA","emp":"Hays Healthcare","ctry":"au","ctry_f":"استرالیا","title":"Midwife","url":"hays.com.au","email":None,"phone":None,"form":"hays.com.au/register-your-cv","sponsor":"Possible","intl":True,"free":True,"reloc":False,"lang_train":False,"url_ok":True,"has_email":False,"has_form":True,"has_phone":False,"fresh":True,"green":False,"lang_req":"IELTS 7","visa":"482/189/190","notes":"آژانس بزرگ"},
        {"id":8,"app":"TOHID","emp":"Arbeitnow","ctry":"de","ctry_f":"آلمان","title":"IT Manager","url":"arbeitnow.com","email":None,"phone":None,"form":None,"sponsor":"Likely","intl":True,"free":True,"reloc":False,"lang_train":False,"url_ok":True,"has_email":False,"has_form":False,"has_phone":False,"fresh":True,"green":False,"lang_req":"English","visa":"Blue Card","notes":"IT"},
    ]

    def score(j):
        s = 0
        if j["sponsor"]=="Confirmed": s+=20
        elif j["sponsor"]=="Likely": s+=12
        elif j["sponsor"]=="Possible": s+=6
        if j["intl"]: s+=15
        if j["free"]: s+=5
        if j["reloc"]: s+=5
        if j["lang_train"]: s+=5
        if j["url_ok"]: s+=10
        elif j["url_ok"]==False: s-=5
        if j["has_email"]: s+=8
        if j["has_form"]: s+=7
        if j["has_phone"]: s+=5
        if j["fresh"]: s+=10
        return min(s,100)

    # ── شیت ۱: داشبورد ──
    ws1 = wb.active
    ws1.title = "Dashboard | داشبورد"
    ws1.sheet_view.rightToLeft = True
    ws1.merge_cells('A1:G1')
    ws1.cell(1, 1, f"📊 داشبورد جامع — {DATE_STR}").font = FA_T

    h = ["شاخص | Metric", "مقدار | Value", "توضیح | Notes"]
    for i, x in enumerate(h, 1): ws1.cell(3, i, x)
    sh(ws1, 3, 3)

    scores = [score(j) for j in ALL_JOBS]
    data = [
        ("کل آگهی‌ها", len(ALL_JOBS), ""),
        ("امتیاز ۸۰+", sum(1 for s in scores if s>=80), "بالاترین اولویت"),
        ("امتیاز ۶۰-۷۹", sum(1 for s in scores if 60<=s<80), "متوسط"),
        ("زیر ۶۰", sum(1 for s in scores if s<60), "نیاز بررسی"),
        ("کشورها", 4, "nz,de,ca,au"),
        ("ایمیل‌ها", sum(1 for j in ALL_JOBS if j["has_email"]), ""),
        ("فرم‌ها", sum(1 for j in ALL_JOBS if j["has_form"]), ""),
        ("تلفن‌ها", sum(1 for j in ALL_JOBS if j["has_phone"]), ""),
        ("لینکدین", 2, "ندا + توحید"),
        ("نقاط ضعف", 20, "۱۰ ندا + ۱۰ توحید"),
    ]
    for r, (l, v, n) in enumerate(data, 4):
        sc(ws1, r, 1, l); sc(ws1, r, 2, v); sc(ws1, r, 3, n)
    ws1.column_dimensions['A'].width = 30
    ws1.column_dimensions['B'].width = 15
    ws1.column_dimensions['C'].width = 25

    # ── شیت ۲: ۱۰ برتر ──
    ws2 = wb.create_sheet("Top 10 | ۱۰ برتر")
    ws2.sheet_view.rightToLeft = True
    ws2.merge_cells('A1:I1')
    ws2.cell(1, 1, "🏆 ۱۰ آگهی برتر").font = FA_T
    h2 = ["#","کارفرما","کشور","شغل","امتیاز","لینک","ایمیل","فرم","وضعیت"]
    for i, x in enumerate(h2, 1): ws2.cell(3, i, x)
    sh(ws2, 3, 9)
    scored_jobs = sorted([(j,score(j)) for j in ALL_JOBS], key=lambda x:x[1], reverse=True)[:10]
    for r, (j, s) in enumerate(scored_jobs, 4):
        sc(ws2,r,1,r-3); sc(ws2,r,2,j["emp"]); sc(ws2,r,3,j["ctry_f"]); sc(ws2,r,4,j["title"])
        c=sc(ws2,r,5,s); c.fill=GF if s>=80 else YF if s>=60 else OF
        sc(ws2,r,6,j["url"]); sc(ws2,r,7,j.get("email") or "—")
        sc(ws2,r,8,j.get("form") or "—")
        sc(ws2,r,9,"🟢" if j["url_ok"] else "🔴")
    for col in 'ABCDEFGHI': ws2.column_dimensions[col].width = 24

    # ── شیت ۳: ایمیل و لینک ──
    ws3 = wb.create_sheet("Emails & Links | ایمیل و لینک")
    ws3.sheet_view.rightToLeft = True
    ws3.merge_cells('A1:H1')
    ws3.cell(1, 1, "📧 ایمیل و لینک آگهی‌ها").font = FA_T
    h3 = ["#","کارفرما","کشور","ایمیل","تلفن","لینک","فرم","وضعیت"]
    for i, x in enumerate(h3, 1): ws3.cell(3, i, x)
    sh(ws3, 3, 8)
    for r, j in enumerate(ALL_JOBS, 4):
        sc(ws3,r,1,r-3); sc(ws3,r,2,j["emp"]); sc(ws3,r,3,j["ctry_f"])
        ec=sc(ws3,r,4,j.get("email") or "—")
        if j.get("email"): ec.fill=GR
        sc(ws3,r,5,j.get("phone") or "—")
        uc=sc(ws3,r,6,j["url"])
        if j["url_ok"]: uc.fill=GR
        fc=sc(ws3,r,7,j.get("form") or "—")
        if j.get("form"): fc.fill=GR
        sc(ws3,r,8,"✅" if j["url_ok"] else "❌")
    for col in 'ABCDEFGH': ws3.column_dimensions[col].width = 26

    # ── شیت ۴: ندا ──
    ws4 = wb.create_sheet("Neda | ندا")
    ws4.sheet_view.rightToLeft = True
    ws4.merge_cells('A1:H1')
    ws4.cell(1, 1, "👩 فرصت‌های مامایی ندا").font = FA_T
    h4 = ["#","کارفرما","کشور","شغل","حمایت","امتیاز","زبان","ویزا"]
    for i, x in enumerate(h4, 1): ws4.cell(3, i, x)
    sh(ws4, 3, 8)
    neda = [(j,score(j)) for j in ALL_JOBS if j["app"]=="NEDA"]
    for r,(j,s) in enumerate(neda,4):
        sc(ws4,r,1,r-3); sc(ws4,r,2,j["emp"]); sc(ws4,r,3,j["ctry_f"])
        sc(ws4,r,4,j["title"]); sc(ws4,r,5,j["sponsor"])
        c=sc(ws4,r,6,s); c.fill=GF if s>=80 else YF if s>=60 else OF
        sc(ws4,r,7,j["lang_req"]); sc(ws4,r,8,j["visa"])
    for col in 'ABCDEFGH': ws4.column_dimensions[col].width = 24

    # ── شیت ۵: توحید ──
    ws5 = wb.create_sheet("Tohid | توحید")
    ws5.sheet_view.rightToLeft = True
    ws5.merge_cells('A1:H1')
    ws5.cell(1, 1, "👨 فرصت‌های IT توحید").font = FA_T
    for i, x in enumerate(h4, 1): ws5.cell(3, i, x)
    sh(ws5, 3, 8)
    tohid = [(j,score(j)) for j in ALL_JOBS if j["app"]=="TOHID"]
    for r,(j,s) in enumerate(tohid,4):
        sc(ws5,r,1,r-3); sc(ws5,r,2,j["emp"]); sc(ws5,r,3,j["ctry_f"])
        sc(ws5,r,4,j["title"]); sc(ws5,r,5,j["sponsor"])
        c=sc(ws5,r,6,s); c.fill=GF if s>=80 else YF if s>=60 else OF
        sc(ws5,r,7,j["lang_req"]); sc(ws5,r,8,j["visa"])
    for col in 'ABCDEFGH': ws5.column_dimensions[col].width = 24

    # ── شیت ۶: مقایسه کشورها ──
    ws6 = wb.create_sheet("Countries | کشورها")
    ws6.sheet_view.rightToLeft = True
    ws6.merge_cells('A1:E1')
    ws6.cell(1, 1, "🌍 مقایسه کشورها").font = FA_T
    h6 = ["کشور","تعداد آگهی","بهترین امتیاز","زبان","ویزا"]
    for i, x in enumerate(h6, 1): ws6.cell(3, i, x)
    sh(ws6, 3, 5)
    cs = {}
    for j in ALL_JOBS:
        c = j["ctry"]
        if c not in cs: cs[c]={"name":j["ctry_f"],"n":0,"best":0}
        cs[c]["n"]+=1; cs[c]["best"]=max(cs[c]["best"],score(j))
    ci={"nz":("English","AEWV/Green List"),"de":("German B1-B2","Blue Card"),"ca":("CLB 7","PNP"),"au":("IELTS 7","482/189")}
    for r,(c,st) in enumerate(cs.items(),4):
        info=ci.get(c,("—","—"))
        sc(ws6,r,1,f"{st['name']} ({c.upper()})"); sc(ws6,r,2,st["n"])
        sc2=sc(ws6,r,3,st["best"]); sc2.fill=GF if st["best"]>=80 else YF
        sc(ws6,r,4,info[0]); sc(ws6,r,5,info[1])
    for col in 'ABCDE': ws6.column_dimensions[col].width = 28

    # ── شیت ۷: نقاط ضعف ──
    ws7 = wb.create_sheet("Weaknesses | نقاط ضعف")
    ws7.sheet_view.rightToLeft = True
    ws7.merge_cells('A1:E1')
    ws7.cell(1, 1, "⚠️ نقاط ضعف لینکدین").font = FA_T
    h7 = ["متقاضی","دسته","مشکل","راه‌حل","اولویت"]
    for i, x in enumerate(h7, 1): ws7.cell(3, i, x)
    sh(ws7, 3, 5)
    weak = [
        ("ندا","عکس","عدم عکس حرفه‌ای","عکس پرتره ساده","🔴 فوری"),
        ("ندا","Headline","عنوان ساده","Registered Midwife | International","🔴 فوری"),
        ("ندا","About","ندارد","۳-۵ پاراگراف","🔴 فوری"),
        ("ندا","Languages","فقط فارسی","English A2 + German A1","🔴 فوری"),
        ("ندا","Skills","کمتر از ۵","۱۰-۱۵ مهارت","🟠 مهم"),
        ("توحید","عکس","عدم عکس حرفه‌ای","عکس پرتره","🔴 فوری"),
        ("توحید","Headline","غیراستاندارد","IT Operations Manager","🔴 فوری"),
        ("توحید","About","فنی و خشک","خلاصه حرفه‌ای","🔴 فوری"),
        ("توحید","Recommendations","ندارد","درخواست از مدیران","🔴 فوری"),
        ("توحید","Open to Work","فعال نیست","فعال کردن badge","🔴 فوری"),
    ]
    for r,(a,cat,iss,fix,pri) in enumerate(weak,4):
        sc(ws7,r,1,a); sc(ws7,r,2,cat); sc(ws7,r,3,iss); sc(ws7,r,4,fix); sc(ws7,r,5,pri)
    for col in 'ABCDE': ws7.column_dimensions[col].width = 28

    # ── شیت ۸: اقدامات ──
    ws8 = wb.create_sheet("Actions | اقدامات")
    ws8.sheet_view.rightToLeft = True
    ws8.merge_cells('A1:D1')
    ws8.cell(1, 1, "🎯 ۱۰ اقدام برتر").font = FA_T
    h8 = ["#","اقدام","کشور","اولویت"]
    for i, x in enumerate(h8, 1): ws8.cell(3, i, x)
    sh(ws8, 3, 4)
    acts = [
        (1,"ارسال CV Saskatchewan HA","🇨🇦 کانادا","🔴 فوری"),
        (2,"ثبت‌نام Hays Healthcare","🇦🇺 استرالیا","🔴 فوری"),
        (3,"ایمیل Holalemania + فرم","🇩🇪 آلمان","🔴 فوری"),
        (4,"پیگیری Working In Health NZ","🇳🇿 نیوزیلند","🟠 مهم"),
        (5,"اصلاح لینکدین ندا","🌍","🟠 مهم"),
        (6,"اصلاح لینکدین توحید","🌍","🟠 مهم"),
        (7,"شروع آمادگی OET","🌍","🟠 مهم"),
        (8,"جستجوی IT اروپا","🇩🇪🇦🇹","🟡 متوسط"),
        (9,"ثبت‌نام TalentOrange","🇩🇪","🟡 متوسط"),
        (10,"بررسی پاسخ ایمیل‌ها","🌍","🟡 متوسط"),
    ]
    for r,(n,a,c,p) in enumerate(acts,4):
        sc(ws8,r,1,n); sc(ws8,r,2,a); sc(ws8,r,3,c); sc(ws8,r,4,p)
    for col in 'ABCD': ws8.column_dimensions[col].width = 35

    # ذخیره
    DASHBOARD_DIR.mkdir(exist_ok=True)
    fn = f"MigrationHunter_v2_{NOW.strftime('%Y%m%d_%H%M')}.xlsx"
    fp = DASHBOARD_DIR / fn
    wb.save(fp)
    print(f"  📊 داشبورد اصلی: {fn} ({len(wb.sheetnames)} شیت)")

# ════════════════════════════════════════════════════════════
# ۴. ساخت Excel Saskatchewan
# ════════════════════════════════════════════════════════════

def build_saskatchewan_excel():
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    wb = Workbook()
    FA = Font(name='B Mitra', size=11)
    FA_T = Font(name='B Mitra', size=14, bold=True)
    HD = Font(name='B Mitra', size=11, bold=True, color='FFFFFF')
    GF = PatternFill(start_color='92D050', end_color='92D050', fill_type='solid')
    YF = PatternFill(start_color='FFD966', end_color='FFD966', fill_type='solid')
    HF = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    GR = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
    BF = PatternFill(start_color='9BC2E6', end_color='9BC2E6', fill_type='solid')
    BD = Border(left=Side('thin'), right=Side('thin'), top=Side('thin'), bottom=Side('thin'))

    def sh(ws, r, c):
        cell = ws.cell(row=r, column=c)
        cell.font = HD; cell.fill = HF
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = BD

    def sc(ws, r, c, v=None, f=FA, fill=None):
        cell = ws.cell(row=r, column=c)
        if v is not None: cell.value = v
        cell.font = f
        cell.alignment = Alignment(vertical='center', wrap_text=True, horizontal='right')
        cell.border = BD
        if fill: cell.fill = fill
        return cell

    # شیت ۱
    ws1 = wb.active
    ws1.title = "Saskatchewan | خلاصه"
    ws1.sheet_view.rightToLeft = True
    ws1.merge_cells('A1:D1')
    ws1.cell(1, 1, f"🇨🇦 Saskatchewan Health Authority — {DATE_STR}").font = FA_T
    h1 = ["فیلد","مقدار","توضیح","وضعیت"]
    for i, x in enumerate(h1, 1): ws1.cell(3, i, x)
    sh(ws1, 3, 4)
    info = [
        ("نام","Saskatchewan Health Authority","سازمان دولتی","✅"),
        ("کشور","🇨🇦 کانادا","استان Saskatchewan","✅"),
        ("ایمیل","SHAInternational@...","استخدام بین‌المللی","✅"),
        ("ایمیل عمومی","careers@saskhealthauthority.ca","مشاغل","✅"),
        ("لینک","saskhealthauthority.ca/careers","صفحه شغل","✅"),
        ("EOI","Government of Saskatchewan","Expression of Interest","✅"),
        ("ثبت‌نام ماما","College of Midwives","نیاز به ثبت‌نام","⚠️"),
        ("ثبت‌نام IT","نیاز ندارد","IT free","✅"),
        ("مسکن","$289,600","ارزان‌ترین","✅"),
        ("آفتاب","۲,۰۰۰+ ساعت","آفتابی‌ترین","✅"),
    ]
    for r,(f,v,n,s) in enumerate(info,4):
        sc(ws1,r,1,f); c=sc(ws1,r,2,v); c.fill=GR if "✅" in s else YF; sc(ws1,r,3,n); sc(ws1,r,4,s)
    for col in 'ABCD': ws1.column_dimensions[col].width = 30

    # شیت ۲
    ws2 = wb.create_sheet("Neda — ایمیل")
    ws2.sheet_view.rightToLeft = True
    ws2.merge_cells('A1:B1')
    ws2.cell(1, 1, "👩 ایمیل ندا").font = FA_T
    en = [("گیرنده","SHAInternational@..."),("موضوع","International Midwife — EOI"),("معرفی","ماما ۱۲+ سال تجربه"),("تخصص","آنتناتال، لیبر، پست‌ناتال"),("زبان","English A2"),("دلیل","مسکن ارزان، آفتاب"),("درخواست","راهنمایی EOI"),("پیوست","CV")]
    for r,(f,v) in enumerate(en,3):
        sc(ws2,r,1,f).fill=BF; sc(ws2,r,2,v)
    ws2.column_dimensions['A'].width = 20; ws2.column_dimensions['B'].width = 45

    # شیت ۳
    ws3 = wb.create_sheet("Tohid — ایمیل")
    ws3.sheet_view.rightToLeft = True
    ws3.merge_cells('A1:B1')
    ws3.cell(1, 1, "👨 ایمیل توحید").font = FA_T
    et = [("گیرنده","SHAInternational@..."),("موضوع","IT Operations Manager"),("معرفی","۱۹+ سال تجربه IT"),("تخصص","VMware, Cisco, Windows Server"),("ارتباط","IT بهداشت High Availability"),("زبان","English A2"),("دلیل","بخش IT رو به رشد"),("پیوست","CV")]
    for r,(f,v) in enumerate(et,3):
        sc(ws3,r,1,f).fill=BF; sc(ws3,r,2,v)
    ws3.column_dimensions['A'].width = 20; ws3.column_dimensions['B'].width = 45

    # شیت ۴
    ws4 = wb.create_sheet("ایمیل‌ها | Emails")
    ws4.sheet_view.rightToLeft = True
    ws4.merge_cells('A1:D1')
    ws4.cell(1, 1, "📧 ایمیل‌های مفید").font = FA_T
    h4 = ["سازمان","ایمیل","حوزه","وضعیت"]
    for i, x in enumerate(h4, 1): ws4.cell(3, i, x)
    sh(ws4, 3, 4)
    ems = [("SHA International","SHAInternational@...","استخدام بین‌المللی","🟢"),("SHA Careers","careers@...","مشاغل","🟢"),("HHR","HHR@health.gov.sk.ca","منابع انسانی","🟢"),("eHealth Sask","careers@ehealthsask.ca","IT بهداشت","🟢")]
    for r,(o,e,f,s) in enumerate(ems,4):
        sc(ws4,r,1,o); c=sc(ws4,r,2,e); c.fill=GR; sc(ws4,r,3,f); sc(ws4,r,4,s)
    for col in 'ABCD': ws4.column_dimensions[col].width = 30

    # شیت ۵
    ws5 = wb.create_sheet("ثبت‌نام | Registration")
    ws5.sheet_view.rightToLeft = True
    ws5.merge_cells('A1:C1')
    ws5.cell(1, 1, "📋 فرآیند ثبت‌نام").font = FA_T
    h5 = ["مرحله","توضیح","وضعیت"]
    for i, x in enumerate(h5, 1): ws5.cell(3, i, x)
    sh(ws5, 3, 3)
    stps = [("۱. EOI","ثبت‌نام Expression of Interest","🔵"),("۲. مدارک","ارزیابی مدارک","🔵"),("۳. ثبت‌نام","College of Midwives","🔵"),("۴. زبان","IELTS/OET","🔵"),("۵. Job Offer","پیشنهاد شغل","🔵"),("۶. LMIA","Labour Market Assessment","🔵"),("۷. ویزا","Work Permit","🔵")]
    for r,(s,d,st) in enumerate(stps,4):
        sc(ws5,r,1,s).fill=BF; sc(ws5,r,2,d); sc(ws5,r,3,st)
    for col in 'ABC': ws5.column_dimensions[col].width = 35

    # شیت ۶
    ws6 = wb.create_sheet("مقایسه | Comparison")
    ws6.sheet_view.rightToLeft = True
    ws6.merge_cells('A1:E1')
    ws6.cell(1, 1, "📈 مقایسه استان‌ها").font = FA_T
    h6 = ["استان","مسکن","آفتاب","استخدام بین‌المللی","امتیاز"]
    for i, x in enumerate(h6, 1): ws6.cell(3, i, x)
    sh(ws6, 3, 5)
    provs = [("Saskatchewan","$289,600","۲,۰۰۰+","✅ فعال",90),("Manitoba","$350,000","خوب","✅",80),("Alberta","$450,000","خوب","✅",75),("Ontario","$1,052,920","متوسط","✅",70),("BC","$1,089,600","متوسط","✅",65)]
    for r,(p,h,s,i,sc2) in enumerate(provs,4):
        sc(ws6,r,1,p); c=sc(ws6,r,2,h); c.fill=GR if "$289" in h else YF; sc(ws6,r,3,s); sc(ws6,r,4,i)
        c2=sc(ws6,r,5,sc2); c2.fill=GF if sc2>=80 else YF
    for col in 'ABCDE': ws6.column_dimensions[col].width = 25

    DASHBOARD_DIR.mkdir(exist_ok=True)
    fn = f"Saskatchewan_{NOW.strftime('%Y%m%d_%H%M')}.xlsx"
    fp = DASHBOARD_DIR / fn
    wb.save(fp)
    print(f"  🇨🇦 Saskatchewan: {fn} ({len(wb.sheetnames)} شیت)")

# ════════════════════════════════════════════════════════════
# ۵. گزارش فارسی
# ════════════════════════════════════════════════════════════

def build_reports():
    OUTPUT_DIR.mkdir(exist_ok=True)
    
    # DAILY_ACTIONS
    da = OUTPUT_DIR / "DAILY_ACTIONS.md"
    da.write_text(f"""# 🎯 ۱۰ اقدام برتر — {DATE_STR}

| # | اقدام | کشور | اولویت |
|---|-------|------|--------|
| ۱ | ارسال CV Saskatchewan HA | 🇨🇦 | 🔴 فوری |
| ۲ | ثبت‌نام Hays Healthcare | 🇦🇺 | 🔴 فوری |
| ۳ | ایمیل Holalemania + فرم | 🇩🇪 | 🔴 فوری |
| ۴ | پیگیری Working In Health NZ | 🇳🇿 | 🟠 مهم |
| ۵ | اصلاح لینکدین ندا | 🌍 | 🟠 مهم |
| ۶ | اصلاح لینکدین توحید | 🌍 | 🟠 مهم |
| ۷ | شروع آمادگی OET | 🌍 | 🟠 مهم |
| ۸ | جستجوی IT اروپا | 🇩🇪🇦🇹 | 🟡 متوسط |
| ۹ | ثبت‌نام TalentOrange | 🇩🇪 | 🟡 متوسط |
| ۱۰ | بررسی پاسخ ایمیل‌ها | 🌍 | 🟡 متوسط |

---
آخرین بروزرسانی: {DATE_STR}
""", encoding='utf-8')
    print("  📝 DAILY_ACTIONS")

# ════════════════════════════════════════════════════════════
# MAIN
# ════════════════════════════════════════════════════════════

if __name__ == "__main__":
    print("=" * 60)
    print("🔄 اجرای کامل تمام آپدیت‌ها")
    print(f"📅 {DATE_STR}")
    print("=" * 60)
    
    print("\n📦 ۱. آرشیو فایل‌های قدیمی...")
    archive_old_excels()
    
    print("\n📝 ۲. بروزرسانی حافظه‌ها...")
    update_memory()
    
    print("\n📊 ۳. ساخت داشبورد اصلی...")
    build_main_dashboard()
    
    print("\n🇨🇦 ۴. ساخت Excel Saskatchewan...")
    build_saskatchewan_excel()
    
    print("\n📝 ۵. تولید گزارش‌ها...")
    build_reports()
    
    print("\n" + "=" * 60)
    print("✅ تمام آپدیت‌ها تکمیل شد!")
    print("=" * 60)
