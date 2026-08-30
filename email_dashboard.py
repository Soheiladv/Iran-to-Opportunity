#!/usr/bin/env python3
"""
MigrationHunter — Email Analysis Excel — نسخه بازسازی شده
جداول مرتب + عنوان ایمیل + جزئیات شغل + RTL + B Mitra
"""
import os, sys, json, io
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import PieChart, Reference
from config_loader import get_applicant_label, get_all_applicant_labels

# Fix Windows console encoding for emoji support
if sys.platform == 'win32':
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8', errors='replace')
    os.environ['PYTHONIOENCODING'] = 'utf-8'

BASE = os.path.dirname(os.path.abspath(__file__))
MEM = os.path.join(BASE, "memory")
DASH = os.path.join(BASE, "dashboard")

FONT_FA = "B Mitra"
FONT_EN = "Times New Roman"
NOW = datetime.now()
DATE_STR = NOW.strftime("%Y-%m-%d %H:%M")
FILE_DATE = NOW.strftime("%Y%m%d_%H%M")

# ─── رنگ‌ها ───
C = {
    "dark": "1B4F72", "med": "2E86C1", "light": "D6EAF8",
    "green": "27AE60", "lgreen": "D5F5E3",
    "yellow": "F39C12", "lyellow": "FEF9E7",
    "red": "E74C3C", "lred": "FADBD8",
    "purple": "8E44AD", "lpurple": "E8DAEF",
    "gray": "95A5A6", "lgray": "F2F3F4",
    "white": "FFFFFF", "dark2": "2C3E50",
    "orange": "E67E22", "lorange": "FDEBD0",
}

thin = Border(left=Side("thin"), right=Side("thin"),
              top=Side("thin"), bottom=Side("thin"))

# ─── توابع کمکی ───
def rtl(ws):
    ws.sheet_view.rightToLeft = True

def fa(sz=10, bold=False, color="000000"):
    return Font(name=FONT_FA, size=sz, bold=bold, color=color)

def en(sz=10, bold=False, color="000000"):
    return Font(name=FONT_EN, size=sz, bold=bold, color=color)

def fl(c):
    return PatternFill(start_color=c, end_color=c, fill_type="solid")

def ac(h="right", v="center", wrap=True):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def ctr():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def wc(ws, r, c, val, font=None, bg=None, align=None):
    cell = ws.cell(row=r, column=c, value=val)
    if font: cell.font = font
    if bg: cell.fill = fl(bg)
    cell.alignment = align or ac()
    cell.border = thin
    return cell

def hdr(ws, row, cols, bg=C["dark"]):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = fa(sz=9, bold=True, color=C["white"])
        cell.fill = fl(bg)
        cell.alignment = ctr()
        cell.border = thin

def aw(ws, col, w):
    ws.column_dimensions[get_column_letter(col)].width = w

def freeze(ws, cell="A2"):
    ws.freeze_panes = cell

# ─── دسته‌بندی ایمیل‌ها ───
CAT = {
    "interview":      {"fa": "🗣️ مصاحبه",        "color": C["green"],  "bg": C["lgreen"]},
    "offer":          {"fa": "🎉 پیشنهاد کار",    "color": C["purple"], "bg": C["lpurple"]},
    "rejection":      {"fa": "❌ رد شده",          "color": C["red"],    "bg": C["lred"]},
    "follow_up":      {"fa": "⏰ پیگیری",         "color": C["yellow"], "bg": C["lyellow"]},
    "inquiry":        {"fa": "💬 استعلام",         "color": C["med"],    "bg": C["light"]},
    "acknowledgment": {"fa": "📩 تأیید دریافت",   "color": C["gray"],   "bg": C["lgray"]},
    "spam":           {"fa": "🗑️ اسپم",           "color": C["gray"],   "bg": C["lgray"]},
    "unknown":        {"fa": "❓ نامشخص",          "color": C["dark2"],  "bg": C["lgray"]},
}

# ─── بارگذاری داده ───
def load_data():
    fp = os.path.join(MEM, "EMAIL_ANALYSIS.json")
    if not os.path.exists(fp):
        print("❌ EMAIL_ANALYSIS.json پیدا نشد — ابتدا python email_analyzer.py")
        return None
    with open(fp, "r", encoding="utf-8") as f:
        return json.load(f)

# ─── شیت ۱: داشبورد ───
def sheet_dashboard(wb, data):
    ws = wb.create_sheet("داشبورد")
    rtl(ws)

    emails = data.get("emails", [])
    total = data.get("total_emails", 0)
    job_rel = data.get("job_related", 0)
    per_acc = data.get("per_account", [])

    # عنوان
    wc(ws, 1, 1, f"تحلیل ایمیل شغلی — {job_rel} ایمیل از {total} کل — {DATE_STR}",
       font=fa(sz=14, bold=True, color=C["dark"]))
    ws.merge_cells("A1:F1")

    # KPI
    r = 3
    wc(ws, r, 1, "کل ایمیل‌ها", font=fa(sz=8, bold=True, color=C["white"]), bg=C["dark"], align=ctr())
    wc(ws, r, 2, total, font=fa(sz=16, bold=True), align=ctr())
    wc(ws, r, 3, "مرتبط با کار", font=fa(sz=8, bold=True, color=C["white"]), bg=C["green"], align=ctr())
    pct = round(job_rel / total * 100) if total else 0
    wc(ws, r, 4, f"{job_rel} ({pct}%)", font=fa(sz=16, bold=True, color=C["green"]), align=ctr())

    # دسته‌بندی
    r = 5
    wc(ws, r, 1, "دسته‌بندی", font=fa(sz=9, bold=True, color=C["white"]), bg=C["dark"], align=ctr())
    wc(ws, r, 2, "تعداد", font=fa(sz=9, bold=True, color=C["white"]), bg=C["dark"], align=ctr())
    wc(ws, r, 3, "درصد", font=fa(sz=9, bold=True, color=C["white"]), bg=C["dark"], align=ctr())
    r += 1

    by_cat = {}
    for e in emails:
        cat = e.get("category", "unknown")
        by_cat[cat] = by_cat.get(cat, 0) + 1

    for ck in ["interview", "offer", "rejection", "follow_up", "inquiry"]:
        info = CAT.get(ck, CAT["unknown"])
        cnt = by_cat.get(ck, 0)
        p = round(cnt / job_rel * 100) if job_rel else 0
        wc(ws, r, 1, info["fa"], font=fa(sz=9, bold=True), bg=info["bg"], align=ctr())
        wc(ws, r, 2, cnt, font=fa(sz=11, bold=True), align=ctr())
        wc(ws, r, 3, f"{p}%", font=fa(sz=11, bold=True, color=info["color"]), align=ctr())
        r += 1

    # حساب‌ها
    r += 1
    wc(ws, r, 1, "حساب ایمیل", font=fa(sz=9, bold=True, color=C["white"]), bg=C["dark"], align=ctr())
    wc(ws, r, 2, "شخص", font=fa(sz=9, bold=True, color=C["white"]), bg=C["dark"], align=ctr())
    wc(ws, r, 3, "LinkedIn", font=fa(sz=9, bold=True, color=C["white"]), bg=C["dark"], align=ctr())
    wc(ws, r, 4, "ایمیل شغلی", font=fa(sz=9, bold=True, color=C["white"]), bg=C["dark"], align=ctr())
    wc(ws, r, 5, "درصد", font=fa(sz=9, bold=True, color=C["white"]), bg=C["dark"], align=ctr())
    r += 1

    for pa in per_acc:
        p = round(pa.get("job_related", 0) / job_rel * 100) if job_rel else 0
        person = pa.get("person", "?")
        app = get_applicant_label(person) if person in ['NEDA', 'TOHID'] else person
        wc(ws, r, 1, pa.get("email", ""), font=en(sz=9))
        wc(ws, r, 2, app, font=fa(sz=9), align=ctr())
        wc(ws, r, 3, pa.get("linkedin", ""), font=en(sz=8, color="0563C1"))
        wc(ws, r, 4, pa.get("job_related", 0), font=fa(sz=10, bold=True), align=ctr())
        wc(ws, r, 5, f"{p}%", font=fa(sz=10, bold=True, color=C["med"]), align=ctr())
        r += 1

    for i, w in enumerate([35, 12, 40, 12, 10]):
        aw(ws, i + 1, w)
    freeze(ws)

# ─── شیت ۲: تمام ایمیل‌ها ───
def sheet_all_emails(wb, data):
    ws = wb.create_sheet("ایمیل‌های شغلی")
    rtl(ws)

    emails = data.get("emails", [])
    job_rel = data.get("job_related", 0)

    wc(ws, 1, 1, f"ایمیل‌های شغلی — {job_rel} ایمیل", font=fa(sz=14, bold=True, color=C["dark"]))
    ws.merge_cells("A1:I1")

    # هدرها
    headers = ["#", "تاریخ", "فرستنده", "عنوان ایمیل", "دسته", "متقاضی", "کارفرما", "کشور", "اقدام"]
    r = 3
    hdr(ws, r, len(headers))
    for i, h in enumerate(headers):
        ws.cell(row=r, column=i+1).value = h
    r += 1

    # داده‌ها — فقط شغلی‌ها
    job_emails = [e for e in emails if e.get("category", "unknown") not in ["spam", "unknown"]]
    
    for idx, e in enumerate(sorted(job_emails, key=lambda x: x.get("date", ""), reverse=True), 1):
        cat = e.get("category", "unknown")
        info = CAT.get(cat, CAT["unknown"])
        app = e.get("applicant", "?")
        app_label = get_applicant_label(app) if app else "?"
        
        employer = e.get("employer", "")
        country = ""
        if employer:
            # استخراج کشور از نام کارفرما
            emp_countries = {
                "Health New Zealand": "🇳🇿 نیوزیلند", "RGH Global": "🇳🇿 نیوزیلند",
                "Saskatchewan HA": "🇨🇦 کانادا", "Alberta Health Services": "🇨🇦 کانادا",
                "Hays Healthcare": "🇦🇺 استرالیا", "Holalemania": "🇩🇪 آلمان",
                "TalentOrange": "🇩🇪 آلمان", "Kate Cowhig": "🇮🇪 ایرلند",
                "CPL Healthcare": "🇮🇪 ایرلند", "Work in Austria": "🇦🇹 اتریش",
                "IND Netherlands": "🇳🇱 هلند", "Finncare": "🇫🇮 فنلاند",
            }
            country = emp_countries.get(employer, "")

        # اقدام
        action_map = {
            "interview": "پاسخ + حضور",
            "offer": "بررسی + تصمیم",
            "rejection": "بایگانی",
            "follow_up": "پیگیری",
            "inquiry": "بررسی",
            "acknowledgment": "تأیید",
        }
        action = action_map.get(cat, "")

        vals = [
            idx,
            e.get("date", "")[:10],
            e.get("from", "").split("<")[0].strip().strip('"')[:30],
            e.get("subject", "")[:60],
            info["fa"],
            app_label,
            employer,
            country,
            action,
        ]

        for ci, v in enumerate(vals):
            bg = info["bg"] if ci == 4 else None
            wc(ws, r, ci + 1, v, font=fa(sz=9), bg=bg)
        r += 1

    widths = [5, 12, 25, 55, 14, 10, 20, 16, 14]
    for i, w in enumerate(widths):
        aw(ws, i + 1, w)
    freeze(ws, "A4")
    ws.auto_filter.ref = f"A3:{get_column_letter(len(headers))}{r-1}"

# ─── شیت ۳: بر اساس متقاضی ───
def sheet_by_applicant(wb, data):
    emails = data.get("emails", [])
    per_acc = data.get("per_account", [])

    for pa in per_acc:
        person = pa.get("person", "?")
        person_fa = "ندا" if person == "NEDA" else "توحید" if person == "TOHID" else person
        color = C["purple"] if person == "NEDA" else C["med"]
        linkedin = pa.get("linkedin", "")
        email_addr = pa.get("email", "")

        # فیلتر ایمیل‌های این شخص
        person_emails = [e for e in emails if e.get("applicant") == person 
                        or e.get("account_id", "").startswith(person.lower())]
        
        if not person_emails:
            person_emails = [e for e in emails]  # fallback: show all

        ws = wb.create_sheet(f"{person_fa} — ایمیل‌ها")
        rtl(ws)

        # عنوان + لینکدین
        wc(ws, 1, 1, f"{'👩' if person == 'NEDA' else '👨'} {person_fa} — ایمیل‌های شغلی",
           font=fa(sz=14, bold=True, color=color))
        ws.merge_cells("A1:H1")

        wc(ws, 2, 1, f"📧 {email_addr}", font=en(sz=9))
        wc(ws, 2, 3, f"🔗 {linkedin}", font=en(sz=9, color="0563C1"))
        ws.merge_cells("C2:H2")

        # آمار دسته‌بندی
        r = 4
        by_cat = {}
        for e in person_emails:
            cat = e.get("category", "unknown")
            by_cat[cat] = by_cat.get(cat, 0) + 1

        wc(ws, r, 1, "دسته", font=fa(sz=9, bold=True, color=C["white"]), bg=color, align=ctr())
        wc(ws, r, 2, "تعداد", font=fa(sz=9, bold=True, color=C["white"]), bg=color, align=ctr())
        wc(ws, r, 3, "درصد", font=fa(sz=9, bold=True, color=C["white"]), bg=color, align=ctr())
        r += 1

        total_job = sum(v for k, v in by_cat.items() if k not in ["spam", "unknown"])
        for ck in ["interview", "offer", "rejection", "follow_up", "inquiry"]:
            info = CAT.get(ck, CAT["unknown"])
            cnt = by_cat.get(ck, 0)
            p = round(cnt / total_job * 100) if total_job else 0
            wc(ws, r, 1, info["fa"], font=fa(sz=9, bold=True), bg=info["bg"], align=ctr())
            wc(ws, r, 2, cnt, font=fa(sz=10, bold=True), align=ctr())
            wc(ws, r, 3, f"{p}%", font=fa(sz=10, bold=True, color=info["color"]), align=ctr())
            r += 1

        # جدول ایمیل‌ها
        r += 1
        headers = ["#", "تاریخ", "فرستنده", "عنوان ایمیل", "دسته", "کارفرما", "کشور", "اقدام"]
        hdr(ws, r, len(headers), bg=color)
        for i, h in enumerate(headers):
            ws.cell(row=r, column=i+1).value = h
        r += 1

        for idx, e in enumerate(sorted(person_emails, key=lambda x: x.get("date", ""), reverse=True), 1):
            cat = e.get("category", "unknown")
            if cat in ["spam", "unknown"]:
                continue
            info = CAT.get(cat, CAT["unknown"])
            
            action_map = {
                "interview": "پاسخ + حضور", "offer": "بررسی + تصمیم",
                "rejection": "بایگانی", "follow_up": "پیگیری", "inquiry": "بررسی",
            }

            vals = [
                idx, e.get("date", "")[:10],
                e.get("from", "").split("<")[0].strip().strip('"')[:30],
                e.get("subject", "")[:60],
                info["fa"], e.get("employer", ""), "", action_map.get(cat, ""),
            ]
            for ci, v in enumerate(vals):
                bg = info["bg"] if ci == 4 else None
                wc(ws, r, ci + 1, v, font=fa(sz=9), bg=bg)
            r += 1

        widths = [5, 12, 25, 55, 14, 20, 14, 14]
        for i, w in enumerate(widths):
            aw(ws, i + 1, w)
        freeze(ws, "A6")

# ─── شیت ۴: بر اساس دسته ───
def sheet_by_category(wb, data):
    emails = data.get("emails", [])

    for ck in ["interview", "offer", "rejection", "follow_up", "inquiry"]:
        cat_emails = [e for e in emails if e.get("category") == ck]
        if not cat_emails:
            continue

        info = CAT[ck]
        ws = wb.create_sheet(info["fa"].split(" ")[-1])
        rtl(ws)

        wc(ws, 1, 1, f"{info['fa']} — {len(cat_emails)} ایمیل",
           font=fa(sz=14, bold=True, color=info["color"]))
        ws.merge_cells("A1:H1")

        headers = ["#", "تاریخ", "فرستنده", "عنوان ایمیل", "متقاضی", "کارفرما", "کشور", "اقدام"]
        r = 3
        hdr(ws, r, len(headers), bg=info["color"])
        for i, h in enumerate(headers):
            ws.cell(row=r, column=i+1).value = h
        r += 1

        action_map = {
            "interview": "پاسخ + حضور", "offer": "بررسی + تصمیم",
            "rejection": "بایگانی", "follow_up": "پیگیری", "inquiry": "بررسی",
        }

        for idx, e in enumerate(sorted(cat_emails, key=lambda x: x.get("date", ""), reverse=True), 1):
            app = e.get("applicant", "?")
            app_label = get_applicant_label(app) if app in ['NEDA', 'TOHID'] else "?"

            vals = [
                idx, e.get("date", "")[:10],
                e.get("from", "").split("<")[0].strip().strip('"')[:30],
                e.get("subject", "")[:60],
                app_label, e.get("employer", ""), "", action_map.get(ck, ""),
            ]
            for ci, v in enumerate(vals):
                wc(ws, r, ci + 1, v, font=fa(sz=9))
            r += 1

        widths = [5, 12, 25, 55, 10, 20, 14, 14]
        for i, w in enumerate(widths):
            aw(ws, i + 1, w)
        freeze(ws, "A4")

# ─── شیت ۵: آمار و نمودار ───
def sheet_stats(wb, data):
    ws = wb.create_sheet("آمار")
    rtl(ws)

    emails = data.get("emails", [])
    job_rel = data.get("job_related", 0)

    wc(ws, 1, 1, "آمار تحلیل ایمیل", font=fa(sz=14, bold=True, color=C["dark"]))
    ws.merge_cells("A1:D1")

    # فرستنده‌های پرتکرار
    r = 3
    wc(ws, r, 1, "فرستنده", font=fa(sz=9, bold=True, color=C["white"]), bg=C["dark"], align=ctr())
    wc(ws, r, 2, "تعداد", font=fa(sz=9, bold=True, color=C["white"]), bg=C["dark"], align=ctr())
    wc(ws, r, 3, "درصد", font=fa(sz=9, bold=True, color=C["white"]), bg=C["dark"], align=ctr())
    r += 1

    senders = {}
    for e in emails:
        s = e.get("from", "").split("<")[0].strip().strip('"')[:40]
        if s: senders[s] = senders.get(s, 0) + 1

    for s, cnt in sorted(senders.items(), key=lambda x: -x[1])[:15]:
        p = round(cnt / job_rel * 100) if job_rel else 0
        wc(ws, r, 1, s, font=fa(sz=9))
        wc(ws, r, 2, cnt, font=fa(sz=10, bold=True), align=ctr())
        wc(ws, r, 3, f"{p}%", font=fa(sz=10, color=C["med"]), align=ctr())
        r += 1

    # Pie chart
    try:
        pie = PieChart()
        pie.title = "دسته‌بندی ایمیل‌ها"
        pie.style = 10
        # Find the category table (rows 5-9 in dashboard)
        # For now, create a mini data table
        r += 2
        by_cat = {}
        for e in emails:
            cat = e.get("category", "unknown")
            by_cat[cat] = by_cat.get(cat, 0) + 1
        
        chart_start = r
        for ck in ["interview", "offer", "rejection", "follow_up", "inquiry"]:
            info = CAT.get(ck, CAT["unknown"])
            cnt = by_cat.get(ck, 0)
            wc(ws, r, 1, info["fa"], font=fa(sz=9))
            wc(ws, r, 2, cnt, font=fa(sz=9), align=ctr())
            r += 1

        data_ref = Reference(ws, min_col=2, min_row=chart_start - 1, max_row=r - 1)
        cats_ref = Reference(ws, min_col=1, min_row=chart_start, max_row=r - 1)
        pie.add_data(data_ref, titles_from_data=True)
        pie.set_categories(cats_ref)
        pie.width = 16
        pie.height = 10
        ws.add_chart(pie, f"D{chart_start}")
    except:
        pass

    aw(ws, 1, 35)
    aw(ws, 2, 10)
    aw(ws, 3, 10)
    aw(ws, 4, 10)
    freeze(ws)

# ─── اجرا ───
def main():
    print("=" * 50)
    print("MigrationHunter — Email Excel (بازسازی شده)")
    print("=" * 50)

    data = load_data()
    if not data:
        return

    job_rel = data.get("job_related", 0)
    print(f"\n📂 {job_rel} ایمیل شغلی")

    print("\n📊 ساخت Excel...")
    wb = Workbook()
    wb.remove(wb.active)

    print("  📝 داشبورد...")
    sheet_dashboard(wb, data)

    print("  📝 ایمیل‌های شغلی...")
    sheet_all_emails(wb, data)

    print("  📝 بر اساس متقاضی...")
    sheet_by_applicant(wb, data)

    print("  📝 بر اساس دسته...")
    sheet_by_category(wb, data)

    print("  📝 آمار...")
    sheet_stats(wb, data)

    os.makedirs(DASH, exist_ok=True)
    fn = f"Email_Analysis_{FILE_DATE}.xlsx"
    fp = os.path.join(DASH, fn)
    wb.save(fp)

    print(f"\n✅ ذخیره شد: {fn}")
    print(f"📋 شیت‌ها: {len(wb.sheetnames)}")
    for s in wb.sheetnames:
        print(f"  - {s}")

if __name__ == "__main__":
    main()
