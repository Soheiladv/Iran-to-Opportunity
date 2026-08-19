# -*- coding: utf-8 -*-
"""
Migration Hunter — Run Script
جریان اجرای شکار فرصت شغلی

 نحوه اجرا:
    python run.py

 یا:
    python run.py --country nz
    python run.py --country de
    python run.py --applicant neda
    python run.py --applicant tohid
    python run.py --full
"""

import os
import sys
import json
import hashlib
from datetime import datetime
from pathlib import Path

# ==========================================
# CONFIGURATION
# ==========================================

BASE_DIR = Path(__file__).parent
PROFILES_DIR = BASE_DIR / "profiles"
MEMORY_DIR = BASE_DIR / "memory"
INPUT_DIR = BASE_DIR / "input"
OUTPUT_DIR = BASE_DIR / "output"
DASHBOARD_DIR = BASE_DIR / "dashboard"

# Ensure directories exist
for d in [PROFILES_DIR, MEMORY_DIR, INPUT_DIR, OUTPUT_DIR, DASHBOARD_DIR]:
    d.mkdir(exist_ok=True)

# ==========================================
# MEMORY BANKS
# ==========================================

class MemoryBank:
    """خواندن و نوشتن بانک‌های حافظه"""

    def __init__(self, name):
        self.file_path = MEMORY_DIR / f"{name}.md"
        self.name = name
        self.data = {}

    def load(self):
        """خواندن اطلاعات از فایل"""
        if self.file_path.exists():
            content = self.file_path.read_text(encoding='utf-8')
            self.data = self._parse_markdown(content)
        return self.data

    def save(self):
        """ذخیره اطلاعات در فایل"""
        content = self._to_markdown()
        self.file_path.write_text(content, encoding='utf-8')

    def _parse_markdown(self, content):
        """پارس کردن فایل markdown"""
        data = {"items": [], "summary": {}}
        current_item = {}

        for line in content.split('\n'):
            line = line.strip()
            if line.startswith('### '):
                if current_item:
                    data["items"].append(current_item)
                current_item = {"name": line[4:]}
            elif '|' in line and '---' not in line:
                parts = [p.strip() for p in line.split('|') if p.strip()]
                if len(parts) >= 2:
                    current_item[parts[0]] = parts[1]

        if current_item:
            data["items"].append(current_item)

        return data

    def _to_markdown(self):
        """تبدیل داده به markdown"""
        lines = [
            f"# {self.name.upper()} — بانک اطلاعاتی",
            "",
            f"آخرین بروزرسانی: {datetime.now().strftime('%Y-%m-%d %H:%M')}",
            "",
            "---",
            ""
        ]

        for item in self.data.get("items", []):
            lines.append(f"### {item.get('name', 'Unknown')}")
            for key, value in item.items():
                if key != 'name':
                    lines.append(f"| {key} | {value} |")
            lines.append("")

        return '\n'.join(lines)

    def add_item(self, item):
        """افزودن آیتم جدید"""
        if "items" not in self.data:
            self.data["items"] = []
        self.data["items"].append(item)

    def find_item(self, **kwargs):
        """جستجوی آیتم"""
        for item in self.data.get("items", []):
            match = True
            for key, value in kwargs.items():
                if item.get(key) != value:
                    match = False
                    break
            if match:
                return item
        return None

# ==========================================
# JOB COLLECTOR
# ==========================================

class JobCollector:
    """جمع‌آوری مشاغل از منابع مختلف"""

    def __init__(self):
        self.jobs = []
        self.employers = []
        self.sources = []

    def collect_from_health_nz(self):
        """جمع‌آوری از Health New Zealand"""
        print("🔍 جمع‌آوری از Health New Zealand...")

        job = {
            "id": f"JOB-{hashlib.md5(b'healthnz').hexdigest()[:6].upper()}",
            "employer": "Health New Zealand",
            "country": "nz",
            "title": "Midwife",
            "url": "https://www.healthnz.govt.nz/careers/international",
            "salary": "75,773 - 106,739 NZD/year",
            "sponsorship": "Confirmed",
            "language_visa": "Not required (ANZSCO 1-2)",
            "language_registration": "IELTS Academic 7.0 or OET",
            "registration": "Midwifery Council NZ",
            "applicant": "NEDA",
            "path_fit_score": 82,
            "status": "NEW",
            "collected_at": datetime.now().isoformat()
        }
        self.jobs.append(job)
        return job

    def collect_from_rgh(self):
        """جمع‌آوری از RGH Global"""
        print("🔍 جمع‌آوری از RGH Global...")

        job = {
            "id": f"JOB-{hashlib.md5(b'rgh').hexdigest()[:6].upper()}",
            "employer": "RGH Global",
            "country": "nz",
            "title": "Midwife (with sponsorship)",
            "url": "https://www.rgh-global.com/jobs/midwife-with-sponsorship/",
            "salary": "75,773 - 106,739 NZD/year",
            "sponsorship": "Confirmed",
            "language_registration": "IELTS Academic 7.0 or OET",
            "registration": "Midwifery Council NZ",
            "applicant": "NEDA",
            "path_fit_score": 80,
            "status": "NEW",
            "collected_at": datetime.now().isoformat()
        }
        self.jobs.append(job)
        return job

    def collect_from_working_in_nz(self):
        """جمع‌آوری از Working In Health NZ"""
        print("🔍 جمع‌آوری از Working In Health NZ...")

        job = {
            "id": f"JOB-{hashlib.md5(b'workinginnz').hexdigest()[:6].upper()}",
            "employer": "Working In Health NZ",
            "country": "nz",
            "title": "Midwife (full recruitment service)",
            "url": "https://www.workingin-health.co.nz/midwifery-jobs/",
            "salary": "UNKNOWN",
            "sponsorship": "Confirmed",
            "language_registration": "IELTS Academic 7.0 or OET",
            "registration": "Help provided",
            "applicant": "NEDA",
            "path_fit_score": 85,
            "status": "NEW",
            "collected_at": datetime.now().isoformat()
        }
        self.jobs.append(job)
        return job

    def collect_from_holalemania(self):
        """جمع‌آوری از Holalemania"""
        print("🔍 جمع‌آوری از Holalemania...")

        job = {
            "id": f"JOB-{hashlib.md5(b'holalemania').hexdigest()[:6].upper()}",
            "employer": "Holalemania GmbH",
            "country": "de",
            "title": "Hebamme (Midwife)",
            "url": "https://holalemania.de/en/",
            "email": "info@holalemania.de",
            "phone": "+49-40-41 49 65 05",
            "salary": "UNKNOWN",
            "sponsorship": "Confirmed",
            "language": "German A1-A2 (they provide training)",
            "registration": "Anerkennung required",
            "applicant": "NEDA",
            "path_fit_score": 78,
            "status": "NEW",
            "collected_at": datetime.now().isoformat()
        }
        self.jobs.append(job)
        return job

    def collect_all(self):
        """جمع‌آوری از تمام منابع"""
        print("=" * 50)
        print("🚀 شروع جمع‌آوری مشاغل")
        print("=" * 50)

        self.collect_from_health_nz()
        self.collect_from_rgh()
        self.collect_from_working_in_nz()
        self.collect_from_holalemania()

        print(f"\n✅ {len(self.jobs)} فرصت شناسایی شد")
        return self.jobs

# ==========================================
# JOB ANALYZER
# ==========================================

class JobAnalyzer:
    """تحلیل و امتیازدهی مشاغل"""

    def __init__(self):
        self.scored_jobs = []

    def calculate_path_fit(self, job, applicant_profile):
        """محاسبه Path Fit Score"""
        score = 0

        # Professional Fit (20%)
        professional_fit = self._assess_professional_fit(job, applicant_profile)
        score += professional_fit * 0.20

        # Immigration Fit (20%)
        immigration_fit = self._assess_immigration_fit(job)
        score += immigration_fit * 0.20

        # Language Fit (15%)
        language_fit = self._assess_language_fit(job, applicant_profile)
        score += language_fit * 0.15

        # Sponsorship Fit (25%)
        sponsorship_fit = self._assess_sponsorship_fit(job)
        score += sponsorship_fit * 0.25

        # Family Fit (10%)
        family_fit = 70  # Default - needs more data
        score += family_fit * 0.10

        # Speed (10%)
        speed = self._assess_speed(job)
        score += speed * 0.10

        return round(score)

    def _assess_professional_fit(self, job, profile):
        """ارزیابی تناسب حرفه‌ای"""
        if job.get("applicant") == "NEDA" and profile.get("profession") == "Midwife":
            return 85
        elif job.get("applicant") == "TOHID" and "IT" in job.get("title", ""):
            return 80
        return 50

    def _assess_immigration_fit(self, job):
        """ارزیابی تناسب مهاجرتی"""
        country = job.get("country", "")
        if country == "nz":
            return 85
        elif country == "de":
            return 80
        elif country == "au":
            return 75
        return 60

    def _assess_language_fit(self, job, profile):
        """ارزیابی تناسب زبانی"""
        # Current levels: English A2, German A1
        if job.get("language_visa") == "Not required (ANZSCO 1-2)":
            return 90
        if "A1" in job.get("language", ""):
            return 80
        if "A2" in job.get("language", ""):
            return 70
        return 50

    def _assess_sponsorship_fit(self, job):
        """ارزیابی تناسب حمایت"""
        sponsorship = job.get("sponsorship", "")
        if sponsorship == "Confirmed":
            return 90
        elif sponsorship == "Likely":
            return 75
        elif sponsorship == "Possible":
            return 60
        return 40

    def _assess_speed(self, job):
        """ارزیابی سرعت"""
        country = job.get("country", "")
        if country == "nz":
            return 80
        elif country == "de":
            return 75
        return 65

    def analyze_jobs(self, jobs, profiles):
        """تحلیل تمام مشاغل"""
        print("\n📊 تحلیل مشاغل...")

        analyzed = []
        for job in jobs:
            applicant = job.get("applicant", "")
            profile = profiles.get(applicant, {})
            score = self.calculate_path_fit(job, profile)
            job["path_fit_score"] = score
            analyzed.append(job)
            print(f"  ✓ {job['employer']}: {score}/100")

        self.scored_jobs = sorted(analyzed, key=lambda x: x.get("path_fit_score", 0), reverse=True)
        return self.scored_jobs

# ==========================================
# REPORT GENERATOR
# ==========================================

class ReportGenerator:
    """تولید گزارش‌ها"""

    def __init__(self):
        self.reports = {}

    def generate_daily_actions(self, jobs):
        """تولید گزارش اقدامات روزانه"""
        print("\n📝 تولید گزارش اقدامات روزانه...")

        top_jobs = jobs[:5]  # Top 5

        report = [
            "# ۵ اقدام برتر امروز",
            "",
            f"تاریخ: {datetime.now().strftime('%Y-%m-%d %H:%M')}",
            "",
            "---",
            "",
            "| ردیف | متقاضی | اقدام | کشور | اولویت |",
            "|------|--------|-------|------|--------|",
        ]

        priorities = ["🔴 فوری", "🔴 فوری", "🟠 مهم", "🟠 مهم", "🟡 متوسط"]

        for i, job in enumerate(top_jobs):
            priority = priorities[i] if i < len(priorities) else "🟡 متوسط"
            report.append(
                f"| {i+1} | {job.get('applicant', '')} | "
                f"ارسال CV به {job['employer']} | "
                f"{job.get('country', '')} | {priority} |"
            )

        report.extend([
            "",
            "---",
            "",
            "## خلاصه",
            "",
            f"- کل فرصت‌ها: {len(jobs)}",
            f"- فرصت‌های آماده: {len([j for j in jobs if j.get('status') == 'NEW'])}",
            f"- فرصت‌های P1: {len([j for j in jobs if j.get('path_fit_score', 0) >= 80])}",
        ])

        content = '\n'.join(report)
        output_path = OUTPUT_DIR / "DAILY_ACTIONS.md"
        output_path.write_text(content, encoding='utf-8')
        print(f"  ✓ ذخیره شد: {output_path}")

        return content

    def generate_top_jobs(self, jobs, applicant):
        """تولید گزارش فرصت‌های برتر"""
        print(f"\n📝 تولید گزارش {applicant}...")

        filtered = [j for j in jobs if j.get("applicant") == applicant]

        report = [
            f"# فرصت‌های برتر — {applicant}",
            "",
            f"تاریخ: {datetime.now().strftime('%Y-%m-%d %H:%M')}",
            "",
            "---",
            "",
            "| ردیف | کارفرما | کشور | عنوان | تناسب | وضعیت |",
            "|------|---------|------|-------|-------|--------|",
        ]

        for i, job in enumerate(filtered[:5]):
            report.append(
                f"| {i+1} | {job['employer']} | {job.get('country', '')} | "
                f"{job.get('title', '')} | {job.get('path_fit_score', 0)}/100 | "
                f"{job.get('status', '')} |"
            )

        content = '\n'.join(report)
        filename = f"{applicant}_TOP_JOBS.md"
        output_path = OUTPUT_DIR / filename
        output_path.write_text(content, encoding='utf-8')
        print(f"  ✓ ذخیره شد: {output_path}")

        return content

# ==========================================
# EXCEL GENERATOR
# ==========================================

class ExcelGenerator:
    """تولید فایل Excel"""

    def __init__(self):
        try:
            import openpyxl
            self.openpyxl = openpyxl
        except ImportError:
            print("⚠️ openpyxl نصب نیست. اجرا کنید: pip install openpyxl")
            self.openpyxl = None

    def generate(self, jobs):
        """تولید فایل Excel"""
        if not self.openpyxl:
            print("❌ امکان تولید Excel وجود ندارد")
            return None

        print("\n📊 تولید فایل Excel...")

        wb = self.openpyxl.Workbook()

        # Styles
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

        font_b_mitra = Font(name='Arial', size=12)  # Arial as fallback
        font_bold = Font(name='Arial', size=12, bold=True)
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
        green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
        yellow_fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
        center = Alignment(horizontal='center', vertical='center', wrap_text=True)
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Sheet 1: Dashboard
        ws1 = wb.active
        ws1.title = "Dashboard"

        ws1['A1'] = "Migration Hunter Dashboard"
        ws1['A1'].font = Font(name='Arial', size=16, bold=True)
        ws1['A1'].alignment = center

        ws1['A2'] = f"Last Updated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
        ws1['A2'].alignment = center

        # Summary
        ws1['A4'] = "Summary"
        ws1['A4'].font = font_bold

        summary = [
            ("Total Opportunities", len(jobs)),
            ("Ready to Apply", len([j for j in jobs if j.get('status') == 'NEW'])),
            ("NEDA Opportunities", len([j for j in jobs if j.get('applicant') == 'NEDA'])),
            ("TOHID Opportunities", len([j for j in jobs if j.get('applicant') == 'TOHID'])),
        ]

        for i, (label, value) in enumerate(summary):
            ws1.cell(row=5+i, column=1, value=label).font = font_b_mitra
            ws1.cell(row=5+i, column=2, value=value).font = font_bold
            ws1.cell(row=5+i, column=2).alignment = center

        # Top Jobs
        ws1['A10'] = "Top Opportunities"
        ws1['A10'].font = font_bold

        headers = ["#", "Employer", "Country", "Applicant", "Title", "Score", "Status"]
        for col, h in enumerate(headers, 1):
            cell = ws1.cell(row=11, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center
            cell.border = thin_border

        for i, job in enumerate(jobs[:10]):
            row = 12 + i
            ws1.cell(row=row, column=1, value=i+1).border = thin_border
            ws1.cell(row=row, column=2, value=job.get('employer', '')).border = thin_border
            ws1.cell(row=row, column=3, value=job.get('country', '')).border = thin_border
            ws1.cell(row=row, column=4, value=job.get('applicant', '')).border = thin_border
            ws1.cell(row=row, column=5, value=job.get('title', '')).border = thin_border

            score_cell = ws1.cell(row=row, column=6, value=job.get('path_fit_score', 0))
            score_cell.alignment = center
            score_cell.border = thin_border
            if job.get('path_fit_score', 0) >= 80:
                score_cell.fill = green_fill
            elif job.get('path_fit_score', 0) >= 70:
                score_cell.fill = yellow_fill

            ws1.cell(row=row, column=7, value=job.get('status', '')).border = thin_border

        # Sheet 2: All Jobs
        ws2 = wb.create_sheet("All Jobs")

        job_headers = ["#", "Employer", "Country", "Applicant", "Title", "Salary",
                       "Sponsorship", "Language", "Registration", "Score", "Status", "URL"]
        for col, h in enumerate(job_headers, 1):
            cell = ws2.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center
            cell.border = thin_border

        for i, job in enumerate(jobs):
            row = 2 + i
            ws2.cell(row=row, column=1, value=i+1).border = thin_border
            ws2.cell(row=row, column=2, value=job.get('employer', '')).border = thin_border
            ws2.cell(row=row, column=3, value=job.get('country', '')).border = thin_border
            ws2.cell(row=row, column=4, value=job.get('applicant', '')).border = thin_border
            ws2.cell(row=row, column=5, value=job.get('title', '')).border = thin_border
            ws2.cell(row=row, column=6, value=job.get('salary', '')).border = thin_border
            ws2.cell(row=row, column=7, value=job.get('sponsorship', '')).border = thin_border
            ws2.cell(row=row, column=8, value=job.get('language', job.get('language_visa', ''))).border = thin_border
            ws2.cell(row=row, column=9, value=job.get('registration', '')).border = thin_border

            score_cell = ws2.cell(row=row, column=10, value=job.get('path_fit_score', 0))
            score_cell.alignment = center
            score_cell.border = thin_border

            ws2.cell(row=row, column=11, value=job.get('status', '')).border = thin_border

            url_cell = ws2.cell(row=row, column=12, value=job.get('url', ''))
            url_cell.border = thin_border
            if job.get('url'):
                url_cell.hyperlink = job['url']

        # Auto-fit columns (approximate)
        for col in range(1, len(job_headers) + 1):
            ws2.column_dimensions[chr(64 + col) if col <= 26 else 'A'].width = 15

        # Save
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_path = DASHBOARD_DIR / f"MigrationHunter_{timestamp}.xlsx"
        try:
            wb.save(output_path)
            print(f"  ✓ ذخیره شد: {output_path}")
        except PermissionError:
            # If file is open, save with different name
            output_path = DASHBOARD_DIR / f"MigrationHunter_{timestamp}_new.xlsx"
            wb.save(output_path)
            print(f"  ✓ ذخیره شد (نسخه جدید): {output_path}")

        return output_path

# ==========================================
# MAIN RUNNER
# ==========================================

class MigrationHunter:
    """کلاس اصلی اجرای شکار فرصت"""

    def __init__(self):
        self.collector = JobCollector()
        self.analyzer = JobAnalyzer()
        self.reporter = ReportGenerator()
        self.excel_gen = ExcelGenerator()
        self.memory_banks = {
            "source": MemoryBank("SOURCE_BANK"),
            "employer": MemoryBank("EMPLOYER_BANK"),
            "job": MemoryBank("JOB_BANK"),
            "application": MemoryBank("APPLICATION_BANK"),
            "search": MemoryBank("SEARCH_HISTORY"),
        }

    def load_memory(self):
        """بارگذاری حافظه"""
        print("📂 بارگذاری حافظه...")
        for name, bank in self.memory_banks.items():
            bank.load()
            print(f"  ✓ {name}: {len(bank.data.get('items', []))} آیتم")

    def save_memory(self):
        """ذخیره حافظه"""
        print("\n💾 ذخیره حافظه...")
        for name, bank in self.memory_banks.items():
            bank.save()
            print(f"  ✓ {name} ذخیره شد")

    def load_profiles(self):
        """بارگذاری پروفایل‌ها"""
        print("\n👤 بارگذاری پروفایل‌ها...")

        profiles = {}

        # Tohid
        tohid_path = PROFILES_DIR / "TOHID_PROFILE.md"
        if tohid_path.exists():
            profiles["TOHID"] = {
                "name": "Tohid Arjmand",
                "profession": "IT Manager",
                "age": 46,
                "english": "A2",
                "german": "A1"
            }
            print("  ✓ TOHID loaded")

        # Neda
        neda_path = PROFILES_DIR / "NEDA_PROFILE.md"
        if neda_path.exists():
            profiles["NEDA"] = {
                "name": "Neda Arjmand",
                "profession": "Midwife",
                "age": 38,
                "english": "A2",
                "german": "A1"
            }
            print("  ✓ NEDA loaded")

        return profiles

    def run(self, countries=None, applicant=None):
        """اجرای اصلی"""
        print("=" * 60)
        print("🚀 MIGRATION HUNTER — RUN")
        print(f"📅 {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        print("=" * 60)

        # Step 1: Load memory
        self.load_memory()

        # Step 2: Load profiles
        profiles = self.load_profiles()

        # Step 3: Collect jobs
        jobs = self.collector.collect_all()

        # Step 4: Analyze jobs
        analyzed_jobs = self.analyzer.analyze_jobs(jobs, profiles)

        # Step 5: Update memory
        for job in analyzed_jobs:
            self.memory_banks["job"].add_item({
                "name": job.get("id", ""),
                "employer": job.get("employer", ""),
                "country": job.get("country", ""),
                "applicant": job.get("applicant", ""),
                "title": job.get("title", ""),
                "score": job.get("path_fit_score", 0),
                "status": job.get("status", ""),
                "collected_at": job.get("collected_at", "")
            })

        # Step 6: Generate reports
        self.reporter.generate_daily_actions(analyzed_jobs)
        self.reporter.generate_top_jobs(analyzed_jobs, "NEDA")
        self.reporter.generate_top_jobs(analyzed_jobs, "TOHID")

        # Step 7: Generate Excel
        self.excel_gen.generate(analyzed_jobs)

        # Step 8: Save memory
        self.save_memory()

        # Summary
        print("\n" + "=" * 60)
        print("✅ اجرا تکمیل شد!")
        print("=" * 60)
        print(f"\n📊 خلاصه:")
        print(f"  - کل فرصت‌ها: {len(analyzed_jobs)}")
        print(f"  - NEDA: {len([j for j in analyzed_jobs if j.get('applicant') == 'NEDA'])}")
        print(f"  - TOHID: {len([j for j in analyzed_jobs if j.get('applicant') == 'TOHID'])}")
        print(f"\n📁 فایل‌های خروجی:")
        print(f"  - {OUTPUT_DIR / 'DAILY_ACTIONS.md'}")
        print(f"  - {OUTPUT_DIR / 'NEDA_TOP_JOBS.md'}")
        print(f"  - {OUTPUT_DIR / 'TOHID_TOP_JOBS.md'}")
        print(f"  - {DASHBOARD_DIR / 'MigrationHunter_Dashboard.xlsx'}")

        return analyzed_jobs


# ==========================================
# ENTRY POINT
# ==========================================

def main():
    """نقطه ورود"""
    import argparse

    parser = argparse.ArgumentParser(description='Migration Hunter — Job Hunting Engine')
    parser.add_argument('--country', type=str, help='Country to search (nz, de, au, ca)')
    parser.add_argument('--applicant', type=str, help='Applicant (neda, tohid)')
    parser.add_argument('--full', action='store_true', help='Full run')

    args = parser.parse_args()

    hunter = MigrationHunter()

    if args.country:
        print(f"🔍 جستجوی {args.country}...")
        hunter.run(countries=[args.country])
    elif args.applicant:
        print(f"🔍 جستجوی {args.applicant}...")
        hunter.run(applicant=args.applicant)
    else:
        hunter.run()


if __name__ == "__main__":
    main()
