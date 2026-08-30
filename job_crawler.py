#!/usr/bin/env python3
"""
MigrationHunter — جستجوی خودکار کاریابی
جستجو در سایت‌های مختلف + ذخیره در Excel + تشخیص واقعی/فیک آگهی
اجرای: python job_crawler.py
"""
import os, sys, json, re, io, time
from datetime import datetime
from urllib.request import urlopen, Request
from urllib.error import URLError, HTTPError
from html.parser import HTMLParser
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from config_loader import get_applicant_label, get_applicants, detect_applicant_from_text

# Fix Windows console encoding
if sys.platform == 'win32':
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8', errors='replace')
    os.environ['PYTHONIOENCODING'] = 'utf-8'

BASE = os.path.dirname(os.path.abspath(__file__))
MEM = os.path.join(BASE, "memory")
DASH = os.path.join(BASE, "dashboard")
OUT = os.path.join(BASE, "output")

NOW = datetime.now()
DATE_STR = NOW.strftime("%Y-%m-%d %H:%M")
FILE_DATE = NOW.strftime("%Y%m%d_%H%M")
OK = "✅"

# ════════════════════════════════════════════════════
#SOURCE DEFINITIONS - تعریف منابع جستجو (قابل تغییر از config)
# ════════════════════════════════════════════════════
SEARCH_SOURCES = [
    {
        "name": "Seek NZ", "country": "NZ", "type": "job_board",
        "url": "https://www.seek.co.nz/midwife-jobs",
        "search_urls": [
            "https://www.seek.co.nz/midwife-jobs",
            "https://www.seek.co.nz/it-manager-jobs",
        ],
        "trust": 85,
    },
    {
        "name": "Trade Me Jobs NZ", "country": "NZ", "type": "job_board",
        "url": "https://www.trademe.co.nz/jobs",
        "search_urls": [
            "https://www.trademe.co.nz/jobs/healthcare",
        ],
        "trust": 80,
    },
    {
        "name": "Seek AU", "country": "AU", "type": "job_board",
        "url": "https://www.seek.com.au/midwife-jobs",
        "search_urls": [
            "https://www.seek.com.au/midwife-jobs",
            "https://www.seek.com.au/it-infrastructure-jobs",
        ],
        "trust": 85,
    },
    {
        "name": "Job Bank Canada", "country": "CA", "type": "government",
        "url": "https://www.jobbank.gc.ca",
        "search_urls": [
            "https://www.jobbank.gc.ca/jobsearch/jobsearch?searchstring=midwife&locationstring=",
            "https://www.jobbank.gc.ca/jobsearch/jobsearch?searchstring=it+manager&locationstring=",
        ],
        "trust": 90,
    },
    {
        "name": "Indeed Canada", "country": "CA", "type": "job_board",
        "url": "ca.indeed.com",
        "search_urls": [
            "https://ca.indeed.com/jobs?q=midwife&l=",
            "https://ca.indeed.com/jobs?q=it+infrastructure+manager&l=",
        ],
        "trust": 75,
    },
    {
        "name": "StepStone DE", "country": "DE", "type": "job_board",
        "url": "https://www.stepstone.de",
        "search_urls": [
            "https://www.stepstone.de/jobs/it-manager",
        ],
        "trust": 80,
    },
    {
        "name": "Indeed DE", "country": "DE", "type": "job_board",
        "url": "de.indeed.com",
        "search_urls": [
            "https://de.indeed.com/jobs?q=it+manager&l=",
        ],
        "trust": 75,
    },
    {
        "name": "IrishJobs", "country": "IE", "type": "job_board",
        "url": "https://www.irishjobs.ie",
        "search_urls": ["https://www.irishjobs.ie"],
        "trust": 70,
    },
    {
        "name": "Indeed NL", "country": "NL", "type": "job_board",
        "url": "nl.indeed.com",
        "search_urls": ["https://nl.indeed.com"],
        "trust": 70,
    },
]

# ═══════════════════════════════════════════════════
# جستجوهای δυναmiک - هیچ Hardcode ندیده
# ═══════════════════════════════════════════════════
# کلیدواژه‌های کلی که در اکثر آگهی‌ها ظاهر می‌شوند
COMMON_JOB_KEYWORDS = [
    "job", "position", "vacancy", "career", "opportunity",
    "مشغل", "شغل", "لوظه", "وظیفه", " Verwendung"
]

# الگوهای Generic برای پیدا کردن لینک‌های آگهی
JOB_LINK_PATTERNS_GENERIC = [
    r'href="[^"]*job[^"]*"',       # لینک‌های حاوی job
    r'href="[^"]*position[^"]*"',  # لینک‌های حاوی position
    r'href="[^"]*vacancy[^"]*"',   # لینک‌های vacancy
    r'href="[^"]*career[^"]*"',    # لینک‌های career
]

# الگوها برای استخراج عنوان (بدون کلمات خاص)
TITLE_EXTRACT_PATTERNS = [
    r'<a[^>]*>([^<]{10,100})</a>',  # کلیک‌های داخل تگ a
    r'<h[1-6][^>]*>([^<]{10,100})</h[1-6]>',  # تیترهای headings
    r'<div[^>]*class="[^"]*title[^"]*"[^>]*>([^<]{10,100})</div>',  # دیوهای Titre
]


# ═══════════════════════════════════════════════════
# پارس HTML دینامیک
# ═══════════════════════════════════════════════════
class DynamicJobParser(HTMLParser):
    """پارس HTML که هیچ کلمه خاصی پیش‌فرض ندارد"""
    def __init__(self):
        super().__init__()
        self.possible_links = []
        self.possible_titles = []
        self._capture_text = False
        self._current_text = ""
        self._tag_stack = []

    def handle_starttag(self, tag, attrs):
        attrs_dict = dict(attrs)
        self._tag_stack.append(tag)
        
        # هر تگ possibul anchor
        if tag == "a" and attrs_dict.get("href"):
            href = attrs_dict["href"]
            text = self._current_text.strip()
            if len(text) > 8 and len(text) < 200:
                self.possible_links.append({
                    "text": text,
                    "href": href,
                    "source": "dynamic"
                })
            self._current_text = ""
        
        # هر تگ ممکن برای عنوان
        if tag in ("h1", "h2", "h3", "h4", "h5", "h6", "span", "div", "p"):
            self._capture_text = True
            self._current_text = ""

    def handle_endtag(self, tag):
        self._tag_stack.pop()
        if tag in ("h1", "h2", "h3", "h4", "h5", "h6", "span", "div", "p"):
            self._capture_text = False
            if self._current_text.strip() and len(self._current_text.strip()) > 8:
                self.possible_titles.append(self._current_text.strip())
            self._current_text = ""

    def handle_data(self, data):
        if self._capture_text:
            self._current_text += data


# ═══════════════════════════════════════════════════
# تشخیص واقعی/فیک آگهی
# ═══════════════════════════════════════════════════
def detect_job_realness(job_title, job_company, applicants_config):
    """
    تشخیص آیا آگهی واقعی است یاFake
    بر اساس تطبیق با پروفایل متقاضیان
    """
    score = 0
    max_possible = 0
    
    for app in applicants_config:
        keywords = app.get("keywords", [])
        title_lower = (job_title or "").lower()
        company_lower = (job_company or "").lower()
        
        # امتیاز از عنوان
        for kw in keywords:
            if kw.lower() in title_lower:
                score += 2
        max_possible += len(keywords) * 2
        
        # امتیاز از نام شرکت
        for kw in keywords:
            if kw.lower() in company_lower:
                score += 1
        max_possible += len(keywords) * 1
    
    # اگر هیچ matching tudi، بازرسی ساده
    if score == 0 and job_title:
        # الگوها برای Arbeitsamte pers i
        generic_words = ["manager", "director", "engineer", "developer", "analyst", "coordinator"]
        title_lower = job_title.lower()
        for gw in generic_words:
            if gw in title_lower:
                score += 1
                break
    
    #_real if score > 20% of max possible or has strong indicators
    realness_ratio = score / max_possible if max_possible > 0 else 0
    is_real = realness_ratio > 0.2 or score >= 2
    
    return {
        "is_real": is_real,
        "realness_score": int(realness_ratio * 100),
        "match_score": score,
        "max_possible": max_possible
    }


# ═══════════════════════════════════════════════════
# استخراج آگهی از HTML
# ═══════════════════════════════════════════════════
def extract_jobs_from_html(html, source_info):
    """استخراج آگهی‌ها از HTML به صورت دینامیک"""
    parser = DynamicJobParser()
    try:
        parser.feed(html)
    except Exception:
        return []
    
    raw_jobs = []
    
    # ۱. استخراج از mahdollیconnects
    for link_info in parser.possible_links:
        title = link_info["text"]
        href = link_info["href"]
        
        # تمیز کردن لینک
        if href.startswith("/"):
            # لینک نسبی - مسیر项目
            from urllib.parse import urljoin
            base = source_info.get("url", "").split("/")[0] + "//" + urlparse(url).netloc if 'urlparse' else href
            href = urljoin(source_info.get("url", ""), href)
        
        # استخراج kompani (اگر قابلی هست)
        company = ""
        
        job_data = {
            "title": title[:150] if title else "بدون عنوان",
            "company": company[:100] if company else "نامشخص",
            "source": source_info.get("name", "نامشخص"),
            "country": source_info.get("country", "؟"),
            "url": href[:300] if href else "",
            "found_at": DATE_STR,
            "raw_html_snippet": ""
        }
        raw_jobs.append(job_data)
    
    # ۲. استخراج از titles پارس شده
    for title in parser.possible_titles[:10]:  # حداکثر 10 title
        if len(title) > 8:
            company = ""
            # تلاش برای پیدا کردن kompanی近邻
            job_data = {
                "title": title[:150],
                "company": company[:100] if company else "نامشخص",
                "source": source_info.get("name", "نامشخص"),
                "country": source_info.get("country", "؟"),
                "url": source_info.get("url", "")[:300],
                "found_at": DATE_STR,
                "raw_html_snippet": ""
            }
            raw_jobs.append(job_data)
    
    # حذف doublé
    seen = set()
    unique_jobs = []
    for j in raw_jobs:
        key = (j["title"][:50], j["company"][:30], j["source"])
        if key not in seen:
            seen.add(key)
            unique_jobs.append(j)
    
    return unique_jobs


# ═══════════════════════════════════════════════════
# ساخت Excel با جزئیات کامل و پیگیری ایمیل
# ═══════════════════════════════════════════════════
def build_jobs_excel(jobs, applicants_config):
    """ساخت Excel با جزئیات کامل والأీگهی‌ها"""
    wb = Workbook()
    ws = wb.active
    ws.title = "نتایج جستجو"
    ws.sheet_view.rightToLeft = True
    
    # ستايل‌ها
    thin_border = Border(
        left=Side("thin"), right=Side("thin"),
        top=Side("thin"), bottom=Side("thin")
    )
    header_font = Font(name="B Mitra", size=9, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1B4F72", end_color="1B4F72", fill_type="solid")
    cell_font = Font(name="B Mitra", size=9)
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    # هدرها - افزودن ستون‌های پیگیری ایمیل
    headers = [
        "#", "عنوان شغل", "شرکت", "کشور", "منبع", "لینک آگهی", 
        "تاریخ افزوده", "امتیاز تطبیق", "حقيقي/فیک", "متقاضی پیشنهادی", 
        "ایمیلながيت", "وضعیت پیگیری"
    ]
    
    for i, h in enumerate(headers):
        cell = ws.cell(row=1, column=i+1, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border
    
    # داده‌ها
    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 45
    ws.column_dimensions["C"].width = 25
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 20
    ws.column_dimensions["F"].width = 45
    ws.column_dimensions["G"].width = 18
    ws.column_dimensions["H"].width = 12
    ws.column_dimensions["I"].width = 15
    ws.column_dimensions["J"].width = 30
    ws.column_dimensions["K"].width = 22
    
    # تزریق applicant info
    applicant_ids = [a["id"].upper() for a in applicants_config]
    
    row = 2
    for idx, job in enumerate(jobs, 1):
        # تشخیص genuine
        realness = detect_job_realness(job.get("title", ""), job.get("company", []), applicants_config)
        
        # تشخیص متقاضی پیشنهادی
        applicant_suggestion = "?"
        for app_id in applicant_ids:
            applicants = get_applicants()
            for a in applicants:
                if app_id == a["id"].upper():
                    #Matching based on keywords in title
                    title_lower = (job.get("title", "") or "").lower()
                    match_count = sum(1 for k in a.get("keywords", []) if k.lower() in title_lower)
                    if match_count > 0:
                        applicant_suggestion = get_applicant_label(a["id"])
                        break
        
        # وضعیت پیگیری از حافظه
        email_status = "—"
        email_file = os.path.join(MEM, "EMAIL_ANALYSIS.json")
        if os.path.exists(email_file):
            try:
                with open(email_file, "r", encoding="utf-8") as f:
                    email_data = json.load(f)
                # Check if this employer/title appears in email analysis
                for e in email_data.get("emails", []):
                    if e.get("employer") and e.get("employer").lower() in (job.get("company") or "").lower():
                        email_status = "✅已 analysed"
                        break
            except:
                pass
        
        # مقادیر سطر
        values = [
            idx,
            job.get("title", "")[:80],
            job.get("company", "")[:60],
            job.get("country", "") or "؟",
            job.get("source", "")[:30],
            job.get("url", "")[:40],
            DATE_STR,
            realness["match_score"],
            "✅ واقعی" if realness["is_real"] else "❌ فیک",
            applicant_suggestion,
            email_status,
            "—"
        ]
        
        for ci, v in enumerate(values):
            cell = ws.cell(row=row, column=ci+1, value=v)
            cell.font = cell_font
            cell.border = thin_border
            # رنگگذاری وضعیت
            if ci == 8:  # واقعیت smear
                if "✅" in str(v):
                    cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                else:
                    cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        
        row += 1
    
    # اضافه کردن ردیف Gesamme
    ws.cell(row=row, column=1, value=len(jobs)).font = header_font
    ws.cell(row=row, column=2, value=f"جمع کل آگهی‌ها").font = header_font
    
    # Auto filter
    ws.auto_filter.ref = f"A1:K{row}"
    
    return wb


# ═══════════════════════════════════════════════════
#fetch page - با HEADERS صحیح
# ═══════════════════════════════════════════════════
def fetch_page(url, timeout=15):
    """دریافت محتوای صفحه با تنظیمات مناسب"""
    try:
        req = Request(
            url,
            headers={
                "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
                "Accept": "text/html,application/xhtml+xml,application/xml",
                "Accept-Language": "en-US,en;q=0.5",
            }
        )
        with urlopen(req, timeout=timeout) as resp:
            if resp.status == 200:
                content = resp.read()
                #Decode with multiple encodings
                for enc in ["utf-8", "cp1252", "latin-1"]:
                    try:
                        return content.decode(enc, errors="ignore")
                    except:
                        continue
                return content.decode("utf-8", errors="ignore")
    except (HTTPError, URLError, OSError) as e:
        pass
    return None


# ═══════════════════════════════════════════════════
# جستجو در منابع
# ═══════════════════════════════════════════════════
def search_source(source):
    """جستجو در یک منبع -return jobs"""
    all_jobs = []
    
    for url in source.get("search_urls", [source["url"]]):
        html = fetch_page(url)
        if html:
            jobs = extract_jobs_from_html(html, source)
            all_jobs.extend(jobs)
            # Be respectful to servers
            time.sleep(0.5)
    
    return all_jobs


# ═══════════════════════════════════════════════════
# اجرا
# ═══════════════════════════════════════════════════
def main():
    print("═" * 60)
    print("MigrationHunter — جستجوی خودکار کاریابی")
    print(f"📅 {DATE_STR}")
    print("═" * 60)
    
    applicants_config = get_applicants()
    print(f"  👥 {len(applicants_config)} متقاضی پیکربندی شده")
    
    all_jobs = []
    crawl_log = []
    
    print(f"\n🔍 جستجو در {len(SEARCH_SOURCES)} منبع...")
    
    for idx, source in enumerate(SEARCH_SOURCES, 1):
        print(f"  📡 {source['name']} ({source['country']}) - در حال بررسی...")
        
        jobs = search_source(source)
        all_jobs.extend(jobs)
        
        status = "✅" if jobs else "⚠️"
        print(f"    {status} {len(jobs)} آگهی یافت شد")
    
    # Remove duplicates based on title+company+source
    seen_keys = set()
    unique_jobs = []
    for j in all_jobs:
        key = (j.get("title", "")[:50].lower(), j.get("company", "")[:30].lower(), j.get("source", "").lower())
        if key not in seen_keys:
            seen_keys.add(key)
            unique_jobs.append(j)
    
    # Sort by relevance to applicants
    unique_jobs.sort(key=lambda j: detect_job_realness(
        j.get("title", ""), j.get("company", []), applicants_config
    )["match_score"], reverse=True)
    
    # Build Excel
    print(f"\n📊 ساخت Excel با {len(unique_jobs)} آگهی...")
    wb = build_jobs_excel(unique_jobs, applicants_config)
    
    os.makedirs(DASH, exist_ok=True)
    fn = f"Job_Crawler_{FILE_DATE}.xlsx"
    fp = os.path.join(DASH, fn)
    wb.save(fp)
    
    # Save results JSON
    os.makedirs(MEM, exist_ok=True)
    with open(os.path.join(MEM, "CRAWLER_RESULTS.json"), "w", encoding="utf-8") as f:
        json.dump({"jobs": unique_jobs, "found": len(unique_jobs)}, f, ensure_ascii=False, indent=2)
    
    # Summary
    print(f"\n  📊 خلاصه")
    print(f"  ├─ مجموع 찾은: {len(all_jobs)}")
    print(f"  ├─rerasunique: {len(unique_jobs)}")
    
    # Stats per country
    countries = {}
    for j in unique_jobs:
        c = j.get("country", "؟")
        countries[c] = countries.get(c, 0) + 1
    
    for c, cnt in sorted(countries.items(), key=lambda x: -x[1])[:5]:
        print(f"  │  {c}: {cnt} آگهی")
    
    print(f"  └─ترین اختصاصی: {unique_jobs[0].get('title', '')[:50] if unique_jobs else '—'}...")
    print(f"  📁 ذخیره شده: {fn}")
    print("═" * 60)
    
    return len(unique_jobs)


if __name__ == "__main__":
    main()